# keiba_bot.py
import time
import re
import os
import math
import html
import json
import requests
import streamlit as st
import pandas as pd
from datetime import datetime

from selenium import webdriver
import statistics
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException

from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from urllib.parse import urljoin
import networkx as nx

# ==================================================
# 0. ユーティリティ
# ==================================================
_CIRCLED_NUMS = '⓪①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳'

def _to_circled(n):
    """数値を丸数字に変換（0〜20）"""
    try:
        n = int(n)
        return _CIRCLED_NUMS[n] if 0 <= n <= 20 else f'({n})'
    except (ValueError, TypeError):
        return ''
        
def parse_date(date_str):
    """日付文字列をdatetimeオブジェクトに変換する"""
    try:
        for sep in ('.', '/'):
            if sep in date_str:
                parts = date_str.split(sep)
                yy = int(parts[0])
                if yy < 100:
                    yy += 2000
                return datetime(yy, int(parts[1]), int(parts[2]))
    except Exception:
            pass
    return datetime.min

MAX_RELATIVE_DIST_DIFF = 600
LONG_DISTANCE_MIN = 1700
SPRINT_DISTANCE_MAX = 1200
# 最後の手段の相対比較: 競馬場さえ同じなら距離不問で採用する時の優先度。
# 整数の通常優先度(1〜5)より必ず下に来るよう 0.5 とし、max()で常に負ける=実質「最後の砦」。
SAME_PLACE_DISTFREE_PRIORITY = 0.5

# 相対評価で参照する過去レースの古さ上限。
# 基本は 9ヶ月以内。ただし (a) 同場同距離 もしくは
# (b) 同場・同レイアウト・距離差±200m以内 の場合は 12ヶ月まで延長して採用する。
RELATIVE_PAST_DAYS_DEFAULT = 274   # ≒ 9ヶ月
RELATIVE_PAST_DAYS_EXTENDED = 365  # 12ヶ月
RELATIVE_EXTENDED_DIST_DIFF = 200

# 隠れ馬経由(bridge)の追加制限。各レッグについて 半年(180日)以内 のみ採用。
# ただし距離 ≥1800m のレッグはレース数が少ないため例外として制限なし。
BRIDGE_RECENT_DAYS = 180
BRIDGE_OLD_OK_DIST_MIN = 1800

def _dist_to_int(dist):
    try:
        return int(str(dist).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0

def _is_relative_distance_allowed(dist, current_dist):
    """相対評価では今回距離から600m以上離れた過去レースを比較に使わない。"""
    d = _dist_to_int(dist)
    cur_d = _dist_to_int(current_dist)
    if d <= 0 or cur_d <= 0:
        return True
    if d >= LONG_DISTANCE_MIN and cur_d >= LONG_DISTANCE_MIN:
        return True
    return abs(d - cur_d) < MAX_RELATIVE_DIST_DIFF

def _rank_to_int(value, default=99):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default

def _apply_first_place_margin_bonus(raw_diff, h1_rank, h2_rank, bonus=0.2):
    """1着馬は流している可能性を見て、1着側の優位差を少し広げる。"""
    r1 = _rank_to_int(h1_rank)
    r2 = _rank_to_int(h2_rank)
    if r1 == 1 and r2 != 1:
        return raw_diff - bonus
    if r2 == 1 and r1 != 1:
        return raw_diff + bonus
    return raw_diff

def _is_long_distance(dist):
    return _dist_to_int(dist) >= LONG_DISTANCE_MIN

def _is_sprint_distance(dist):
    d = _dist_to_int(dist)
    return 0 < d <= SPRINT_DISTANCE_MAX


# ==================================================
# 南関・共通実戦速度指数（総合点へは未加算の独立検証軸）
# ==================================================
# 保存32レースの時系列比較で、上位3頭捕捉と大敗型の抑制が両立した配合。
SPEED_INDEX_TOP2_WEIGHT = 0.50
SPEED_INDEX_RECENCY_WEIGHT = 0.50
SPEED_INDEX_VOLATILITY_PENALTY = 0.20
SPEED_INDEX_AVERAGE_SCORE = 70.0
SPEED_INDEX_RECORD_SCORE = 100.0
SPEED_INDEX_MIN_SCORE = 20.0
SPEED_INDEX_MAX_SCORE = 105.0
SPEED_INDEX_DAILY_VARIANT_MIN_RACES = 3
SPEED_INDEX_DAILY_VARIANT_MIN_DISTANCES = 2
SPEED_INDEX_DAILY_VARIANT_CAP = 1.5  # 秒/1000m
SPEED_INDEX_DAILY_VARIANT_SHRINKAGE = 0.10  # 32レース時系列検証で過補正を防いだ採用率
_SPEED_META_CACHE = {"signature": None, "rows": []}
_SPEED_HISTORY_CACHE = {"signature": None, "rows": []}
_NANKAN_SPEED_PLACES = {'川崎', '船橋', '浦和', '大井'}
_JRA_SPEED_PLACES = {'JRA', '札幌', '函館', '福島', '新潟', '東京', '中山', '中京', '京都', '阪神', '小倉'}


def _winner_clock_to_seconds(value):
    """対戦表の勝時計(例: 549/1324/2163)を秒へ変換する。"""
    digits = re.sub(r'\D', '', str(value or ''))
    if len(digits) < 3:
        return None
    try:
        tenths = int(digits[-1])
        whole = int(digits[:-1])
        minutes, seconds = divmod(whole, 100)
        if seconds >= 60:
            return None
        return minutes * 60 + seconds + tenths / 10.0
    except (TypeError, ValueError):
        return None


def _speed_class_bucket(title):
    """日別馬場差をクラス差から分離するための粗いクラス帯。"""
    text = str(title or '').upper().replace(' ', '')
    for label in ('A1', 'A2', 'B1', 'B2', 'B3', 'C1', 'C2', 'C3'):
        if label in text:
            return label
    if '2歳' in text:
        return '2歳'
    if '3歳' in text:
        return '3歳'
    return 'その他'


def _load_speed_race_meta(cache_dir='cache'):
    """保存済み対戦表から日付・場・馬場・距離・勝時計・クラスを読む。"""
    try:
        paths = sorted(
            os.path.join(cache_dir, name)
            for name in os.listdir(cache_dir)
            if name.startswith('taisen_') and name.endswith('.html')
        )
    except OSError:
        return []
    signature = (
        len(paths),
        max((os.path.getmtime(path) for path in paths), default=0),
        sum((os.path.getsize(path) for path in paths), 0),
    )
    if _SPEED_META_CACHE.get('signature') == signature:
        return list(_SPEED_META_CACHE.get('rows') or [])

    pattern = re.compile(
        r"openMovieForStrobeMediaPlayback\('(\d+)'\).*?"
        r'<p class="nk23_c-table08__detail">(.*?)</p>',
        re.S,
    )
    by_race = {}
    for path in paths:
        try:
            with open(path, encoding='utf-8', errors='ignore') as fh:
                page = fh.read()
        except OSError:
            continue
        for match in pattern.finditer(page):
            race_id, body = match.groups()
            plain = re.sub(r'<br\s*/?>', '|', body)
            plain = re.sub(r'<[^>]+>', '', plain)
            parts = [re.sub(r'\s+', ' ', html.unescape(x)).strip() for x in plain.split('|')]
            if len(parts) < 3:
                continue
            date = parse_date(parts[0])
            venue_parts = parts[1].split()
            nums = re.findall(r'\d+', parts[2])
            if date == datetime.min or len(venue_parts) < 2 or len(nums) < 3:
                continue
            place, going = venue_parts[0], venue_parts[1]
            if place not in {'川崎', '船橋', '浦和', '大井'}:
                continue
            winner_time = _winner_clock_to_seconds(nums[2])
            dist = _dist_to_int(nums[1])
            if winner_time is None or dist <= 0:
                continue
            title = parts[3] if len(parts) >= 4 else ''
            by_race[race_id] = {
                'race_id': race_id,
                'date': date.strftime('%Y.%m.%d'),
                'place': place,
                'going': going,
                'dist': dist,
                'winner_time': winner_time,
                'class_bucket': _speed_class_bucket(title),
                'title': title,
            }

    rows = list(by_race.values())
    _SPEED_META_CACHE['signature'] = signature
    _SPEED_META_CACHE['rows'] = rows
    return list(rows)


def _median_or_none(values):
    vals = [float(v) for v in values if v is not None and math.isfinite(float(v))]
    return statistics.median(vals) if vals else None


def _normalize_speed_surface(place, surface=''):
    surface = str(surface or '').strip()
    if surface in {'芝', 'ダ', '障'}:
        return surface
    return '' if str(place or '') in _JRA_SPEED_PLACES else 'ダ'


def _speed_reference_category(place):
    place = str(place or '')
    if place in _NANKAN_SPEED_PLACES:
        return '南関'
    if place in _JRA_SPEED_PLACES:
        return 'JRA'
    return '他地方'


def _load_speed_history_rows(cache_dir='backtest_data'):
    """保存済み近走時計を重複除去し、全競馬場の基準時計素材として読む。"""
    try:
        paths = sorted(
            os.path.join(cache_dir, name)
            for name in os.listdir(cache_dir)
            if name.endswith('.json')
        )
    except OSError:
        paths = []
    persistent_path = os.path.join('cache', 'speed_history_reference.json')
    if os.path.exists(persistent_path):
        paths.append(persistent_path)
    signature = (
        len(paths),
        max((os.path.getmtime(path) for path in paths), default=0),
        sum((os.path.getsize(path) for path in paths), 0),
    )
    if _SPEED_HISTORY_CACHE.get('signature') == signature:
        return [dict(row) for row in (_SPEED_HISTORY_CACHE.get('rows') or [])]

    rows = []
    seen = set()
    for path in paths:
        try:
            with open(path, encoding='utf-8') as fh:
                payload = json.load(fh)
        except (OSError, ValueError, TypeError):
            continue
        if isinstance(payload.get('rows'), list):
            stored_horses = {
                str(idx): {'name': row.get('horse_name', ''), 'hist': [row]}
                for idx, row in enumerate(payload.get('rows') or [])
                if isinstance(row, dict)
            }
        else:
            stored_horses = payload.get('horses') or {}
        for horse in stored_horses.values():
            horse_name = str((horse or {}).get('name') or '')
            for run in (horse or {}).get('hist', []):
                place = str((run or {}).get('source_place') or (run or {}).get('place') or '')
                if not place or place == '不明':
                    continue
                dist = _dist_to_int((run or {}).get('dist'))
                try:
                    run_time = float((run or {}).get('time') or 0)
                except (TypeError, ValueError):
                    continue
                if dist <= 0 or run_time <= 0:
                    continue
                run_date = parse_date(str((run or {}).get('date') or ''))
                if run_date == datetime.min:
                    continue
                surface = _normalize_speed_surface(place, (run or {}).get('surface'))
                dedupe_key = (horse_name, run_date, place, surface, dist, round(run_time, 2))
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                rows.append({
                    'horse_name': horse_name,
                    'date': run_date,
                    'place': place,
                    'surface': surface,
                    'dist': dist,
                    'time': run_time,
                })

    _SPEED_HISTORY_CACHE.update({'signature': signature, 'rows': rows})
    return [dict(row) for row in rows]


def _persist_speed_history_rows(horses_data, path=os.path.join('cache', 'speed_history_reference.json')):
    """新しく取得した近走を競馬場・芝ダート別の基準素材としてローカル蓄積する。"""
    existing = []
    try:
        with open(path, encoding='utf-8') as fh:
            payload = json.load(fh)
            if isinstance(payload.get('rows'), list):
                existing = payload['rows']
    except (OSError, ValueError, TypeError):
        existing = []

    by_key = {}
    for row in existing:
        if not isinstance(row, dict):
            continue
        key = (
            str(row.get('horse_name') or ''), str(row.get('date') or ''),
            str(row.get('source_place') or row.get('place') or ''),
            str(row.get('surface') or ''), _dist_to_int(row.get('dist')),
            str(row.get('time') or ''),
        )
        by_key[key] = row
    for horse in (horses_data or {}).values():
        horse_name = str((horse or {}).get('name') or '')
        for hist in (horse or {}).get('hist', []):
            place = str((hist or {}).get('source_place') or (hist or {}).get('place') or '')
            dist = _dist_to_int((hist or {}).get('dist'))
            try:
                run_time = float((hist or {}).get('time') or 0)
            except (TypeError, ValueError):
                continue
            if not place or place == '不明' or dist <= 0 or run_time <= 0:
                continue
            surface = _normalize_speed_surface(place, (hist or {}).get('surface'))
            row = {
                'horse_name': horse_name,
                'date': str((hist or {}).get('date') or ''),
                'place': str((hist or {}).get('place') or ''),
                'source_place': place,
                'surface': surface,
                'dist': dist,
                'time': run_time,
            }
            key = (horse_name, row['date'], place, surface, dist, str(run_time))
            by_key[key] = row
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        tmp_path = f'{path}.tmp'
        with open(tmp_path, 'w', encoding='utf-8') as fh:
            json.dump({'version': 1, 'rows': list(by_key.values())[-20000:]}, fh, ensure_ascii=False)
        os.replace(tmp_path, path)
        _SPEED_HISTORY_CACHE['signature'] = None
    except OSError:
        pass


def _speed_record_scale(values, dist):
    """平均級=70、レコード級=100に使う頑健な2基準を返す。"""
    vals = sorted(float(v) for v in values if v is not None and math.isfinite(float(v)))
    if not vals:
        return None
    average_time = statistics.median(vals)
    # 大標本では最速1件の誤抽出を避けて2番目、小標本では最速値を使う。
    fastest_credible = vals[1] if len(vals) >= 20 else vals[0]
    min_gap = 2.5 * float(dist) / 1000.0
    max_gap = 10.0 * float(dist) / 1000.0
    record_time = max(fastest_credible, average_time - max_gap)
    record_time = min(record_time, average_time - min_gap)
    return {
        'average_time': average_time,
        'record_time': record_time,
        'samples': len(vals),
    }


def _build_speed_scales(history_rows, daily_variants, as_of):
    exact_surface = {}
    exact_distance = {}
    category_surface = {}
    category_distance = {}
    groups = ({}, {}, {}, {})
    for row in history_rows or []:
        run_date = row.get('date')
        if isinstance(as_of, datetime) and isinstance(run_date, datetime) and run_date >= as_of:
            continue
        place = str(row.get('place') or '')
        surface = _normalize_speed_surface(place, row.get('surface'))
        dist = _dist_to_int(row.get('dist'))
        try:
            run_time = float(row.get('time') or 0)
        except (TypeError, ValueError):
            continue
        if not place or place == '不明' or dist <= 0 or run_time <= 0:
            continue
        if place in _NANKAN_SPEED_PLACES and isinstance(run_date, datetime):
            variant = daily_variants.get((run_date.strftime('%Y.%m.%d'), place), 0.0)
            run_time -= variant * dist / 1000.0
        category = _speed_reference_category(place)
        keys = (
            (place, surface, dist),
            (place, dist),
            (category, surface, dist),
            (category, dist),
        )
        for group, key in zip(groups, keys):
            group.setdefault(key, []).append(run_time)

    for source, target in zip(groups, (exact_surface, exact_distance, category_surface, category_distance)):
        for key, values in source.items():
            scale = _speed_record_scale(values, key[-1])
            if scale:
                target[key] = scale
    return {
        'exact_surface': exact_surface,
        'exact_distance': exact_distance,
        'category_surface': category_surface,
        'category_distance': category_distance,
    }


def _blend_speed_scales(exact, fallback, label):
    if not exact:
        if not fallback:
            return None
        return dict(fallback, source=label, confidence='共通')
    if not fallback:
        return dict(exact, source=label, confidence='競馬場別')
    n = int(exact.get('samples', 0) or 0)
    reliability = n / (n + 8.0)
    return {
        'average_time': reliability * exact['average_time'] + (1.0 - reliability) * fallback['average_time'],
        'record_time': reliability * exact['record_time'] + (1.0 - reliability) * fallback['record_time'],
        'samples': n,
        'source': label,
        'confidence': '高' if n >= 20 else '中' if n >= 8 else '暫定',
    }


def _speed_scale_for_run(place, surface, dist, scales):
    category = _speed_reference_category(place)
    surface = _normalize_speed_surface(place, surface)
    exact = scales['exact_surface'].get((place, surface, dist))
    if exact is None:
        exact = scales['exact_distance'].get((place, dist))
    fallback = scales['category_surface'].get((category, surface, dist))
    if fallback is None:
        fallback = scales['category_distance'].get((category, dist))
    if exact or fallback:
        return _blend_speed_scales(exact, fallback, f'{place}{surface}{dist}m基準')

    candidates = [
        (abs(int(other_dist) - int(dist)), int(other_dist), scale)
        for (cat, other_dist), scale in scales['category_distance'].items()
        if cat == category
    ]
    if not candidates:
        return None
    _gap, base_dist, base = min(candidates)
    ratio = float(dist) / float(base_dist)
    return {
        'average_time': base['average_time'] * ratio,
        'record_time': base['record_time'] * ratio,
        'samples': base.get('samples', 0),
        'source': f'{category}{base_dist}m近似',
        'confidence': '近似',
    }


def _build_speed_reference(meta_rows):
    """共通時計基準と開催日別の高速・低速馬場差を作る。

    日別馬場差は、各勝時計を別日の同場・同距離・同馬場状態・同クラス帯と比較し、
    秒/1000mへ揃えて同日複数距離の中央値を取る。最低3競走かつ2距離を必要とする。
    推定誤差を抑えるため、実際に指数へ使う量は推定値の10%へ縮約する。
    負値=高速、正値=低速。
    """
    rows = [r for r in (meta_rows or []) if r.get('winner_time') and r.get('dist')]
    by_pd = {}
    by_pdg = {}
    by_pdgc = {}
    for row in rows:
        pd = (row['place'], int(row['dist']))
        pdg = pd + (row.get('going') or '',)
        pdgc = pdg + (row.get('class_bucket') or 'その他',)
        by_pd.setdefault(pd, []).append(row)
        by_pdg.setdefault(pdg, []).append(row)
        by_pdgc.setdefault(pdgc, []).append(row)

    # 能力比較用の共通基準。クラスでは割らず、速いクラスの時計を高く評価できるようにする。
    common_par = {}
    for key, group in by_pd.items():
        if len(group) >= 8:
            common_par[key] = _median_or_none(r['winner_time'] for r in group)
    for key, group in by_pdg.items():
        if len(group) >= 6:
            common_par[key] = _median_or_none(r['winner_time'] for r in group)

    daily_residuals = {}
    daily_distances = {}
    for row in rows:
        row_date = row['date']
        group = by_pdgc.get(
            (row['place'], int(row['dist']), row.get('going') or '', row.get('class_bucket') or 'その他'),
            [],
        )
        candidates = [r['winner_time'] for r in group if r.get('date') != row_date]
        if len(candidates) < 4:
            continue
        expected = _median_or_none(candidates)
        if expected is None:
            continue
        residual_per_1000 = (float(row['winner_time']) - expected) * 1000.0 / int(row['dist'])
        key = (row_date, row['place'])
        daily_residuals.setdefault(key, []).append(residual_per_1000)
        daily_distances.setdefault(key, set()).add(int(row['dist']))

    daily_variants = {}
    for key, residuals in daily_residuals.items():
        if (len(residuals) < SPEED_INDEX_DAILY_VARIANT_MIN_RACES
                or len(daily_distances.get(key, set())) < SPEED_INDEX_DAILY_VARIANT_MIN_DISTANCES):
            continue
        variant = _median_or_none(residuals)
        if variant is None:
            continue
        raw_variant = max(
            -SPEED_INDEX_DAILY_VARIANT_CAP,
            min(SPEED_INDEX_DAILY_VARIANT_CAP, variant),
        )
        daily_variants[key] = raw_variant * SPEED_INDEX_DAILY_VARIANT_SHRINKAGE
    return common_par, daily_variants


def _speed_variant_label(variant_per_1000):
    if variant_per_1000 <= -0.03:
        return '高速'
    if variant_per_1000 >= 0.03:
        return '低速'
    return '標準'


def calculate_common_speed_indices(horses_data, current_course='', current_dist='', cache_dir='cache', as_of=None):
    """近5走時計を平均70・レコード級100の共通指数へ変換する。

    index=現在能力、race_index=出走馬内最高100、same_condition_max=同場同距離自己最高。
    すべて独立検証軸で、総合点には加算しない。
    """
    meta_rows = _load_speed_race_meta(cache_dir)
    _unused_common_par, daily_variants = _build_speed_reference(meta_rows)
    meta_by_race = {r['race_id']: r for r in meta_rows}
    if not isinstance(as_of, datetime):
        _persist_speed_history_rows(horses_data)
    as_of_dt = as_of if isinstance(as_of, datetime) else datetime.now()
    history_rows = _load_speed_history_rows()

    # 今回取得した近走も基準素材へ加える。過去保存分と重なる走りは除外する。
    seen_rows = {
        (row.get('horse_name'), row.get('date'), row.get('place'), row.get('surface'), row.get('dist'), round(float(row.get('time') or 0), 2))
        for row in history_rows
    }
    for horse in (horses_data or {}).values():
        horse_name = str((horse or {}).get('name') or '')
        for hist in (horse or {}).get('hist', []):
            place = str((hist or {}).get('source_place') or (hist or {}).get('place') or '')
            dist = _dist_to_int((hist or {}).get('dist'))
            run_date = parse_date(str((hist or {}).get('date') or ''))
            try:
                run_time = float((hist or {}).get('time') or 0)
            except (TypeError, ValueError):
                continue
            if not place or place == '不明' or dist <= 0 or run_time <= 0 or run_date == datetime.min:
                continue
            surface = _normalize_speed_surface(place, (hist or {}).get('surface'))
            key = (horse_name, run_date, place, surface, dist, round(run_time, 2))
            if key in seen_rows:
                continue
            seen_rows.add(key)
            history_rows.append({
                'horse_name': horse_name, 'date': run_date, 'place': place,
                'surface': surface, 'dist': dist, 'time': run_time,
            })
    scales = _build_speed_scales(history_rows, daily_variants, as_of_dt)
    current_dist_int = _dist_to_int(current_dist)
    result = {}

    for umaban, horse in (horses_data or {}).items():
        run_scores = []
        for hist in (horse or {}).get('hist', [])[:5]:
            try:
                raw_time = float(hist.get('time'))
            except (TypeError, ValueError):
                continue
            dist = _dist_to_int(hist.get('dist'))
            place = str(hist.get('source_place') or hist.get('place') or '')
            surface = _normalize_speed_surface(place, hist.get('surface'))
            if raw_time <= 0 or dist <= 0 or not place or place == '不明':
                continue
            url = str(hist.get('url') or '')
            race_match = re.search(r'/result/(\d+)\.do', url)
            race_id = race_match.group(1) if race_match else ''
            meta = meta_by_race.get(race_id, {})
            is_nankan_run = place in _NANKAN_SPEED_PLACES
            scale = _speed_scale_for_run(place, surface, dist, scales)
            if not scale:
                continue
            run_date = parse_date(str(hist.get('date') or meta.get('date') or ''))
            date_key = run_date.strftime('%Y.%m.%d') if run_date != datetime.min else meta.get('date', '')
            # 開催日馬場差は南関4場だけ。JRA・他地方の転入走は競馬場/距離基準のみ。
            variant = daily_variants.get((date_key, place), 0.0) if is_nankan_run else 0.0
            # 高速馬場(variant<0)では時計を遅い側へ戻し、低速馬場では速い側へ戻す。
            adjusted_time = raw_time - variant * dist / 1000.0
            scale_range = max(0.5, float(scale['average_time']) - float(scale['record_time']))
            run_index = (
                SPEED_INDEX_AVERAGE_SCORE
                + (SPEED_INDEX_RECORD_SCORE - SPEED_INDEX_AVERAGE_SCORE)
                * (float(scale['average_time']) - adjusted_time) / scale_range
            )
            run_index = max(SPEED_INDEX_MIN_SCORE, min(SPEED_INDEX_MAX_SCORE, run_index))
            days = max(0, (as_of_dt - run_date).days) if run_date != datetime.min else 180
            run_scores.append({
                'index': run_index,
                'raw_time': raw_time,
                'adjusted_time': adjusted_time,
                'variant_per_1000': variant,
                'variant_label': _speed_variant_label(variant),
                'date': date_key,
                'place': place,
                'surface': surface,
                'dist': dist,
                'days': days,
                'race_id': race_id,
                'par_source': scale.get('source', ''),
                'reference_confidence': scale.get('confidence', ''),
                'reference_samples': scale.get('samples', 0),
                'average_time': scale.get('average_time'),
                'record_time': scale.get('record_time'),
                'is_transfer': not is_nankan_run,
            })
        if not run_scores:
            result[str(umaban)] = {'index': None, 'rank': None, 'runs': 0, 'corrected_runs': 0}
            continue
        values = sorted((r['index'] for r in run_scores), reverse=True)
        top2 = statistics.mean(values[:2])
        recency_weights = [math.exp(-r['days'] / 100.0) for r in run_scores]
        recency = sum(r['index'] * w for r, w in zip(run_scores, recency_weights)) / sum(recency_weights)
        volatility = statistics.pstdev(values) if len(values) >= 2 else 0.0
        horse_index = (
            SPEED_INDEX_TOP2_WEIGHT * top2
            + SPEED_INDEX_RECENCY_WEIGHT * recency
            - SPEED_INDEX_VOLATILITY_PENALTY * volatility
        )
        best_run = max(run_scores, key=lambda r: r['index'])
        same_condition_runs = [
            r for r in run_scores
            if current_course and current_dist_int > 0
            and r.get('place') == current_course and int(r.get('dist') or 0) == current_dist_int
        ]
        same_condition_best = max(same_condition_runs, key=lambda r: r['index']) if same_condition_runs else None
        result[str(umaban)] = {
            'index': round(horse_index, 1),
            'rank': None,
            'runs': len(run_scores),
            'corrected_runs': sum(abs(r['variant_per_1000']) >= 0.03 for r in run_scores),
            'best_run': best_run,
            'same_condition_max': round(same_condition_best['index'], 1) if same_condition_best else None,
            'same_condition_runs': len(same_condition_runs),
            'same_condition_best_run': same_condition_best,
            'top2': round(top2, 1),
            'recency': round(recency, 1),
            'volatility': round(volatility, 1),
        }

    valid = [info['index'] for info in result.values() if info.get('index') is not None]
    race_best = max(valid) if valid else None
    for info in result.values():
        if info.get('index') is not None:
            info['rank'] = 1 + sum(other > info['index'] for other in valid)
            info['total'] = len(valid)
            # レース内最速馬を必ず100とし、絶対指数の点差はそのまま残す。
            info['race_index'] = round(max(0.0, 100.0 - (race_best - info['index'])), 1)
    return result

def _is_kankan_distance(dist):
    """根幹距離 (1200/1600/2000m)。これ以外は同場・同距離成績を絶対視する。"""
    return _dist_to_int(dist) in (1200, 1600, 2000)

def _is_ooi_inner(dist):
    d_int = _dist_to_int(dist)
    return d_int in [1500, 1600, 1650]

def _is_ooi_outer(dist):
    d_int = _dist_to_int(dist)
    return d_int > 0 and d_int not in [1500, 1600, 1650]

def _is_one_turn(place, dist):
    """南関東の完全な1ターン（コーナー2回）コースを定義"""
    d = _dist_to_int(dist)
    if place == "川崎" and d == 900: return True
    if place == "浦和" and d == 800: return True
    if place == "船橋" and d in [1000, 1200]: return True
    if place == "大井" and d in [1000, 1200, 1400]: return True
    return False

def _is_same_track_layout(place, dist1, dist2):
    """競馬場と距離から、コース形態（大井の内/外、コーナー数など）が一致するか判定"""
    d1 = _dist_to_int(dist1)
    d2 = _dist_to_int(dist2)
    
    if place == "大井":
        def get_ooi_layout(d):
            if d <= 1400: return "outer_1turn"
            if d <= 1650: return "inner_2turn"
            return "outer_2turn"
        return get_ooi_layout(d1) == get_ooi_layout(d2)
        
    elif place == "川崎":
        def get_kawasaki_layout(d):
            if d == 900: return "1turn"
            if d <= 1600: return "2turn"
            return "multi_turn"
        return get_kawasaki_layout(d1) == get_kawasaki_layout(d2)

    elif place == "船橋":
        def get_funabashi_layout(d):
            if d <= 1200: return "1turn"
            if d <= 1800: return "2turn"
            return "multi_turn"
        return get_funabashi_layout(d1) == get_funabashi_layout(d2)

    elif place == "浦和":
        def get_urawa_layout(d):
            if d <= 800: return "1turn"
            if d <= 1500: return "2turn"
            return "multi_turn"
        return get_urawa_layout(d1) == get_urawa_layout(d2)

    return d1 == d2

def _race_condition_priority(place, dist, current_place, current_dist):
    """現在条件に対する比較レースの優先度。5=同場同距離、4=同場同系統。"""
    d = _dist_to_int(dist)
    cur_d = _dist_to_int(current_dist)
    if not place or not current_place or d <= 0 or cur_d <= 0:
        return 0
    if not _is_relative_distance_allowed(d, cur_d):
        # 最後の手段: 600m窓で弾かれても、同場なら距離不問で最弱採用する。
        # ガード(ターン数一致/大井内外一致)は _is_same_place_distfree_ok 内で担保。
        if _is_same_place_distfree_ok(place, d, current_place, current_dist):
            return SAME_PLACE_DISTFREE_PRIORITY
        return 0
    if _is_long_distance(cur_d):
        if place == current_place and d == cur_d:
            return 5
        # 長距離でも「同じ競馬場の同じコース形態」を最優先する。
        # 例: 船橋1800では、別形態の船橋2200より船橋1600/1500を近い条件として扱う。
        if place == current_place and _is_same_track_layout(place, d, cur_d):
            return 4
        same_turn = _is_turn_match_to_current(place, d, current_place, cur_d)
        if place == current_place and same_turn:
            return 3
        if place == current_place and _is_long_distance(d):
            return 2
        if place != current_place and same_turn and _is_long_distance(d):
            return 2
        if place != current_place and same_turn:
            return 1
        return 0
    if place == current_place:
        if d == cur_d:
            return 4
        if _is_same_track_layout(place, d, cur_d):
            return 3
        return 2
    if d == cur_d:
        return 1
    return 0

def _ooi_track_side(dist):
    d_int = _dist_to_int(dist)
    if _is_ooi_inner(d_int):
        return "inner"
    if _is_ooi_outer(d_int):
        return "outer"
    return ""

_BRIDGE_DISCOUNT_BY_PRIORITY = {5: 1.0, 4: 0.7, 3: 0.5, 2: 0.4, 1: 0.35, SAME_PLACE_DISTFREE_PRIORITY: 0.25}

# 場別「人気-着順」プロファイル: 南関4場でのみ集計し、走った場の信頼度を測る。
# 未経験は中立(影響なし)、走って裏切っている場合のみディスカウントをかける。
_NANKAN_PROFILE_PLACES = ("大井", "船橋", "川崎", "浦和")
_TRACK_PROFILE_SHRINKAGE_K = 3  # N<3 の場は0方向に縮約して暴れ防止

def _track_aptitude_score(rank_i, pop_i):
    """1走分の (人気-着順) スコア。上位人気で勝ち切ったレースに加点。"""
    base = pop_i - rank_i
    if rank_i == 1:
        if pop_i == 1:
            return base + 2.0
        if pop_i == 2:
            return base + 1.5
        if pop_i == 3:
            return base + 1.0
        return base
    if rank_i == 2 and pop_i <= 3:
        return base + 0.5
    return base

def _compute_track_profile(hist):
    """馬の出走履歴から場別 avg(人気-着順) プロファイルを作る。
    返り値: {place: {"avg": shrunk_avg, "n": run_count, "raw_avg": raw}}
    """
    runs_by_place = {}
    for pr in hist or []:
        if not isinstance(pr, dict):
            continue
        place = pr.get("place", "")
        if place not in _NANKAN_PROFILE_PLACES:
            continue
        try:
            rank_i = int(pr.get("rank"))
            pop_i = int(pr.get("pop"))
        except (TypeError, ValueError):
            continue
        if rank_i <= 0 or pop_i <= 0:
            continue
        runs_by_place.setdefault(place, []).append(_track_aptitude_score(rank_i, pop_i))

    profile = {}
    for place, scores in runs_by_place.items():
        n = len(scores)
        raw_avg = sum(scores) / n
        shr_avg = (n * raw_avg) / (n + _TRACK_PROFILE_SHRINKAGE_K)
        profile[place] = {"avg": shr_avg, "n": n, "raw_avg": raw_avg}
    return profile

def _track_discount_factor(profile, place):
    """場別プロファイルから比較ディスカウント係数を返す。
    - 未経験 / 期待通り以上(avg>=0): 1.0 (影響なし)
    - 経験場が1場しかない: 1.0 (比較材料が単一場のみで「裏切り」と断定できないため)
    - 裏切り(avg<0): max(0.75, 1 + avg * 0.03)
    """
    if not profile or place not in _NANKAN_PROFILE_PLACES:
        return 1.0
    prof = profile.get(place)
    if not prof or prof["n"] == 0:
        return 1.0
    avg = prof["avg"]
    if avg >= 0:
        return 1.0
    exp_count = sum(1 for p in _NANKAN_PROFILE_PLACES if profile.get(p, {}).get("n", 0) > 0)
    if exp_count <= 1:
        return 1.0
    return max(0.75, 1.0 + avg * 0.03)

def _is_turn_match_to_current(place, dist, current_place, current_dist):
    """過去レースのコーナー数(1turn/2turn+)が今回と一致するか。
    priority=0 の参考値採用時にギリギリの品質担保として使う。"""
    cur_turn = _layout_turn_class(current_place, current_dist)
    leg_turn = _layout_turn_class(place, dist)
    return _turn_classes_match(cur_turn, leg_turn)

def _is_same_place_distfree_ok(place, dist, current_place, current_dist):
    """最後の手段: 競馬場が同じなら距離不問で相対比較を許可するか。
    600m窓(_is_relative_distance_allowed)で弾かれた後だけ使うフォールバック。
    ガードは必ず維持する:
      - コーナー数(ターン数)が今回と一致すること
      - 大井は内回り/外回りが今回と一致すること
    """
    if not place or not current_place or place != current_place:
        return False
    d = _dist_to_int(dist)
    cur_d = _dist_to_int(current_dist)
    if d <= 0 or cur_d <= 0:
        return False
    if not _is_turn_match_to_current(place, d, current_place, current_dist):
        return False
    if place == "大井" and _ooi_track_side(d) != _ooi_track_side(cur_d):
        return False
    return True

def _is_pop_drop_outlier(rank, pop):
    """1〜3人気で着順を3つ以上落とした(R-P>=3)= 不本意な結果(外れ値)。
    本来強い馬の凡走には理由(出遅れ・揉まれる・落鉄等)がある可能性が高く、
    同条件で他に対戦データがあるなら、こちらは比較材料として優先しない。"""
    if rank is None or pop is None:
        return False
    try:
        r, p = int(rank), int(pop)
    except (TypeError, ValueError):
        return False
    return 1 <= p <= 3 and (r - p) >= 3

def _is_bridge_leg_recent_enough(leg_date, leg_dist, today=None):
    """隠れ馬経由のレッグとして採用可否を日付で判定。
    - 半年(180日)以内なら採用
    - 距離 ≥1800m なら日付制限なし(長距離レースはレース数が少ないため)
    leg_date が解釈不能なら True (情報不足で弾かない側に倒す)。"""
    if _dist_to_int(leg_dist) >= BRIDGE_OLD_OK_DIST_MIN:
        return True
    dt = parse_date(leg_date) if isinstance(leg_date, str) else leg_date
    if not isinstance(dt, datetime) or dt == datetime.min:
        return True
    base = today or datetime.now()
    return (base - dt).days <= BRIDGE_RECENT_DAYS

def _track_marker_symbol(profile, place, experienced_place_count=None):
    """場別プロファイルから ◎/○/△/- のマーカーを返す。
    experienced_place_count: 南関4場のうち出走経験のある場数。1場のみの場合は△を出さず最低○とする
    (比較材料が単一場しかない時に「悪い」と判定するのは不当なため)。"""
    if not profile:
        return "-"
    prof = profile.get(place)
    if not prof or prof["n"] == 0:
        return "-"
    avg = prof["avg"]
    if avg >= 1.5:
        return "◎"
    if avg >= 0:
        return "○"
    if experienced_place_count is not None and experienced_place_count <= 1:
        return "○"
    return "△"


def _detect_post_rest_improvement(hist):
    """3ヶ月(90日)以上の休み明けで、その後の平均着順が休み前より2.0以上向上していれば
    休み明け初戦の日付を閾値として返す。該当しなければ None。
    複数の休みがある場合は、条件を満たす中で最も新しい休みを採用する。"""
    valid = []
    for h in hist or []:
        if not isinstance(h, dict):
            continue
        d_str = h.get("date")
        if not d_str:
            continue
        dt = parse_date(d_str) if isinstance(d_str, str) else d_str
        if not isinstance(dt, datetime) or dt == datetime.min:
            continue
        try:
            r = int(h.get("rank"))
        except (TypeError, ValueError):
            continue
        valid.append((dt, r))
    if len(valid) < 3:
        return None
    valid.sort(key=lambda x: x[0])  # 古い→新しい

    threshold = None
    for i in range(len(valid) - 1, 0, -1):
        gap_days = (valid[i][0] - valid[i - 1][0]).days
        if gap_days < 90:
            continue
        before = valid[:i]
        after = valid[i:]
        if len(after) < 2:
            continue
        avg_before = sum(b[1] for b in before) / len(before)
        avg_after = sum(a[1] for a in after) / len(after)
        if avg_before - avg_after >= 2.0:
            threshold = valid[i][0]
            break  # 最も新しい該当休みを優先
    return threshold


def _detect_last_run_surge(hist):
    """前走で急に走り出した馬を検出し、前走URLを相対評価の優先材料にする。

    条件:
      - 前走が3着以内、かつ2〜5走前の有効着順が全て4着以下
      - または前走1着、かつ2〜5走前の有効人気が全て5人気以下
    """
    runs = [h for h in (hist or []) if isinstance(h, dict)]
    if len(runs) < 3:
        return None

    latest = runs[0]
    latest_url = latest.get("url", "")
    if not latest_url:
        return None

    latest_rank = _rank_to_int(latest.get("rank"))
    older = runs[1:5]
    reasons = []

    older_ranks = [_rank_to_int(h.get("rank")) for h in older if h.get("rank") is not None]
    if latest_rank <= 3 and len(older_ranks) >= 2 and all(r >= 4 for r in older_ranks):
        reasons.append("prev_in_money_after_older_4plus")

    older_pops = [_rank_to_int(h.get("pop")) for h in older if h.get("pop") is not None]
    if latest_rank == 1 and len(older_pops) >= 2 and all(p >= 5 for p in older_pops):
        reasons.append("prev_win_after_older_low_pop")

    if not reasons:
        return None

    return {
        "url": latest_url,
        "date": latest.get("date", ""),
        "rank": latest_rank,
        "reasons": reasons,
    }


def _front_position_in_field(pas, field_size, fraction=1.0 / 3.0):
    """道中の通過順が出走頭数の前 fraction(既定1/3)以内に入っていたかを返す。
    先行して粘った形を捉えたいので、各コーナー通過順のうち最も前(最小)の位置で判定する。
    通過順・頭数が不明な場合は False(=先行の確証なし)。"""
    try:
        fs = int(field_size)
    except (TypeError, ValueError):
        return False
    if fs <= 0:
        return False
    positions = [int(x) for x in re.findall(r'\d+', str(pas or ""))]
    if not positions:
        return False
    threshold = max(1, math.ceil(fs * fraction))
    return min(positions) <= threshold


def _detect_tataki_second(hist, rest_days=75, min_fade_positions=3):
    """前走が2.5ヶ月以上の休み明けで、先行後にばてた馬の叩き2戦目を検出する。
    条件: 前走と前々走が75日以上空いている /
          前走の最初の通過順が出走頭数の前1/3 /
          その通過順からゴールで着順を3つ以上落とした。"""
    valid = []
    for h in hist or []:
        if not isinstance(h, dict):
            continue
        d_str = h.get("date")
        dt = parse_date(d_str) if isinstance(d_str, str) else d_str
        if not isinstance(dt, datetime) or dt == datetime.min:
            continue
        try:
            rank = int(h.get("rank"))
        except (TypeError, ValueError):
            continue
        pop = h.get("pop")
        try:
            pop = int(pop)
        except (TypeError, ValueError):
            pop = None
        valid.append({
            "date": dt, "rank": rank, "pop": pop,
            "pas": h.get("pas", ""), "field_size": h.get("field_size"),
        })

    if len(valid) < 2:
        return None

    # parse_nankankeiba_detail の近走は直近順だが、念のため日付で並べ直す。
    valid.sort(key=lambda x: x["date"], reverse=True)
    last = valid[0]
    before_last = valid[1]
    gap_days = (last["date"] - before_last["date"]).days
    if gap_days < rest_days:
        return None
    positions = [int(x) for x in re.findall(r'\d+', str(last.get("pas") or ""))]
    try:
        field_size = int(last.get("field_size"))
    except (TypeError, ValueError):
        return None
    if not positions or field_size <= 0:
        return None
    front_position = positions[0]
    if front_position > max(1, math.ceil(field_size / 3.0)):
        return None
    fade_positions = last["rank"] - front_position
    if fade_positions < min_fade_positions:
        return None

    return {
        "gap_days": gap_days,
        "pop": last["pop"],
        "rank": last["rank"],
        "front_position": front_position,
        "fade_positions": fade_positions,
        "detail": f"前走{gap_days}日明け・{front_position}番手から{last['rank']}着",
    }


def _is_past_race_within_relative_window(r_date, r_place, r_dist_str, current_place, current_dist, today=None):
    """相対評価で過去レースを採用してよいか日付ベースで判定。
    基本は 9ヶ月以内。9〜12ヶ月の場合は (a) 同場×同距離 もしくは
    (b) 同場×同レイアウト×距離差±200m以内 のみコース適性として救済。
    日付が取れない場合は情報枯渇回避のため True を返す。"""
    if isinstance(r_date, str):
        dt = parse_date(r_date)
    elif isinstance(r_date, datetime):
        dt = r_date
    else:
        dt = datetime.min
    if dt == datetime.min:
        return True

    today = today or datetime.now()
    days_diff = (today - dt).days
    if days_diff < 0:
        return True
    if days_diff <= RELATIVE_PAST_DAYS_DEFAULT:
        return True
    if days_diff > RELATIVE_PAST_DAYS_EXTENDED:
        return False

    r_d = _dist_to_int(r_dist_str)
    c_d = _dist_to_int(current_dist)
    if not r_place or not current_place or r_d <= 0 or c_d <= 0:
        return False
    if r_place != current_place:
        return False
    if r_d == c_d:
        return True
    if _is_same_track_layout(r_place, r_d, c_d) and abs(r_d - c_d) <= RELATIVE_EXTENDED_DIST_DIFF:
        return True
    return False

def _layout_turn_class(place, dist):
    """競馬場をまたいでコース形態(コーナー数)を揃えるためのターン分類。"""
    d = _dist_to_int(dist)
    if d <= 0 or not place:
        return ""
    if place == "大井":
        if d <= 1400: return "1turn"
        if d <= 1650: return "2turn_inner"
        return "2turn_outer"
    if place == "川崎":
        if d == 900: return "1turn"
        if d <= 1600: return "2turn"
        return "multi_turn"
    if place == "船橋":
        if d <= 1200: return "1turn"
        if d <= 1800: return "2turn"
        return "multi_turn"
    if place == "浦和":
        if d <= 800: return "1turn"
        if d <= 1500: return "2turn"
        return "multi_turn"
    return ""

def _turn_classes_match(c1, c2):
    """ターン数(コーナー数)が同じなら一致と見なす。大井の内外回りは他場の2turnと同等。"""
    if not c1 or not c2:
        return False
    base1 = c1.split("_")[0]
    base2 = c2.split("_")[0]
    return base1 == base2

def _hidden_bridge_priority(place1, dist1, place2, dist2, current_dist=None, current_place=None):
    """隠れ馬経由で比較してよいレース条件の優先度を返す。
    5: 両レッグ今回と同場・同距離
    4: 両レッグ今回と同場・同レイアウト
    3: 片レッグ同場同レイアウト + 片レッグ同場(レイアウト違い)
    2: 両レッグ今回と同場(レイアウト違いを含む)
    1: 両レッグ間で同場・同レイアウト、両レッグ他場、今回と同ターン数、距離差<400m
    ※全優先度で、今回距離から600m以上離れるレッグは不採用
    0: 不採用
    """
    d1 = _dist_to_int(dist1)
    d2 = _dist_to_int(dist2)
    cur_d = _dist_to_int(current_dist)
    if not place1 or not place2 or d1 <= 0 or d2 <= 0:
        return 0

    if current_place and cur_d > 0:
        if not (_is_relative_distance_allowed(d1, cur_d) and _is_relative_distance_allowed(d2, cur_d)):
            # 最後の手段: 両レッグとも今回と同場なら距離不問で最弱採用する。
            # ガード(ターン数一致/大井内外一致)は _is_same_place_distfree_ok 内で担保。
            if (_is_same_place_distfree_ok(place1, d1, current_place, current_dist)
                    and _is_same_place_distfree_ok(place2, d2, current_place, current_dist)):
                return SAME_PLACE_DISTFREE_PRIORITY
            return 0

        # 各レッグのコーナー数 (ターン数) が今回と一致しないと採用しない。
        # 例: 船橋1800(2turn) に対して川崎2000/浦和2000(multi) のレッグは混ぜない。
        if not _is_turn_match_to_current(place1, d1, current_place, current_dist):
            return 0
        if not _is_turn_match_to_current(place2, d2, current_place, current_dist):
            return 0

        if _is_long_distance(cur_d):
            # 各レッグは 2turn 以上(=1turnでない)であることを必須とする (ターン一致チェックで実質吸収済み)
            leg1_base = _layout_turn_class(place1, d1).split("_")[0]
            leg2_base = _layout_turn_class(place2, d2).split("_")[0]
            if leg1_base in ("1turn", "") or leg2_base in ("1turn", ""):
                return 0
            # 長距離は同場(距離違い)を他場(長距離)より優先する。
            if place1 == current_place and place2 == current_place:
                if d1 == cur_d and d2 == cur_d:
                    return 5
                if _is_long_distance(d1) and _is_long_distance(d2) and d1 == d2:
                    return 4
                # 同場・2turn+ ・非長距離(同距離同士)
                if d1 == d2:
                    return 3
                # 距離違いでも両レッグが今回距離と同レイアウト (例: 船橋1600+1800 vs 船橋1800)。
                # 距離違いの直接対決(2200等・別レイアウト)より、同レイアウト bridge を優先したい。
                if (_is_same_track_layout(current_place, d1, cur_d)
                    and _is_same_track_layout(current_place, d2, cur_d)):
                    # どちらかのレッグが今回距離と一致なら少し優遇
                    if d1 == cur_d or d2 == cur_d:
                        return 4
                    return 3
                return 0
            # 他場・距離違いは混ぜない (他場版)
            if place1 == place2 and d1 != d2:
                return 0
            if place1 != current_place and place2 != current_place:
                if _is_long_distance(d1) and _is_long_distance(d2):
                    if d1 == cur_d and d2 == cur_d:
                        return 3
                    if place1 == place2 and d1 == d2:
                        return 2
                    return 3
                # 他場・2turn+ ・非長距離(同距離同士)
                if place1 == place2 and d1 == d2:
                    return 1
                return 0
            return 0

        # 両レッグが今回と同場
        if place1 == current_place and place2 == current_place:
            p1 = _race_condition_priority(place1, d1, current_place, cur_d)
            p2 = _race_condition_priority(place2, d2, current_place, cur_d)
            if p1 >= 4 and p2 >= 4:
                return 5
            if p1 >= 3 and p2 >= 3:
                return 4
            if (p1 >= 3 and p2 >= 2) or (p1 >= 2 and p2 >= 3):
                return 3
            return 2

        # 両レッグ同士は同場(今回とは違う場)・同レイアウト・今回と同ターン数・距離差<400m
        if place1 == place2 and place1 != current_place:
            if _is_same_track_layout(place1, d1, d2):
                cur_turn = _layout_turn_class(current_place, cur_d)
                leg_turn = _layout_turn_class(place1, d1)
                if _turn_classes_match(cur_turn, leg_turn) and abs(d1 - cur_d) < 400 and abs(d2 - cur_d) < 400:
                    return 1

        # 拡張弱bridge: 場も距離も揃わない組み合わせ(片レッグ同場+他場、別々の他場、
        # 同他場のレイアウト違い等)も priority=1 の弱比較として採用し「比較不可」を減らす。
        # 前提条件は上で確認済み: 各レッグの距離窓(今回±600m未満)・ターン数一致。
        # 勝敗判定は弱bridge専用の厳格閾値(大差のみ＞/＜、それ未満は＝)を使う。
        # 大井の内外回りまたぎだけは禁止を維持する。
        if current_place == "大井":
            if place1 == "大井" and _ooi_track_side(d1) != _ooi_track_side(cur_d):
                return 0
            if place2 == "大井" and _ooi_track_side(d2) != _ooi_track_side(cur_d):
                return 0
        if place1 == "大井" and place2 == "大井" and _ooi_track_side(d1) != _ooi_track_side(d2):
            return 0
        return 1

    # current_place が無いケース(汎用): 既存挙動を概ね維持
    if place1 == place2:
        if d1 == d2:
            return 3
        if place1 == "大井" and _ooi_track_side(d1) != _ooi_track_side(d2):
            return 0
        return 2
    return 0

def _hidden_bridge_note(priority, is_long=False):
    if priority == SAME_PLACE_DISTFREE_PRIORITY:
        return "同場(距離不問)"
    if is_long:
        if priority >= 5:
            return "長距離同場同距離"
        if priority == 4:
            return "長距離同場同系統"
        if priority == 3:
            return "長距離同場別距離"
        if priority == 2:
            return "長距離他場別距離"
        return ""
    if priority >= 5:
        return "同場同距離"
    if priority == 4:
        return "同場同系統"
    if priority == 3:
        return "同場"
    if priority == 2:
        return "同場(レイアウト違)"
    if priority == 1:
        return "弱比較(他場/混成)"
    return ""

# 隠れ馬の合議制パラメータ(バックテストで調整可能)
_BRIDGE_CONSENSUS_MIN_VOTES = 1     # 1件でも着差基準を満たせば決着。複数レースは合議精度を高める証拠として扱う
_BRIDGE_CONSENSUS_AGREE = 0.70      # 同方向の重み比率がこれ以上で決着
# 合議には最良priorityに加えて1段下まで使う。
# 1段下は補助票として減量し、最良priorityの結論を逆方向へ覆す用途には使わない。
_BRIDGE_CONSENSUS_INCLUDE_ONE_LOWER = True
_BRIDGE_CONSENSUS_LOWER_WEIGHT = 0.6   # best-1の補助票係数（最良priority票=1.0）

# 確信度ゲート(A): tier層化(dom_G)の昇降エッジに使う最小確信度。
#  bridge勝敗は採用レース/経路の票数で判定する。同じ隠れ馬でも複数レースがあれば
#  精度を上げる材料として票に含める。
#  直接対決・推移補完は実測/論理なのでゲート対象外。
_REL_CONFIDENCE_GATE = True            # False で旧挙動(全エッジ採用)に戻せる(A/B用)
_REL_DOM_EDGE_MIN_BRIDGE_VOTES = 2     # bridgeが層化エッジになるのに必要な合議票数(採用レース/経路数)
_REL_S_FLOOR = False                   # 最上位グループは必ずS(Sフロアoff)。Sを必ず立てる方を優先(ユーザー採用2026/06/24、6/19のonを撤回)。S複勝率は38%維持
_REL_C_CEILING = True                  # 確信できる負けが無い馬はCに沈めずBへ(C=ほぼ馬券圏外狙い)
_REL_C_INTERNAL_WIN_PROMOTE = True     # 三軍内で複数馬に勝ち、三軍内負けなしなら二軍へ救済
_REL_C_INTERNAL_WIN_MIN = 2
_REL_MATCHUP_TIER_CONSISTENCY = True   # 上位への実比較と矛盾する二軍/三軍の所属・相対点を救済

# 相対tier多層化(A): tier層化のレベルを4段階(S/A/B/C)にクランプせず、層の数だけ
#  S→A→B→C→D→E… と割り当てて能力の上下関係の解像度を上げる。点数も4値固定では
#  なく、そのレース内のレベル位置で最上位50〜最下位15に線形配点する。
#  Falseで旧4段階クランプ・固定配点に戻す(backtest_relative.py で A/B 比較)。
_REL_MULTI_TIER = False                 # S/A/B/C+判定不能の5グループ(C以下は受け皿)。多層化D/Eは不要(ユーザー要望2026/06/24)
_TIER_LABELS = "SABCDEFGHIJKLMNOPQRSTUVWXYZ"  # レベル0→S, 1→A, 2→B, 3→C, 4→D...
_REL_TOP_SCORE = 50.0                   # 最上位レベルの相対点
_REL_BOTTOM_SCORE = 15.0                # 最下位レベルの相対点
_REL_FOURTIER_SCORE = {"S": 50, "A": 40, "B": 25, "C": 15}  # 旧4段階の固定配点
_REL_TIER_DISPLAY_LABELS = {"S": "主力", "A": "一軍", "B": "二軍", "C": "三軍"}

def _rel_label_display(label):
    """相対評価の表示名。内部キー(S/A/B/C)は総合ランクと分けて保持する。"""
    if label is None:
        return "不明"
    return _REL_TIER_DISPLAY_LABELS.get(label, str(label))

def _rel_level_to_label(level):
    """層レベル(0=最上位)→tierラベル。多層ON時は層数だけS→Z、OFF時はS/A/B/Cにクランプ。"""
    if _REL_MULTI_TIER:
        return _TIER_LABELS[min(level, len(_TIER_LABELS) - 1)]
    return ["S", "A", "B", "C"][min(level, 3)]

def _rel_label_to_score(label, race_max_idx):
    """tierラベル→相対点。多層ON時はレース内のレベル位置で50〜15を線形補間。OFF時は旧4値。"""
    if not _REL_MULTI_TIER:
        return _REL_FOURTIER_SCORE.get(label)
    idx = _TIER_LABELS.find(label)
    if idx < 0:
        return None
    if race_max_idx <= 0:
        return _REL_TOP_SCORE
    return round(_REL_TOP_SCORE - (_REL_TOP_SCORE - _REL_BOTTOM_SCORE) * idx / race_max_idx, 1)


def _apply_matchup_tier_consistency(tiers, matchup_matrix, matchup_source):
    """上位グループとの採用済み比較に矛盾する二軍・三軍表示を救済する。

    比較記録がない相手は「負けていない」に含めない。direct/bridge/mixed の
    実比較だけを対象にする。上位tierへ勝っており、昇格で飛び越える各tierに
    明示的な負けがなければ、勝った相手のtierまで上げる。例えば二軍が一軍に
    勝っていても別の一軍に負けていれば、矛盾を無理に一方向へ解かず据え置く。
    従来の主力ケア（主力に非敗なら二軍/三軍に置かず、互角は一軍）も残す。

    相対点は救済後tierの既存配点（S=50/A=40/B=25/C=15）を使う。
    配点表そのものは変更しない。
    """
    corrected = dict(tiers or {})
    promotions = {}
    accepted_sources = {"direct", "bridge", "mixed"}
    tier_idx = {label: idx for idx, label in enumerate(("S", "A", "B", "C"))}

    def _symbols_against(horse, opponents):
        symbols = []
        for opponent in opponents:
            source = (matchup_source.get(horse, {}) or {}).get(opponent)
            symbol = (matchup_matrix.get(horse, {}) or {}).get(opponent)
            if source in accepted_sources and symbol in (">", "=", "<"):
                symbols.append(symbol)
        return symbols

    for horse, tier in list(corrected.items()):
        if tier not in ("B", "C"):
            continue

        original_idx = tier_idx[tier]
        target = None
        reason = None
        relations = []

        # 主力への非敗ケア。比較なしは対象外で、主力への負けが1つでもあれば発火しない。
        top_horses = {h for h, label in tiers.items() if label == "S"}
        top_symbols = _symbols_against(horse, top_horses)
        if top_symbols and "<" not in top_symbols:
            target = "S" if ">" in top_symbols else "A"
            reason = "top_nonloss"
            relations = top_symbols

        # 一軍/二軍への明示的な勝ちをtierの下限として使う。
        # 最も高い勝利先から調べ、飛び越える層への負けがあれば昇格を見送る。
        if target is None:
            for target_idx in range(original_idx):
                target_label = ("S", "A", "B", "C")[target_idx]
                target_horses = {h for h, label in tiers.items() if label == target_label}
                target_symbols = _symbols_against(horse, target_horses)
                if ">" not in target_symbols:
                    continue
                jumped_horses = {
                    h for h, label in tiers.items()
                    if target_idx <= tier_idx.get(label, 99) < original_idx
                }
                jumped_symbols = _symbols_against(horse, jumped_horses)
                if "<" in jumped_symbols:
                    continue
                target = target_label
                reason = "upper_tier_win"
                relations = jumped_symbols
                break

        if target is not None:
            corrected[horse] = target
            promotions[horse] = {
                "from": tier,
                "to": target,
                "reason": reason,
                "relations": tuple(relations),
            }
    return corrected, promotions

def _rel_tier_color(label):
    """tierラベル→表示色。S/A/B/Cは従来色、多層のD以降はレベルで青→灰へ。"""
    base = {"S": "#e74c3c", "A": "#e67e22", "B": "#f1c40f", "C": "#3498db"}
    if label in base:
        return base[label]
    return "#7f8c8d"  # D以降(多層)はグレー系

# ＝のみ馬の救済(A): 勝ち負けが付かず「＝」だけの馬でも、同場直接対決の着差で先着して
#  いる相手がいれば、その相手と同格tierまで引き上げる。S_FLOOR(確信勝ちないSをAへ)の
#  後に適用するため、＝先着でもSになり得る(ユーザー指摘 2026/06/23)。
_REL_EQUAL_LEAD_PROMOTE = True
_REL_EQUAL_LEAD_MIN = 0.0    # 先着とみなす最小着差(秒)。lead>0(わずかでも先着)なら相手tierへ
_REL_EQUAL_LEAD_LOOSE = True  # 確信度ガードを外す(bridge1票も採用 + ＝のみ馬同士の連鎖引き上げ)。⑤採用2026/06/24
_REL_C_DEMOTE = False  # 最下位C純化(過半数全敗/＝後着→C)は検証でC複勝8%→11%と精度悪化のため不採用(2026/06/24)。Trueで再検証可

# 展開予想: テン速度を今回枠と同じ内/外バケツで採用する(False で旧=通算プール+一律枠補正)。
_TEN_SPEED_DRAW_BUCKET = True
# 展開予想: テン速度に、今回と同じ内/外側での1コーナー位置・ハナ率を組み合わせる。
_PACE_FIRST_CORNER_DRAW = True
# 展開予想: 短縮/延長×逃げ経験、外枠での過去挙動(無理逃げ/番手)、内枠先手確保 を est_pos に反映。
_PACE_DRAW_INTENT = True
# 砂被り耐性: 砂を嫌う馬(談話 or 内で被ると凡走実績)は内枠でも後方でなく番手外付け、外枠なら好材料。
# 内で被っても溜められる実績馬は内枠を好位確保とみる(旧:内枠揉まれ嫌い一律+1.5の逆挙動を修正)。
_KICKBACK_LOGIC = True
# ペース推定の精緻化: 逃げ経験のある馬だけを「本当に逃げる馬」として数え、まくり質はペース加圧、
# JRA/重賞の通過は自己条件で1〜2列前に出るとみて前へ補正。
_PACE_REFINE = True
# ペースは番手の馬・騎手が作る。下手(低P)同士が前で乗ると煽って加圧、上手い(高P)騎手の番手は壊さない(沈静)。
_PACE_JOCKEY = True
_PACE_JP_HIGH = 7   # これ以上を上手い騎手(番手で沈静)
_PACE_JP_LOW = 3    # これ以下を下手な騎手(前で煽り)
# ペース加点: S/Mペース(前残り想定)では想定隊列の前5頭に+5点。
_PACE_POS_BONUS = True
_PACE_POS_BONUS_PTS = 5

# 競馬場×距離の逃げ・先行馬勝率。想定隊列の前30%(小数点以下切り捨て、最低1頭)に
# 勝率50%以上=+7、45%以上=+6、40%以上=+5、35%以上=+4、35%未満=+3を加える。
# 既存のS/Mペース前5頭+5とは重複させ、コース固有の前有利はH/M~Hでも残す。
_COURSE_FRONT_WIN_RATES = {
    ("浦和", 800, ""): 57.0,
    ("浦和", 1300, ""): 47.5,
    ("浦和", 1400, ""): 45.3,
    ("浦和", 1500, ""): 43.1,
    ("浦和", 1900, ""): 41.5,
    ("浦和", 2000, ""): 41.0,
    ("船橋", 1000, ""): 49.5,
    ("船橋", 1200, ""): 43.0,
    ("船橋", 1500, ""): 40.5,
    ("船橋", 1600, ""): 39.3,
    ("船橋", 1700, ""): 40.7,
    ("船橋", 1800, ""): 43.5,
    ("船橋", 2200, ""): 39.5,
    ("大井", 1000, ""): 45.0,
    ("大井", 1200, "内"): 38.7,
    ("大井", 1400, "内"): 36.8,
    ("大井", 1600, "内"): 37.7,
    ("大井", 1600, "外"): 32.8,
    ("大井", 1800, "外"): 34.0,
    ("大井", 2000, "外"): 31.3,
    ("大井", 2600, "外"): 32.3,
    ("川崎", 900, ""): 53.5,
    ("川崎", 1400, ""): 44.3,
    ("川崎", 1500, ""): 43.5,
    ("川崎", 1600, ""): 40.3,
    ("川崎", 2000, ""): 41.8,
    ("川崎", 2100, ""): 41.3,
}


def _course_front_bonus(course_text, distance):
    """今回条件に対応する (勝率, 加点) を返す。未登録条件は (None, 0)。"""
    text = str(course_text or "")
    place = next((p for p in ("浦和", "船橋", "大井", "川崎") if p in text), "")
    layout = "内" if "内" in text else ("外" if "外" in text else "")
    rate = _COURSE_FRONT_WIN_RATES.get((place, int(distance), layout))
    if rate is None:
        # 内外の指定がない距離はレイアウト文字が付いていても共通値を使う。
        rate = _COURSE_FRONT_WIN_RATES.get((place, int(distance), ""))
    if rate is None:
        return None, 0
    if rate >= 50.0:
        return rate, 7
    if rate >= 45.0:
        return rate, 6
    if rate >= 40.0:
        return rate, 5
    if rate >= 35.0:
        return rate, 4
    return rate, 3

# 優劣判定(＜/＞)の基準。
#  直接対決: 1400m未満=0.3秒 / 1400〜1699m=0.6秒 / 1700m以上=0.7秒。
#  隠れ馬経由: 生の2レッグ合算推定着差。全priorityで直接/既存基準より+0.2秒厳しくする。

# A/B検証用: False で隠れ馬・直接対決ともに旧固定閾値に戻す。本番True。
_REL_NEW_RULE = True
# 合議パスの新旧切替(None=_REL_NEW_RULEに従う)。
_REL_NEW_CONSENSUS = None
_BRIDGE_CONSENSUS_DEADZONE = 0.10  # 旧ロジック用(中立扱いの極小差)

# 隠れ馬経由の事前フィルタ ON/OFF(合議制があればノイズは多数決で吸収できる前提でA/B検証用)
_BRIDGE_FILTER_YARDSTICK = True     # 物差し馬フィルタ(勝ち馬から1.5秒&下位25%の隠れ馬を除外)。A/Bで残す方が僅かに良
_BRIDGE_FILTER_UV_RANK10 = False    # 比較2頭U/Vが経由レースで10着以下なら不採用。A/Bで外す方が明確に良→OFF

def _rel_direct_draw_th(cur_dist):
    """直接対決の優劣判定基準(実着差・秒)。"""
    d = _dist_to_int(cur_dist)
    if d < 1400:
        return 0.3
    if d < 1700:
        return 0.6
    return 0.7

def _bridge_priority_threshold(priority, cur_dist=None):
    """隠れ馬経由の優劣判定基準(生の2レッグ合算推定着差・秒)。
    ノイズ対策として全priorityで+0.2秒厳しくする。
    priority5(今回と同場・同距離)=直接対決の距離別基準+0.2秒。
    priority1/0.5は一律1.2秒。低優先度ほど確かな差が無いと優劣を付けない。"""
    priority = float(priority or 0)
    if priority >= 5:
        base = _rel_direct_draw_th(cur_dist)
    elif priority <= 1:
        base = 1.0
    else:
        base = 0.5 + 0.1 * max(0.0, 5 - priority)
    return round(base + 0.2, 2)

def _relative_entry_datetime(value):
    """相対評価エントリの日付を datetime に揃える。"""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        return parse_date(value)
    return datetime.min

def _relative_recency_weight(value, now=None):
    """新しい証拠を重くする緩やかな重み（0.60〜1.00）。"""
    dt = _relative_entry_datetime(value)
    if dt == datetime.min:
        return 0.70
    now = now or datetime.now()
    days = max(0, (now - dt).days)
    return max(0.60, 1.0 - min(days, 730) / 1825.0)

def _relative_source_quality(entry, current_course="", current_dist=0):
    """
    直接対決と隠れ馬bridgeを同じ品質軸で比較する。
    return: (condition_quality, source_priority)。
    ランク判定では condition_quality を主軸にし、同品質の直接対決とbridgeは同じ母集団で扱う。
    """
    if not isinstance(entry, dict):
        return (-1.0, -1)
    if "direct_priority" in entry:
        quality = float(entry.get("direct_priority", 0) or 0)
        if (
            entry.get("place") == current_course
            and _dist_to_int(entry.get("dist", "")) == _dist_to_int(current_dist)
        ):
            quality = 5.0
        return (quality, 2)
    if "bridge_priority" in entry:
        return (float(entry.get("bridge_priority", 0) or 0), 1)
    return (-1.0, 0)

def _relative_condition_weight(entry, current_course="", current_dist=0, now=None):
    """同priorityの根拠を集約する際の、条件・新しさ・場適性の重み。"""
    quality = max(0.0, _relative_source_quality(entry, current_course, current_dist)[0])
    condition_weight = 0.55 + 0.45 * min(quality, 5.0) / 5.0
    track_weight = max(0.55, min(1.0, float(entry.get("track_factor", 1.0) or 1.0)))
    surge_weight = 1.8 if entry.get("last_run_surge") else 1.0
    return condition_weight * _relative_recency_weight(entry.get("date"), now=now) * track_weight * surge_weight

def _drop_stale_direct_reversal(entries, draw_th, cur_dist):
    """直近で力関係が改善した時、古い大敗を1件だけ落とす。"""
    entries = list(entries or [])
    if len(entries) < 3:
        return entries
    ordered = sorted(entries, key=lambda e: _relative_entry_datetime(e.get("date")))
    newest = ordered[-1]
    newest_diff = float(newest.get("diff", 0) or 0)
    improve_th = 0.35 if _dist_to_int(cur_dist) <= 1600 else 0.5
    drop_id = None

    # diffは正なら本馬優勢、負なら相手優勢。
    if newest_diff >= -draw_th:
        for old in ordered[:-1]:
            old_diff = float(old.get("diff", 0) or 0)
            if old_diff <= -draw_th and (newest_diff - old_diff) >= improve_th:
                drop_id = id(old)
                break
    if drop_id is None and newest_diff <= draw_th:
        for old in ordered[:-1]:
            old_diff = float(old.get("diff", 0) or 0)
            if old_diff >= draw_th and (old_diff - newest_diff) >= improve_th:
                drop_id = id(old)
                break
    if drop_id is None:
        return entries
    kept = [e for e in entries if id(e) != drop_id]
    return kept or entries

def _aggregate_direct_entries(entries, cur_dist, current_course="", now=None):
    """
    最良条件の直接対決を加重平均+加重勝敗で集約する。
    return: (symbol, selected_entries, weighted_diff, wins, losses)
    """
    entries = list(entries or [])
    if not entries:
        return "=", [], 0.0, 0, 0
    draw_th = _rel_direct_draw_th(cur_dist)
    selected = _drop_stale_direct_reversal(entries, draw_th, cur_dist)
    selected = sorted(
        selected,
        key=lambda e: (
            -_relative_condition_weight(e, current_course, cur_dist, now=now),
            -_relative_entry_datetime(e.get("date")).timestamp()
            if _relative_entry_datetime(e.get("date")) != datetime.min else 0,
        ),
    )[:5]

    weighted_total = weighted_diff = 0.0
    pos_weight = neg_weight = 0.0
    wins = losses = 0
    for entry in selected:
        weight = _relative_condition_weight(entry, current_course, cur_dist, now=now)
        diff = float(entry.get("diff", 0) or 0)
        weighted_total += weight
        weighted_diff += diff * weight
        if diff >= draw_th:
            wins += 1
            pos_weight += weight
        elif diff <= -draw_th:
            losses += 1
            neg_weight += weight

    avg_diff = weighted_diff / weighted_total if weighted_total else float(selected[0].get("diff", 0) or 0)
    if pos_weight > neg_weight and avg_diff >= draw_th:
        symbol = ">"
    elif neg_weight > pos_weight and avg_diff <= -draw_th:
        symbol = "<"
    else:
        symbol = "="
    return symbol, selected, avg_diff, wins, losses

def _apply_strong_comparison_locks(
    tier_map,
    matchup_matrix,
    matchup_source,
    matchup_quality,
    matchup_evidence_count,
):
    """二重記号廃止前の互換用。現在は追加ロックを作らない。"""
    return dict(tier_map or {}), [], {}

    # 以下は旧キャッシュ解析時のみに備えた到達不能コード。
    adjusted = dict(tier_map or {})
    rank_order = {"S": 0, "A": 1, "B": 2, "C": 3}
    pos_to_rank = {v: k for k, v in rank_order.items()}
    horses = list((matchup_matrix or {}).keys())

    relation_g = nx.DiGraph()
    relation_g.add_nodes_from(horses)
    for u in horses:
        for v, sym in (matchup_matrix.get(u, {}) or {}).items():
            if sym != ">":
                continue
            source = (matchup_source.get(u, {}) or {}).get(v)
            quality = float((matchup_quality.get(u, {}) or {}).get(v, -1) or -1)
            count = int((matchup_evidence_count.get(u, {}) or {}).get(v, 0) or 0)
            # 反証経路も低品質の単1本bridgeでは作らない。
            if source == "direct" and quality >= 2:
                relation_g.add_edge(u, v)
            elif source == "bridge" and quality >= 3 and count >= 2:
                relation_g.add_edge(u, v)

    locks = []
    seen_pairs = set()
    for u in horses:
        for v, sym in (matchup_matrix.get(u, {}) or {}).items():
            if sym != ">":
                continue
            pair_key = tuple(sorted((u, v)))
            if pair_key in seen_pairs:
                continue
            seen_pairs.add(pair_key)
            source = (matchup_source.get(u, {}) or {}).get(v)
            quality = float((matchup_quality.get(u, {}) or {}).get(v, -1) or -1)
            count = int((matchup_evidence_count.get(u, {}) or {}).get(v, 0) or 0)
            eligible = (
                (source == "direct" and quality >= 3)
                or (source == "bridge" and quality >= 4 and count >= 2)
            )
            if not eligible:
                continue

            counter_g = relation_g.copy()
            if counter_g.has_edge(u, v):
                counter_g.remove_edge(u, v)
            if counter_g.has_edge(v, u):
                counter_g.remove_edge(v, u)
            try:
                counter_path = nx.shortest_path(counter_g, v, u)
            except (nx.NetworkXNoPath, nx.NodeNotFound):
                counter_path = []
            locks.append({
                "winner": u,
                "loser": v,
                "source": source,
                "quality": quality,
                "evidence_count": count,
                "enforce": not bool(counter_path),
                "counter_path": counter_path,
            })

    notes = {}
    for _ in range(len(locks) + 2):
        changed = False
        for lock in locks:
            winner, loser = lock["winner"], lock["loser"]
            if not lock["enforce"]:
                notes.setdefault(loser, "※完敗比較あり。別ルートに相手以上を示す反証があるため降格保留。")
                continue
            winner_rank = adjusted.get(winner)
            loser_rank = adjusted.get(loser)
            if winner_rank not in rank_order or loser_rank not in rank_order:
                continue
            required_pos = min(rank_order[winner_rank] + 1, rank_order["C"])
            if rank_order[loser_rank] < required_pos:
                adjusted[loser] = pos_to_rank[required_pos]
                notes[loser] = "※高品質の完敗比較により、勝った馬より1段下へ調整。"
                changed = True
        if not changed:
            break
    return adjusted, locks, notes

def _bridge_consensus_symbol(entries, cur_dist):
    """ペアを繋ぐ複数の隠れ馬経由比較を、優先度で重み付けした多数決で集約する。
    各エントリの「生の2レッグ合算推定着差(raw_diff, u視点, 正なら u 優勢)」を、
    そのエントリの priority に応じた閾値(priority5=直接対決基準+0.2秒, priority1/0.5=1.2秒)で勝/敗/分に振り分け、
    最良priorityを主票、1段下を0.6倍の補助票として重み付けする。
    同じ隠れ馬でも複数レース/複数経路があれば、それぞれを精度向上の票として扱う。
    1件でも着差基準を満たし、同方向の重み比率が十分なら >/< 。
    最良priorityだけで決まった方向を1段下の票だけで逆転させず、強く競合する場合は = とする。
    着差基準を満たす票が無ければ = 。記号は > / < / = の3種類だけを返す。
    返り値: (記号, 本馬優勢票数, 相手優勢票数)
    """
    valid = [
        e for e in (entries or [])
        if isinstance(e.get("diff"), (int, float)) and (e.get("bridge_priority", 0) or 0) > 0
    ]
    if not valid:
        return "=", 0, 0

    # 最良priorityを主票とし、設定時は1段下まで補助票として採用する。
    best_priority = max(float(e.get("bridge_priority", 0) or 0) for e in valid)
    allowed_priorities = {best_priority}
    if _BRIDGE_CONSENSUS_INCLUDE_ONE_LOWER and best_priority - 1 >= 1:
        allowed_priorities.add(best_priority - 1)
    valid = [
        e for e in valid
        if float(e.get("bridge_priority", 0) or 0) in allowed_priorities
    ]

    pos_w = neg_w = 0.0
    pos_n = neg_n = 0
    pos_strong = neg_strong = 0
    _use_new = _REL_NEW_RULE if _REL_NEW_CONSENSUS is None else _REL_NEW_CONSENSUS
    if not _use_new:
        # 旧ロジック: deadzoneで割引後diffの符号投票。
        dz = _BRIDGE_CONSENSUS_DEADZONE
        for e in valid:
            d = e["diff"]; w = (e.get("bridge_priority", 0) or 0) * float(e.get("consensus_weight_factor", 1.0) or 1.0)
            if d > dz:
                pos_w += w; pos_n += 1
            elif d < -dz:
                neg_w += w; neg_n += 1
        n = pos_n + neg_n
        tot = pos_w + neg_w
        if n < _BRIDGE_CONSENSUS_MIN_VOTES or tot <= 0:
            return "=", pos_n, neg_n
        if pos_w > neg_w and pos_w / tot >= _BRIDGE_CONSENSUS_AGREE:
            return ">", pos_n, neg_n
        if neg_w > pos_w and neg_w / tot >= _BRIDGE_CONSENSUS_AGREE:
            return "<", pos_n, neg_n
        return "=", pos_n, neg_n

    # 新ロジック: 採用された各レース/経路をそのまま票にする。
    # 同じ隠れ馬経由でも、何度も近い条件で比較できていれば精度向上の材料として数える。
    votes = []
    for e in valid:
        entry_priority = float(e.get("bridge_priority", 0) or 0)
        entry_weight = _relative_condition_weight(e, "", cur_dist)
        if entry_weight <= 0:
            continue
        d = float(e.get("raw_diff", e.get("diff", 0)) or 0)
        level_factor = 1.0 if entry_priority == best_priority else _BRIDGE_CONSENSUS_LOWER_WEIGHT
        w = max(0.10, entry_weight * level_factor)
        th = _bridge_priority_threshold(entry_priority, cur_dist)
        if d >= th:
            votes.append({"direction": 1, "weight": w, "priority": entry_priority})
        elif d <= -th:
            votes.append({"direction": -1, "weight": w, "priority": entry_priority})

    def _vote_direction(selected_votes):
        positive = sum(v["weight"] for v in selected_votes if v["direction"] > 0)
        negative = sum(v["weight"] for v in selected_votes if v["direction"] < 0)
        total = positive + negative
        if total <= 0:
            return "="
        if positive > negative and positive / total >= _BRIDGE_CONSENSUS_AGREE:
            return ">"
        if negative > positive and negative / total >= _BRIDGE_CONSENSUS_AGREE:
            return "<"
        return "="

    best_votes = [v for v in votes if v["priority"] == best_priority]
    best_direction = _vote_direction(best_votes)
    combined_direction = _vote_direction(votes)

    # 1段下は支持・補完には使うが、最良priorityだけで出た結論を逆方向へは覆さない。
    # 競合が強く、合算で元の方向の70%を維持できなければ「＝」へ慎重化する。
    if best_direction in (">", "<"):
        final_direction = best_direction if combined_direction == best_direction else "="
    else:
        final_direction = combined_direction

    pos_w = sum(v["weight"] for v in votes if v["direction"] > 0)
    neg_w = sum(v["weight"] for v in votes if v["direction"] < 0)
    pos_n = sum(1 for v in votes if v["direction"] > 0)
    neg_n = sum(1 for v in votes if v["direction"] < 0)
    n = pos_n + neg_n
    tot = pos_w + neg_w
    if n < _BRIDGE_CONSENSUS_MIN_VOTES or tot <= 0:
        return "=", pos_n, neg_n
    if final_direction == ">":
        return ">", pos_n, neg_n
    if final_direction == "<":
        return "<", pos_n, neg_n
    return "=", pos_n, neg_n

def _rank_floor_e(rank):
    """最終表示用ランクはEを下限に丸める。"""
    return "E" if rank in ("F", "G") else rank

_DISPLAY_MARKERS = ("[[RB]]", "[[/RB]]", "[[B]]", "[[/B]]", "[[SMALL]]", "[[/SMALL]]")

def strip_display_markers(text):
    out = str(text or "")
    for marker in _DISPLAY_MARKERS:
        out = out.replace(marker, "")
    return out

def _format_small_text_markers(escaped_text):
    return (
        str(escaped_text or "")
        .replace("[[SMALL]]", '<span class="pace-stable-note">')
        .replace("[[/SMALL]]", "</span>")
    )

def _style_from_pas(pas):
    nums = [int(x) for x in re.findall(r'\d+', str(pas or ''))]
    if not nums:
        return "位置取り不明"

    first = nums[0]
    last = nums[-1]
    if first <= 1:
        return "逃げ"
    if first <= 4:
        return "番手"
    if first >= 5 and last <= 4 and (first - last) >= 3:
        return "まくり"
    if first <= 8:
        return "差し"
    return "追い込み"

def _frame_group(frame):
    try:
        f = int(frame)
    except (TypeError, ValueError):
        return "枠不明"
    if f <= 2:
        return "内枠"
    if f <= 5:
        return "中枠"
    return "外枠"

def _gate_group(field_size, gate_no):
    try:
        n = int(field_size)
        g = int(gate_no)
    except (TypeError, ValueError):
        return "位置不明"
    if n <= 0 or g <= 0:
        return "位置不明"

    pos = (g - 0.5) / n
    if pos <= 1 / 3:
        return "内目"
    if pos >= 2 / 3:
        return "外目"
    return "中目"

def _draw_side(field_size, gate_no):
    """テン速度の内外バケツ用に、枠を「内/外」へざっくり2分割する。
    今回枠と過去走枠を同じ基準で分け、今回枠に対応するバケツのテン速度を採用する。"""
    try:
        n = int(field_size)
        g = int(gate_no)
    except (TypeError, ValueError):
        return None
    if n <= 0 or g <= 0:
        return None
    return "内" if g <= (n + 1) / 2 else "外"

def _select_cycle_draw_entries(entries, current_field_size, current_u_gate, current_v_gate):
    """三すくみ解消用。今回と過去の内/外サイドが最も多く一致する直接対決を選ぶ。

    両馬の枠情報があれば2馬分を比較し、片方しか取れない場合も
    判明した馬のサイドだけで判定する。一致を0件しか確認できない時は
    根拠を捕てず、従来の集約へ戻す。
    """
    entries = list(entries or [])
    current_u_side = _draw_side(current_field_size, current_u_gate)
    current_v_side = _draw_side(current_field_size, current_v_gate)
    if not entries or (current_u_side is None and current_v_side is None):
        return entries

    scored = []
    for entry in entries:
        past_field_size = entry.get("field_size")
        comparisons = []
        past_u_side = _draw_side(past_field_size, entry.get("u_gate"))
        past_v_side = _draw_side(past_field_size, entry.get("v_gate"))
        if current_u_side is not None and past_u_side is not None:
            comparisons.append(current_u_side == past_u_side)
        if current_v_side is not None and past_v_side is not None:
            comparisons.append(current_v_side == past_v_side)
        if comparisons:
            scored.append((sum(comparisons), len(comparisons), entry))

    if not scored:
        return entries
    best_matches = max(matches for matches, _, _ in scored)
    if best_matches <= 0:
        return entries
    best_known = max(known for matches, known, _ in scored if matches == best_matches)
    selected = [
        entry for matches, known, entry in scored
        if matches == best_matches and known == best_known
    ]
    return selected or entries

def _first_corner_draw_profile(hist, today_side, current_place="", current_dist=0):
    """今回と同じ内/外側での1コーナー位置傾向を返す。

    近走を3/2/1.5/1/1倍、同場同距離を1.5倍で重み付けする。同じ枠側の
    サンプルが少ない時は、全走の傾向を2走分の事前値として縮約し、
    「2走未満なら通算へ全戻し」にしない。
    """
    runs = []
    recency_weights = (3.0, 2.0, 1.5, 1.0, 1.0)
    cur_dist = _dist_to_int(current_dist)
    for idx, run in enumerate((hist or [])[:5]):
        if not isinstance(run, dict):
            continue
        positions = [int(x) for x in re.findall(r'\d+', str(run.get("pas") or ""))]
        if not positions:
            continue
        first_pos = positions[0]
        side = _draw_side(run.get("field_size"), run.get("gate_no"))
        weight = recency_weights[min(idx, len(recency_weights) - 1)]
        run_dist = _dist_to_int(run.get("dist", ""))
        if (
            current_place
            and run.get("place") == current_place
            and run_dist == cur_dist
        ):
            weight *= 1.5
        elif run_dist > 0 and cur_dist > 0 and abs(run_dist - cur_dist) > 300:
            # 800mの最初の通過順を1400/1500mの1Cと同価値にしない。
            weight *= 0.6
        runs.append({
            "first_pos": first_pos,
            "side": side,
            "weight": weight,
            "exact": bool(
                current_place
                and run.get("place") == current_place
                and run_dist == cur_dist
            ),
        })

    if not runs:
        return {
            "avg_first_pos": None, "lead_rate": 0.0, "front2_rate": 0.0,
            "side_count": 0, "exact_side_leads": 0, "exact_any_leads": 0,
        }

    def _weighted_metrics(items):
        total = sum(x["weight"] for x in items)
        if total <= 0:
            return None, 0.0, 0.0
        avg = sum(x["first_pos"] * x["weight"] for x in items) / total
        lead = sum(x["weight"] for x in items if x["first_pos"] == 1) / total
        front2 = sum(x["weight"] for x in items if x["first_pos"] <= 2) / total
        return avg, lead, front2

    pooled_avg, pooled_lead, pooled_front2 = _weighted_metrics(runs)
    side_runs = [x for x in runs if today_side is not None and x["side"] == today_side]
    if side_runs:
        side_avg, side_lead, side_front2 = _weighted_metrics(side_runs)
        alpha = len(side_runs) / (len(side_runs) + 2.0)
        avg_first_pos = side_avg * alpha + pooled_avg * (1.0 - alpha)
        lead_rate = side_lead * alpha + pooled_lead * (1.0 - alpha)
        front2_rate = side_front2 * alpha + pooled_front2 * (1.0 - alpha)
    else:
        avg_first_pos, lead_rate, front2_rate = pooled_avg, pooled_lead, pooled_front2

    return {
        "avg_first_pos": avg_first_pos,
        "lead_rate": lead_rate,
        "front2_rate": front2_rate,
        "side_count": len(side_runs),
        "exact_side_leads": sum(1 for x in side_runs if x["exact"] and x["first_pos"] == 1),
        "exact_any_leads": sum(1 for x in runs if x["exact"] and x["first_pos"] == 1),
    }

def _draw_label_and_group(run):
    frame = run.get("frame") or ""
    if frame:
        return f"{frame}枠", _frame_group(frame)

    field_size = run.get("field_size")
    gate_no = run.get("gate_no")
    if field_size and gate_no:
        group = _gate_group(field_size, gate_no)
        return f"{field_size}頭{gate_no}番({group})", group

    return "位置不明", "位置不明"

def _draw_short_label(run):
    field_size = run.get("field_size")
    gate_no = run.get("gate_no")
    if field_size and gate_no:
        return f"{_to_circled(gate_no)}/{field_size}頭"

    frame = run.get("frame") or ""
    if frame:
        return f"{frame}枠"

    return "位置不明"

def _draw_detail_label(run):
    field_size = run.get("field_size")
    gate_no = run.get("gate_no")
    if field_size and gate_no:
        return f"{gate_no}番({field_size})"

    frame = run.get("frame") or ""
    if frame:
        return f"{frame}枠"

    return "位置不明"

def _extract_frame_no(z, z_full_text=""):
    """近走セルから枠番をできるだけ拾う。HTML改修に備えてテキスト/属性/classを横断する。"""
    m = re.search(r'([1-8])\s*枠', z_full_text or "")
    if m:
        return m.group(1)

    for el in [z] + list(z.find_all(True)):
        attrs = " ".join(
            str(v)
            for k, v in el.attrs.items()
            if k in ("class", "title", "alt", "aria-label")
        )
        if isinstance(el.get("class"), list):
            attrs += " " + " ".join(el.get("class", []))
        m = re.search(r'(?:waku|frame|枠)[-_ ]?0?([1-8])(?:\b|[^0-9])', attrs, re.I)
        if m:
            return m.group(1)
        short_text = el.get_text(" ", strip=True)
        if len(short_text) <= 12:
            m = re.search(r'([1-8])\s*枠', short_text)
            if m:
                return m.group(1)
    return ""

def _extract_field_gate(z_full_text=""):
    """近走セルの「14頭 4番」から頭数と馬番を拾う。馬番は頭数比で内/中/外に相対化する。"""
    text = re.sub(r'\s+', ' ', str(z_full_text or ""))
    m = re.search(r'(\d{1,2})\s*頭\s*(\d{1,2})\s*番', text)
    if not m:
        return None, None
    field_size, gate_no = int(m.group(1)), int(m.group(2))
    if 1 <= gate_no <= field_size <= 18:
        return field_size, gate_no
    return None, None

def build_race_content_summary(history):
    if not history:
        return "【近走】過去戦績なし。"

    notable = []
    positives = []
    negatives = []

    for run in history or []:
        if not isinstance(run, dict):
            continue
        rank = run.get("rank")
        pop = run.get("pop")
        is_positive = rank is not None and (rank <= 2 or (pop is not None and pop - rank >= 5))
        is_negative = rank is not None and pop is not None and pop <= 3 and rank >= 6
        if not (is_positive or is_negative):
            continue

        style = _style_from_pas(run.get("pas", ""))
        draw_txt, draw_grp = _draw_label_and_group(run)
        draw_short = _draw_short_label(run)
        draw_detail = _draw_detail_label(run)
        rank_txt = f"{rank}着" if rank is not None else "着不明"
        pop_txt = f"{pop}人" if pop is not None else "人気不明"
        pas_txt = str(run.get("pas", "") or "").strip()
        pas_part = f"({pas_txt})" if pas_txt else ""
        item = {
            "style": style,
            "draw_txt": draw_txt,
            "draw_short": draw_short,
            "draw_detail": draw_detail,
            "draw_grp": draw_grp,
            "rank": rank,
            "pop": pop,
            "text": f"{draw_detail}で{style}{pas_part}で{rank_txt}({pop_txt})"
        }
        notable.append(item)
        if is_positive:
            positives.append(item)
        if is_negative:
            negatives.append(item)

    if not notable:
        return "【近走】近5走では、連対または人気を大きく上回る好走/上位人気での大敗パターンは目立たない。"

    def _uniq(seq):
        seen = set()
        out = []
        for x in seq:
            if x in seen:
                continue
            seen.add(x)
            out.append(x)
        return out

    parts = []
    if positives:
        styles = _uniq([p["style"] for p in positives])
        pos_desc = "、".join(p["text"] for p in positives[:3])
        if len(styles) == 1:
            style_intro = f"{styles[0]}。"
        elif len(styles) == 2:
            style_intro = f"{styles[0]}か{styles[1]}。"
        else:
            style_intro = f"{'・'.join(styles)}。"
        if len(positives) == 1:
            p = positives[0]
            parts.append(f"【近走】{p['style']}。好走形は{p['text']}。{p['draw_grp']}で{p['style']}を取れる形が鍵。")
        else:
            parts.append(f"【近走】{style_intro}好走形は{pos_desc}。")
    else:
        parts.append("【近走】近5走に連対または人気を5つ以上上回る明確な好走形はない。")

    if negatives:
        neg_desc = "、".join(n["text"] for n in negatives[:2])
        parts.append(f"一方、上位人気で崩れた形は{neg_desc}。同じ枠/位置取りの再現は割引。")

    if positives:
        positive_styles = {p["style"] for p in positives}
        positive_draw_groups = {p["draw_grp"] for p in positives}
        if len(positive_styles) == 1:
            parts.append(f"勝ち負けには{next(iter(positive_styles))}の再現が重要。")
        if len(positive_draw_groups) == 1 and "位置不明" not in positive_draw_groups:
            parts.append(f"{next(iter(positive_draw_groups))}で形を作れた時に信頼しやすい。")

    return "".join(parts)

def ensure_race_content_in_ai_detail(ai_text, horses_data):
    if not ai_text or "🐎レース内容" in ai_text:
        return ai_text

    sep = "----------------------------------------"
    blocks = ai_text.split(sep)
    out = []
    for block in blocks:
        m = re.search(r'^\s*(?:([\u2460-\u2473])|\[(\d{1,2})\])', block)
        umaban = ""
        if m:
            if m.group(1):
                umaban = str(ord(m.group(1)) - 0x245f)
            elif m.group(2):
                umaban = m.group(2)
        if umaban and umaban in horses_data and "【調教】" in block:
            summary = build_race_content_summary(horses_data[umaban].get("hist", []))
            block = block.rstrip() + f"\n🐎レース内容\n{summary}\n"
        out.append(block)
    return sep.join(out).strip()

# ==================================================
# 1. 設定 & 定数
# ==================================================
DATA_DIR = "2025data"
JOCKEY_FILE = os.path.join(DATA_DIR, "2025_NARJockey.csv")
TRAINER_FILE = os.path.join(DATA_DIR, "2025_NankanTrainer.csv")
POWER_FILE = os.path.join(DATA_DIR, "2025_騎手パワー.xlsx")
SCORE_POLICY_VERSION = "v20260714g_speed70_record100_exact_track_surface"

# 当日各コース横断の調教上位(赤字)への加点。
# 今走調教(中間追い切り含む)だけを対象に、同日・同じ調教場所・同じメトリクスで10頭以上いる場合のみ、3位以内を対象。
# 配点は順位別: 1位+5 / 2位+3 / 3位+2。複数本の追い切りが対象になっても、
# その馬の最も高い全体順位1件だけを採用する。同タイムは同順位=同点。
CHOKYO_TOP_POINTS = {1: 5, 2: 3, 3: 2}
CHOKYO_RACE_POINTS = {1: 3, 2: 2, 3: 1}
CHOKYO_RANK_BONUS_CAP = 8
CHOKYO_TOP_BONUS_CAP = CHOKYO_RANK_BONUS_CAP  # 旧名互換
CHOKYO_TOP_BONUS = 5  # 1位の点(互換用)
CHOKYO_TOP_MIN_GROUP_SIZE = 10
CHOKYO_RANK_DISPLAY_MIN_GROUP_SIZE = 3
CHOKYO_RACE_MIN_HORSES = 5

# レース内順位は異なる調教場の生時計を直接比べず、コース・メトリクス別の
# 基準タイムに対する差（負荷補正後）で比較する。3Fは従来の基準超判定と同じ基準。
CHOKYO_STANDARD_3F = {
    '大井': 37.5, '川崎': 37.5, '船橋': 37.0, '小林': 37.0, '浦和': 39.0,
    '小林坂': 37.0, '園田': 37.5, '美W': 37.0, '栗CW': 37.0, 'CW': 37.0,
}
# 1Fは保存済み調教の主な分布（美W/栗CW）に合わせた終い基準。
CHOKYO_STANDARD_1F = {
    '大井': 12.5, '川崎': 12.5, '船橋': 12.5, '小林': 12.5, '浦和': 12.8,
    '小林坂': 12.5, '園田': 12.5, '美W': 12.1, '栗CW': 12.3, 'CW': 12.3,
}

# ==================================================
# secrets 読み込み
# ==================================================
try:
    KEIBA_ID = st.secrets["keibabook"]["login_id"]
except Exception:
    KEIBA_ID = ""

try:
    KEIBA_PASS = st.secrets["keibabook"]["password"]
except Exception:
    KEIBA_PASS = ""


# ==================================================
# 0. 名前正規化
# ==================================================
def normalize_name(abbrev, full_list, priority_set=None):
    if not abbrev:
        return ""

    clean = re.sub(r"[ 　▲△☆◇★\d\.]+", "", str(abbrev)).strip()
    if not clean:
        return ""

    if not full_list:
        return clean

    if clean in full_list:
        return clean

    candidates = []
    for full in full_list:
        if clean in full:
            diff = len(full) - len(clean)
            is_priority = 1 if (priority_set and full in priority_set) else 0
            candidates.append((0, -is_priority, diff, full))
        elif all(c in full for c in clean):
            diff = len(full) - len(clean)
            is_priority = 1 if (priority_set and full in priority_set) else 0
            candidates.append((1, -is_priority, diff, full))

    if candidates:
        candidates.sort(key=lambda x: (x[0], x[1], x[2]))
        return candidates[0][3]

    return clean


def lookup_power(place_name, jockey_name, power_data):
    """競馬場一致＋騎手名（完全一致→苗字＋名前の一部文字一致）でpower_dataを検索"""
    if not place_name or not jockey_name:
        return None
    clean = re.sub(r"[ 　▲△☆◇★\d\.]+", "", str(jockey_name)).strip()
    if not clean:
        return None
    # 1. 完全一致
    result = power_data.get((place_name, clean))
    if result:
        return result
    # 2. 部分一致: 同じ競馬場で名前が含まれる、または全文字が含まれる
    for (venue, name), data in power_data.items():
        if venue != place_name:
            continue
        if clean in name or name in clean:
            return data
        if all(c in name for c in clean):
            return data
    return None


# ==================================================
# 2. 共通関数
# ==================================================
@st.cache_resource
def get_http_session() -> requests.Session:
    sess = requests.Session()
    sess.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
    })
    retry = Retry(total=3, backoff_factor=1, status_forcelist=(500, 502, 503, 504))
    adapter = HTTPAdapter(max_retries=retry)
    sess.mount("https://", adapter)
    sess.mount("http://", adapter)
    return sess


def get_driver():
    ops = Options()
    ops.add_argument("--headless=new")
    ops.add_argument("--no-sandbox")
    ops.add_argument("--disable-dev-shm-usage")
    ops.add_argument("--disable-gpu")
    ops.add_argument("--window-size=1280,2200")
    ops.page_load_strategy = 'eager'
    ops.add_argument("--lang=ja-JP")
    ops.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(options=ops)
    driver.set_page_load_timeout(30)
    return driver

def login_keibabook_robust(driver) -> bool:
    if not KEIBA_ID or not KEIBA_PASS:
        print("❌ [エラー] ログインIDかパスワードがsecretsに設定されていません。")
        return False

    driver.get("https://s.keibabook.co.jp/login/login")
    
    try:
        WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.NAME, "login_id"))
        ).send_keys(KEIBA_ID)
        
        WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "input[type='password']"))
        ).send_keys(KEIBA_PASS)
        
        WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='submit'], .btn-login"))
        ).click()
        
        time.sleep(2.0)
        
        if driver.find_elements(By.CSS_SELECTOR, "a.logout") or driver.find_elements(By.CSS_SELECTOR, "a.mypage"):
            print("✅ ログインに成功しました")
            return True
            
        print("⚠️ ログイン後画面の確認はできませんでしたが、処理を継続します")
        return True
        
    except Exception as e:
        print(f"❌ ログイン処理中にエラー発生: {e}")
        return False


# ==================================================
# 3. 相対評価（直接対決＋隠れ馬）
# ==================================================
_RACE_RESULT_CACHE = {}
# 結果ページの 馬名→uma_id（メモ連携用。戻り値タプルは変えずに副作用で保持）
_RESULT_UMA_IDS = {}


def result_uma_ids(url):
    """直近 fetch_race_all_horses(url) で取得した 馬名→uma_id を返す。"""
    return _RESULT_UMA_IDS.get(url, {})


def fetch_race_all_horses(url):
    if url in _RACE_RESULT_CACHE:
        return _RACE_RESULT_CACHE[url]

    sess = get_http_session()
    try:
        time.sleep(0.5)
        res = sess.get(url, timeout=8)
        res.encoding = "cp932"
        soup = BeautifulSoup(res.text, "html.parser")

        tbl = soup.find("table", class_=re.compile(r"nk23_c-table01"))
        if not tbl:
            for t in soup.find_all("table"):
                if t.find("a", href=re.compile(r"/uma_info/")):
                    tbl = t
                    break
        if not tbl:
            result = ({}, "", "", {}, {}, {})
            _RACE_RESULT_CACHE[url] = result
            return result

        race_place, race_dist = "", ""
        page_text = soup.get_text(" ", strip=True)
        for kp in ("大井", "川崎", "船橋", "浦和", "門別", "盛岡", "水沢", "金沢", "笠松", "名古屋", "園田", "姫路", "高知", "佐賀"):
            if kp in page_text[:800]:
                race_place = kp
                break
        dm = re.search(r"(\d{3,4})\s*[mｍ]", page_text[:800])
        if dm:
            race_dist = dm.group(1)

        headers = [th.get_text(strip=True) for th in tbl.find_all("th")]
        pop_idx = next((i for i, h in enumerate(headers) if "人気" in h), None)
        gate_idx = next((i for i, h in enumerate(headers) if "馬番" in h), None)

        horses = {}
        ranks = {}
        popularities = {}
        gates = {}
        uma_ids = {}
        winner_sec = None
        for tr in tbl.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) < 4:
                continue

            name_link = tr.find("a", href=re.compile(r"/uma_info/"))
            if not name_link:
                continue
            horse_name = name_link.get_text(strip=True)
            _uid_m = re.search(r"/uma_info/(\d+)", name_link.get("href") or "")
            if _uid_m:
                uma_ids[horse_name] = _uid_m.group(1)
            rank_txt = tds[0].get_text(strip=True) if tds else ""
            rank_m = re.search(r"\d+", rank_txt)
            rank_val = int(rank_m.group(0)) if rank_m else None
            gate_val = None
            gate_cell = tr.select_one("td.umaban")
            if gate_cell:
                gate_m = re.search(r"\d+", gate_cell.get_text(strip=True))
                if gate_m:
                    gate_val = int(gate_m.group(0))
            if gate_val is None and gate_idx is not None and len(tds) > gate_idx:
                gate_m = re.search(r"\d+", tds[gate_idx].get_text(strip=True))
                if gate_m:
                    gate_val = int(gate_m.group(0))
            if gate_val is None:
                # 公式結果表は多くの場合「着順 / 枠 / 馬番 / 馬名...」の順。
                for td in tds[1:4]:
                    classes = " ".join(td.get("class", []))
                    if "waku" in classes:
                        continue
                    gate_m = re.fullmatch(r"\d{1,2}", td.get_text(strip=True))
                    if gate_m:
                        gate_val = int(gate_m.group(0))
                        break
            pop_val = None
            if pop_idx is not None and len(tds) > pop_idx:
                pop_m = re.search(r"\d+", tds[pop_idx].get_text(strip=True))
                if pop_m:
                    pop_val = int(pop_m.group(0))
            if pop_val is None:
                for td in tds:
                    pop_m = re.search(r"(\d+)\s*(?:人気|人)$", td.get_text(strip=True))
                    if pop_m:
                        pop_val = int(pop_m.group(1))
                        break

            time_sec = None
            for td in tds:
                txt = td.get_text(strip=True)
                tm = re.match(r"^(\d{1,2}):(\d{2}\.\d)$", txt)
                if tm:
                    time_sec = int(tm.group(1)) * 60 + float(tm.group(2))
                    break
            if time_sec is None and len(tds) > 10:
                txt10 = tds[10].get_text(strip=True)
                tm2 = re.match(r"^(\d{2,3}\.\d)$", txt10)
                if tm2:
                    candidate = float(tm2.group(1))
                    if 30.0 <= candidate <= 120.0:
                        time_sec = candidate
            if time_sec is None:
                continue

            if winner_sec is None:
                winner_sec = time_sec

            horses[horse_name] = round(time_sec - winner_sec, 1)
            if rank_val is not None:
                ranks[horse_name] = rank_val
            if pop_val is not None:
                popularities[horse_name] = pop_val
            if gate_val is not None:
                gates[horse_name] = gate_val

        result = (horses, race_place, race_dist, ranks, popularities, gates)
        _RACE_RESULT_CACHE[url] = result
        _RESULT_UMA_IDS[url] = uma_ids
        return result
    except Exception as e:
        print(f"  [相対評価] スクレイピングエラー: {url}: {e}")
        return {}, "", "", {}, {}, {}

def _is_nige(pas_str):
    """通過順位の最初が1なら逃げと判定"""
    if not pas_str:
        return False
    first = pas_str.split("-")[0]
    return first.strip() == "1"

def _pick_3f_time(times):
    """調教欄の数値列からラスト3Fタイムを選ぶ。10〜16秒台の最終1Fは除外する。"""
    vals = []
    for t in times or []:
        try:
            v = float(t)
        except (TypeError, ValueError):
            continue
        if v >= 33.0:
            vals.append(v)
    return vals[-1] if vals else None

def _extract_training_course(text):
    m = re.search(r'(小林坂|小林|園田|美[ＷW]|栗[ＣC][ＷW]|ＣＷ|CW|浦和|川崎|大井|船橋)', str(text or ""))
    if not m:
        return "", False
    course = m.group(1).replace('Ｗ', 'W').replace('Ｃ', 'C')
    tail = str(text or "")[m.end():m.end() + 8]
    is_main_track = course in {"浦和", "川崎", "大井", "船橋"} and bool(re.match(r'\s*(?:本馬場|本)', tail))
    return course, is_main_track

def _chokyo_load_penalty(line):
    """負荷・併せによる『タイムの出やすさ』を割り引くための横比較用ペナルティ秒。
    併せ馬 or 一杯 → +0.2秒、強め → +0.1秒、馬なり(単走) → 0。
    本人の負荷は時計より前に書かれ、' / '以降は併せ相手の情報なので、
    負荷語(一杯/強め)は ' / ' より前の本人区間だけで判定する。
    併せ相手の区切りは ' / '(前後スペース付き)。日付の '6/26' はスペースが無いので除外される。
    ' / ' があれば併せ馬ありとみなす(単走には付かない)。
    返す値は比較用にのみ加算し、表示タイムには反映しない。"""
    s = str(line or "")
    self_part = s.split(' / ', 1)[0]   # 本人の負荷区間（時計を含む。日付スラッシュは無視）
    has_awase = ' / ' in s             # 併せ相手の記載 = 併せ馬
    if has_awase or ('一杯' in self_part):
        return 0.2
    if '強め' in self_part:
        return 0.1
    return 0.0


def _training_judge_time(t_3f, is_main_track):
    try:
        t = float(t_3f)
    except (TypeError, ValueError):
        return None
    return t + 1.0 if is_main_track else t

def _training_threshold(table, course, is_main_track):
    threshold = table.get(course, 99.9)
    if course in {"浦和", "川崎", "大井", "船橋"} and is_main_track:
        threshold -= 1.0
    return threshold


def _compute_day_chokyo_red(entries):
    """当日生成した全レースを横断し、同一(日付・調教コース・メトリクス)グループ内で
    負荷補正タイム(adj_time)が速い順に順位付け。同一グループ10頭以上の時だけ対象。
    配点は 1位+5/2位+3/3位+2。同タイム(adj_time一致)は同順位=同点(競争順位: 自分より速い頭数+1)。
    1頭が複数本の追い切りで複数グループの上位に入っても、最も高い点の1件だけを採用する。
    entries: [{r_num, umaban, key, time, adj_time, disp, line}] / key=(date,(course,is_main),metric)
    Returns: {(r_num, umaban): {"rank": best_rank, "total": group_size, "line", "disp",
                                "time", "metric", "points": 全体順位点(最大5)}}"""
    groups = {}
    for e in entries or []:
        if e.get("key") is None or e.get("time") is None:
            continue
        groups.setdefault(e["key"], []).append(e)

    def _adj(e):
        return float(e.get("adj_time", e["time"]))

    # 同一グループに同じ馬の同メトリクスが複数ある場合は、最速の1本だけに集約する。
    horse_candidates = {}  # (r_num, umaban) -> [(points, info)]
    for key, grp in groups.items():
        deduped = {}
        for e in grp:
            horse_key = (int(e["r_num"]), str(e["umaban"]))
            prev = deduped.get(horse_key)
            if prev is None or _adj(e) < _adj(prev):
                deduped[horse_key] = e
        grp = list(deduped.values())
        group_size = len(grp)
        # 「全体」は少なくとも2レースを横断した時だけ成立させる。
        # 1レースだけの生成では、後段のレース内TOP3だけを適用する。
        race_count = len({int(e["r_num"]) for e in grp})
        if group_size < CHOKYO_TOP_MIN_GROUP_SIZE or race_count < 2:
            continue
        for e in grp:
            # 競争順位(同タイム=同順位): 自分より adj_time が小さい(速い)頭数 + 1。
            rank = 1 + sum(1 for o in grp if _adj(o) < _adj(e))
            pts = CHOKYO_TOP_POINTS.get(rank, 0)
            if pts <= 0:
                continue
            k = (int(e["r_num"]), str(e["umaban"]))
            info = {
                "rank": rank,
                "total": group_size,
                "line": e.get("line"),
                "disp": e.get("disp"),
                "time": e.get("time"),
                "metric": e.get("metric"),
            }
            horse_candidates.setdefault(k, []).append((pts, info))

    red = {}
    for k, candidates in horse_candidates.items():
        # 点が高い順、同点なら順位・負荷補正タイムが良い順。
        total_points, best_info = min(
            candidates,
            key=lambda x: (-int(x[0]), int(x[1]["rank"]), float(x[1].get("time") or 999)),
        )
        out = dict(best_info)
        out["points"] = total_points
        out["global_points"] = total_points
        red[k] = out
    return red


def _chokyo_standard_margin(entry):
    """レース内比較用に、負荷補正タイムが基準より何秒上下かを返す。小さいほど優秀。"""
    key = entry.get("key")
    if not key or len(key) < 3:
        return None
    _date, course_key, metric = key
    try:
        course, is_main = course_key
    except (TypeError, ValueError):
        return None
    standards = CHOKYO_STANDARD_1F if metric == "1F" else CHOKYO_STANDARD_3F if metric == "3F" else {}
    threshold = standards.get(course)
    if threshold is None:
        return None
    if metric == "3F":
        threshold = _training_threshold(standards, course, bool(is_main))
    try:
        return float(entry.get("adj_time", entry.get("time"))) - float(threshold)
    except (TypeError, ValueError):
        return None


def _compute_race_chokyo_bonus(entries):
    """各レースの比較可能馬を基準タイム差で並べ、5頭以上なら上位3頭へ+3/+2/+1。

    1頭に複数本・複数メトリクスがある場合は、その馬が基準を最も大きく上回った1本だけを採用する。
    Returns: {(r_num, umaban): {rank,total,points,margin,disp,line,metric}}
    """
    race_horses = {}
    for e in entries or []:
        if (e.get("label") or "今走調教") != "今走調教":
            continue
        margin = _chokyo_standard_margin(e)
        if margin is None:
            continue
        k = (int(e["r_num"]), str(e["umaban"]))
        candidate = dict(e)
        candidate["margin"] = margin
        prev = race_horses.get(k)
        if prev is None or margin < float(prev["margin"]):
            race_horses[k] = candidate

    by_race = {}
    for (rn, ub), e in race_horses.items():
        by_race.setdefault(rn, []).append((ub, e))

    result = {}
    for rn, grp in by_race.items():
        total = len(grp)
        if total < CHOKYO_RACE_MIN_HORSES:
            continue
        for ub, e in grp:
            margin = float(e["margin"])
            rank = 1 + sum(1 for _other_ub, o in grp if float(o["margin"]) < margin)
            pts = CHOKYO_RACE_POINTS.get(rank, 0)
            if pts <= 0:
                continue
            result[(rn, ub)] = {
                "rank": rank,
                "total": total,
                "points": pts,
                "race_points": pts,
                "margin": margin,
                "disp": e.get("disp"),
                "line": e.get("line"),
                "metric": e.get("metric"),
                "label": "今走調教",
            }
    return result


def _combine_chokyo_rank_points(global_points, race_points):
    """全体順位点とレース内順位点を合算し、順位ボーナス全体を最大8点に収める。"""
    return min(
        CHOKYO_RANK_BONUS_CAP,
        max(0, int(global_points or 0)) + max(0, int(race_points or 0)),
    )


def _compute_day_chokyo_ranks(entries):
    """表示用の同日調教順位を全頭ぶん作る。

    同じ日・調教コース・メトリクスで3頭以上いるグループだけを対象にし、
    今走/前走調教を別グループとして扱う。加点はここでは決めず、表示用の
    rank/total を返す。
    """
    groups = {}
    for e in entries or []:
        if e.get("key") is None or e.get("time") is None:
            continue
        label = e.get("label") or "今走調教"
        key = (label, e["key"])
        groups.setdefault(key, []).append(e)

    def _adj(e):
        return float(e.get("adj_time", e["time"]))

    ranks = {}
    for (_label, _key), grp in groups.items():
        group_size = len(grp)
        if group_size < CHOKYO_RANK_DISPLAY_MIN_GROUP_SIZE:
            continue
        for e in grp:
            rank = 1 + sum(1 for o in grp if _adj(o) < _adj(e))
            line = str(e.get("line") or "").rstrip()
            k = (int(e["r_num"]), str(e["umaban"]), e.get("label") or "今走調教", line)
            info = {
                "rank": rank,
                "total": group_size,
                "line": line,
                "label": e.get("label") or "今走調教",
                "disp": e.get("disp"),
                "time": e.get("time"),
                "metric": e.get("metric"),
            }
            prev = ranks.get(k)
            if prev is None or int(rank) < int(prev.get("rank") or 999):
                ranks[k] = info
    return ranks


def _rank_infos_for_race(rank_map, r_num):
    out = {}
    for (rn, ub, _label, _line), info in (rank_map or {}).items():
        if int(rn) != int(r_num):
            continue
        out.setdefault(str(ub), []).append(info)
    for infos in out.values():
        infos.sort(key=lambda x: (0 if x.get("label") == "今走調教" else 1, int(x.get("rank") or 999)))
    return out


def _apply_chokyo_rank_to_ai(ai_text, rank_here):
    """出走馬分析テキスト内の今走/前走調教行へ (N位/M頭) を付ける。"""
    if not rank_here or not ai_text:
        return ai_text
    out = []
    cur = None
    for line in ai_text.split('\n'):
        hm = re.match(r'^([①-⑳])', line)
        if hm:
            cur = str(ord(hm.group(1)) - 0x245f)
        else:
            bm = re.match(r'^\[(\d{1,2})\]', line)
            if bm:
                cur = bm.group(1)
        if cur in rank_here and ('今走調教' in line or '前走調教' in line) and not re.search(r'\(\d+位/\d+頭\)', line):
            clean_line = strip_display_markers(line).rstrip()
            match_info = None
            for info in rank_here.get(cur, []):
                label = info.get("label") or "今走調教"
                target_line = str(info.get("line") or "").rstrip()
                if label not in clean_line:
                    continue
                if target_line and clean_line != target_line:
                    continue
                match_info = info
                break
            if match_info:
                rank = match_info.get("rank")
                total = match_info.get("total")
                rank_note = f"{rank}位/{total}頭" if total else f"{rank}位"
                if match_info.get("points", 0):
                    out.append(f"[[RB]]{clean_line} ({rank_note})[[/RB]]")
                else:
                    out.append(f"{clean_line} ({rank_note})")
                continue
        out.append(line)
    return '\n'.join(out)


def _apply_chokyo_red_to_ai(ai_text, red_here):
    """後方互換用: TOP3情報だけを渡された場合も調教順位反映として扱う。"""
    if not red_here:
        return ai_text
    normalized = {}
    for ub, info in red_here.items():
        item = dict(info) if isinstance(info, dict) else {"rank": info}
        item.setdefault("label", "今走調教")
        item["points"] = item.get("points", CHOKYO_TOP_POINTS.get(int(item.get("rank") or 99), 0))
        normalized.setdefault(str(ub), []).append(item)
    return _apply_chokyo_rank_to_ai(ai_text, normalized)


def _relabel_match_grades(match_txt, horses, grades):
    """対戦表テキスト内、現出走馬の末尾ランク表記 (A)/(B)… を最新gradesへ差し替える。
    調教加点の反映で総合評価のランクが変わっても、対戦表(生成時に旧gradesで作成)が
    食い違わないようにする。各セグメントは『N着 [u]馬名(枠)…(ランク)』形式で、
    ランクは単一の英字カッコなので枠等の他カッコと衝突しない。[u] を起点に照合する。"""
    if not match_txt:
        return match_txt
    out = match_txt
    for u, h in (horses or {}).items():
        newg = grades.get(_norm_horse_name((h or {}).get("name", "")))
        if not newg:
            continue
        pat = re.compile(r'(\[' + re.escape(str(u)) + r'\][^/\n]*?)\([SABCDEFG\-]\)')
        out = pat.sub(lambda m: f"{m.group(1)}({newg})", out)
    return out


def _merge_chokyo_rank_infos(red_here=None, rank_here=None):
    merged = {str(ub): [dict(info) for info in infos] for ub, infos in (rank_here or {}).items()}
    for ub, info in (red_here or {}).items():
        item = dict(info) if isinstance(info, dict) else {"rank": info}
        item.setdefault("label", "今走調教")
        item["points"] = item.get("points", CHOKYO_TOP_POINTS.get(int(item.get("rank") or 99), 0))
        target_line = str(item.get("line") or "").rstrip()
        bucket = merged.setdefault(str(ub), [])
        matched = False
        for existing in bucket:
            if existing.get("label") != item.get("label"):
                continue
            existing_line = str(existing.get("line") or "").rstrip()
            if target_line and existing_line and target_line != existing_line:
                continue
            existing.update({k: v for k, v in item.items() if v not in (None, "")})
            matched = True
            break
        if not matched:
            bucket.append(item)
    return merged


def _rebuild_race_with_chokyo(p, red_here, rank_here=None):
    """調教順位加点を反映して、1レース分の grades/評価一覧/出走馬分析/HTML/テキストを再生成する。
    ループ内の確定処理と同じ順序を踏む（scored_data には全体/レース内順位点が設定済み）。"""
    horses = p["horses"]
    scored_data = p["scored_data"]
    pace_text = p["pace_text"]
    if red_here:
        pace_text, pace_speed_list, pace_bonus_map = predict_pace_python(
            horses,
            p["danwa"],
            p.get("course_text") or p["header1"],
            horse_patterns=p.get("horse_patterns"),
            chokyo_front_umabans={str(ub) for ub in red_here},
        )
        for sd in scored_data.values():
            sd.pop("score_pace", None)
            sd.pop("score_pattern", None)
        for _u, _b in (pace_bonus_map or {}).items():
            if str(_u) in scored_data:
                scored_data[str(_u)]["score_pace"] = _b
            elif _u in scored_data:
                scored_data[_u]["score_pace"] = _b
        apply_pattern_bonus(scored_data, horses, pace_speed_list, p.get("horse_patterns"))

    ai_out_clean, grades = format_deterministic_evaluation(horses, p["cyokyo"], scored_data, p["danwa"])
    grades = apply_internal_rank_locks(grades, p["rank_locks"])
    grades = adjust_grades_by_jockey_choices(grades, horses)
    grades = apply_score_floor_to_grades(grades, horses, scored_data)
    grades = ensure_top_grade_presence(grades)
    grades = apply_display_grade_overrides(grades, horses, scored_data)

    ai_out_clean = _apply_final_grades_to_ai_detail(ai_out_clean, grades)
    ai_out_clean = ensure_race_content_in_ai_detail(ai_out_clean, horses)
    ai_out_clean = _apply_chokyo_rank_to_ai(ai_out_clean, _merge_chokyo_rank_infos(red_here, rank_here))

    # 対戦表のランク表記も最新gradesへ揃える（総合評価/出走馬分析と食い違わないように）。
    match_txt = _relabel_match_grades(p.get("match_txt", ""), horses, grades)

    eval_list_text = build_evaluation_list(grades, horses, scored_data)
    final_html = generate_html_output(
        p["year"], p["month"], p["day"], p["place_name"], p["r_num"], p["header1"],
        pace_text, eval_list_text, match_txt, ai_out_clean, p["details_text"], p["index_table_html"],
        build_speed_index_table(horses, scored_data),
    )
    final_text = (
        f"📅 {p['year']}/{p['month']}/{p['day']} {p['place_name']}{p['r_num']}R\n\n"
        f"{p['header1']}\n\n{pace_text}\n\n{eval_list_text}\n\n{match_txt}\n\n"
        f"【相対評価】\n{p['index_table_html']}\n\n【出走馬分析】\n{ai_out_clean}\n\n{p['details_text']}"
    )
    return {"grades": grades, "ai": ai_out_clean, "eval": eval_list_text, "html": final_html, "text": final_text}


def _awase_partner_rank(a_part, base_rank, class_ranks):
    """併せ相手のクラスを class_ranks の序列値に変換する。
    調教テキストのクラスは全角(例「（Ｂ２）」)なので半角化してから照合する
    (旧実装は全角のまま class_ranks.get に渡し格上判定が常に失敗していた)。
    JRAの古馬クラス(古オープン/古馬N勝)や2-3歳も南関クラス序列へ近似マップする。
    照合不能時は base_rank(同級扱い)を返す。"""
    m = re.search(r'（([^）]*)）', str(a_part or ""))
    if not m:
        return base_rank
    raw = m.group(1)
    norm = raw.translate(str.maketrans('ＡＢＣＤＥ０１２３４５６７８９', 'ABCDE0123456789'))
    if norm in class_ranks:
        return class_ranks[norm]
    if 'オープン' in norm:
        return class_ranks.get('A1', 7)
    if '３勝' in norm or '3勝' in norm:
        return class_ranks.get('A2', 6)
    if '２勝' in norm or '2勝' in norm:
        return class_ranks.get('B1', 5)
    if '１勝' in norm or '1勝' in norm:
        return class_ranks.get('B3', 3)
    if '歳' in norm:  # 2歳/3歳の若駒は最下級相当
        return class_ranks.get('C3', 0)
    return base_rank


def _parse_kon_chokyo_for_ranking(line):
    """『今走調教：…』行から (date_key, (course, is_main_track), time_value, metric) を取り出す。
    metric は 1F(9.5–13.9) があればそれを優先、なければ 3F(33.0–41.0) を用いる。"""
    entries = _parse_kon_chokyo_entries_for_ranking(line)
    if entries:
        e = entries[0]
        return (e["date"], e["course_key"], e["time"], e["metric"])
    return None


def _parse_kon_chokyo_entries_for_ranking(line):
    """『今走調教：…』行から横断比較用の候補を返す。
    同じ日・同じ調教コースで、1F(9.5–13.9) と 3F(33.0–41.0) は別グループで比較する。
    """
    return _parse_chokyo_entries_for_ranking(line, allowed_labels=("今走調教",))


def _parse_chokyo_entries_for_ranking(line, allowed_labels=("今走調教", "前走調教")):
    """調教行から横断比較用の候補を返す。allowed_labels で今走/前走を選べる。"""
    label = next((x for x in allowed_labels if x in str(line or "")), "")
    if not line or not label:
        return []
    m_d = re.search(r'(\d{1,2})/(\d{1,2})', line)
    if not m_d:
        return []
    date_key = f"{int(m_d.group(1)):02d}/{int(m_d.group(2)):02d}"

    course, is_main = _extract_training_course(line)
    if not course:
        return []

    # ' / ' 以降は併せ相手の情報なので、本馬の時計だけを拾う。
    self_part = str(line).split(' / ', 1)[0]
    nums = []
    for s in re.findall(r'(\d{1,3}\.\d)', self_part):
        try:
            nums.append(float(s))
        except ValueError:
            continue

    one_f_val = None
    three_f_val = None
    for n in nums:
        if 9.5 <= n <= 13.9:
            one_f_val = n
        elif 33.0 <= n <= 41.0:
            three_f_val = n

    penalty = _chokyo_load_penalty(line)  # 併せ/一杯+0.2, 強め+0.1（比較用のみ）

    out = []
    if one_f_val is not None:
        out.append({
            "date": date_key,
            "course_key": (course, is_main),
            "time": one_f_val,                 # 表示・保存用は実計測タイム
            "adj_time": one_f_val + penalty,   # 横比較ランキング用の負荷補正タイム
            "metric": "1F",
            "disp": f"{course}{one_f_val:.1f}",
            "line": line.rstrip(),
            "label": label,
        })
    if three_f_val is not None:
        out.append({
            "date": date_key,
            "course_key": (course, is_main),
            "time": three_f_val,
            "adj_time": three_f_val + penalty,
            "metric": "3F",
            "disp": f"{course}{three_f_val:.1f}",
            "line": line.rstrip(),
            "label": label,
        })
    return out


def _compute_prev_chokyo_comparison(prev_lines, current_lines):
    """同じ馬の前走/今走調教を同コース・同メトリクスで対応比較する。

    各条件では最後に記載された調教を採用し、3Fの対応比較を優先、なければ1Fを使う。
    比較には負荷補正後タイムを使う。
    """
    def _latest_by_condition(lines, label):
        latest = {}
        for idx, raw in enumerate(lines or []):
            full_line = f"{label}：{raw}"
            for e in _parse_chokyo_entries_for_ranking(full_line, allowed_labels=(label,)):
                key = (e.get("course_key"), e.get("metric"))
                item = dict(e)
                item["line_index"] = idx
                latest[key] = item
        return latest

    prev_map = _latest_by_condition(prev_lines, "前走調教")
    current_map = _latest_by_condition(current_lines, "今走調教")
    shared = set(prev_map) & set(current_map)
    if not shared:
        return {"points": 0, "detail": "", "flag": "", "delta": None}

    # 今走側で最後に出た条件を優先し、同じ行に1F/3Fがあれば3Fを採用する。
    selected_key = max(
        shared,
        key=lambda k: (
            int(current_map[k].get("line_index", -1)),
            1 if k[1] == "3F" else 0,
            int(prev_map[k].get("line_index", -1)),
        ),
    )
    prev_e = prev_map[selected_key]
    current_e = current_map[selected_key]
    delta = round(
        float(current_e.get("adj_time", current_e["time"]))
        - float(prev_e.get("adj_time", prev_e["time"])),
        3,
    )

    if delta <= -0.5:
        points, flag = 5, "【前回比大幅更新】"
    elif delta <= -0.2:
        points, flag = 3, "【前回比更新】"
    elif -0.1 <= delta <= 0.1:
        points, flag = 1, ""
    elif delta >= 0.6:
        points, flag = -3, "【調教下落】"
    else:
        points, flag = 0, ""

    course = selected_key[0][0]
    metric = selected_key[1]
    signed_delta = f"{delta:+.1f}秒"
    detail = f"{course}{metric} {signed_delta}"
    return {
        "points": points,
        "detail": detail,
        "flag": flag,
        "delta": delta,
        "metric": metric,
        "course": course,
    }


def _parse_training_line_full(line):
    """今走調教行を1本ぶん構造化する（Supabase/SQLite保存用）。前走調教は対象外。
    時計は本人区間（' / ' より前）からのみ拾い、併せ相手のタイムは拾わない。
    保存するタイムは実計測値（負荷補正は加えない）。"""
    if not line or "今走調教" not in line:
        return None
    m_d = re.search(r'(\d{1,2})/(\d{1,2})', line)
    train_date = f"{int(m_d.group(1)):02d}/{int(m_d.group(2)):02d}" if m_d else ""
    course, is_main = _extract_training_course(line)
    self_part = str(line).split(' / ', 1)[0]
    load_type = next((w for w in ("一杯", "強め", "馬なり") if w in self_part), "")
    awase = ' / ' in str(line)
    one_f = three_f = None
    for s in re.findall(r'(\d{1,3}\.\d)', self_part):
        try:
            v = float(s)
        except ValueError:
            continue
        if 9.5 <= v <= 13.9:
            one_f = v
        elif 33.0 <= v <= 41.0:
            three_f = v
    return {
        "train_date": train_date, "course": course, "is_main": is_main,
        "load_type": load_type, "awase": awase,
        "time_1f": one_f, "time_3f": three_f,
        "line_text": str(line).rstrip(),
    }


def _build_training_rows(horses, cyokyo, race_key):
    """1レース分の今走調教（中間追い切り含む全行）を uma_id 紐付けの保存行へ変換する。
    horses: {umaban: {uma_id, name, ...}}, cyokyo: {umaban: 調教テキスト}。"""
    rows = []
    for umaban, h in (horses or {}).items():
        uid = (h or {}).get("uma_id")
        if not uid:
            continue
        text = (cyokyo or {}).get(str(umaban)) or (cyokyo or {}).get(umaban) or ""
        for ln in re.findall(r'今走調教：[^\n]+', text):
            parsed = _parse_training_line_full(ln)
            if not parsed:
                continue
            parsed.update({"uma_id": uid, "horse_name": h.get("name", ""), "race_key": race_key})
            rows.append(parsed)
    return rows


def _build_kon_chokyo_marks(per_race_ai_text, num_races_total):
    """各レースの ai_text から調教ラインを抽出し、順位表示/太字/赤太字マーカーを決定する。
    Returns: dict {race_num: modified_ai_text}
    """
    all_entries = []
    per_race_lines = {}

    for r_num, text in per_race_ai_text.items():
        lines = (text or "").split('\n')
        per_race_lines[r_num] = lines
        for idx, line in enumerate(lines):
            parsed_entries = _parse_chokyo_entries_for_ranking(
                line, allowed_labels=("今走調教", "前走調教")
            )
            if not parsed_entries:
                continue
            for parsed in parsed_entries:
                all_entries.append({
                    "race_num": r_num,
                    "idx": idx,
                    "date": parsed["date"],
                    "course_key": parsed["course_key"],
                    "time": parsed["time"],
                    "adj_time": parsed.get("adj_time", parsed["time"]),
                    "metric": parsed["metric"],
                    "label": parsed.get("label", "今走調教"),
                })

    if not all_entries:
        return {r: '\n'.join(per_race_lines[r]) for r in per_race_lines}

    global_groups = {}
    for e in all_entries:
        key = (e["label"], e["date"], e["course_key"], e["metric"])
        global_groups.setdefault(key, []).append(e)
    for grp in global_groups.values():
        grp.sort(key=lambda x: x.get("adj_time", x["time"]))

    per_race_groups = {}
    for e in all_entries:
        key = (e["race_num"], e["label"], e["date"], e["course_key"], e["metric"])
        per_race_groups.setdefault(key, []).append(e)
    for grp in per_race_groups.values():
        grp.sort(key=lambda x: x.get("adj_time", x["time"]))

    def _rank_in(group, target):
        target_adj = target.get("adj_time", target["time"])
        for e in group:
            if e["race_num"] == target["race_num"] and e["idx"] == target["idx"]:
                return 1 + sum(1 for o in group if o.get("adj_time", o["time"]) < target_adj)
        return None

    line_marks = {}
    priority = {"RB": 0, "B": 1, "RANK": 2}
    for e in all_entries:
        gkey = (e["label"], e["date"], e["course_key"], e["metric"])
        pkey = (e["race_num"], e["label"], e["date"], e["course_key"], e["metric"])
        g_group = global_groups[gkey]
        p_group = per_race_groups[pkey]
        g_rank = _rank_in(g_group, e)
        p_rank = _rank_in(p_group, e)
        g_total = len(g_group)
        p_total = len(p_group)

        kind = "RANK" if g_total >= CHOKYO_RANK_DISPLAY_MIN_GROUP_SIZE and g_rank is not None else None
        rank_disp = None
        total_disp = None
        if kind:
            rank_disp = g_rank
            total_disp = g_total
        if (e.get("label") == "今走調教" and num_races_total >= 2 and g_total >= CHOKYO_TOP_MIN_GROUP_SIZE
                and g_rank is not None and g_rank <= 3):
            kind = 'RB'
            rank_disp = g_rank
            total_disp = g_total
        elif e.get("label") == "今走調教" and p_total >= 3 and p_rank == 1:
            kind = 'B'
            rank_disp = rank_disp or g_rank
            total_disp = total_disp or g_total
        if kind:
            old = line_marks.get((e["race_num"], e["idx"]))
            if (not old
                    or priority.get(kind, 9) < priority.get(old[0], 9)
                    or ((rank_disp or 99) < (old[1] or 99))):
                line_marks[(e["race_num"], e["idx"])] = (kind, rank_disp, total_disp)

    out = {}
    for r_num, lines in per_race_lines.items():
        new_lines = list(lines)
        for idx, original in enumerate(lines):
            if '[[' in original or re.search(r'\(\d+位/\d+頭\)', original):
                # 既に当日調教上位(赤字)や順位が入っている行は二重化しない。
                continue
            mark = line_marks.get((r_num, idx))
            if not mark:
                continue
            kind, rank, total = mark
            text = original.rstrip()
            rank_note = f" ({rank}位/{total}頭)" if rank and total else ""
            if kind == 'RB':
                text = f"[[RB]]{text}{rank_note}[[/RB]]"
            elif kind == 'B':
                text = f"[[B]]{text}{rank_note}[[/B]]"
            else:
                text = f"{text}{rank_note}"
            new_lines[idx] = text
        out[r_num] = '\n'.join(new_lines)
    return out


def _strip_chokyo_jockey_lines(text):
    """details_text 用: 『調:…』ブロックと『【騎手】…』行を除去する。
    『調:』以降は次の見出し（【…】や『話:』『騎:』『師:』など）が現れるまで連結行として捨てる。"""
    if not text:
        return text
    lines = text.split('\n')
    out = []
    skip_chokyo = False
    for ln in lines:
        s = ln.lstrip()
        if skip_chokyo:
            if s.startswith(('【', '騎:', '師:', '話:', '🐎', '[', '①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩', '⑪', '⑫', '⑬', '⑭', '⑮', '⑯', '⑰', '⑱', '⑲', '⑳')) or re.match(r'^\d{1,2}[\)）]', s):
                skip_chokyo = False
            else:
                continue
        if s.startswith('調:') or s.startswith('調：'):
            skip_chokyo = True
            continue
        if s.startswith('【騎手】'):
            continue
        out.append(ln)
    return '\n'.join(out)

def apply_no_relative_scores(scored_data):
    """比較不可馬の⑤相対を、出走馬内の騎手P順位で50〜25点に補完する。

    1位=50点、2位=47点…の3点刻み。9位以下は25点。
    同じ騎手Pは同順位・同点とする。順位は比較不可馬だけではなく、
    出走馬全頭の jockey_p_base を基準に決める。
    """
    rows = list((scored_data or {}).values())
    if not rows:
        return
    all_p = [int(sc.get("jockey_p_base", 0) or 0) for sc in rows]
    rank_by_p = {p: 1 + sum(other > p for other in all_p) for p in set(all_p)}
    for score_info in rows:
        if not score_info.get("has_no_relative"):
            continue
        jockey_p = int(score_info.get("jockey_p_base", 0) or 0)
        jockey_rank = rank_by_p.get(jockey_p, len(all_p) or 1)
        score_info["score_5"] = max(25, 50 - 3 * (jockey_rank - 1))
        score_info["no_relative_grade"] = f"相対不明(騎手P{jockey_rank}位)"

# 検証用: None以外の dict を入れておくと calculate_relative_scores が
# 内部の対戦行列・着差・tier をそこへ書き出す(本番は None で完全に無効)。
_RELATIVE_DEBUG = None

def calculate_relative_scores(horses_data, current_course, current_dist, r_num):
    # 相対tier→点数(100点満点の最重量・50点枠)。判定不能(tier=None)は
    # apply_no_relative_scores が騎手Pの出走馬内順位により25〜50点で補完するため、
    # NO_DATA_SCORE は上書き前の中立プレースホルダにすぎない。
    NO_DATA_SCORE = 20

    cur_dist = _dist_to_int(current_dist)
    # 優劣判定の基準着差: 1400m未満=0.3秒 / 1400〜1699m=0.6秒 / 1700m以上=0.7秒。
    DRAW_TH = _rel_direct_draw_th(cur_dist)

    umaban_to_name = {}
    name_to_umaban = {}

    race_url_hist_data = {}
    race_url_pas_data = {}
    race_url_meta = {}
    HIDDEN_HORSE_MAX_RUNS = 5
    direct_only_urls = set()

    # 3ヶ月以上休み明けで明確に成績向上した馬は、休み前のデータを比較対象から落とす（矛盾解消用）。
    post_rest_threshold = {}
    # 前走で急に走り出した馬は、使えるなら前走の比較材料を中心に評価する。
    last_run_surge = {}
    # 馬ごとの南関4場プロファイル（avg(人気-着順)＋上位人気勝利ボーナス、N<3 縮約）。
    # 「走ったことあるけど人気を裏切っている場」の比較データだけを軽く割引するために使う。
    track_profiles = {}
    for umaban, h in horses_data.items():
        hname = h.get("name", "")
        thr = _detect_post_rest_improvement(h.get("hist", []))
        if thr:
            post_rest_threshold[hname] = thr
        surge = _detect_last_run_surge(h.get("hist", []))
        if surge:
            last_run_surge[hname] = surge
        track_profiles[hname] = _compute_track_profile(h.get("hist", []))

    def _pair_track_factor(name_u, name_v, place):
        """比較1件のディスカウント係数。両馬の場別アプティチュードのうち弱い方に合わせる。"""
        fu = _track_discount_factor(track_profiles.get(name_u), place)
        fv = _track_discount_factor(track_profiles.get(name_v), place)
        return min(fu, fv)

    def _last_run_surge_names_for_url(url, *names):
        if not url:
            return []
        matched = []
        for name in names:
            info = last_run_surge.get(name)
            if info and info.get("url") == url:
                matched.append(name)
        return matched

    def _mark_last_run_surge_entry(entry, url, *names):
        surge_names = _last_run_surge_names_for_url(url, *names)
        if surge_names:
            entry["last_run_surge"] = True
            entry["last_run_surge_horses"] = list(surge_names)
        return entry

    for umaban, h in horses_data.items():
        hname = h.get("name", "")
        umaban_to_name[umaban] = hname
        name_to_umaban[hname] = umaban
        valid_idx = 0
        for pr in h.get("hist", []):
            if not isinstance(pr, dict): continue
            if not pr.get("pas") or pr.get("rank") is None: continue
            past_url = pr.get("url", "")
            if not past_url: continue
            p_time = pr.get("time", 0.0)
            if p_time <= 0: continue
            valid_idx += 1
            p_date = pr.get("date", "")
            p_place = pr.get("place", "")
            p_dist = pr.get("dist", "")
            if past_url not in race_url_meta:
                race_url_meta[past_url] = {"date": p_date, "place": p_place, "dist": p_dist}
            if past_url not in race_url_hist_data:
                race_url_hist_data[past_url] = {}
            race_url_hist_data[past_url][hname] = p_time
            # 逃げ馬判定用にpas/rankも保持
            if past_url not in race_url_pas_data:
                race_url_pas_data[past_url] = {}
            race_url_pas_data[past_url][hname] = {"pas": pr.get("pas", ""), "rank": pr.get("rank"), "gate_no": pr.get("gate_no")}

            if valid_idx > HIDDEN_HORSE_MAX_RUNS:
                direct_only_urls.add(past_url)

    direct_urls = {url for url, horses in race_url_hist_data.items() if len(horses) >= 2}
    urls_to_scrape = (direct_urls | (set(race_url_hist_data.keys()) - direct_only_urls))

    race_url_to_data = {}
    race_url_rank_data = {}
    race_url_pop_data = {}
    race_url_gate_data = {}
    for past_url in urls_to_scrape:
        all_horses, scraped_place, scraped_dist, scraped_ranks, scraped_pops, scraped_gates = fetch_race_all_horses(past_url)
        if all_horses:
            race_url_to_data[past_url] = all_horses
            race_url_rank_data[past_url] = scraped_ranks
            race_url_pop_data[past_url] = scraped_pops
            race_url_gate_data[past_url] = scraped_gates
            if scraped_place and not race_url_meta.get(past_url, {}).get("place"):
                race_url_meta[past_url]["place"] = scraped_place
            if scraped_dist and not race_url_meta.get(past_url, {}).get("dist"):
                race_url_meta[past_url]["dist"] = scraped_dist

    G = nx.DiGraph()
    current_names = set(umaban_to_name.values())

    def _safe_rank(rank):
        try:
            return int(rank)
        except (TypeError, ValueError):
            return 99

    def _hist_rank_for(hi, hname):
        if hi.get("h1_name") == hname:
            return _safe_rank(hi.get("h1_rank"))
        if hi.get("h2_name") == hname:
            return _safe_rank(hi.get("h2_rank"))
        return 99

    def _hist_pas_for(hi, hname):
        if hi.get("h1_name") == hname:
            return hi.get("h1_pas", "") or ""
        if hi.get("h2_name") == hname:
            return hi.get("h2_pas", "") or ""
        return ""

    def _hist_gate_for_relative(hi, hname):
        if hi.get("h1_name") == hname:
            return hi.get("h1_gate")
        if hi.get("h2_name") == hname:
            return hi.get("h2_gate")
        return None

    def _rank_sort_key(rank, dt):
        dt_obj = parse_date(dt) if isinstance(dt, str) else (dt if isinstance(dt, datetime) else datetime.min)
        return (_safe_rank(rank), -dt_obj.timestamp() if dt_obj != datetime.min else 0)

    def _direct_race_priority(place, dist):
        d = _dist_to_int(dist)
        # コーナー数 (ターン数) が今回と一致しなければ直接対決でも採用しない。
        # 例: 船橋1800(2turn) 現条件で 川崎/浦和2000(multi) の直接対決は採用しない。
        if not _is_turn_match_to_current(place, d, current_course, cur_dist):
            return 0
        return _race_condition_priority(place, d, current_course, cur_dist)

    def _fallback_condition_score(place, dist):
        """比較相手が少ない馬だけに使う緩和条件スコア。通常優先度より常に下。"""
        d = _dist_to_int(dist)
        if not place or d <= 0 or cur_dist <= 0:
            return -999
        if not _is_relative_distance_allowed(d, cur_dist):
            return -999
        if _is_long_distance(cur_dist):
            return -999
        score = 0
        if place == current_course:
            score += 4
        if d == cur_dist:
            score += 3
        elif abs(d - cur_dist) <= 100:
            score += 2
        elif abs(d - cur_dist) <= 300:
            score += 1
        try:
            if _is_same_track_layout(place, d, cur_dist):
                score += 2
        except Exception:
            pass
        return score

    def _fallback_bridge_score(place1, dist1, place2, dist2):
        """A案: 同場経由の比較が一切できない馬のみ使う最終fallback。
        両レッグとも他場・両レッグ同距離=今回距離 の場合のみスコアを返す。
        (同場・レイアウト違いなどは _hidden_bridge_priority のP2-1で扱う前提)
        """
        d1 = _dist_to_int(dist1)
        d2 = _dist_to_int(dist2)
        if not place1 or not place2 or d1 <= 0 or d2 <= 0 or cur_dist <= 0:
            return -999
        if place1 == current_course or place2 == current_course:
            return -999
        if d1 != cur_dist or d2 != cur_dist:
            return -999
        score = 50
        if place1 == place2:
            score += 5
        return score

    def _add_edge(h1_name, h2_name, raw_diff, r_date, r_place, r_dist_str, race_url, base_cost, is_direct, h1_rank=None, h2_rank=None, h1_gate=None, h2_gate=None, h1_pas=None, h2_pas=None, h1_pop=None, h2_pop=None):
        if h1_name > h2_name:
            h1_name, h2_name = h2_name, h1_name
            raw_diff = -raw_diff
            h1_rank, h2_rank = h2_rank, h1_rank
            h1_gate, h2_gate = h2_gate, h1_gate
            h1_pas, h2_pas = h2_pas, h1_pas
            h1_pop, h2_pop = h2_pop, h1_pop

        raw_diff = _apply_first_place_margin_bonus(raw_diff, h1_rank, h2_rank)
        capped_diff = max(-1.2, min(1.2, raw_diff))
        r_dist_int = _dist_to_int(r_dist_str)
        dist_diff_val = abs(r_dist_int - cur_dist) if r_dist_int > 0 and cur_dist > 0 else 9999

        is_same_place = (r_place == current_course)
        is_exact_cond = is_same_place and (dist_diff_val == 0)

        badge = ""
        if is_exact_cond: badge = "[場×距]"
        elif is_same_place: badge = "[場]"
        elif dist_diff_val == 0: badge = "[距]"

        edge_cost = base_cost + (0 if is_direct else 100)

        history_item = {
            "date": r_date,
            "date_str": str(r_date),
            "place": r_place,
            "dist": r_dist_str,
            "raw_diff": capped_diff,
            "badge": badge,
            "url": race_url,
            "h1_name": h1_name,
            "h2_name": h2_name,
            "h1_rank": _safe_rank(h1_rank),
            "h2_rank": _safe_rank(h2_rank),
            "h1_pop": _safe_rank(h1_pop),
            "h2_pop": _safe_rank(h2_pop),
            "h1_gate": h1_gate,
            "h2_gate": h2_gate,
            "field_size": len(race_url_to_data.get(race_url, {}) or {}),
            "h1_pas": h1_pas or "",
            "h2_pas": h2_pas or "",
        }

        if G.has_edge(h1_name, h2_name):
            ed = G[h1_name][h2_name]
            ed["diffs"].append(capped_diff)
            ed["history"].append(history_item)
            ed["rank_diff"] = sum(ed["diffs"]) / len(ed["diffs"])
            if edge_cost < ed["explore_cost"]:
                ed["explore_cost"] = edge_cost
        else:
            G.add_edge(
                h1_name, h2_name,
                diffs=[capped_diff],
                history=[history_item],
                rank_diff=capped_diff,
                explore_cost=edge_cost
            )

    for race_url in direct_urls:
        meta = race_url_meta.get(race_url, {})
        r_place = meta.get("place", "")
        r_dist_str = meta.get("dist", "")
        r_date = meta.get("date", "")
        r_dist = _dist_to_int(r_dist_str)
        if not _is_relative_distance_allowed(r_dist, cur_dist):
            continue
        if not _is_past_race_within_relative_window(r_date, r_place, r_dist_str, current_course, current_dist):
            continue

        is_same_place = (r_place == current_course)
        is_exact_cond = is_same_place and (r_dist == cur_dist)
        is_same_layout = _is_same_track_layout(r_place, r_dist_str, current_dist)

        if is_exact_cond: base_cost = 0.5
        elif is_same_place and is_same_layout: base_cost = 2
        elif is_same_place: base_cost = 5
        elif is_same_layout: base_cost = 8
        else: base_cost = 15

        runners_in_race = race_url_hist_data[race_url]
        names = list(runners_in_race.keys())

        # 直接対決はレースの生タイム差をそのまま採用する（逃げ調整は実施しない）。
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                h1, h2 = names[i], names[j]
                t1 = runners_in_race[h1]
                t2 = runners_in_race[h2]
                raw_diff = t1 - t2
                h1_rank = race_url_pas_data.get(race_url, {}).get(h1, {}).get("rank")
                h2_rank = race_url_pas_data.get(race_url, {}).get(h2, {}).get("rank")
                h1_gate = race_url_pas_data.get(race_url, {}).get(h1, {}).get("gate_no")
                h2_gate = race_url_pas_data.get(race_url, {}).get(h2, {}).get("gate_no")
                h1_pas = race_url_pas_data.get(race_url, {}).get(h1, {}).get("pas", "")
                h2_pas = race_url_pas_data.get(race_url, {}).get(h2, {}).get("pas", "")
                _add_edge(h1, h2, raw_diff, r_date, r_place, r_dist_str, race_url, base_cost, True, h1_rank, h2_rank, h1_gate, h2_gate, h1_pas, h2_pas)

    for race_url, race_horses in race_url_to_data.items():
        meta = race_url_meta.get(race_url, {})
        r_place = meta.get("place", "")
        r_dist_str = meta.get("dist", "")
        r_date = meta.get("date", "")
        r_dist = _dist_to_int(r_dist_str)
        if not _is_relative_distance_allowed(r_dist, cur_dist):
            continue
        if not _is_past_race_within_relative_window(r_date, r_place, r_dist_str, current_course, current_dist):
            continue

        current_in_race = [h for h in race_horses if h in current_names]
        if not current_in_race: continue

        race_ranks = race_url_rank_data.get(race_url, {})
        race_pops = race_url_pop_data.get(race_url, {})
        def _is_bad_against_popularity(hname):
            # 隠れ馬経由のみ適用: 人気より5つ以上着順が悪かった馬は経由馬から除外。
            # 例) 1人気6着以下、3人気8着以下、8人気13着以下 → 物差し馬として不採用。
            rank = race_ranks.get(hname)
            pop = race_pops.get(hname)
            if rank is None or pop is None:
                return False
            return _safe_rank(rank) - _safe_rank(pop) >= 5

        hidden_horses = [
            (h, t) for h, t in race_horses.items()
            if h not in current_names and not _is_bad_against_popularity(h)
        ]
        # 物差し馬フィルタ: 勝ち馬との着差>1.5秒 かつ そのレース着順が下位25%の馬は隠れ馬から除外
        # （出走8頭以上のレースのみ適用。少頭数レースは情報枯渇回避のためスルー）
        if _BRIDGE_FILTER_YARDSTICK and len(race_horses) >= 8:
            sorted_race = sorted(race_horses.items(), key=lambda x: x[1])
            bottom_count = max(1, int(len(sorted_race) * 0.25))
            bottom_names = {nm for nm, _ in sorted_race[-bottom_count:]}
            hidden_horses = [(h, t) for h, t in hidden_horses if not (t > 1.5 and h in bottom_names)]
        if not hidden_horses: continue

        is_same_place = (r_place == current_course)
        is_exact_cond = is_same_place and (r_dist == cur_dist)
        is_same_layout = _is_same_track_layout(r_place, r_dist_str, current_dist)
        
        if is_exact_cond: base_cost = 0.5
        elif is_same_place and is_same_layout: base_cost = 2
        elif is_same_place: base_cost = 5
        elif is_same_layout: base_cost = 8
        else: base_cost = 15

        # 隠れ馬経由もスクレイピング生の着差をそのまま使う（逃げ調整は実施しない）。
        for curr_name in current_in_race:
            curr_time = race_horses.get(curr_name)
            if curr_time is None: continue
            for hid_name, hid_time in hidden_horses:
                raw_diff = curr_time - hid_time
                curr_rank = race_url_pas_data.get(race_url, {}).get(curr_name, {}).get("rank")
                hid_rank = race_url_rank_data.get(race_url, {}).get(hid_name)
                curr_gate = race_url_pas_data.get(race_url, {}).get(curr_name, {}).get("gate_no")
                hid_gate = race_url_gate_data.get(race_url, {}).get(hid_name)
                curr_pas = race_url_pas_data.get(race_url, {}).get(curr_name, {}).get("pas", "")
                _add_edge(curr_name, hid_name, raw_diff, r_date, r_place, r_dist_str, race_url, base_cost, False, curr_rank, hid_rank, curr_gate, hid_gate, curr_pas, "")

        for i in range(len(hidden_horses)):
            for j in range(i + 1, len(hidden_horses)):
                h1_name, h1_time = hidden_horses[i]
                h2_name, h2_time = hidden_horses[j]
                raw_diff = h1_time - h2_time
                h1_rank = race_url_rank_data.get(race_url, {}).get(h1_name)
                h2_rank = race_url_rank_data.get(race_url, {}).get(h2_name)
                h1_gate = race_url_gate_data.get(race_url, {}).get(h1_name)
                h2_gate = race_url_gate_data.get(race_url, {}).get(h2_name)
                _add_edge(h1_name, h2_name, raw_diff, r_date, r_place, r_dist_str, race_url, base_cost, False, h1_rank, h2_rank, h1_gate, h2_gate)

    for u, v, d in G.edges(data=True):
        def _get_dt(x):
            dt_str = x.get("date", "")
            return parse_date(dt_str) if isinstance(dt_str, str) else (x["date"] if isinstance(x["date"], datetime) else datetime.min)
        d["history"].sort(
            key=lambda hi: (
                min(_safe_rank(hi.get("h1_rank")), _safe_rank(hi.get("h2_rank"))),
                -(_get_dt(hi).timestamp() if _get_dt(hi) != datetime.min else 0),
            )
        )
        # 同場同距離で複数レースがある場合、最も着順が良いもの（同着順なら最新）のみ残す。
        # ただし今回開催場×今回距離(=同条件)のレースは全件残す: 同じ条件で複数回対戦している
        # 情報自体が貴重で、表示でも全比較を見たいため。
        seen_cond = set()
        deduped = []
        for hi in d["history"]:
            cond_key = (hi.get("place", ""), hi.get("dist", ""))
            keep_exact_repeat = (
                hi.get("place", "") == current_course
                and _dist_to_int(hi.get("dist", "")) == cur_dist
            )
            if cond_key in seen_cond and not keep_exact_repeat:
                continue
            if not keep_exact_repeat:
                seen_cond.add(cond_key)
            deduped.append(hi)
        d["history"] = deduped[:8]
        d["diffs"] = [hi["raw_diff"] for hi in d["history"]]
        d["rank_diff"] = sum(d["diffs"]) / len(d["diffs"]) if d["diffs"] else 0

    current_list = list(current_names)
    pair_net = {u: {v: [] for v in current_list} for u in current_list}

    for u in current_list:
        for v in current_list:
            if u == v: continue

            direct_entries = []
            if G.has_edge(u, v):
                for hi in G[u][v]["history"]:
                    u_rank_v = _hist_rank_for(hi, u)
                    v_rank_v = _hist_rank_for(hi, v)
                    direct_rank = min(u_rank_v, v_rank_v)
                    direct_entries.append((-hi["raw_diff"], hi.get("date", ""), hi.get("place", ""), hi.get("dist", ""), hi.get("url", ""), direct_rank, _hist_pas_for(hi, u), _hist_pas_for(hi, v), u_rank_v, v_rank_v, _hist_gate_for_relative(hi, u), _hist_gate_for_relative(hi, v), hi.get("field_size")))
            if G.has_edge(v, u):
                for hi in G[v][u]["history"]:
                    u_rank_v = _hist_rank_for(hi, u)
                    v_rank_v = _hist_rank_for(hi, v)
                    direct_rank = min(u_rank_v, v_rank_v)
                    direct_entries.append((hi["raw_diff"], hi.get("date", ""), hi.get("place", ""), hi.get("dist", ""), hi.get("url", ""), direct_rank, _hist_pas_for(hi, u), _hist_pas_for(hi, v), u_rank_v, v_rank_v, _hist_gate_for_relative(hi, u), _hist_gate_for_relative(hi, v), hi.get("field_size")))

            # 休み明け改善フィルタ: u もしくは v が「3ヶ月休み→明確に向上」型なら、
            # その馬の休み前の直接対戦は採用しない（フィルタ後に何も残らない場合は元のまま）。
            u_thr = post_rest_threshold.get(u)
            v_thr = post_rest_threshold.get(v)
            if direct_entries and (u_thr or v_thr):
                filtered_direct = []
                for entry in direct_entries:
                    dt_str = entry[1]
                    dt_obj = parse_date(dt_str) if isinstance(dt_str, str) else dt_str
                    if isinstance(dt_obj, datetime) and dt_obj != datetime.min:
                        if u_thr and dt_obj < u_thr:
                            continue
                        if v_thr and dt_obj < v_thr:
                            continue
                    filtered_direct.append(entry)
                if filtered_direct:
                    direct_entries = filtered_direct

            if direct_entries:
                priority_entries = []
                for diff, dt, place, dist, url, rank, u_pas, v_pas, u_rank_v, v_rank_v, u_gate, v_gate, field_size in direct_entries:
                    priority = _direct_race_priority(place, dist)
                    if priority:
                        priority_entries.append((priority, diff, dt, place, dist, url, rank, u_pas, v_pas, u_rank_v, v_rank_v, u_gate, v_gate, field_size))

                if priority_entries:
                    surge_urls = {
                        info.get("url") for name in (u, v)
                        for info in [last_run_surge.get(name)]
                        if info and info.get("url")
                    }
                    surge_direct = [x for x in priority_entries if x[5] in surge_urls]
                    if surge_direct:
                        target_direct = surge_direct
                        best_priority = max(x[0] for x in target_direct)
                    else:
                        best_priority = max(x[0] for x in priority_entries)
                        target_direct = [x for x in priority_entries if x[0] == best_priority]
                    target_direct.sort(key=lambda x: _rank_sort_key(x[6], x[2]))

                    def _direct_entry_dt(entry):
                        dt = entry[2]
                        parsed = parse_date(dt) if isinstance(dt, str) else dt
                        return parsed if isinstance(parsed, datetime) else datetime.min

                    is_same_cond_best = (
                        (_is_long_distance(cur_dist) and best_priority >= 5)
                        or (not _is_long_distance(cur_dist) and best_priority >= 4)
                    )

                    # 同条件で複数レースある場合の外れ値フィルタ:
                    # 1〜3人気で3着以上落とした馬がいるレースは「不本意な結果」として優先しない。
                    # clean(両馬OK)が1件でもあれば clean の中から各馬ベストを選ぶ。
                    if is_same_cond_best and len(target_direct) >= 2:
                        def _entry_has_outlier(entry):
                            url_e = entry[5]
                            u_rank_e, v_rank_e = entry[9], entry[10]
                            pops = race_url_pop_data.get(url_e, {})
                            return (
                                _is_pop_drop_outlier(u_rank_e, pops.get(u))
                                or _is_pop_drop_outlier(v_rank_e, pops.get(v))
                            )
                        clean_direct = [e for e in target_direct if not _entry_has_outlier(e)]
                        if clean_direct:
                            target_direct = clean_direct

                    # 最良条件の複数対戦を後段で新しさ・場適性により加重集約する。
                    # 両馬それぞれの「最も良かった1戦」の抽出は、時系列と安定度を
                    # 消すため廃止。最新側から最大5件を候補とする。
                    target_direct.sort(key=_direct_entry_dt, reverse=True)
                    selected_direct = target_direct[:5]

                    for priority, d, dt, p, dist, url, rank, u_pas, v_pas, _u_rank_v, _v_rank_v, u_gate, v_gate, field_size in selected_direct:
                        track_factor = _pair_track_factor(u, v, p)
                        entry = {
                            "diff": d * track_factor,
                            "is_strict": priority >= 3,
                            "place": p,
                            "dist": dist,
                            "date": dt,
                            "url": url,
                            "rank": rank,
                            "direct_priority": priority,
                            "track_factor": track_factor,
                            "u_pas": u_pas,
                            "v_pas": v_pas,
                            "u_gate": u_gate,
                            "v_gate": v_gate,
                            "field_size": field_size,
                            "per_horse_best": False,
                        }
                        pair_net[u][v].append(_mark_last_run_surge_entry(entry, url, u, v))

                if pair_net[u][v]:
                    direct_has_current_exact = any(
                        e.get("place") == current_course
                        and _dist_to_int(e.get("dist", "")) == cur_dist
                        for e in pair_net[u][v]
                    )
                    # 直接対決があっても、同じ条件品質のbridgeはランク判定に併用する。
                    # ただし同場同距離ではない直接対決しか無い非根幹距離だけは、
                    # 従来通り高品質bridge(priority>=3)に絞って救済採用する。
                    bridge_rescue_mode = not (direct_has_current_exact or _is_kankan_distance(cur_dist))
                else:
                    bridge_rescue_mode = False
            else:
                bridge_rescue_mode = False

            hidden_nodes = [n for n in G.nodes() if n not in current_names]
            for h in hidden_nodes:
                u_h_hist = []
                if G.has_edge(u, h): u_h_hist.extend([(-hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, u), hi.get("url", "")) for hi in G[u][h]["history"]])
                if G.has_edge(h, u): u_h_hist.extend([(hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, u), hi.get("url", "")) for hi in G[h][u]["history"]])

                h_v_hist = []
                if G.has_edge(h, v): h_v_hist.extend([(-hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, v), hi.get("url", "")) for hi in G[h][v]["history"]])
                if G.has_edge(v, h): h_v_hist.extend([(hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, v), hi.get("url", "")) for hi in G[v][h]["history"]])

                bridge_diffs = []
                u_thr_b = post_rest_threshold.get(u)
                v_thr_b = post_rest_threshold.get(v)
                bridge_pre_rest_fallback = []  # 休み明けフィルタで全部消えた場合のフォールバック
                for diff_uh, p_uh, d_uh, dt_uh, rank_uh, url_uh in u_h_hist:
                    for diff_hv, p_hv, d_hv, dt_hv, rank_hv, url_hv in h_v_hist:
                        # 隠れ馬経由のみ適用: 本馬や相手が10着以下のレースは参考外
                        if _BRIDGE_FILTER_UV_RANK10 and (_safe_rank(rank_uh) >= 10 or _safe_rank(rank_hv) >= 10):
                            continue

                        # 隠れ馬経由: 各レッグが半年以内(or ≥1800m)でないと採用しない
                        if not _is_bridge_leg_recent_enough(dt_uh, d_uh):
                            continue
                        if not _is_bridge_leg_recent_enough(dt_hv, d_hv):
                            continue

                        dt_uh_obj = parse_date(dt_uh) if isinstance(dt_uh, str) else dt_uh
                        dt_hv_obj = parse_date(dt_hv) if isinstance(dt_hv, str) else dt_hv
                        dt_combined = min(dt_uh_obj, dt_hv_obj) if dt_uh_obj and dt_hv_obj else (dt_uh_obj or dt_hv_obj)

                        bridge_priority = _hidden_bridge_priority(p_uh, d_uh, p_hv, d_hv, cur_dist, current_course)
                        if not bridge_priority:
                            continue

                        # 休み明け改善フィルタ: u-leg は u の閾値、v-leg は v の閾値で
                        is_pre_rest = False
                        if u_thr_b and isinstance(dt_uh_obj, datetime) and dt_uh_obj != datetime.min and dt_uh_obj < u_thr_b:
                            is_pre_rest = True
                        if v_thr_b and isinstance(dt_hv_obj, datetime) and dt_hv_obj != datetime.min and dt_hv_obj < v_thr_b:
                            is_pre_rest = True
                        is_current_same = (
                            p_uh == current_course
                            and p_hv == current_course
                            and str(d_uh) == str(cur_dist)
                            and str(d_hv) == str(cur_dist)
                        )
                        is_current_venue = (p_uh == current_course and p_hv == current_course)
                        is_current_layout = (
                            is_current_venue
                            and _is_same_track_layout(current_course, d_uh, cur_dist)
                            and _is_same_track_layout(current_course, d_hv, cur_dist)
                        )
                        bridge_rank = min(_safe_rank(rank_uh), _safe_rank(rank_hv))
                        surge_names = []
                        if _last_run_surge_names_for_url(url_uh, u):
                            surge_names.append(u)
                        if _last_run_surge_names_for_url(url_hv, v):
                            surge_names.append(v)
                        entry_tuple = (bridge_priority, diff_uh + diff_hv, p_uh, d_uh, dt_combined, is_current_same, bridge_rank, is_current_venue, is_current_layout, surge_names)
                        if is_pre_rest:
                            bridge_pre_rest_fallback.append(entry_tuple)
                        else:
                            bridge_diffs.append(entry_tuple)

                # 休み明けフィルタで全部弾かれたらフォールバック
                if not bridge_diffs and bridge_pre_rest_fallback:
                    bridge_diffs = bridge_pre_rest_fallback

                if bridge_diffs:
                    surge_required = [name for name in (u, v) if name in last_run_surge]
                    if surge_required:
                        surge_required_all = [
                            x for x in bridge_diffs
                            if set(surge_required).issubset(set(x[9]))
                        ]
                        if surge_required_all:
                            bridge_diffs = surge_required_all
                        else:
                            surge_any = [x for x in bridge_diffs if x[9]]
                            if surge_any:
                                bridge_diffs = surge_any

                    best_priority = max(x[0] for x in bridge_diffs)
                    # 救済モード: 同条件以外の直接対決が既にある状態。
                    # 同場・同レイアウト以上 (priority>=3) の bridge のみ採用する
                    # (例: 船橋2200直接対決しか無いペアに、船橋1600+1800 bridge を救済追加)
                    if bridge_rescue_mode and best_priority < 3:
                        continue
                    # 採用priority集合: 最良 + (合議拡充用に)1段下も投票に含める。1段下は
                    # consensus_weight_factor で重みを下げ、救済モードでは従来通り最良のみ採用する。
                    # 下流でもペア全体の最良priority+1段下だけを残し、合議票として集約する。
                    accept_priorities = [best_priority]
                    if (_BRIDGE_CONSENSUS_INCLUDE_ONE_LOWER and not bridge_rescue_mode
                            and (best_priority - 1) >= 1):
                        accept_priorities.append(best_priority - 1)
                    for accept_p in accept_priorities:
                        tier_diffs = [x for x in bridge_diffs if x[0] == accept_p]
                        if not tier_diffs:
                            continue
                        tier_diffs.sort(key=lambda x: _rank_sort_key(x[6], x[4]))
                        # 同じ隠れ馬でも複数レースがあるなら全件を合議へ渡す。
                        # edge履歴側で最大8件に絞っているため、ここでは追加で1件化しない。
                        selected_diffs = tier_diffs
                        is_lower_tier = (accept_p < best_priority)
                        for best in selected_diffs:
                            bridge_is_strict = accept_p >= 4
                            discount = _BRIDGE_DISCOUNT_BY_PRIORITY.get(accept_p, 0.3)
                            track_factor = _pair_track_factor(u, v, best[2])
                            discounted_diff = best[1] * discount * track_factor
                            entry = {
                                "diff": discounted_diff,
                                "raw_diff": best[1],
                                "is_strict": bridge_is_strict,
                                "place": best[2],
                                "dist": best[3],
                                "date": best[4],
                                "url": "",
                                "bridge_priority": accept_p,
                                "via_hidden": h,
                                "track_factor": track_factor,
                                "rank": best[6],
                            }
                            if best[9]:
                                entry["last_run_surge"] = True
                                entry["last_run_surge_horses"] = list(best[9])
                            if is_lower_tier:
                                entry["consensus_weight_factor"] = _BRIDGE_CONSENSUS_LOWER_WEIGHT
                            pair_net[u][v].append(entry)

    fallback_sparse_names = set()

    def _comparison_counts():
        return {
            h: sum(1 for opp in current_list if opp != h and (pair_net[h].get(opp) or pair_net[opp].get(h)))
            for h in current_list
        }

    def _reverse_pair_entry(entry):
        rev = dict(entry)
        rev["diff"] = -float(rev.get("diff", 0) or 0)
        if "raw_diff" in entry:
            rev["raw_diff"] = -float(entry.get("raw_diff", 0) or 0)
        rev["u_pas"] = entry.get("v_pas", "")
        rev["v_pas"] = entry.get("u_pas", "")
        rev["u_gate"] = entry.get("v_gate")
        rev["v_gate"] = entry.get("u_gate")
        return rev

    def _add_pair_entry_both_directions(u, v, entry):
        pair_net[u][v] = [entry]
        pair_net[v][u] = [_reverse_pair_entry(entry)]
        fallback_sparse_names.update([u, v])

    def _direct_fallback_candidate(u, v):
        direct_entries = []
        if G.has_edge(u, v):
            for hi in G[u][v]["history"]:
                direct_rank = min(_hist_rank_for(hi, u), _hist_rank_for(hi, v))
                direct_entries.append((-hi["raw_diff"], hi.get("date", ""), hi.get("place", ""), hi.get("dist", ""), hi.get("url", ""), direct_rank, _hist_pas_for(hi, u), _hist_pas_for(hi, v)))
        if G.has_edge(v, u):
            for hi in G[v][u]["history"]:
                direct_rank = min(_hist_rank_for(hi, u), _hist_rank_for(hi, v))
                direct_entries.append((hi["raw_diff"], hi.get("date", ""), hi.get("place", ""), hi.get("dist", ""), hi.get("url", ""), direct_rank, _hist_pas_for(hi, u), _hist_pas_for(hi, v)))
        if not direct_entries:
            return None

        candidates = []
        for diff, dt, place, dist, url, rank, u_pas, v_pas in direct_entries:
            priority = _direct_race_priority(place, dist)
            if priority:
                score = 200 + priority * 10
                adj_diff = diff
            else:
                # priority=0 の直接対決fallback もコーナー数一致を必須にする
                if not _is_turn_match_to_current(place, dist, current_course, cur_dist):
                    continue
                relaxed_score = _fallback_condition_score(place, dist)
                if relaxed_score <= -999:
                    continue
                score = 200 + relaxed_score
                adj_diff = diff * 0.55
            track_factor = _pair_track_factor(u, v, place)
            adj_diff = adj_diff * track_factor
            entry = {
                "diff": adj_diff,
                "is_strict": priority >= 3,
                "place": place,
                "dist": dist,
                "date": dt,
                "url": url,
                "rank": rank,
                "direct_priority": priority,
                "fallback_relaxed": priority == 0,
                "track_factor": track_factor,
                "u_pas": u_pas,
                "v_pas": v_pas,
            }
            entry = _mark_last_run_surge_entry(entry, url, u, v)
            if entry.get("last_run_surge"):
                score += 60
            candidates.append((score, rank, dt, entry))
        if not candidates:
            return None
        candidates.sort(key=lambda x: (-x[0], *_rank_sort_key(x[1], x[2])))
        return candidates[0][0], candidates[0][3]

    def _has_same_venue_link(name):
        # A案: 同場経由比較(bridge_priority>=2)もしくは直接対決(direct_priority>=1)があるか
        for entries in pair_net[name].values():
            for entry in (entries or []):
                if not isinstance(entry, dict):
                    continue
                if entry.get("bridge_priority", 0) >= 2:
                    return True
                if entry.get("direct_priority", 0) >= 1:
                    return True
        return False

    def _hidden_fallback_candidate(u, v):
        allow_other_venue_fallback = (not _has_same_venue_link(u)) and (not _has_same_venue_link(v))

        hidden_nodes = [n for n in G.nodes() if n not in current_names]
        candidates = []
        for h in hidden_nodes:
            u_h_hist = []
            if G.has_edge(u, h): u_h_hist.extend([(-hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, u)) for hi in G[u][h]["history"]])
            if G.has_edge(h, u): u_h_hist.extend([(hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, u)) for hi in G[h][u]["history"]])

            h_v_hist = []
            if G.has_edge(h, v): h_v_hist.extend([(-hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, v)) for hi in G[h][v]["history"]])
            if G.has_edge(v, h): h_v_hist.extend([(hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, v)) for hi in G[v][h]["history"]])

            for diff_uh, p_uh, d_uh, dt_uh, rank_uh in u_h_hist:
                for diff_hv, p_hv, d_hv, dt_hv, rank_hv in h_v_hist:
                    if not _is_bridge_leg_recent_enough(dt_uh, d_uh):
                        continue
                    if not _is_bridge_leg_recent_enough(dt_hv, d_hv):
                        continue
                    dt_uh_obj = parse_date(dt_uh) if isinstance(dt_uh, str) else dt_uh
                    dt_hv_obj = parse_date(dt_hv) if isinstance(dt_hv, str) else dt_hv
                    dt_combined = min(dt_uh_obj, dt_hv_obj) if dt_uh_obj and dt_hv_obj else (dt_uh_obj or dt_hv_obj)

                    bridge_priority = _hidden_bridge_priority(p_uh, d_uh, p_hv, d_hv, cur_dist, current_course)
                    is_current_same = (
                        p_uh == current_course
                        and p_hv == current_course
                        and str(d_uh) == str(cur_dist)
                        and str(d_hv) == str(cur_dist)
                    )
                    is_current_venue = (p_uh == current_course and p_hv == current_course)
                    is_current_layout = (
                        is_current_venue
                        and _is_same_track_layout(current_course, d_uh, cur_dist)
                        and _is_same_track_layout(current_course, d_hv, cur_dist)
                    )

                    raw_est = diff_uh + diff_hv
                    if bridge_priority:
                        discount = _BRIDGE_DISCOUNT_BY_PRIORITY.get(bridge_priority, 0.3)
                        score = 100 + bridge_priority * 10
                    else:
                        if not allow_other_venue_fallback:
                            continue
                        # priority=0 採用時は各レッグのコーナー数(1turn vs 2turn+)一致を必須
                        if not _is_turn_match_to_current(p_uh, d_uh, current_course, cur_dist):
                            continue
                        if not _is_turn_match_to_current(p_hv, d_hv, current_course, cur_dist):
                            continue
                        relaxed_score = _fallback_bridge_score(p_uh, d_uh, p_hv, d_hv)
                        if relaxed_score <= -999:
                            continue
                        discount = 0.3
                        score = relaxed_score

                    track_factor = _pair_track_factor(u, v, p_uh)
                    discounted_diff = raw_est * discount * track_factor
                    if abs(discounted_diff) >= 2.0:
                        continue

                    bridge_rank = min(_safe_rank(rank_uh), _safe_rank(rank_hv))
                    candidates.append((score, bridge_rank, dt_combined, {
                        "diff": discounted_diff,
                        "raw_diff": raw_est,
                        "is_strict": bridge_priority >= 4,
                        "place": p_uh,
                        "dist": d_uh,
                        "date": dt_combined,
                        "url": "",
                        "bridge_priority": bridge_priority,
                        "via_hidden": h,
                        "rank": bridge_rank,
                        "fallback_relaxed": bridge_priority == 0,
                        "track_factor": track_factor,
                    }))
        if not candidates:
            return None
        candidates.sort(key=lambda x: (-x[0], *_rank_sort_key(x[1], x[2])))
        return candidates[0][0], candidates[0][3]

    def _reference_bridge_candidate(u, v):
        """4頭以上比較不能時の参考値bridge: priority=0 の隠れ馬経由も強制採用する。
        通常のbridge優先度フィルタは外すが、各レッグの半年制限/着順制限は維持。
        強い割引(×0.25)で「参考値」扱い。"""
        hidden_nodes_local = [n for n in G.nodes() if n not in current_names]
        candidates_local = []
        for h in hidden_nodes_local:
            u_h_hist = []
            if G.has_edge(u, h):
                u_h_hist.extend([(-hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, u)) for hi in G[u][h]["history"]])
            if G.has_edge(h, u):
                u_h_hist.extend([(hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, u)) for hi in G[h][u]["history"]])
            h_v_hist = []
            if G.has_edge(h, v):
                h_v_hist.extend([(-hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, v)) for hi in G[h][v]["history"]])
            if G.has_edge(v, h):
                h_v_hist.extend([(hi["raw_diff"], hi["place"], hi["dist"], hi["date"], _hist_rank_for(hi, v)) for hi in G[v][h]["history"]])

            for diff_uh, p_uh, d_uh, dt_uh, rank_uh in u_h_hist:
                for diff_hv, p_hv, d_hv, dt_hv, rank_hv in h_v_hist:
                    if _BRIDGE_FILTER_UV_RANK10 and (_safe_rank(rank_uh) >= 10 or _safe_rank(rank_hv) >= 10):
                        continue
                    if not _is_bridge_leg_recent_enough(dt_uh, d_uh):
                        continue
                    if not _is_bridge_leg_recent_enough(dt_hv, d_hv):
                        continue
                    dt_uh_obj = parse_date(dt_uh) if isinstance(dt_uh, str) else dt_uh
                    dt_hv_obj = parse_date(dt_hv) if isinstance(dt_hv, str) else dt_hv
                    dt_combined = min(dt_uh_obj, dt_hv_obj) if dt_uh_obj and dt_hv_obj else (dt_uh_obj or dt_hv_obj)

                    bridge_priority = _hidden_bridge_priority(p_uh, d_uh, p_hv, d_hv, cur_dist, current_course)
                    # priority=0 は参考値として採用するが、コーナー数(1turn vs 2turn+)が
                    # 今回と異なるレッグは品質担保のため禁止する
                    if not bridge_priority:
                        if not _is_turn_match_to_current(p_uh, d_uh, current_course, cur_dist):
                            continue
                        if not _is_turn_match_to_current(p_hv, d_hv, current_course, cur_dist):
                            continue
                    # 参考値: priority=0 は強い割引、priority>=1 は通常割引
                    discount = _BRIDGE_DISCOUNT_BY_PRIORITY.get(bridge_priority, 0.25)
                    track_factor = _pair_track_factor(u, v, p_uh)
                    discounted_diff = (diff_uh + diff_hv) * discount * track_factor
                    if abs(discounted_diff) >= 2.0:
                        continue
                    bridge_rank = min(_safe_rank(rank_uh), _safe_rank(rank_hv))
                    score = (bridge_priority * 10) if bridge_priority else 1
                    candidates_local.append((score, bridge_rank, dt_combined, {
                        "diff": discounted_diff,
                        "raw_diff": diff_uh + diff_hv,
                        "is_strict": False,
                        "place": p_uh,
                        "dist": d_uh,
                        "date": dt_combined,
                        "url": "",
                        "bridge_priority": bridge_priority,
                        "via_hidden": h,
                        "rank": bridge_rank,
                        "track_factor": track_factor,
                        "is_reference": True,
                    }))
        if not candidates_local:
            return None
        candidates_local.sort(key=lambda x: (-x[0], *_rank_sort_key(x[1], x[2])))
        return candidates_local[0][3]

    comparison_counts = _comparison_counts()
    for sparse_name in sorted(current_list, key=lambda h: int(name_to_umaban.get(h, 99))):
        if comparison_counts.get(sparse_name, 0) >= 2:
            continue
        if comparison_counts.get(sparse_name, 0) == 0:
            # 直接対決も隠れ馬経由（主ループ判定）も成立しない馬は
            # フォールバックで補完せず「比較不可」として残す。
            # （表示根拠ゼロのまま S/A ランクに昇格してしまう事故を防ぐ）
            continue

        candidates = []
        for opp in current_list:
            if opp == sparse_name or pair_net[sparse_name].get(opp):
                continue
            direct_candidate = _direct_fallback_candidate(sparse_name, opp)
            hidden_candidate = _hidden_fallback_candidate(sparse_name, opp)
            best_candidate = None
            if direct_candidate and hidden_candidate:
                best_candidate = direct_candidate if direct_candidate[0] >= hidden_candidate[0] else hidden_candidate
            else:
                best_candidate = direct_candidate or hidden_candidate
            if best_candidate:
                candidates.append((best_candidate[0], opp, best_candidate[1]))

        candidates.sort(key=lambda x: (-x[0], int(name_to_umaban.get(x[1], 99))))
        for _, opp, entry in candidates:
            if comparison_counts.get(sparse_name, 0) >= 2:
                break
            if pair_net[sparse_name].get(opp) or pair_net[opp].get(sparse_name):
                continue
            _add_pair_entry_both_directions(sparse_name, opp, entry)
            comparison_counts = _comparison_counts()

    # 2頭以上比較不能なら priority=0 の bridge も「参考値」として採用する。
    # (通常は priority>=1 のみ採用するが、比較不能が複数いると展開が読めないため緩和)
    comparison_counts = _comparison_counts()
    no_compare_pre = [h for h in current_list if comparison_counts.get(h, 0) == 0]
    reference_horses = set()
    if len(no_compare_pre) >= 2:
        for u in no_compare_pre:
            for v in current_list:
                if u == v:
                    continue
                if pair_net[u].get(v) or pair_net[v].get(u):
                    continue
                ref_entry = _reference_bridge_candidate(u, v)
                if ref_entry:
                    pair_net[u][v].append(ref_entry)
                    reverse = _reverse_pair_entry(ref_entry)
                    reverse["is_reference"] = True
                    pair_net[v][u].append(reverse)
                    reference_horses.update([u, v])

    matchup_matrix = {u: {} for u in current_list}
    # ＝のみ馬救済用: 直接対決の符号付き着差(u視点, 正なら u が先着)。
    matchup_margin = {u: {} for u in current_list}
    matchup_rank_priority = {u: {} for u in current_list}
    matchup_avg = {u: {} for u in current_list}
    matchup_source = {u: {} for u in current_list}
    matchup_quality = {u: {} for u in current_list}
    matchup_evidence_count = {u: {} for u in current_list}
    matchup_selected_entries = {u: {} for u in current_list}
    # 隠れ馬合議の内訳 (u,v)->(sym, 本馬優勢票, 相手優勢票)。HTML表示で「N件合議」を出すのに使う。
    bridge_consensus_detail = {}
    nige_dependent = set()  # 逃げた時の好走で総合判定を救済された馬
    now = datetime.now()

    def _store_matchup_avg(u_name, v_name, entries):
        if not entries:
            return
        avg = sum(e["diff"] for e in entries) / len(entries)
        matchup_avg[u_name][v_name] = avg
        matchup_avg[v_name][u_name] = -avg

    def inverse_sym(s):
        if s == ">": return "<"
        if s == "=": return "="
        if s == "<": return ">"
        return "="

    def _entry_rank_priority(entry):
        """ランキング用の条件優先度。小さいほど信頼度が高い。"""
        bridge_priority = entry.get("bridge_priority", 0)
        if bridge_priority:
            return 5 - bridge_priority
        direct_priority = entry.get("direct_priority", 0)
        if direct_priority:
            return 5 - direct_priority
        if entry.get("is_strict"):
            return 1
        return 4

    def _entry_is_current_exact(entry):
        bridge_priority = entry.get("bridge_priority", 0)
        if bridge_priority:
            return bridge_priority >= 5
        return (
            entry.get("direct_priority", 0) >= 4
            or (
                entry.get("place") == current_course
                and _dist_to_int(entry.get("dist", "")) == cur_dist
            )
        )

    def _prioritize_strong_conditions_in_cycles(dominance_G):
        """循環内では、より優先度の高い比較条件を残して弱い条件のエッジを外す。"""
        adjusted = dominance_G.copy()
        changed = True
        while changed:
            changed = False
            for horses in list(nx.strongly_connected_components(adjusted)):
                if len(horses) < 3:
                    continue
                internal_edges = [
                    (u, v, adjusted[u][v])
                    for u, v in adjusted.edges()
                    if u in horses and v in horses
                ]
                if not internal_edges:
                    continue
                best_priority = min(data.get("rank_priority", 99) for _, _, data in internal_edges)
                for u, v, data in internal_edges:
                    if data.get("rank_priority", 99) > best_priority:
                        adjusted.remove_edge(u, v)
                        changed = True
        return adjusted

    def _fill_equal_chain_matchups():
        """「A=B、B>CならA>C」を相対ランク用の補完比較として反映する。"""
        MAX_HOPS = 3
        equal_G = nx.Graph()
        equal_G.add_nodes_from(current_list)
        chain_G = nx.DiGraph()
        chain_G.add_nodes_from(current_list)
        base_win_edges = []

        def _add_win_edge(src, dst, points):
            if src == dst:
                return
            points = abs(points) or 1
            if chain_G.has_edge(src, dst):
                chain_G[src][dst]["points"] = max(chain_G[src][dst].get("points", 1), points)
            else:
                chain_G.add_edge(src, dst, points=points)

        for i, h1 in enumerate(current_list):
            for h2 in current_list[i + 1:]:
                rel = matchup_matrix.get(h1, {}).get(h2)
                if rel == "=":
                    equal_G.add_edge(h1, h2)
                elif rel == ">":
                    _add_win_edge(h1, h2, 1)
                    base_win_edges.append((h1, h2, 1))
                elif rel == "<":
                    _add_win_edge(h2, h1, 1)
                    base_win_edges.append((h2, h1, 1))

        equal_groups = list(nx.connected_components(equal_G))
        equal_group_by_horse = {}
        for idx, horses in enumerate(equal_groups):
            for horse in horses:
                equal_group_by_horse[horse] = idx

        # 同等グループの外部勝敗だけをグループ全体に伝播する。
        for src, dst, points in base_win_edges:
            src_group = equal_group_by_horse.get(src)
            dst_group = equal_group_by_horse.get(dst)
            if src_group is None or dst_group is None or src_group == dst_group:
                continue
            for eq_src in equal_groups[src_group]:
                for eq_dst in equal_groups[dst_group]:
                    _add_win_edge(eq_src, eq_dst, points)

        inferred = {h: {} for h in current_list}
        for i, h1 in enumerate(current_list):
            for h2 in current_list[i + 1:]:
                if matchup_matrix.get(h1, {}).get(h2) is not None or matchup_matrix.get(h2, {}).get(h1) is not None:
                    continue

                h1_to_h2 = nx.has_path(chain_G, h1, h2)
                h2_to_h1 = nx.has_path(chain_G, h2, h1)
                if h1_to_h2 == h2_to_h1:
                    continue

                try:
                    if h1_to_h2:
                        path = nx.shortest_path(chain_G, h1, h2)
                        winner, loser = h1, h2
                    else:
                        path = nx.shortest_path(chain_G, h2, h1)
                        winner, loser = h2, h1
                except nx.NetworkXNoPath:
                    continue
                if len(path) - 1 > MAX_HOPS:
                    continue

                win_sym = ">"
                lose_sym = inverse_sym(win_sym)

                if winner == h1 and loser == h2:
                    h1_sym, h2_sym = win_sym, lose_sym
                    h1_chain, h2_chain = list(path), list(reversed(path))
                else:
                    h1_sym, h2_sym = lose_sym, win_sym
                    h1_chain, h2_chain = list(reversed(path)), list(path)

                matchup_matrix[h1][h2] = h1_sym
                matchup_matrix[h2][h1] = h2_sym
                matchup_rank_priority[h1][h2] = 98
                matchup_rank_priority[h2][h1] = 98
                inferred[h1][h2] = {"symbol": h1_sym, "chain": h1_chain}
                inferred[h2][h1] = {"symbol": h2_sym, "chain": h2_chain}

        return inferred

    def _mixed_comparison_symbol(entries, cur_dist, current_course):
        """同じ条件品質の直接対決とbridgeを、同じ票として集約する。"""
        entries = list(entries or [])
        if not entries:
            return "=", [], 0.0, 0, 0

        best_quality = max(
            (_relative_source_quality(e, current_course, cur_dist)[0] for e in entries),
            default=-1.0,
        )
        votes = []
        selected = sorted(
            entries,
            key=lambda e: (
                -_relative_condition_weight(e, current_course, cur_dist),
                -_relative_entry_datetime(e.get("date")).timestamp()
                if _relative_entry_datetime(e.get("date")) != datetime.min else 0,
            ),
        )[:8]

        margin_total = margin_weight = 0.0
        for entry in selected:
            quality = _relative_source_quality(entry, current_course, cur_dist)[0]
            level_factor = 1.0 if quality == best_quality else _BRIDGE_CONSENSUS_LOWER_WEIGHT
            weight = max(0.10, _relative_condition_weight(entry, current_course, cur_dist) * level_factor)
            if "bridge_priority" in entry:
                diff = float(entry.get("raw_diff", entry.get("diff", 0)) or 0)
                threshold = _bridge_priority_threshold(entry.get("bridge_priority", 0), cur_dist)
            else:
                diff = float(entry.get("diff", 0) or 0)
                threshold = _rel_direct_draw_th(cur_dist)

            margin_total += diff * weight
            margin_weight += weight
            if diff >= threshold:
                votes.append({"direction": 1, "weight": weight})
            elif diff <= -threshold:
                votes.append({"direction": -1, "weight": weight})

        pos_w = sum(vote["weight"] for vote in votes if vote["direction"] > 0)
        neg_w = sum(vote["weight"] for vote in votes if vote["direction"] < 0)
        pos_n = sum(1 for vote in votes if vote["direction"] > 0)
        neg_n = sum(1 for vote in votes if vote["direction"] < 0)
        total_w = pos_w + neg_w
        if total_w <= 0:
            symbol = "="
        elif pos_w > neg_w and pos_w / total_w >= _BRIDGE_CONSENSUS_AGREE:
            symbol = ">"
        elif neg_w > pos_w and neg_w / total_w >= _BRIDGE_CONSENSUS_AGREE:
            symbol = "<"
        else:
            symbol = "="
        margin = (margin_total / margin_weight) if margin_weight else 0.0
        return symbol, selected, margin, pos_n, neg_n

    for i, u in enumerate(current_list):
        for j, v in enumerate(current_list):
            if i >= j: continue

            all_pair_entries = [e for e in (pair_net[u].get(v, []) or []) if isinstance(e, dict)]
            if not all_pair_entries:
                continue
            surge_pair_entries = [e for e in all_pair_entries if e.get("last_run_surge")]
            if surge_pair_entries:
                all_pair_entries = surge_pair_entries

            direct_entries = [e for e in all_pair_entries if "direct_priority" in e]
            bridge_entries = [e for e in all_pair_entries if "bridge_priority" in e]
            best_direct_quality = max(
                (_relative_source_quality(e, current_course, cur_dist)[0] for e in direct_entries),
                default=-1.0,
            )
            best_bridge_quality = max(
                (_relative_source_quality(e, current_course, cur_dist)[0] for e in bridge_entries),
                default=-1.0,
            )
            best_quality = max(best_direct_quality, best_bridge_quality)
            if best_quality < 0:
                continue

            target_direct_entries = [
                e for e in direct_entries
                if _relative_source_quality(e, current_course, cur_dist)[0] == best_quality
            ] if best_direct_quality == best_quality else []

            target_bridge_entries = []
            if best_bridge_quality == best_quality:
                accepted_bridge_qualities = {best_quality}
                if _BRIDGE_CONSENSUS_INCLUDE_ONE_LOWER and best_quality - 1 >= 1:
                    accepted_bridge_qualities.add(best_quality - 1)
                target_bridge_entries = [
                    e for e in bridge_entries
                    if _relative_source_quality(e, current_course, cur_dist)[0] in accepted_bridge_qualities
                ]

            # 条件品質が同じなら、直接対決とbridgeを平等に同じ母集団で評価する。
            # 条件品質が違う場合は、従来通り今回条件に近い高品質の材料だけを採用する。
            if target_direct_entries and target_bridge_entries:
                target_entries = target_direct_entries + target_bridge_entries
                source_kind = "mixed"
            elif target_direct_entries:
                target_entries = target_direct_entries
                source_kind = "direct"
            elif target_bridge_entries:
                target_entries = target_bridge_entries
                source_kind = "bridge"
            else:
                continue
            if not target_entries:
                continue

            # ===== 直接対決 + 隠れ馬経由の混合集約 =====
            if source_kind == "mixed":
                final_sym_u, target_entries, weighted_diff, cons_pos, cons_neg = _mixed_comparison_symbol(
                    target_entries, cur_dist, current_course
                )
                draw_th = _rel_direct_draw_th(cur_dist)
                if final_sym_u == "<":
                    nige_win_u = any(
                        "direct_priority" in e
                        and float(e.get("diff", 0) or 0) >= draw_th
                        and _is_nige(e.get("u_pas", ""))
                        for e in target_entries
                    )
                    if nige_win_u:
                        final_sym_u = ">"
                        nige_dependent.add(u)
                elif final_sym_u == ">":
                    nige_win_v = any(
                        "direct_priority" in e
                        and -float(e.get("diff", 0) or 0) >= draw_th
                        and _is_nige(e.get("v_pas", ""))
                        for e in target_entries
                    )
                    if nige_win_v:
                        final_sym_u = "<"
                        nige_dependent.add(v)

                matchup_matrix[u][v] = final_sym_u
                matchup_matrix[v][u] = inverse_sym(final_sym_u)
                bridge_consensus_detail[(u, v)] = (final_sym_u, cons_pos, cons_neg)
                bridge_consensus_detail[(v, u)] = (inverse_sym(final_sym_u), cons_neg, cons_pos)
                _store_matchup_avg(u, v, target_entries)
                rank_priority = min(_entry_rank_priority(e) for e in target_entries)
                matchup_rank_priority[u][v] = rank_priority
                matchup_rank_priority[v][u] = rank_priority
                matchup_source[u][v] = matchup_source[v][u] = "mixed"
                matchup_margin[u][v] = weighted_diff
                matchup_margin[v][u] = -weighted_diff
                matchup_quality[u][v] = matchup_quality[v][u] = best_quality
                matchup_evidence_count[u][v] = matchup_evidence_count[v][u] = cons_pos + cons_neg
                matchup_selected_entries[u][v] = list(target_entries)
                matchup_selected_entries[v][u] = [_reverse_pair_entry(e) for e in target_entries]
                continue

            # ===== 隠れ馬の合議制（採用レース/経路=1票） =====
            if source_kind == "bridge":
                best_bridge_priority = max(float(e.get("bridge_priority", 0) or 0) for e in target_entries)
                if best_bridge_priority > 0:
                    final_sym_u, cons_pos, cons_neg = _bridge_consensus_symbol(target_entries, cur_dist)
                else:
                    # priority=0の参考値は、強い割引後diffで保守的に判定する。
                    weights = [_relative_condition_weight(e, current_course, cur_dist) for e in target_entries]
                    total_w = sum(weights)
                    avg_ref = (
                        sum(float(e.get("diff", 0) or 0) * w for e, w in zip(target_entries, weights)) / total_w
                        if total_w else 0.0
                    )
                    ref_th = _rel_direct_draw_th(cur_dist)
                    final_sym_u = ">" if avg_ref >= ref_th else ("<" if avg_ref <= -ref_th else "=")
                    cons_pos = 1 if final_sym_u == ">" else 0
                    cons_neg = 1 if final_sym_u == "<" else 0

                matchup_matrix[u][v] = final_sym_u
                matchup_matrix[v][u] = inverse_sym(final_sym_u)
                bridge_consensus_detail[(u, v)] = (final_sym_u, cons_pos, cons_neg)
                bridge_consensus_detail[(v, u)] = (inverse_sym(final_sym_u), cons_neg, cons_pos)
                _store_matchup_avg(u, v, target_entries)
                rank_priority = min(_entry_rank_priority(e) for e in target_entries)
                matchup_rank_priority[u][v] = rank_priority
                matchup_rank_priority[v][u] = rank_priority
                matchup_source[u][v] = matchup_source[v][u] = "bridge"
                # ＝のみ馬救済用: bridge推定着差(raw_diffを条件重みで加重平均, u視点, 正でu先着)
                _bw = _bd = 0.0
                for _e in target_entries:
                    _ew = _relative_condition_weight(_e, current_course, cur_dist)
                    _bd += float(_e.get("raw_diff", _e.get("diff", 0)) or 0) * _ew
                    _bw += _ew
                matchup_margin[u][v] = (_bd / _bw) if _bw > 0 else 0.0
                matchup_margin[v][u] = -matchup_margin[u][v]
                matchup_quality[u][v] = matchup_quality[v][u] = best_quality
                matchup_evidence_count[u][v] = matchup_evidence_count[v][u] = cons_pos + cons_neg
                matchup_selected_entries[u][v] = list(target_entries)
                matchup_selected_entries[v][u] = [_reverse_pair_entry(e) for e in target_entries]
                continue

            # ===== 直接対決の加重集約 =====
            best_bridge_priority = 0

            is_forgiven_u = False
            is_forgiven_v = False
            if current_course == "大井" and _is_ooi_outer(cur_dist):
                for entry in target_entries:
                    if entry["place"] == "大井" and _is_ooi_inner(entry["dist"]):
                        if entry["diff"] < 0: is_forgiven_u = True
                        if -entry["diff"] < 0: is_forgiven_v = True

            final_sym_u, target_entries, weighted_diff, wins, losses = _aggregate_direct_entries(
                target_entries, cur_dist, current_course=current_course
            )
            draw_th = _rel_direct_draw_th(cur_dist)

            # 逃げ救済: u が逃げて差を付けた（diff>=draw_th）レースがあるのに、
            # 逃げれなかった時に大敗して通算が＜/＜＜になっている場合、＞へ昇格させ ※逃 を付ける。
            if final_sym_u == "<":
                nige_win_u = any(
                    float(e.get("diff", 0) or 0) >= draw_th and _is_nige(e.get("u_pas", ""))
                    for e in target_entries
                )
                if nige_win_u:
                    final_sym_u = ">"
                    nige_dependent.add(u)
            # v 視点の逃げ救済: v が逃げて差を付けたのに、通算では u＞v になっている場合、
            # v 側を = に戻し（u から見て＝）、v に ※逃 を付ける。
            elif final_sym_u == ">":
                nige_win_v = any(
                    -float(e.get("diff", 0) or 0) >= draw_th and _is_nige(e.get("v_pas", ""))
                    for e in target_entries
                )
                if nige_win_v:
                    final_sym_u = "<"
                    nige_dependent.add(v)

            if is_forgiven_u and final_sym_u == "<": final_sym_u = "="
            if is_forgiven_v and final_sym_u == ">": final_sym_u = "="

            matchup_matrix[u][v] = final_sym_u
            matchup_matrix[v][u] = inverse_sym(final_sym_u)
            _store_matchup_avg(u, v, target_entries)
            rank_priority = min(_entry_rank_priority(e) for e in target_entries)
            matchup_rank_priority[u][v] = rank_priority
            matchup_rank_priority[v][u] = rank_priority
            matchup_source[u][v] = matchup_source[v][u] = "direct"
            matchup_margin[u][v] = weighted_diff
            matchup_margin[v][u] = -weighted_diff
            matchup_quality[u][v] = matchup_quality[v][u] = best_quality
            matchup_evidence_count[u][v] = matchup_evidence_count[v][u] = len(target_entries)
            matchup_selected_entries[u][v] = list(target_entries)
            matchup_selected_entries[v][u] = [_reverse_pair_entry(e) for e in target_entries]

    chain_matchups = _fill_equal_chain_matchups()

    def _resolve_priority_cycles():
        """3頭循環 (a>b>c>a) を検出し、まず今回と近い内/外サイドの直接対決で再判定する。
        それでも循環が残る場合は、最も信頼度の低い (rank_priority が大きい)
        エッジを = に降格する従来処理へフォールバックする。
        """
        def _is_win(sym):
            return sym == ">"

        draw_adjusted_pairs = set()
        current_field_size = len(current_list)
        for _ in range(len(current_list) * 3 + 2):
            downgrade_edges = []
            seen_cycles = set()
            adjusted_this_round = False
            for a in current_list:
                for b in current_list:
                    if b == a or not _is_win(matchup_matrix.get(a, {}).get(b)):
                        continue
                    for c in current_list:
                        if c in (a, b):
                            continue
                        if not _is_win(matchup_matrix.get(b, {}).get(c)):
                            continue
                        if not _is_win(matchup_matrix.get(c, {}).get(a)):
                            continue
                        cycle_key = tuple(sorted([a, b, c]))
                        if cycle_key in seen_cycles:
                            continue
                        seen_cycles.add(cycle_key)
                        edges = [
                            (a, b, matchup_rank_priority.get(a, {}).get(b, 99)),
                            (b, c, matchup_rank_priority.get(b, {}).get(c, 99)),
                            (c, a, matchup_rank_priority.get(c, {}).get(a, 99)),
                        ]

                        # 同条件の複数直接対決が内外で食い違う時は、今回の側に近いレースを優先。
                        # 対称なペアは1回だけ再集約し、更新後の行列で循環を検査し直す。
                        for src, dst, _priority in edges:
                            pair_key = tuple(sorted((src, dst)))
                            if pair_key in draw_adjusted_pairs:
                                continue
                            if matchup_source.get(src, {}).get(dst) != "direct":
                                continue
                            original_entries = matchup_selected_entries.get(src, {}).get(dst, [])
                            exact_entries = [
                                e for e in original_entries
                                if e.get("place") == current_course
                                and _dist_to_int(e.get("dist", "")) == cur_dist
                            ]
                            if len(exact_entries) < 2:
                                draw_adjusted_pairs.add(pair_key)
                                continue
                            draw_entries = _select_cycle_draw_entries(
                                exact_entries,
                                current_field_size,
                                name_to_umaban.get(src),
                                name_to_umaban.get(dst),
                            )
                            draw_adjusted_pairs.add(pair_key)
                            if len(draw_entries) >= len(exact_entries):
                                continue
                            draw_sym, draw_entries, _avg_diff, _wins, _losses = _aggregate_direct_entries(
                                draw_entries, cur_dist, current_course=current_course
                            )
                            matchup_matrix[src][dst] = draw_sym
                            matchup_matrix[dst][src] = inverse_sym(draw_sym)
                            _store_matchup_avg(src, dst, draw_entries)
                            matchup_evidence_count[src][dst] = len(draw_entries)
                            matchup_evidence_count[dst][src] = len(draw_entries)
                            matchup_selected_entries[src][dst] = list(draw_entries)
                            matchup_selected_entries[dst][src] = [
                                _reverse_pair_entry(e) for e in draw_entries
                            ]
                            adjusted_this_round = True

                        if adjusted_this_round:
                            break
                        if max(e[2] for e in edges) == min(e[2] for e in edges):
                            continue
                        weak = max(edges, key=lambda e: e[2])
                        downgrade_edges.append((weak[0], weak[1]))
                    if adjusted_this_round:
                        break
                if adjusted_this_round:
                    break
            if adjusted_this_round:
                continue
            if not downgrade_edges:
                break
            for src, dst in downgrade_edges:
                if matchup_matrix.get(src, {}).get(dst) == ">":
                    matchup_matrix[src][dst] = "="
                    matchup_matrix[dst][src] = "="
                    matchup_rank_priority[src][dst] = 97
                    matchup_rank_priority[dst][src] = 97
                    chain_matchups.get(src, {}).pop(dst, None)
                    chain_matchups.get(dst, {}).pop(src, None)

    _resolve_priority_cycles()

    comparable_horses = set()
    for u in current_list:
        for v in current_list:
            if u != v and pair_net[u][v]:
                comparable_horses.add(u)
                comparable_horses.add(v)

    all_tiers = {}
    for u in current_list:
        if u not in comparable_horses:
            all_tiers[u] = None

    pool = sorted(comparable_horses, key=lambda h: int(name_to_umaban.get(h, 99)))
    if pool:
        def _tiers_from_matchups(horses):
            if not horses:
                return {}

            tiers = {h: None for h in horses}

            # 勝敗グラフを構築する。u→v は「u が v に勝ち」。データのある馬だけ tier 対象。
            # 「=」はエッジを張らず、データ有り判定にのみ使う。
            dom_G = nx.DiGraph()
            dom_G.add_nodes_from(horses)
            has_data = set()
            for u in horses:
                for v in horses:
                    if u == v:
                        continue
                    rel = matchup_matrix.get(u, {}).get(v)
                    if rel not in (">", "<", "="):
                        continue
                    has_data.add(u)
                    has_data.add(v)
                    if rel == ">":
                        # 確信度ゲート(A): bridgeの薄い単発勝敗は層化に使わない。
                        # has_data には残るので、薄エッジのみの馬は孤立せず中位(B)へ落ち着く。
                        if _REL_CONFIDENCE_GATE:
                            src = matchup_source.get(u, {}).get(v)
                            ev = matchup_evidence_count.get(u, {}).get(v, 0) or 0
                            if src == "bridge" and ev < _REL_DOM_EDGE_MIN_BRIDGE_VOTES:
                                continue
                        dom_G.add_edge(
                            u, v,
                            rank_priority=matchup_rank_priority.get(u, {}).get(v, 99),
                        )

            if not has_data:
                return tiers

            # 循環(三すくみ・4頭以上の循環)内では、信頼度の低い(rank_priority が大きい)
            # エッジを外してから、残った強連結成分(SCC)を1ブロックに圧縮する。
            # これにより任意サイズの循環をまとめて同一 tier へ落とし込める
            # (旧 min_losses ヒューリスティックより厳密。JRA app の _rank_component 相当)。
            pruned_G = _prioritize_strong_conditions_in_cycles(dom_G)

            scc_blocks = list(nx.strongly_connected_components(pruned_G))
            block_by_horse = {}
            for idx, block in enumerate(scc_blocks):
                for h in block:
                    block_by_horse[h] = idx

            # ブロック同士の DAG を作る (循環は圧縮済みなので必ず非循環)。
            block_G = nx.DiGraph()
            block_G.add_nodes_from(range(len(scc_blocks)))
            for src, dst in pruned_G.edges():
                b_src, b_dst = block_by_horse[src], block_by_horse[dst]
                if b_src != b_dst:
                    block_G.add_edge(b_src, b_dst)

            # ブロック DAG を上から層化。自分に勝つブロック(predecessor)が remaining に
            # 残っていないブロックを最上位レイヤとし、上から順にレベル0,1,2…を振る。
            # レベル→ラベルは _rel_level_to_label が担当(多層ONで S→A→B→C→D…、
            # OFFで S/A/B/C にクランプ)。単一ブロックでは旧トポロジカル層化と一致。
            level_by_block = {}
            remaining_blocks = {
                b for b in block_G.nodes()
                if any(h in has_data for h in scc_blocks[b])
            }
            tier_idx = 0
            while remaining_blocks:
                top = {
                    b for b in remaining_blocks
                    if not any(pred in remaining_blocks for pred in block_G.predecessors(b))
                }
                if not top:
                    top = set(remaining_blocks)
                for b in top:
                    level_by_block[b] = tier_idx   # クランプ撤廃。生レベルを保持(多層化)
                remaining_blocks -= top
                tier_idx += 1

            for h in has_data:
                b = block_by_horse.get(h)
                if b is not None and b in level_by_block:
                    tiers[h] = _rel_level_to_label(level_by_block[b])

            def _has_solid_edge(h, neighbors):
                # h と neighbors の間に「確信できる」(同場直接対決 or 合議2票以上)エッジがあるか。
                return any(
                    matchup_source.get(h, {}).get(n) in ("direct", "mixed")
                    or (matchup_source.get(h, {}).get(n) == "bridge"
                        and (matchup_evidence_count.get(h, {}).get(n, 0) or 0) >= _REL_DOM_EDGE_MIN_BRIDGE_VOTES)
                    for n in neighbors
                )

            # 証拠フロア(任意): Sに「確信できる勝ち」が無い馬をAへ。Sの希薄化は防げるが、
            # 隠れ馬経由でしか勝ちが無い「薄データだが強い馬」(JRA転入1戦級等)も落とすので既定OFF。
            if _REL_CONFIDENCE_GATE and _REL_S_FLOOR:
                for h in has_data:
                    if tiers.get(h) == "S" and not _has_solid_edge(h, dom_G.successors(h)):
                        tiers[h] = "A"

            # C天井(A): 最下位tier(C)は「確信できる負け」が1本以上ある馬に限定する。
            # 確信できる負けが無い(薄い/曖昧なだけの)馬はCに沈めず中位Bへ。
            # → Cを「ほぼ馬券圏外」の精度に寄せ、薄データ馬の不当なD/E落ちも防ぐ。
            if _REL_CONFIDENCE_GATE and _REL_C_CEILING:
                for h in has_data:
                    if tiers.get(h) == "C" and not _has_solid_edge(h, dom_G.predecessors(h)):
                        tiers[h] = "B"

            def _tier_idx(label):
                idx = _TIER_LABELS.find(label or "")
                return idx if idx >= 0 else 999

            def _best_tier(labels):
                labels = [label for label in labels if label]
                return min(labels, key=_tier_idx) if labels else None

            def _one_tier_lower(label):
                idx = _tier_idx(label)
                if idx >= 999:
                    return None
                if _REL_MULTI_TIER:
                    return _rel_level_to_label(idx + 1)
                return _rel_level_to_label(min(idx + 1, 3))

            def _equal_source_allowed(h, opp):
                src = matchup_source.get(h, {}).get(opp)
                if src in ("direct", "mixed"):
                    return True
                if src == "bridge":
                    return (
                        _REL_EQUAL_LEAD_LOOSE
                        or (matchup_evidence_count.get(h, {}).get(opp, 0) or 0) >= _REL_DOM_EDGE_MIN_BRIDGE_VOTES
                    )
                return False

            # ＝のみ馬の配置: B固定にはしない。
            # ＝相手への着差がプラスなら、その相手グループと同格へ。
            # 全てマイナスなら、比較した相手グループの一段下へ置く。
            if _REL_EQUAL_LEAD_PROMOTE:
                equal_only = [h for h in has_data
                              if dom_G.in_degree(h) == 0 and dom_G.out_degree(h) == 0]
                base_tiers = dict(tiers)
                for h in equal_only:
                    plus_labels = []
                    flat_labels = []
                    minus_labels = []
                    for opp in horses:
                        if opp == h or matchup_matrix.get(h, {}).get(opp) != "=":
                            continue
                        if not _equal_source_allowed(h, opp):
                            continue
                        opp_label = base_tiers.get(opp)
                        if not opp_label:
                            continue
                        lead = matchup_margin.get(h, {}).get(opp, 0.0) or 0.0
                        if lead > _REL_EQUAL_LEAD_MIN:
                            plus_labels.append(opp_label)
                        elif lead < -_REL_EQUAL_LEAD_MIN:
                            minus_labels.append(opp_label)
                        else:
                            flat_labels.append(opp_label)

                    target = None
                    if plus_labels:
                        target = _best_tier(plus_labels)
                    elif flat_labels:
                        target = _best_tier(flat_labels)
                    elif minus_labels:
                        target = _one_tier_lower(_best_tier(minus_labels))
                    if target:
                        tiers[h] = target

            # 最下位への降格(C純化, ユーザー要望2026/06/24): 「絶対来なそうな馬」をCへ落とす。
            #  (a) 出走メンバーの過半数と比較があり、その全てに負けている馬。
            #  (b) 勝ちが無く、＝相手に着差マイナス(後着)があって先着が無い馬。
            if _REL_C_DEMOTE:
                _half = len(horses) / 2.0
                for h in has_data:
                    rels = {o: matchup_matrix.get(h, {}).get(o) for o in horses if o != h}
                    cmp_opps = [o for o, r in rels.items() if r in (">", "<", "=")]
                    if not cmp_opps:
                        continue
                    wins = [o for o in cmp_opps if rels[o] == ">"]
                    losses = [o for o in cmp_opps if rels[o] == "<"]
                    eqs = [o for o in cmp_opps if rels[o] == "="]
                    if len(cmp_opps) >= _half and len(losses) == len(cmp_opps):
                        tiers[h] = "C"   # (a) 過半数と比較あり全敗
                    elif not wins and eqs:
                        leads = [matchup_margin.get(h, {}).get(o, 0.0) or 0.0 for o in eqs]
                        if max(leads) <= _REL_EQUAL_LEAD_MIN and min(leads) < -_REL_EQUAL_LEAD_MIN:
                            tiers[h] = "C"   # (b) 勝ちなしで＝後着あり・先着なし

            # 4段階クランプ時の三軍内救済:
            # 生レベルでは上でも、C以下が全部「三軍」に丸められると、明確に先着している馬と
            # 先着を許した馬が同じ三軍に見える。三軍内で複数馬に勝ち、三軍内の負けが無い馬は
            # 一段上の二軍へ上げる。
            if _REL_C_INTERNAL_WIN_PROMOTE and not _REL_MULTI_TIER:
                promoted_from_c = {}
                for h in has_data:
                    if tiers.get(h) != "C":
                        continue
                    c_opps = [o for o in horses if o != h and tiers.get(o) == "C"]
                    real_wins = [
                        o for o in c_opps
                        if matchup_matrix.get(h, {}).get(o) == ">"
                        and matchup_source.get(h, {}).get(o) in ("direct", "bridge", "mixed")
                    ]
                    real_losses = [
                        o for o in c_opps
                        if matchup_matrix.get(h, {}).get(o) == "<"
                        and matchup_source.get(h, {}).get(o) in ("direct", "bridge", "mixed")
                    ]
                    if len(real_wins) >= _REL_C_INTERNAL_WIN_MIN and not real_losses:
                        promoted_from_c[h] = "B"
                tiers.update(promoted_from_c)

            return tiers

        all_tiers.update(_tiers_from_matchups(pool))

        # 参考値bridge を使った馬で 三すくみ (A>B>C>A) や、参考値の影響で
        # 既存comparable馬のtierが動いた(パワーバランスのズレ)場合、
        # 参考値依存の馬を比較不能に戻す。比較ルート自体は pair_net に残り、
        # _build_relative_html では他と同様に表示される(is_reference フラグで脚注を付ける)。
        if reference_horses:
            def _entry_is_reference(entries):
                return any(isinstance(e, dict) and e.get("is_reference") for e in (entries or []))

            cycle_horses = set()
            names_list = list(current_list)
            for a in names_list:
                for b in names_list:
                    if b == a:
                        continue
                    if matchup_matrix.get(a, {}).get(b) != ">":
                        continue
                    for c in names_list:
                        if c == a or c == b:
                            continue
                        if matchup_matrix.get(b, {}).get(c) != ">":
                            continue
                        if matchup_matrix.get(c, {}).get(a) != ">":
                            continue
                        # a > b > c > a の閉路
                        if (_entry_is_reference(pair_net[a].get(b))
                            or _entry_is_reference(pair_net[b].get(c))
                            or _entry_is_reference(pair_net[c].get(a))):
                            cycle_horses.update([a, b, c])

            # パワーバランスのズレ判定: 参考値ぬきで matchup を組み直して
            # 既存comparable馬(no_compare_pre 以外)のtierが変わるか確認する。
            no_compare_pre_set = set(no_compare_pre)
            balance_shift = False
            try:
                baseline_pool = sorted(
                    [h for h in pool if h not in no_compare_pre_set],
                    key=lambda h: int(name_to_umaban.get(h, 99)),
                )
                baseline_matchup = {h: {} for h in baseline_pool}
                for h in baseline_pool:
                    for k in baseline_pool:
                        if h == k:
                            continue
                        # 参考値以外のエントリのみで sym を再判定
                        sym_full = matchup_matrix.get(h, {}).get(k)
                        if sym_full is None:
                            continue
                        non_ref_entries = [
                            e for e in (pair_net[h].get(k) or [])
                            if isinstance(e, dict) and not e.get("is_reference")
                        ]
                        if non_ref_entries:
                            baseline_matchup[h][k] = sym_full
                # 簡易再tier: 単純に各馬の wins/losses/draws を再計算
                baseline_tiers = {}
                stats_b = {h: {"wins": set(), "losses": set(), "draws": set()} for h in baseline_pool}
                for h in baseline_pool:
                    for k in baseline_pool:
                        if h == k:
                            continue
                        rel = baseline_matchup.get(h, {}).get(k)
                        if rel == ">":
                            stats_b[h]["wins"].add(k)
                        elif rel == "<":
                            stats_b[h]["losses"].add(k)
                        elif rel == "=":
                            stats_b[h]["draws"].add(k)
                # ベースラインでの全comparable馬のtierが同じか確認
                for h in baseline_pool:
                    full_tier = all_tiers.get(h)
                    # 参考値が入る前の状態でも各馬は元々pool内なので、
                    # 単純化: comparable馬のtierが None になっていたら shift とみなす
                    if full_tier is None and stats_b[h]["wins"] | stats_b[h]["losses"] | stats_b[h]["draws"]:
                        balance_shift = True
                        break
            except Exception:
                pass

            to_revert = cycle_horses & no_compare_pre_set
            if balance_shift:
                # 既存comparable馬のtierが揺らいだ場合、参考値依存の馬は全て比較不能に戻す
                to_revert = to_revert | (set(reference_horses) & no_compare_pre_set)
            for h in to_revert:
                all_tiers[h] = None

    # 上位tierとの馬対馬比較に反する二軍/三軍を救済する。
    # 所属変更後は、そのtierの既存配点を相対点にも適用する。
    matchup_tier_promotions = {}
    if _REL_MATCHUP_TIER_CONSISTENCY:
        all_tiers, matchup_tier_promotions = _apply_matchup_tier_consistency(
            all_tiers, matchup_matrix, matchup_source
        )

    # 優劣記号は > / < / = のみ。大差専用の追加ロックは使わない。
    strong_comparison_locks = []
    strong_rank_notes = {}

    def _collect_same_condition_locks():
        """今回と同場・同距離の直接対決で明確な勝敗があるペアを集める
        (JRA app の collect_same_condition_locks 相当)。
        最終グレードで負けた馬が勝った馬を逆転しないためのハードガード用。
        対戦表スクレイプ(adjust_grades_by_matchups)と違い、内部の時間差・
        外れ値除去・逃げ救済を通した matchup_matrix の最終判定に基づくため、
        単発の同条件完敗でも発火できる。
        return: [{"winner": 正規化名, "loser": 正規化名, "strength": "clear"|"strong"}]
        """
        locks = []
        seen = set()
        recent_days = RELATIVE_PAST_DAYS_EXTENDED if _is_long_distance(cur_dist) else BRIDGE_RECENT_DAYS
        now_dt = datetime.now()
        for u in current_list:
            for v in current_list:
                if u == v:
                    continue
                sym = matchup_matrix.get(u, {}).get(v)
                if sym != ">":
                    continue
                pair_key = tuple(sorted([u, v]))
                if pair_key in seen:
                    continue
                # u が v に勝ち。今回と同場・同距離の「直接対決」エビデンスがあるか。
                has_evidence = False
                for e in (pair_net[u].get(v) or []):
                    if not isinstance(e, dict) or e.get("bridge_priority"):
                        continue  # bridge は除外、直接対決のみをロック根拠にする
                    if e.get("place") != current_course:
                        continue
                    if _dist_to_int(e.get("dist", "")) != cur_dist:
                        continue
                    if float(e.get("diff", 0) or 0) <= 0:
                        continue  # u 有利(diff>0)のレースのみ
                    d_str = e.get("date")
                    dt = parse_date(d_str) if isinstance(d_str, str) else d_str
                    if isinstance(dt, datetime) and dt != datetime.min and (now_dt - dt).days > recent_days:
                        continue
                    has_evidence = True
                    break
                if not has_evidence:
                    continue
                seen.add(pair_key)
                locks.append({
                    "winner": _norm_horse_name(u),
                    "loser": _norm_horse_name(v),
                    "strength": "clear",
                })

        # 追加ハードガード: 今回と同場・同距離の「最新」直接対決で着順差が
        # POSITION_LOCK_GAP(5)以上付いたペアは、内部判定(逃げ救済・外れ値除去・
        # 通算多数決)で勝敗が緩和されていても「負けた馬が勝った馬を上回らない」を保証する。
        # 同じ2頭がより新しい同条件レースで逆転していれば、その最新結果を採用する。
        POSITION_LOCK_GAP = 5
        for u in current_list:
            for v in current_list:
                if u >= v:
                    continue
                latest = None  # (dt, u_rank, v_rank)
                for a, b in ((u, v), (v, u)):
                    if not G.has_edge(a, b):
                        continue
                    for hi in G[a][b]["history"]:
                        if hi.get("place") != current_course:
                            continue
                        if _dist_to_int(hi.get("dist", "")) != cur_dist:
                            continue
                        u_r = _hist_rank_for(hi, u)
                        v_r = _hist_rank_for(hi, v)
                        if u_r >= 99 or v_r >= 99:
                            continue
                        d_str = hi.get("date")
                        dt = parse_date(d_str) if isinstance(d_str, str) else d_str
                        if not isinstance(dt, datetime) or dt == datetime.min:
                            continue
                        if (now_dt - dt).days > recent_days:
                            continue
                        if latest is None or dt > latest[0]:
                            latest = (dt, u_r, v_r)
                if latest is None:
                    continue
                _, u_r, v_r = latest
                if abs(u_r - v_r) < POSITION_LOCK_GAP:
                    continue
                w, l = (u, v) if u_r < v_r else (v, u)
                locks.append({
                    "winner": _norm_horse_name(w),
                    "loser": _norm_horse_name(l),
                    "strength": "clear",
                    "source": "position",
                })

        # 追加ハードガード3: 今回と同じ競馬場(距離不問)の「最新」直接対戦で
        # 1.0秒以上の大差負けを喫した馬は、そのレースで先着した馬より上の評価にしない。
        # 最新対戦のみを根拠にするため、その後同場で勝ち返していれば発火しない。
        # ※raw_diff は ±1.2秒でキャップ済みのため 1.0秒閾値はそのまま判定できる。
        BORO_MAKE_TIME_GAP = 1.0
        for u in current_list:
            for v in current_list:
                if u >= v:
                    continue
                latest = None  # (dt, adv_u) adv_u>0 = u が先着
                for a, b in ((u, v), (v, u)):
                    if not G.has_edge(a, b):
                        continue
                    for hi in G[a][b]["history"]:
                        if hi.get("place") != current_course:
                            continue
                        d_str = hi.get("date")
                        dt = parse_date(d_str) if isinstance(d_str, str) else d_str
                        if not isinstance(dt, datetime) or dt == datetime.min:
                            continue
                        if (now_dt - dt).days > recent_days:
                            continue
                        adv_u = -hi["raw_diff"] if hi.get("h1_name") == u else hi["raw_diff"]
                        if latest is None or dt > latest[0]:
                            latest = (dt, adv_u)
                if latest is None:
                    continue
                _, adv_u = latest
                if abs(adv_u) < BORO_MAKE_TIME_GAP:
                    continue
                w, l = (u, v) if adv_u > 0 else (v, u)
                locks.append({
                    "winner": _norm_horse_name(w),
                    "loser": _norm_horse_name(l),
                    "strength": "clear",
                    "source": "boro_time",
                })
        return locks

    same_condition_locks = _collect_same_condition_locks()
    # 強比較ロックも最終評価グレードまで伝播させる。
    # 反証経路があるペアは enforce=False なので対象外。
    existing_lock_keys = {
        (lock.get("winner"), lock.get("loser"), lock.get("strength"))
        for lock in same_condition_locks
    }
    for lock in strong_comparison_locks:
        if not lock.get("enforce"):
            continue
        out_lock = {
            "winner": _norm_horse_name(lock.get("winner", "")),
            "loser": _norm_horse_name(lock.get("loser", "")),
            "strength": "strong",
            "source": "strong_comparison",
        }
        key = (out_lock["winner"], out_lock["loser"], out_lock["strength"])
        if key not in existing_lock_keys:
            same_condition_locks.append(out_lock)
            existing_lock_keys.add(key)

    # 内部の最終勝敗判定(タイム差・外れ値除去・最新同条件優先を通した記号)を
    # 正規化名で持ち出す。対戦表スクレイプ(生着順の累積負け)による降格が
    # この判定と矛盾する場合の拒否権として adjust_grades_by_matchups で使う。
    internal_verdicts = {}
    for u in current_list:
        for v, sym in (matchup_matrix.get(u, {}) or {}).items():
            if sym:
                internal_verdicts.setdefault(_norm_horse_name(u), {})[_norm_horse_name(v)] = sym

    relative_scores = {}
    relative_tiers = {}
    # 多層化時はそのレースで実際に使われた最下位レベル(ラベルindex)を基準に
    # 50〜15を線形配点する。4段階時は固定index=3(C)で旧配点と一致。
    if _REL_MULTI_TIER:
        _used_idx = [_TIER_LABELS.find(t) for t in all_tiers.values() if t]
        race_max_idx = max([i for i in _used_idx if i >= 0], default=0)
    else:
        race_max_idx = 3
    for umaban, hname in umaban_to_name.items():
        if hname in all_tiers and all_tiers[hname] is not None:
            relative_tiers[umaban] = all_tiers[hname]
            _sc = _rel_label_to_score(all_tiers[hname], race_max_idx)
            relative_scores[umaban] = _sc if _sc is not None else NO_DATA_SCORE
        else:
            relative_tiers[umaban] = None
            relative_scores[umaban] = NO_DATA_SCORE

    no_same_venue_names = {n for n in current_list if not _has_same_venue_link(n)}

    relative_html = _build_relative_html(
        G, {}, all_tiers, umaban_to_name, name_to_umaban,
        current_course, current_dist, DRAW_TH, r_num, None, matchup_matrix,
        fallback_sparse_names, chain_matchups, no_same_venue_names,
        nige_dependent, post_rest_threshold, track_profiles, reference_horses,
        bridge_consensus_detail, strong_rank_notes, matchup_source, horses_data
    )

    if _RELATIVE_DEBUG is not None:
        _RELATIVE_DEBUG.clear()
        def _slim_entries(entries):
            out = []
            for e in (entries or []):
                if not isinstance(e, dict):
                    continue
                if "direct_priority" in e:
                    kind, pri = "direct", e.get("direct_priority")
                elif "bridge_priority" in e:
                    kind, pri = "bridge", e.get("bridge_priority")
                else:
                    kind, pri = "?", None
                out.append({"kind": kind, "priority": pri, "diff": e.get("diff"),
                            "place": e.get("place"), "dist": e.get("dist")})
            return out
        _RELATIVE_DEBUG.update({
            "matchup_matrix": {u: dict(d) for u, d in matchup_matrix.items()},
            "matchup_avg": {u: dict(d) for u, d in matchup_avg.items()},
            "matchup_source": {u: dict(d) for u, d in matchup_source.items()},
            "matchup_quality": {u: dict(d) for u, d in matchup_quality.items()},
            "matchup_evidence_count": {u: dict(d) for u, d in matchup_evidence_count.items()},
            "strong_comparison_locks": list(strong_comparison_locks),
            "strong_rank_notes": dict(strong_rank_notes),
            "tiers_by_name": dict(all_tiers),
            "matchup_tier_promotions": dict(matchup_tier_promotions),
            "umaban_to_name": dict(umaban_to_name),
            "current_list": list(current_list),
            "pair_net": {u: {v: _slim_entries(es) for v, es in dd.items()}
                         for u, dd in pair_net.items()},
        })

    return relative_scores, relative_tiers, relative_html, same_condition_locks, internal_verdicts

def _build_relative_html(G, all_scores, all_tiers, umaban_to_name, name_to_umaban, current_course, current_dist, draw_th, r_num, runner_G=None, matchup_matrix=None, fallback_sparse_names=None, chain_matchups=None, no_same_venue_names=None, nige_dependent=None, post_rest_threshold=None, track_profiles=None, reference_horses=None, bridge_consensus_detail=None, strong_rank_notes=None, matchup_source=None, horses_data=None):
    bridge_consensus_detail = bridge_consensus_detail or {}
    strong_rank_notes = strong_rank_notes or {}
    matchup_source = matchup_source or {}
    cur_dist_val = _dist_to_int(current_dist)
    current_names = set(umaban_to_name.values())
    fallback_sparse_names = set(fallback_sparse_names or [])
    no_same_venue_names = set(no_same_venue_names or [])
    chain_matchups = chain_matchups or {}
    nige_dependent = set(nige_dependent or [])
    track_profiles = track_profiles or {}
    reference_horses = set(reference_horses or [])
    horses_data = horses_data or {}

    def _track_marker_text(hname):
        """馬名横の小さな脚注「※ 大◎ 船- 川- 浦○」を組み立てる。今回開催場を先頭に。"""
        prof = track_profiles.get(hname)
        if not prof:
            return ""
        order = [current_course] + [p for p in _NANKAN_PROFILE_PLACES if p != current_course]
        # 南関4場以外で今回開催の場合(JRA等)は今回場を先頭から落として表示
        order = [p for p in order if p in _NANKAN_PROFILE_PLACES]
        if not order:
            return ""
        exp_count = sum(1 for p in _NANKAN_PROFILE_PLACES if prof.get(p, {}).get("n", 0) > 0)
        parts = []
        for p in order:
            sym = _track_marker_symbol(prof, p, experienced_place_count=exp_count)
            parts.append(f"{p[0]}{sym}")
        return " ".join(parts)
    post_rest_threshold = dict(post_rest_threshold or {})

    def _safe_rank(rank):
        try:
            return int(rank)
        except (TypeError, ValueError):
            return 99

    def _hist_rank_for(hi, hname):
        if hi.get("h1_name") == hname:
            return _safe_rank(hi.get("h1_rank"))
        if hi.get("h2_name") == hname:
            return _safe_rank(hi.get("h2_rank"))
        return 99

    def _hist_gate_for(hi, hname):
        if hi.get("h1_name") == hname:
            return hi.get("h1_gate")
        if hi.get("h2_name") == hname:
            return hi.get("h2_gate")
        return None

    def _circled_gate(gate):
        try:
            g = int(gate)
        except (TypeError, ValueError):
            return ""
        return _to_circled(g) if 1 <= g <= 20 else f"({g})"

    def _rank_sort_key(rank, dt):
        dt_obj = parse_date(dt) if isinstance(dt, str) else (dt if isinstance(dt, datetime) else datetime.min)
        return (_safe_rank(rank), -dt_obj.timestamp() if dt_obj != datetime.min else 0)

    cycle_pairs = set()
    if matchup_matrix:
        names = list(matchup_matrix.keys())
        for i, a in enumerate(names):
            for j, b in enumerate(names):
                if i == j: continue
                if matchup_matrix[a].get(b) != ">": continue
                for k, c in enumerate(names):
                    if k == i or k == j: continue
                    if (matchup_matrix[b].get(c) == ">" and
                        matchup_matrix[c].get(a) == ">"):
                        cycle_pairs.update({(a, b), (b, c), (c, a)})

    no_compare_horses = [(u, n) for u, n in umaban_to_name.items() if all_tiers.get(n) is None]

    def _last_run_note(umaban):
        horse = horses_data.get(str(umaban)) or horses_data.get(umaban) or {}
        hist = horse.get("hist") or []
        last = next((h for h in hist if isinstance(h, dict)), None)
        if not last:
            return ""
        race_name = str(last.get("race_name", "") or "").strip()
        date = str(last.get("date", "") or "").strip()
        place = str(last.get("place", "") or "").strip()
        dist = str(last.get("dist", "") or "").strip()
        pas = str(last.get("pas", "") or "").strip()
        rank = last.get("rank")
        pop = last.get("pop")
        cond = f"{place}{dist}" if (place or dist) else ""
        parts = []
        if race_name:
            parts.append(race_name)
        if date:
            parts.append(date)
        if cond:
            parts.append(cond)
        if pas:
            parts.append(f"通過{pas}")
        # 人気・着順（取得できていれば）。例: 3人気1着
        res = ""
        if pop is not None:
            res += f"{pop}人気"
        if rank is not None:
            res += f"{rank}着"
        if res:
            parts.append(res)
        if not parts:
            return ""
        return "前走: " + " ".join(parts)

    hidden_cache = {}
    hidden_nodes = [n for n in G.nodes() if n not in current_names]

    def _get_direct_opponents_set(hname):
        direct_opps = set()
        for opp_n in current_names:
            if hname == opp_n: continue
            h_a, h_b = (hname, opp_n) if hname < opp_n else (opp_n, hname)
            if G.has_edge(h_a, h_b): direct_opps.add(opp_n)
        return direct_opps

    for name_a in current_names:
        direct_opps_a = _get_direct_opponents_set(name_a)
        for name_b in current_names:
            if name_a >= name_b: continue
            selected_source = (matchup_source.get(name_a, {}) or {}).get(name_b)
            if selected_source == "direct" or (selected_source is None and name_b in direct_opps_a):
                continue
            allow_other_venue_fallback = (name_a in no_same_venue_names) and (name_b in no_same_venue_names)
            allow_reference_bridge = (name_a in reference_horses or name_b in reference_horses)

            candidates = []
            for h in hidden_nodes:
                u_h_hist = G[name_a][h]["history"] if G.has_edge(name_a, h) else (G[h][name_a]["history"] if G.has_edge(h, name_a) else [])
                h_v_hist = G[h][name_b]["history"] if G.has_edge(h, name_b) else (G[name_b][h]["history"] if G.has_edge(name_b, h) else [])

                if not u_h_hist or not h_v_hist: continue

                bridge_pairs = []
                bridge_pairs_pre_rest = []  # 休み明けフィルタで弾かれた候補（全部弾かれた時のフォールバック）
                a_thr = post_rest_threshold.get(name_a)
                b_thr = post_rest_threshold.get(name_b)
                for hi_uh in u_h_hist:
                    for hi_hv in h_v_hist:
                        # 隠れ馬経由のみ適用: 本馬(name_a)や相手(name_b)が10着以下のレースは参考外
                        if _BRIDGE_FILTER_UV_RANK10 and (_hist_rank_for(hi_uh, name_a) >= 10 or _hist_rank_for(hi_hv, name_b) >= 10):
                            continue

                        # 隠れ馬経由は各レッグが半年以内(or ≥1800m)のみ採用
                        if not _is_bridge_leg_recent_enough(hi_uh.get("date", ""), hi_uh.get("dist", "")):
                            continue
                        if not _is_bridge_leg_recent_enough(hi_hv.get("date", ""), hi_hv.get("dist", "")):
                            continue

                        # 休み明け改善フィルタ: name_a-leg は a の閾値、name_b-leg は b の閾値で
                        is_pre_rest_disp = False
                        if a_thr:
                            dt_a_leg = parse_date(hi_uh.get("date", "")) if isinstance(hi_uh.get("date", ""), str) else hi_uh.get("date")
                            if isinstance(dt_a_leg, datetime) and dt_a_leg != datetime.min and dt_a_leg < a_thr:
                                is_pre_rest_disp = True
                        if b_thr:
                            dt_b_leg = parse_date(hi_hv.get("date", "")) if isinstance(hi_hv.get("date", ""), str) else hi_hv.get("date")
                            if isinstance(dt_b_leg, datetime) and dt_b_leg != datetime.min and dt_b_leg < b_thr:
                                is_pre_rest_disp = True

                        p_uh = hi_uh["place"]
                        d_uh = hi_uh["dist"]
                        p_hv = hi_hv["place"]
                        d_hv = hi_hv["dist"]

                        d_uh_val = -hi_uh["raw_diff"] if G.has_edge(name_a, h) else hi_uh["raw_diff"]
                        d_hv_val = -hi_hv["raw_diff"] if G.has_edge(h, name_b) else hi_hv["raw_diff"]
                        est_diff = d_uh_val + d_hv_val
                        bridge_rank = min(_hist_rank_for(hi_uh, name_a), _hist_rank_for(hi_hv, name_b))

                        bridge_priority = _hidden_bridge_priority(p_uh, d_uh, p_hv, d_hv, current_dist, current_course)
                        # A案: bridge_priority=0 でも、両馬とも同場経由ゼロかつ両レッグ他場・同距離=今回距離 の時のみ通す
                        # 加えて、2頭以上比較不能(reference_horses)時はpriority=0も「参考値」として表示する
                        is_relaxed_fallback = False
                        is_reference_disp = False
                        if not bridge_priority:
                            # priority=0 は各レッグのコーナー数(1turn/2turn+)一致を必須
                            if not _is_turn_match_to_current(p_uh, d_uh, current_course, current_dist):
                                continue
                            if not _is_turn_match_to_current(p_hv, d_hv, current_course, current_dist):
                                continue
                            same_dist_other_venue = (
                                p_uh != current_course and p_hv != current_course
                                and str(d_uh) == str(current_dist) and str(d_hv) == str(current_dist)
                            )
                            if allow_other_venue_fallback and same_dist_other_venue:
                                is_relaxed_fallback = True
                            elif allow_reference_bridge:
                                is_reference_disp = True
                            else:
                                continue
                        is_current_same = (
                            p_uh == current_course
                            and p_hv == current_course
                            and str(d_uh) == str(current_dist)
                            and str(d_hv) == str(current_dist)
                        )
                        is_current_venue = (p_uh == current_course and p_hv == current_course)
                        is_current_layout = (
                            is_current_venue
                            and _is_same_track_layout(current_course, d_uh, current_dist)
                            and _is_same_track_layout(current_course, d_hv, current_dist)
                        )
                        if is_reference_disp:
                            bridge_note = "参考"
                        elif is_relaxed_fallback:
                            bridge_note = "他場同距離"
                        elif is_current_same:
                            bridge_note = ""
                        else:
                            bridge_note = _hidden_bridge_note(bridge_priority, is_long=_is_long_distance(current_dist))
                        entry_dict = {
                            "est": est_diff,
                            "uh_hist": hi_uh,
                            "hv_hist": hi_hv,
                            "leg_diffs": {
                                "a_bridge": d_uh_val,
                                "bridge_b": d_hv_val,
                            },
                            "is_strict": is_current_same,
                            "is_current_layout": is_current_layout,
                            "is_current_venue": is_current_venue,
                            "bridge_priority": bridge_priority,
                            "bridge_note": bridge_note,
                            "is_relaxed_fallback": is_relaxed_fallback,
                            "is_reference_disp": is_reference_disp,
                            "rank": bridge_rank,
                        }
                        if is_pre_rest_disp:
                            bridge_pairs_pre_rest.append(entry_dict)
                        else:
                            bridge_pairs.append(entry_dict)

                # 休み明けフィルタで全部弾かれたらフォールバック
                if not bridge_pairs and bridge_pairs_pre_rest:
                    bridge_pairs = bridge_pairs_pre_rest

                if not bridge_pairs:
                    continue

                best_priority = max(vp["bridge_priority"] for vp in bridge_pairs)
                target_pairs = [vp for vp in bridge_pairs if vp["bridge_priority"] == best_priority]

                def _pair_latest_dt(vp):
                    dt_uh = parse_date(vp["uh_hist"].get("date", "")) if isinstance(vp["uh_hist"].get("date", ""), str) else vp["uh_hist"].get("date", datetime.min)
                    dt_hv = parse_date(vp["hv_hist"].get("date", "")) if isinstance(vp["hv_hist"].get("date", ""), str) else vp["hv_hist"].get("date", datetime.min)
                    return min(dt_uh, dt_hv) if dt_uh and dt_hv else (dt_uh or dt_hv or datetime.min)

                target_pairs.sort(key=lambda vp: _rank_sort_key(vp.get("rank", 99), _pair_latest_dt(vp)))
                best_pair = target_pairs[0]
                if best_pair.get("is_relaxed_fallback"):
                    discount = 0.3
                else:
                    discount = _BRIDGE_DISCOUNT_BY_PRIORITY.get(best_priority, 0.5)
                bp_place = best_pair["uh_hist"].get("place", "")
                fa = _track_discount_factor(track_profiles.get(name_a) if track_profiles else None, bp_place)
                fb = _track_discount_factor(track_profiles.get(name_b) if track_profiles else None, bp_place)
                track_factor = min(fa, fb)
                avg_est = best_pair["est"] * discount * track_factor

                u_h_same = (best_pair["uh_hist"]["place"] == current_course and str(best_pair["uh_hist"]["dist"]) == str(current_dist))
                h_v_same = (best_pair["hv_hist"]["place"] == current_course and str(best_pair["hv_hist"]["dist"]) == str(current_dist))
                u_h_layout = (
                    best_pair["uh_hist"]["place"] == current_course
                    and _is_same_track_layout(current_course, best_pair["uh_hist"]["dist"], current_dist)
                )
                h_v_layout = (
                    best_pair["hv_hist"]["place"] == current_course
                    and _is_same_track_layout(current_course, best_pair["hv_hist"]["dist"], current_dist)
                )
                is_same_for_symbol = (u_h_same and h_v_same) or (u_h_layout and h_v_layout)
                cond_score = (1 if u_h_layout else 0) + (1 if h_v_layout else 0)
                if not best_pair["is_strict"]: cond_score -= 1

                def _ym(date_str):
                    dt = parse_date(date_str) if isinstance(date_str, str) else date_str
                    if isinstance(dt, datetime) and dt != datetime.min:
                        return f"{dt.year % 100:02d}/{dt.month}"
                    return ""

                uh_ym = _ym(best_pair['uh_hist'].get('date', ''))
                hv_ym = _ym(best_pair['hv_hist'].get('date', ''))
                uh_label = f"{uh_ym}{best_pair['uh_hist']['place'][0]}{best_pair['uh_hist']['dist']}"
                hv_label = f"{hv_ym}{best_pair['hv_hist']['place'][0]}{best_pair['hv_hist']['dist']}"
                place_dist = uh_label if uh_label == hv_label else f"{uh_label}/{hv_label}"
                bridge_note = best_pair.get("bridge_note", "") if not best_pair["is_strict"] else ""
                if bridge_note: place_dist = f"{bridge_note}({place_dist})"

                uh_url = best_pair['uh_hist'].get('url', '')
                hv_url = best_pair['hv_hist'].get('url', '')
                best_dt = _pair_latest_dt(best_pair)
                leg_gates = {
                    "name_a": _hist_gate_for(best_pair["uh_hist"], name_a),
                    "bridge_on_a": _hist_gate_for(best_pair["uh_hist"], h),
                    "bridge_on_b": _hist_gate_for(best_pair["hv_hist"], h),
                    "name_b": _hist_gate_for(best_pair["hv_hist"], name_b),
                }
                candidates.append((h, avg_est, cond_score, is_same_for_symbol, place_dist, best_pair["is_strict"], uh_url, hv_url, uh_label, hv_label, bridge_note, best_priority, best_pair.get("rank", 99), best_dt, best_pair.get("is_current_venue", False), leg_gates, best_pair.get("leg_diffs", {})))

            if not candidates: continue
            sprint_hidden_draw = 0.5 if _is_sprint_distance(current_dist) else None
            candidates.sort(
                key=lambda c: (
                    c[11],
                    1 if c[11] == 2 and c[14] else (0 if c[11] == 2 else 1),
                    -_safe_rank(c[12]),
                    c[13] if isinstance(c[13], datetime) else datetime.min,
                    c[2],
                    abs(c[1]) >= ((sprint_hidden_draw or 0.55) if c[5] else (sprint_hidden_draw or 0.75)),
                    -abs(c[1])
                ),
                reverse=True
            )
            best = candidates[0]
            if abs(best[1]) < 2.0:
                hidden_cache[(name_a, name_b)] = best

    html_parts = ["<div style='font-size:0.92em; line-height:1.6;'>"]

    def _get_direct_opponents(hname):
        direct_opps = set()
        for opp_n in current_names:
            if hname == opp_n: continue
            h_a, h_b = (hname, opp_n) if hname < opp_n else (opp_n, hname)
            if G.has_edge(h_a, h_b): direct_opps.add(opp_n)
        return direct_opps

    def _current_horse_label(hname):
        return f"[{name_to_umaban.get(hname, '?')}:{hname}]"

    def _render_matchup_note(opps, label="推移比較"):
        if not opps:
            return ""
        display_sym = {">": "＞", "=": "＝", "<": "＜"}
        parts = []
        for sym, color in [(">", "#27ae60"), ("=", "#777"), ("<", "#c0392b")]:
            names = [
                _current_horse_label(opp)
                for opp, comp in opps
                if comp and comp.get("symbol") == sym
            ]
            if names:
                parts.append(f"<span style='color:{color};'>本馬 {display_sym.get(sym, sym)} {'、'.join(names)}</span>")
        if not parts:
            return ""
        return f"<div style='margin-left:45px; font-size:0.76em; color:#666; line-height:1.35;'>※{label}：{' / '.join(parts)}</div>"

    def _render_horse_block(umaban, hname):
        parts = []
        mark_btn = f'<span class="mark-btn" data-uma="{umaban}" data-race="{r_num}" data-mark="" onclick="cycleMark(this)"></span>'
        nige_mark = " <span style='color:#d35400; font-size:0.85em;'>※逃</span>" if hname in nige_dependent else ""
        tk_txt = _track_marker_text(hname)
        track_mark = f" <span style='color:#888; font-size:0.78em; margin-left:6px;'>※ {tk_txt}</span>" if tk_txt else ""
        ref_mark = " <span style='color:#9b59b6; font-size:0.78em; margin-left:4px;'>※参考値</span>" if hname in reference_horses else ""
        strong_note = strong_rank_notes.get(hname, "")
        strong_mark = f" <span style='color:#c0392b; font-size:0.76em; margin-left:4px;'>{html.escape(strong_note)}</span>" if strong_note else ""
        parts.append(f"<div class='horse-header'>{mark_btn} <strong>[{umaban}] {hname}</strong>{nige_mark}{track_mark}{ref_mark}{strong_mark}</div>")

        race_groups = {}
        for opp_u, opp_n in umaban_to_name.items():
            if hname == opp_n: continue
            h_a, h_b = (hname, opp_n) if hname < opp_n else (opp_n, hname)
            if G.has_edge(h_a, h_b):
                for hi in G[h_a][h_b]["history"]:
                    r_key = (hi['date'], hi['place'], hi['dist'], hi.get('url', ''))
                    if r_key not in race_groups:
                        race_groups[r_key] = []
                    adv = -hi['raw_diff'] if hname == h_a else hi['raw_diff']
                    race_groups[r_key].append((opp_u, adv))


        def race_priority(item):
            (r_date, r_place, r_dist, r_url), _ = item
            priority = _race_condition_priority(r_place, r_dist, current_course, cur_dist_val)
            return (priority, r_date)

        def _diff_symbol_and_color(adv, is_same_condition=True, is_strict=True):
            # 隠れ馬経由表示用フォールバック(通常は matchup_matrix の記号で上書きされる)。
            # 同条件(priority5相当)も隠れ馬経由なので、直接対決基準+0.2秒に揃える。
            aa = abs(adv)
            draw_limit = _bridge_priority_threshold(5, current_dist)
            if aa < draw_limit: return "＝", "#888"
            return ("＞" if adv > 0 else "＜"), ("#27ae60" if adv > 0 else "#e74c3c")

        def fmt_opps_by_tier(opps_list, is_same_condition=True, is_strict=True):
            # 直接対決の実着差基準に揃える。
            draw_limit = _rel_direct_draw_th(current_dist)
            advantage = [(u, a) for u, a in opps_list if a >= draw_limit]
            even      = [(u, a) for u, a in opps_list if abs(a) < draw_limit]
            behind    = [(u, a) for u, a in opps_list if a <= -draw_limit]

            parts = []
            def _fmt(lst, color):
                return "".join([f"<span style='color:{color};'>[{u}:{umaban_to_name.get(u, '')}]({a:+.1f})</span>" for u, a in sorted(lst, key=lambda x: abs(x[1]), reverse=True)])

            if advantage: parts.append(f"本馬 ＞ {_fmt(advantage, '#27ae60')}")
            if even: parts.append(f"本馬 ＝ {_fmt(even, '#888')}")
            if behind: parts.append(f"本馬 ＜ {_fmt(behind, '#e74c3c')}")
            return " / ".join(parts)

        for (r_date, r_place, r_dist, r_url), opps in sorted(race_groups.items(), key=race_priority, reverse=True):
            if not _is_relative_distance_allowed(r_dist, current_dist):
                continue
            opps = [(u, a) for u, a in opps if abs(a) < 2.0]
            if not opps: continue

            is_match = (r_place == current_course and str(r_dist) == str(current_dist))
            is_same_layout = (
                r_place == current_course
                and _is_same_track_layout(current_course, r_dist, current_dist)
            )
            style = "background:#fff9c4; border-left:3px solid #fbc02d; padding-left:5px;" if (is_match or is_same_layout) else ""
            if is_match:
                badge = " <span style='color:#fbc02d; font-weight:bold;'>[同条件]</span>"
            elif is_same_layout:
                badge = " <span style='color:#caa300; font-weight:bold;'>[同系統]</span>"
            else:
                badge = ""

            opp_names_in_race = {umaban_to_name.get(u_opp) for u_opp, _ in opps}
            has_cycle = any(
                (hname == a and b in opp_names_in_race) or (hname == b and a in opp_names_in_race)
                for (a, b) in cycle_pairs
            )
            if has_cycle:
                badge += " <span style='color:#e74c3c; font-weight:bold;'>⚡三すくみ</span>"

            race_label = f"{r_date}の{r_place}{r_dist}"
            link = f"<a href='{r_url}' target='_blank' style='color:#3498db;text-decoration:none;'>🔍{race_label}</a>{badge}" if r_url else f"🔍{race_label}{badge}"

            rel_str = fmt_opps_by_tier(opps, is_same_condition=(is_match or is_same_layout), is_strict=True)

            parts.append(f"<div style='margin-left:25px; font-size:0.85em; {style}'>{link}</div>")
            parts.append(f"<div style='margin-left:45px; font-size:0.85em; {style}'>└ {rel_str}</div>")

        direct_opps = _get_direct_opponents(hname)
        hidden_comparisons = []

        for opp_u, opp_n in umaban_to_name.items():
            if hname == opp_n:
                continue
            selected_source = (matchup_source.get(hname, {}) or {}).get(opp_n)
            if selected_source == "direct" or (selected_source is None and opp_n in direct_opps):
                continue
            if hname < opp_n: cache_key, sign = (hname, opp_n), 1
            else: cache_key, sign = (opp_n, hname), -1

            if cache_key not in hidden_cache: continue
            h_name, est, cond_score, both_same, place_dist, is_strict, uh_url, hv_url, uh_label, hv_label, bridge_note, bridge_priority, bridge_rank, bridge_dt, bridge_current_venue, leg_gates, leg_diffs = hidden_cache[cache_key]
            est_for_display = est * sign
            hidden_comparisons.append((opp_u, opp_n, h_name, est_for_display, both_same, place_dist, is_strict, uh_url, hv_url, uh_label, hv_label, bridge_note, bridge_priority, bridge_rank, bridge_dt, bridge_current_venue, leg_gates, leg_diffs, sign))

        if hidden_comparisons:
            hidden_comparisons.sort(key=lambda x: (x[12], 1 if x[12] == 2 and x[15] else (0 if x[12] == 2 else 1), -_safe_rank(x[13]), x[14] if isinstance(x[14], datetime) else datetime.min, x[3], x[6]), reverse=True)
            parts.append(f"<div style='margin-left:25px; font-size:0.82em; color:#9b59b6; margin-top:3px;'>🔗 隠れ馬経由の比較:</div>")
            for opp_u, opp_n, h_name, est, is_same_cond, place_dist, is_strict, uh_url, hv_url, uh_label, hv_label, bridge_note, bridge_priority, bridge_rank, bridge_dt, bridge_current_venue, leg_gates, leg_diffs, sign in hidden_comparisons:
                # 見出しの勝敗記号は唯一の真実源 matchup_matrix(=tierの根拠)から引く。
                # こうすれば合議経路でも同条件経路でも、表示記号とtierが食い違わない。
                auth_sym = (matchup_matrix.get(hname, {}) or {}).get(opp_n) if matchup_matrix else None
                if auth_sym:
                    sym = {">": "＞", "=": "＝", "<": "＜"}.get(auth_sym, "＝")
                    color = {">": "#27ae60", "=": "#777", "<": "#c0392b"}.get(auth_sym, "#777")
                else:
                    sym, color = _diff_symbol_and_color(est, is_same_condition=is_same_cond, is_strict=is_strict)
                # 合議の票数は注記として表示(複数レース/経路が方向を示した証跡)。
                cons = bridge_consensus_detail.get((hname, opp_n))
                if cons and (cons[1] + cons[2]) >= 1:
                    cons_pos, cons_neg = cons[1], cons[2]
                    magnitude_label = f"合議{cons_pos}-{cons_neg}"
                    consensus_note = f"<span style='color:#8e44ad;'>隠れ馬経由{cons_pos + cons_neg}件合議(本馬優勢{cons_pos}/相手優勢{cons_neg})</span> "
                else:
                    magnitude_label = f"{est:+.1f}"
                    consensus_note = ""

                def _leg(url, label):
                    if url:
                        return f"<a href='{url}' target='_blank' style='color:#3498db;text-decoration:none;'>{label}</a>"
                    return f"<span style='color:#888;'>{label}</span>"

                def _relation(left_label, left_gate, adv, right_label, right_gate):
                    try:
                        adv_val = float(adv)
                    except (TypeError, ValueError):
                        adv_val = 0.0
                    if adv_val > 0.05:
                        rel = "＞"
                    elif adv_val < -0.05:
                        rel = "＜"
                    else:
                        rel = "＝"
                    return f"{left_label}{left_gate}{rel}({abs(adv_val):.1f}){right_label}{right_gate}"

                note_label = bridge_note
                strict_prefix = "" if (is_strict or not note_label) else f"<span style='color:#888;'>※{note_label}</span> "
                if sign == 1:
                    main_gate = _circled_gate(leg_gates.get("name_a"))
                    bridge_main_gate = _circled_gate(leg_gates.get("bridge_on_a"))
                    bridge_opp_gate = _circled_gate(leg_gates.get("bridge_on_b"))
                    opp_gate = _circled_gate(leg_gates.get("name_b"))
                    main_leg = _leg(uh_url, uh_label)
                    opp_leg = _leg(hv_url, hv_label)
                    main_adv = leg_diffs.get("a_bridge", 0)
                    opp_adv = leg_diffs.get("bridge_b", 0)
                else:
                    main_gate = _circled_gate(leg_gates.get("name_b"))
                    bridge_main_gate = _circled_gate(leg_gates.get("bridge_on_b"))
                    bridge_opp_gate = _circled_gate(leg_gates.get("bridge_on_a"))
                    opp_gate = _circled_gate(leg_gates.get("name_a"))
                    main_leg = _leg(hv_url, hv_label)
                    opp_leg = _leg(uh_url, uh_label)
                    main_adv = -leg_diffs.get("bridge_b", 0)
                    opp_adv = -leg_diffs.get("a_bridge", 0)

                main_rel = _relation("本馬", main_gate, main_adv, "経由馬", bridge_main_gate)
                opp_rel = _relation("経由馬", bridge_opp_gate, opp_adv, "相手", opp_gate)

                if uh_url and hv_url and uh_url == hv_url:
                    legs_html = f"参考: {main_rel} / {opp_rel}: {main_leg}"
                else:
                    legs_html = f"{main_rel}: {main_leg} / {opp_rel}: {opp_leg}"

                parts.append(
                    f"<div style='margin-left:25px; font-size:0.82em;'>"
                    f"<span style='color:{color};'>本馬{sym}[{opp_u}:{opp_n}]({magnitude_label})</span></div>"
                    f"<div style='margin-left:45px; font-size:0.8em; color:#666;'>"
                    f"{strict_prefix}{consensus_note}経由(代表例)：<span style='color:#9b59b6;'>{h_name}</span> {legs_html}</div>"
                )

        chain_comparisons = []
        for _, opp_n in umaban_to_name.items():
            if hname == opp_n:
                continue
            comp_data = chain_matchups.get(hname, {}).get(opp_n)
            if comp_data:
                chain_comparisons.append((opp_n, comp_data))
        if chain_comparisons:
            chain_comparisons.sort(key=lambda x: {">": 0, "=": 1, "<": 2}.get(x[1].get("symbol", ""), 9))
            parts.append(_render_matchup_note(chain_comparisons, label="推移比較"))

        return "\n".join(parts)

    def _render_hidden_focus_races():
        race_map = {}
        for (name_a, name_b), best in hidden_cache.items():
            h_name, est, cond_score, both_same, place_dist, is_strict, uh_url, hv_url, uh_label, hv_label, bridge_note, bridge_priority, bridge_rank, bridge_dt, bridge_current_venue, leg_gates, leg_diffs = best
            for url, label, current_name in ((uh_url, uh_label, name_a), (hv_url, hv_label, name_b)):
                if not url:
                    continue
                entry = race_map.setdefault(url, {
                    "label": label,
                    "current": set(),
                    "bridges": set(),
                    "priority": bridge_priority,
                })
                entry["current"].add(current_name)
                entry["bridges"].add(h_name)
                entry["priority"] = max(entry["priority"], bridge_priority)

        focus = [
            (url, data)
            for url, data in race_map.items()
            if len(data["current"]) >= 3
        ]
        if not focus:
            return ""

        def _horse_sort_key(name):
            u = name_to_umaban.get(name, "99")
            return int(u) if str(u).isdigit() else 99

        focus.sort(key=lambda item: (len(item[1]["current"]), item[1]["priority"]), reverse=True)
        rows = [
            "<h4 style='background:#7e57c2; color:#fff; padding:5px; border-radius:4px;'>注目レース</h4>",
            "<div style='font-size:0.86em; color:#555; margin:3px 0 8px;'>隠れ馬経由の参考レースのうち、現出走馬3頭以上に関係するもの。</div>"
        ]
        for url, data in focus[:5]:
            horses = []
            for name in sorted(data["current"], key=_horse_sort_key):
                u = name_to_umaban.get(name, "?")
                horses.append(f"[{u}] {html.escape(name)}")
            bridges = "・".join(html.escape(x) for x in sorted(data["bridges"]))
            label = html.escape(data["label"] or "参考")
            rows.append(
                f"<div style='margin-left:25px; font-size:0.84em; border-left:3px solid #b39ddb; padding-left:6px; margin-bottom:6px;'>"
                f"<a href='{url}' target='_blank' style='color:#3498db;text-decoration:none;'>🔎{label}</a>："
                f"関係馬 {' / '.join(horses)}"
                f"{f'　経由:{bridges}' if bridges else ''}</div>"
            )
        return "\n".join(rows)

    # 出現するtierラベルをレベル順に並べ、比較不可(！)は最上段に置く。
    _present = sorted({t for t in all_tiers.values() if t},
                      key=lambda t: _TIER_LABELS.find(t) if _REL_MULTI_TIER else {"S": 0, "A": 1, "B": 2, "C": 3}.get(t, 99))
    for tier in ((["！"] if no_compare_horses else []) + _present):
        if tier == "！":
            if not no_compare_horses: continue
            html_parts.append(
                "<h4 style='background:#999; color:#fff; padding:5px; border-radius:4px;'>"
                "不明</h4>"
            )
            for umaban, hname in sorted(no_compare_horses, key=lambda x: int(x[0])):
                mark_btn = f'<span class="mark-btn" data-uma="{umaban}" data-race="{r_num}" data-mark="" onclick="cycleMark(this)"></span>'
                last_note = html.escape(_last_run_note(umaban))
                note_html = (
                    f" <span style='font-size:0.76em; color:#777; font-weight:normal; margin-left:6px;'>{last_note}</span>"
                    if last_note else ""
                )
                html_parts.append(
                    f"<div class='horse-header'>{mark_btn} <strong>[{umaban}] {hname}</strong>{note_html}</div>"
                )
        else:
            horses = [(u, n) for u, n in umaban_to_name.items() if all_tiers.get(n) == tier]
            if not horses: continue
            html_parts.append(f"<h4 style='background:{_rel_tier_color(tier)}; color:#fff; padding:5px; border-radius:4px;'>{_rel_label_display(tier)}</h4>")
            for umaban, hname in sorted(horses, key=lambda x: int(x[0])):
                html_parts.append(_render_horse_block(umaban, hname))

    focus_races_html = _render_hidden_focus_races()
    if focus_races_html:
        html_parts.append(focus_races_html)

    html_parts.append("</div>")
    return "\n".join(html_parts)


# ==================================================
# 3b. スコア算出
# ==================================================
def calculate_horse_scores(horses_data, cyokyo_data, current_course, current_dist, is_young, race_class_str, r_num, pace_speed_list=None, is_newcomer=False):
    class_ranks = {'2歳': 0, '3歳': 0, 'C3': 0, 'C2': 1, 'C1': 2, 'B3': 3, 'B2': 4, 'B1': 5, 'A2': 6, 'A1': 7}
    standard_times = CHOKYO_STANDARD_3F

    if is_newcomer:
        relative_scores = {str(u): 0 for u in horses_data.keys()}
        relative_tiers = {str(u): None for u in horses_data.keys()}
        relative_html = "<div style='font-size:0.92em; line-height:1.6; color:#666;'>新馬戦のため相対評価は取得しません。</div>"
        same_condition_locks = []
        internal_verdicts = {}
    else:
        relative_scores, relative_tiers, relative_html, same_condition_locks, internal_verdicts = calculate_relative_scores(
            horses_data, current_course, current_dist, r_num
        )

    # 競馬場・距離・開催日馬場差を補正した独立の共通実戦速度指数。
    # 検証段階のため総合点へは足さず、順位と根拠だけを表示する。
    common_speed_indices = calculate_common_speed_indices(
        horses_data, current_course=current_course, current_dist=current_dist,
    )

    # 想定隊列の後方20%（floor）に該当する馬番を抽出。地方競馬は前有利のため score_3 から -5 する。
    back_20_set = set()
    if pace_speed_list:
        n_pace = len(pace_speed_list)
        cutoff = int(n_pace * 0.2)
        if cutoff > 0:
            sorted_by_pos = sorted(pace_speed_list, key=lambda x: -float(x.get("est_pos", 0) or 0))
            for sd in sorted_by_pos[:cutoff]:
                back_20_set.add(str(sd.get("umaban", "")))

    base_rank = class_ranks.get(race_class_str, 1)
    # 全体TOP3とレース内TOP3は全レースの調教素材を集めた後、ループ後段で加点・再描画する。

    scored_data = {}
    for umaban, horse_data in horses_data.items():
        flags = []
        cyokyo_text = cyokyo_data.get(umaban, "")

        # 当日各コース横断の調教上位（赤字）集計用。今走調教のコース/メトリクス/タイムを控える。
        chokyo_time_disp = ""    # 例: 大井36.5（3F、🕰️表示用）
        chokyo_rank_key = None   # (date, (course,is_main), metric) ＝ 同日同コース同メトリクスで順位付け
        chokyo_rank_time = None  # 順位付けに使う実タイム（1F優先→無ければ3F）
        chokyo_rank_disp = ""    # 例: 大井12.3 / 大井37.0
        chokyo_rank_entries = [] # 今走調教の全行から抽出した同日横断ランキング素材
        zen_chokyo_rank_entries = [] # 前走調教の全行から抽出した同日横断ランキング素材

        chokyo_tanpyo = "なし"
        m_tanpyo = re.search(r'【短評】([^\n]+)', cyokyo_text)
        if m_tanpyo:
            chokyo_tanpyo = m_tanpyo.group(1).strip()
            
        current_jockey = (horse_data.get("jockey") or "").strip()
        
        jockey_raw = horse_data.get("display_power", "")
        p_base, p_prev, win_rate, n_count, jockey_win_rate = 5, 5, 0.0, 0, 0.0
        if jockey_raw:
            m_p = re.search(r'P:(\d+)', jockey_raw)
            m_prev_p = re.search(r'前P:(\d+)', jockey_raw)
            m_aisei = re.search(r'相性:勝([\d\.]+)%\(\d+/(\d+)\)', jockey_raw)
            m_jwin = re.search(r'P:\d+\(勝([\d\.]+)%', jockey_raw)
            if m_p: p_base = int(m_p.group(1))
            if m_prev_p: p_prev = int(m_prev_p.group(1))
            if m_aisei:
                try: win_rate, n_count = float(m_aisei.group(1)), int(m_aisei.group(2))
                except: pass
            if m_jwin:
                try: jockey_win_rate = float(m_jwin.group(1))
                except: pass

        runs = horse_data.get("hist", [])
        valid_runs = []
        for r in runs:
            if not isinstance(r, dict): continue
            rank = r.get("rank")
            if rank is None: continue
            valid_runs.append({
                "course": str(r.get("place", "")),
                "dist": str(r.get("dist", "")),
                "jockey": str(r.get("jockey", "")),
                "positions": str(r.get("pas", "")),
                "agari": r.get("agari", 0),
                "rank": rank,
                "pop": r.get("pop", 99)
        })

        tataki_second = _detect_tataki_second(runs)
        past_jockeys = [r["jockey"] for r in valid_runs]
        is_demodori = (current_jockey != (past_jockeys[0] if past_jockeys else "") and current_jockey in past_jockeys[:3])

        # ②騎手(最大15点):
        #   現在の騎手Pをベースに、乗り替わり時は「現在P-前走P」をそのまま加減する。
        #   馬×騎手の相性勝率が騎手自身の勝率より高ければ+5点(低くても減点なし)。
        #   どの加点後も0〜15点に収める。
        score_2 = p_base
        is_shobu_norikawari = False
        prev_jockey_name = (horse_data.get("prev_jockey") or (past_jockeys[0] if past_jockeys else "")).strip()
        is_norikawari = bool(current_jockey and prev_jockey_name and current_jockey != prev_jockey_name)

        if is_norikawari:
            delta = p_base - p_prev   # 乗り替わりによる騎手Pの増減
            score_2 += delta
            is_shobu_norikawari = delta >= 2

        if win_rate > jockey_win_rate:
            score_2 += 5
        score_2 = max(0, min(15, score_2))

        # ③盛り返し・叩き2戦目ボーナス(最大+5・合計へ上乗せ)。
        # 道中2つ以上下げて4角→着順で2つ以上盛り返した実績を前走+4/2走前+2/3走前+2 で評価。
        # 前走が75日以上の休み明けで、前1/3に先行後3つ以上順位を落とした馬は+5。
        score_3 = 0
        if valid_runs:
            morikaeshi_bonus = 0
            has_morikaeshi_flag = False
            bonus_settings = {0: 4, 1: 2, 2: 2}

            for idx, r in enumerate(valid_runs[:3]):
                p_str = r.get("positions", "")
                p_list = [int(p) for p in p_str.split('-') if p.isdigit()]
                final_rank = r.get("rank", 99)

                if len(p_list) >= 2:
                    first_pos = p_list[0]
                    last_corner_pos = p_list[-1]
                    max_pos = max(p_list)

                    is_dropped = (max_pos - first_pos) >= 2
                    is_recovered = (last_corner_pos - final_rank) >= 2

                    if is_dropped and is_recovered:
                        morikaeshi_bonus += bonus_settings.get(idx, 0)
                        has_morikaeshi_flag = True

            if has_morikaeshi_flag:
                score_3 += morikaeshi_bonus
                flags.append("【盛り返し】")

        if is_young: score_3 = int(score_3 * 0.5)

        # 叩2は若馬の内容点半減後に加え、年齢を問わず最終的に+5を保証する。
        if tataki_second:
            score_3 += 5
            flags.append("【叩2】")

        score_3 = max(0, min(5, score_3))

        score_4 = 0
        score_chokyo_prev = 0
        chokyo_prev_detail = ""
        zen_lines = re.findall(r'前走調教：([^\n]+)', cyokyo_text)
        kon_lines = re.findall(r'今走調教：([^\n]+)', cyokyo_text)
        is_kakuue_senchaku = False
        is_fast_time = False

        if zen_lines:
            for _line in zen_lines:
                for _e in _parse_chokyo_entries_for_ranking("前走調教：" + _line, allowed_labels=("前走調教",)):
                    zen_chokyo_rank_entries.append(_e)
        
        if zen_lines and kon_lines:
            prev_comparison = _compute_prev_chokyo_comparison(zen_lines, kon_lines)
            score_chokyo_prev = int(prev_comparison.get("points", 0) or 0)
            chokyo_prev_detail = prev_comparison.get("detail", "")
            score_4 += score_chokyo_prev
            if prev_comparison.get("flag"):
                flags.append(prev_comparison["flag"])
        
        if kon_lines:
            for _line in kon_lines:
                for _e in _parse_kon_chokyo_entries_for_ranking("今走調教：" + _line):
                    chokyo_rank_entries.append(_e)

            clean_kon = re.sub(r'\d{1,2}/\d{1,2}\s+', '', kon_lines[-1])
            parts_s = clean_kon.split('/')
            
            t_part = parts_s[0].strip()
            a_part = parts_s[1].strip() if len(parts_s) > 1 else "単走"
            times = re.findall(r'\b\d{2}\.\d{1}\b', t_part)
            c_name_norm, is_main_track = _extract_training_course(t_part)
            
            if times and c_name_norm:
                t_3f = _pick_3f_time(times)
                if t_3f is None:
                    t_3f = 999.9
                
                super_fast = {
                    '大井': 36.8, '川崎': 36.8, '船橋': 36.5, '小林': 36.5,
                    '浦和': 38.5, '小林坂': 36.5, '園田': 36.8, '美W': 36.5, '栗CW': 36.5, 'CW': 36.5
                }
                
                super_threshold = _training_threshold(super_fast, c_name_norm, is_main_track)
                standard_threshold = _training_threshold(standard_times, c_name_norm, is_main_track)
                # 一律化: 基準を切れば馬なり/強め/常時速に関係なく一律点。
                if c_name_norm in super_fast and t_3f <= super_threshold:
                    score_4 += 12
                    flags.append("【絶好調教】")
                    is_fast_time = True
                elif t_3f < standard_threshold:
                    score_4 += 8
                    flags.append("【基準超】")
                    is_fast_time = True

                # 当日各コース横断の調教上位ランキング用キーを控える（同コース・同メトリクスで順位付け）。
                # 互換用に代表1件も残すが、実際の横比較は chokyo_rank_entries の全件で行う。
                if t_3f is not None and t_3f < 900 and c_name_norm in standard_times:
                    chokyo_time_disp = f"{c_name_norm}{t_3f:.1f}"
                if chokyo_rank_entries:
                    _best_e = sorted(chokyo_rank_entries, key=lambda e: (0 if e.get("metric") == "1F" else 1, float(e.get("adj_time", e.get("time")) or 999)))[0]
                    chokyo_rank_key = (_best_e["date"], _best_e["course_key"], _best_e["metric"])
                    chokyo_rank_time = _best_e["time"]
                    chokyo_rank_disp = _best_e.get("disp") or f"{_best_e['course_key'][0]}{_best_e['time']:.1f}"

            is_kakushita_awase = False
            if a_part == "単走":
                score_4 += 2
            else:
                # 併せ馬の状況を「相手クラス×結果(先着/同入/遅れ)×内外位置×負荷」で評価。
                p_rank = _awase_partner_rank(a_part, base_rank, class_ranks)
                is_soto = "外" in a_part          # 外=陣営が「強い」と見て置く位置(好材料)
                won = "先着" in a_part
                evened = "同入" in a_part
                lost = "遅れ" in a_part

                if p_rank > base_rank:             # 格上相手
                    if won:
                        score_4 += 7; is_kakuue_senchaku = True
                    elif evened:
                        score_4 += 5; is_kakuue_senchaku = True
                elif p_rank == base_rank:          # 同級相手
                    if won:
                        score_4 += 4
                    elif evened:
                        score_4 += 2
                else:                               # 格下相手
                    is_kakushita_awase = True
                    if won:
                        score_4 += 3
                    elif evened:
                        score_4 += 1

                # 外併せで先着/同入は「陣営がこの馬を上に見ている」自信シグナル。
                if is_soto and (won or evened):
                    score_4 += 2
                    flags.append("【外併せ好結果】")
                # 相手が一杯/強め(本気)なのに同級以上で先着/同入は価値が高い。
                if ("一杯" in a_part or "強め" in a_part) and (won or evened) and p_rank >= base_rank:
                    score_4 += 1
                # 後ろから追走して追いついた前進気勢も加点。
                if "追走" in a_part and (won or evened):
                    score_4 += 1

                if lost:
                    if p_rank > base_rank:
                        score_4 -= 2   # 格上相手の遅れは軽い
                    else:
                        score_4 -= 6

            if is_kakushita_awase and not is_fast_time and ("先着" in a_part or "同入" in a_part):
                score_4 += 4
        
        if is_young and is_kakuue_senchaku and is_shobu_norikawari:
            # 騎手点は「騎手P＋乗り替わりP差＋相性ボーナス」を維持する。
            score_4 = 16
            flags.append("【変身特注】")
            
        score_4 = max(0, min(20, score_4))

        raw_score_4 = score_4
        score_5 = relative_scores.get(umaban, 20)
        has_no_relative = (relative_tiers.get(umaban) is None)
        speed_info = common_speed_indices.get(str(umaban), {})
        # ⑥対戦表(0〜10)は対戦表fetch後に compute_matchup_points で後付けする(初期値0)。

        scored_data[umaban] = {
            "score_2": score_2,
            "score_3": score_3,
            "score_4": score_4,
            "score_4_raw": raw_score_4,
            "score_chokyo_prev": score_chokyo_prev,
            "chokyo_prev_detail": chokyo_prev_detail,
            "score_5": score_5,
            "score_6": 0,
            "flags": list(dict.fromkeys(flags)),
            "chokyo_tanpyo": chokyo_tanpyo,
            "chokyo_time_disp": chokyo_time_disp,
            "chokyo_time_info": chokyo_time_disp,
            "chokyo_rank_key": chokyo_rank_key,
            "chokyo_rank_time": chokyo_rank_time,
            "chokyo_rank_disp": chokyo_rank_disp,
            "chokyo_rank_entries": chokyo_rank_entries,
            "zen_chokyo_rank_entries": zen_chokyo_rank_entries,
            "score_chokyo_top": 0,
            "score_chokyo_global": 0,
            "score_chokyo_race": 0,
            "relative_tier": relative_tiers.get(umaban),
            "common_speed_index": speed_info.get("index"),
            "common_speed_race_index": speed_info.get("race_index"),
            "common_speed_same_max": speed_info.get("same_condition_max"),
            "common_speed_same_runs": speed_info.get("same_condition_runs", 0),
            "common_speed_same_best_run": speed_info.get("same_condition_best_run"),
            "common_speed_rank": speed_info.get("rank"),
            "common_speed_total": speed_info.get("total"),
            "common_speed_runs": speed_info.get("runs", 0),
            "common_speed_corrected_runs": speed_info.get("corrected_runs", 0),
            "common_speed_best_run": speed_info.get("best_run"),
            "has_no_relative": has_no_relative,
            "is_newcomer": is_newcomer,
            "tataki_second": bool(tataki_second),
            "tataki_second_detail": tataki_second.get("detail", "") if tataki_second else "",
            "jockey_current": current_jockey,
            "jockey_prev": past_jockeys[0] if past_jockeys else "",
            "jockey_p_base": p_base,
            "jockey_p_prev": p_prev,
            "compat_win_rate": win_rate,
            "compat_n": n_count
        }

    apply_no_relative_scores(scored_data)

    return scored_data, relative_html, same_condition_locks, internal_verdicts


def _rank_from_totals(totals):
    if not totals:
        return {}

    sorted_scores = sorted([t for _, t in totals], reverse=True)
    max_score = sorted_scores[0]
    second_score = sorted_scores[1] if len(sorted_scores) >= 2 else max_score
    s_condition = (max_score >= 75) and ((max_score - second_score) >= 15)

    def _band_a(diff):
        # 高得点シナリオ(a): D/E のみ +5 拡張
        if diff <= 19:   return "A"
        elif diff <= 29: return "B"
        elif diff <= 39: return "C"
        elif diff <= 54: return "D"
        return "E"

    def _band_b(diff):
        # 抜けたトップシナリオ(b): 全バンドを +5 シフト
        if diff <= 24:   return "A"
        elif diff <= 34: return "B"
        elif diff <= 44: return "C"
        elif diff <= 54: return "D"
        return "E"

    def _band_default(diff):
        # S 無しシナリオ: 従来通り
        if diff <= 19:   return "A"
        elif diff <= 29: return "B"
        elif diff <= 39: return "C"
        elif diff <= 49: return "D"
        return "E"

    ranks = {}
    for umaban, total in totals:
        diff = max_score - total
        if max_score >= 90:
            ranks[umaban] = "S" if total >= 90 else _band_a(diff)
        elif s_condition:
            ranks[umaban] = "S" if total == max_score else _band_b(diff)
        else:
            ranks[umaban] = _band_default(diff)
    return ranks


def apply_score_floor_to_grades(grades, horses_data, scored_data):
    """最終降格後の評価に、素点から見た最低ラインをかける。

    対戦・同条件・騎手選択の降格は残すが、A/B相当の素点馬がEまで沈むと
    表示上の点数と評価が噛み合わないため、素点バンドから最大2段までの
    降格に抑え、あわせて絶対点の下限も軽く保証する。
    """
    if not grades or not horses_data or not scored_data:
        return grades

    rank_val = {"S": 7, "A": 6, "B": 5, "C": 4, "D": 3, "E": 2}
    val_to_rank = {7: "S", 6: "A", 5: "B", 4: "C", 3: "D", 2: "E"}
    totals = []
    total_by_name = {}
    for umaban, data in horses_data.items():
        name_norm = _norm_horse_name(data.get("name", ""))
        if not name_norm:
            continue
        sc = scored_data.get(str(umaban), scored_data.get(umaban, {}))
        total = _deterministic_total(sc)
        totals.append((name_norm, total))
        total_by_name[name_norm] = total

    raw_grades = _rank_from_totals(totals)
    new_grades = {k: _rank_floor_e(v) for k, v in grades.items()}
    for name, grade in list(new_grades.items()):
        cur_val = rank_val.get(_rank_floor_e(grade), 0)
        raw_grade = _rank_floor_e(raw_grades.get(name, grade))
        raw_val = rank_val.get(raw_grade, cur_val)
        min_val = max(2, raw_val - 2)

        total = total_by_name.get(name)
        if total is not None:
            if total >= 80:
                min_val = max(min_val, rank_val["B"])
            elif total >= 70:
                min_val = max(min_val, rank_val["C"])
            elif total >= 60:
                min_val = max(min_val, rank_val["D"])

        if cur_val and cur_val < min_val:
            new_grades[name] = val_to_rank.get(min_val, "E")

    return new_grades


def _grades_from_ai_detail_totals(ai_text):
    """出走馬分析テキストの素点(【結論:計N点】)から評価バンドを再構築する。

    キャッシュ再計算(relativeモード)では前回実行の grades がキャッシュに残っているが、
    それはロック・対戦表・騎手の降格を適用済みの「最終値」。これを起点に再度
    降格パスを通すと再実行のたびに評価が一方向へ沈み(ラチェット)、
    全馬 D/E で S/A 不在の表になる(26/6/11 大井9R: 素点91点の馬がE化)。
    そのため毎回ここで素点からバンドを作り直し、降格パスの起点にする。
    """
    if not ai_text:
        return {}
    header_re = re.compile(r'^([①-⑳]|\[\d{1,2}\])([^\s\t　]+)')
    totals = []
    current_name = ""
    for line in str(ai_text).split("\n"):
        m = header_re.match(line.strip())
        if m:
            current_name = _norm_horse_name(m.group(2))
            continue
        m_total = re.search(r'【結論:計(\d+)点】', line)
        if m_total and current_name:
            totals.append((current_name, int(m_total.group(1))))
            current_name = ""
    if not totals:
        return {}
    return dict(_rank_from_totals(totals))


def ensure_top_grade_presence(grades):
    """最終評価に S/A が1頭も残らなかった場合の補正。

    素点バンドは最高点馬を必ず S/A にするが、降格パス(内部ロック・対戦表・騎手)が
    上位帯の馬を全て降格させると本命不在の評価一覧になる。その場合は降格後の
    序列で最上位の帯にいる馬を A へ引き上げ、「全レースに必ず S か A がいる」
    表示仕様を保証する。帯単位の引き上げなので「負けた馬を勝った馬より
    上にしない」関係は帯内では維持される。
    """
    if not grades:
        return grades
    rank_val = {"S": 7, "A": 6, "B": 5, "C": 4, "D": 3, "E": 2}
    vals = {k: rank_val[v] for k, v in grades.items() if v in rank_val}
    if not vals:
        return grades
    best = max(vals.values())
    if best >= 6:
        return grades
    new_grades = dict(grades)
    for k, v in vals.items():
        if v == best:
            new_grades[k] = "A"
    return new_grades


def apply_display_grade_overrides(grades, horses_data, scored_data):
    """総合評価で表示する最終ランクを grades 自体へ反映する。

    これまで build_evaluation_list() だけで行っていた「騎手Pが下がる乗り替わり時の
    S→A」を、対戦表・出走馬分析・保存gradesにも同じように効かせる。
    """
    if not grades or not horses_data or not scored_data:
        return grades
    new_grades = {k: _rank_floor_e(v) for k, v in grades.items()}
    for u, hd in horses_data.items():
        name_norm = _norm_horse_name((hd or {}).get("name", ""))
        if not name_norm:
            continue
        grade = new_grades.get(name_norm, "")
        if not grade:
            for k_norm, v in new_grades.items():
                if k_norm and (k_norm in name_norm or name_norm in k_norm):
                    name_norm = k_norm
                    grade = v
                    break
        if grade != "S":
            continue
        sd = scored_data.get(str(u), scored_data.get(u, {}))
        curr = sd.get("jockey_current", "")
        prev = sd.get("jockey_prev", "")
        p_base = sd.get("jockey_p_base", 0)
        p_prev = sd.get("jockey_p_prev", 0)
        if prev and curr and curr != prev and p_base < p_prev:
            new_grades[name_norm] = "A"
    return new_grades


def _grade_for_display_name(grades, hname, fallback=""):
    name_norm = _norm_horse_name(hname)
    grade = grades.get(name_norm, "")
    if not grade and name_norm:
        for k_norm, v in grades.items():
            if k_norm and (k_norm in name_norm or name_norm in k_norm):
                grade = v
                break
    return _rank_floor_e(grade or fallback)


def _apply_final_grades_to_ai_detail(ai_text, grades):
    """出走馬分析の馬見出し末尾ランクを最終 grades に揃える。"""
    if not ai_text or not grades:
        return ai_text
    header_pat = r'^([①-⑳]|[\[]\d{1,2}[\]])([^\s\t　]+)[　\s\t]+([SABCDEFG-])'
    new_lines = []
    for line in str(ai_text).split('\n'):
        m = re.match(header_pat, line)
        if m:
            prefix, hname, old_rank = m.groups()
            line = f"{prefix}{hname}　{_grade_for_display_name(grades, hname, old_rank)}"
        new_lines.append(line)
    return '\n'.join(new_lines)


def _clean_danwa_text(text):
    clean = str(text or "").replace("\n", " ").strip()
    clean = re.sub(r'^\s*話[:：]\s*', '', clean)
    clean = re.sub(r'\s+', ' ', clean).strip()
    if not clean or clean in {"なし", "コメントなし", "データなし"}:
        return "コメントなし"
    return clean

def _danwa_leading_mark(text):
    clean = _clean_danwa_text(text)
    if clean == "コメントなし":
        return ""
    m = re.match(r'\s*([◎○◯〇⚪△▲☆★◇◆×✕])', clean)
    return m.group(1) if m else ""


def _deterministic_total(sc):
    """1頭の合計点(最大100＋盛返③上乗せ)。
    通常: ②騎手15 + ③盛返 + ④調教20 + ⑤相対50 + ⑥対戦表10。
    比較不可馬も通常式で合計し、⑤相対には騎手P順位による
    50〜25点の補完点を使う。"""
    s2 = int(sc.get("score_2", 0) or 0)
    s3 = int(sc.get("score_3", 0) or 0)
    s4 = int(sc.get("score_4", 0) or 0)
    s5 = int(sc.get("score_5", 0) or 0)
    s6 = int(sc.get("score_6", 0) or 0)
    sp = int(sc.get("score_pace", 0) or 0)   # ⑦展開加点(ペース加点＋競馬場×距離の前有利加点)
    spt = int(sc.get("score_pattern", 0) or 0)  # 好走パターン(ユーザーメモ)による±5
    sct = int(sc.get("score_chokyo_top", 0) or 0)  # 全体＋レース内の調教順位加点（合計最大8）
    return s2 + s3 + s4 + s5 + s6 + sp + spt + sct


def _format_pattern_short(pattern):
    """好走パターンを「好走傾向：逃◯ 内✕ 中◯」の短い文字列に。空なら ''。"""
    if not pattern:
        return ""
    short = {"逃げ": "逃", "番手": "番", "内枠": "内", "中枠": "中", "外枠": "外"}
    parts = [f"{short[d]}{pattern[d]}" for d in ("逃げ", "番手", "内枠", "中枠", "外枠")
             if pattern.get(d)]
    return ("好走傾向：" + " ".join(parts)) if parts else ""


def _est_pos_position_dim(est_pos):
    """想定位置(est_pos)を 逃げ/番手 の好走パターン次元へ。該当なしは None。"""
    try:
        ep = float(est_pos)
    except (TypeError, ValueError):
        return None
    if ep <= 1.5:
        return "逃げ"
    if ep <= 3.5:
        return "番手"
    return None


def _gate_pattern_dim(field_size, umaban):
    """今回枠を 内枠/中枠/外枠 へ。"""
    return {"内目": "内枠", "中目": "中枠", "外目": "外枠"}.get(_gate_group(field_size, umaban))


def _fetch_horse_patterns(horses_data):
    """出走各馬の好走パターンを {uma_id: {dim: mark}} で取得（DB未設定でも安全）。"""
    try:
        import store as _db
        uma_ids = [h.get("uma_id") for h in horses_data.values() if h.get("uma_id")]
        return _db.get_patterns_map(uma_ids)
    except Exception as e:
        print(f"  [好走パターン] 取得スキップ: {e}")
        return {}


def apply_pattern_bonus(scored_data, horses_data, pace_speed_list, patterns):
    """ユーザーの好走パターン(◯/✕)を今回の想定位置・枠と突き合わせ、
    一致する◯=+5 / ✕=-5 を score_pattern として加点する。"""
    if not patterns:
        return
    est_by_umaban = {str(p.get("umaban")): p.get("est_pos") for p in (pace_speed_list or [])}
    field_size = len(horses_data)
    for umaban, h in horses_data.items():
        uid = h.get("uma_id")
        pat = patterns.get(uid) if uid else None
        if not pat:
            continue
        bonus = 0
        pos_dim = _est_pos_position_dim(est_by_umaban.get(str(umaban)))
        if pos_dim and pat.get(pos_dim) in ("◯", "✕"):
            bonus += 5 if pat[pos_dim] == "◯" else -5
        gate_dim = _gate_pattern_dim(field_size, umaban)
        if gate_dim and pat.get(gate_dim) in ("◯", "✕"):
            bonus += 5 if pat[gate_dim] == "◯" else -5
        sd = scored_data.get(str(umaban)) or scored_data.get(umaban)
        if bonus and sd is not None:
            sd["score_pattern"] = bonus


def format_deterministic_evaluation(horses_data: dict, cyokyo_data: dict, scored_data: dict, danwa_data: dict):
    apply_no_relative_scores(scored_data)
    totals = []
    for umaban in horses_data.keys():
        sc = scored_data.get(str(umaban), {})
        totals.append((str(umaban), _deterministic_total(sc)))

    rank_by_umaban = _rank_from_totals(totals)
    total_by_umaban = dict(totals)
    final_text = ""
    grades = {}

    for umaban, data in horses_data.items():
        horse_name = data.get("name", "")
        circled_num = chr(0x245f + int(umaban)) if str(umaban).isdigit() and 1 <= int(umaban) <= 20 else f"[{umaban}]"

        curr_j = (data.get("jockey") or "").strip()
        prev_j = (data.get("prev_jockey") or "").strip()
        is_misute = bool(data.get("misute_jockey"))
        if prev_j:
            if curr_j == prev_j:
                jockey_disp = f"{curr_j}(同)"
            else:
                prefix = "😢" if is_misute else ""
                jockey_disp = f"{prefix}{curr_j}←{prev_j}"
        else:
            jockey_disp = f"😢{curr_j}" if is_misute else curr_j

        jockey_stats = data.get("display_power", "").replace("【騎手】", "")
        cyokyo_text = cyokyo_data.get(umaban, "データなし")
        cyokyo_lines = re.findall(r'(?:前走調教：|今走調教：)[^\n]+', cyokyo_text)
        cyokyo_display = "\n".join(cyokyo_lines) if cyokyo_lines else "データなし"
        race_content = build_race_content_summary(data.get("hist", []))

        sc = scored_data.get(str(umaban), {})
        total = total_by_umaban.get(str(umaban), 0)
        rank = _rank_floor_e(rank_by_umaban.get(str(umaban), "-"))
        flags = "".join(sc.get("flags", []))
        # 厩舎コメント(談話)を一言一句そのまま「」で表示(要約・言い換えなし)。
        comment = _clean_danwa_text(danwa_data.get(str(umaban), ""))
        grades[_norm_horse_name(horse_name)] = rank

        s2 = int(sc.get("score_2", 0) or 0)
        s3 = int(sc.get("score_3", 0) or 0)
        s4 = int(sc.get("score_4", 0) or 0)
        s5 = int(sc.get("score_5", 0) or 0)
        s6 = int(sc.get("score_6", 0) or 0)
        sp = int(sc.get("score_pace", 0) or 0)
        pace_disp = f" ⑦展開{sp}" if sp else ""
        spt = int(sc.get("score_pattern", 0) or 0)
        pat_disp = f" 好走{'+' if spt > 0 else ''}{spt}" if spt else ""
        sct = int(sc.get("score_chokyo_top", 0) or 0)
        scg = int(sc.get("score_chokyo_global", 0) or 0)
        scr = int(sc.get("score_chokyo_race", 0) or 0)
        scprev = int(sc.get("score_chokyo_prev", 0) or 0)
        prev_detail = sc.get("chokyo_prev_detail", "")
        prev_parts = []
        if prev_detail:
            prev_parts.append(f"前走比{'+' if scprev > 0 else ''}{scprev} {prev_detail}")
        prev_disp = f"({' / '.join(prev_parts)})" if prev_parts else ""
        rank_parts = []
        if scg:
            global_rank = sc.get("chokyo_red_rank")
            rank_parts.append(f"全体{global_rank}位+{scg}" if global_rank else f"全体+{scg}")
        if scr:
            race_rank = sc.get("chokyo_race_rank")
            rank_parts.append(f"レース内{race_rank}位+{scr}" if race_rank else f"レース内+{scr}")
        chokyo_top_disp = ""
        if sct:
            rank_detail = f"({'/'.join(rank_parts)})" if rank_parts else ""
            chokyo_top_disp = f" 調教順位+{sct}{rank_detail}"
        tier = sc.get("relative_tier")
        rel_label = sc.get("no_relative_grade") or (f"相対{_rel_label_display(tier)}" if tier else "相対不明")
        speed_index = sc.get("common_speed_index")
        speed_same_max = sc.get("common_speed_same_max")
        speed_same_runs = int(sc.get("common_speed_same_runs", 0) or 0)
        speed_rank = sc.get("common_speed_rank")
        speed_total = sc.get("common_speed_total")
        speed_runs = int(sc.get("common_speed_runs", 0) or 0)
        speed_best = sc.get("common_speed_best_run") or {}
        speed_line = ""
        if speed_index is not None:
            rank_note = f" {speed_rank}位/{speed_total}頭" if speed_rank and speed_total else ""
            variant = float(speed_best.get("variant_per_1000", 0) or 0)
            variant_note = ""
            if abs(variant) >= 0.03:
                variant_note = (
                    f" / 最良走:{speed_best.get('date', '')}{speed_best.get('place', '')}"
                    f"{speed_best.get('dist', '')}m {speed_best.get('variant_label', '')}馬場"
                    f"{abs(variant):.1f}秒/1000m補正"
                )
            same_note = f"{float(speed_same_max):.1f}（{speed_same_runs}走）" if speed_same_max is not None else "なし"
            ref_note = speed_best.get('par_source', '')
            confidence = speed_best.get('reference_confidence', '')
            ref_disp = f" / {ref_note} 信頼:{confidence}" if ref_note else ""
            speed_line = (
                f"【速度指数】指数:{float(speed_index):.1f}{rank_note}"
                f" / 同場同距離:{same_note}（近走{speed_runs}件）{ref_disp}{variant_note}\n"
            )

        block = f"{circled_num}{horse_name}　{rank}\n"
        block += f"騎:{jockey_disp}　{jockey_stats}\n"
        block += f"【結論:計{total}点】\n"
        block += f"【内訳】⑤{rel_label}={s5} ④調教{s4}{prev_disp} ②騎手{s2} ⑥対戦{s6} ③盛返{s3}{pace_disp}{pat_disp}{chokyo_top_disp}\n"
        block += speed_line
        block += f"{flags} 「{comment}」\n" if flags else f"「{comment}」\n"
        block += f"【調教】\n{cyokyo_display}\n"
        block += f"🐎レース内容\n{race_content}\n"

        final_text += block + "\n----------------------------------------\n\n"

    return final_text.strip(), grades


# ==================================================
# 4. データロード
# ==================================================
@st.cache_resource
def load_resources():
    res = {
        "jockeys": [],
        "trainers": [],
        "power_data": {},
        "power_jockeys": set()
    }

    def get_valid_path(target_path):
        if os.path.exists(target_path): return target_path
        basename = os.path.basename(target_path)
        p2 = os.path.join(DATA_DIR, basename)
        if os.path.exists(p2): return p2
        if os.path.exists(basename): return basename
        return None

    j_path = get_valid_path(JOCKEY_FILE)
    if j_path:
        for enc in ("utf-8-sig", "cp932"):
            try:
                with open(j_path, "r", encoding=enc) as f:
                    res["jockeys"] = [l.strip().replace(" ", "").replace("　", "") for l in f if l.strip()]
                break
            except Exception:
                continue

    t_path = get_valid_path(TRAINER_FILE)
    if t_path:
        for enc in ("utf-8-sig", "cp932"):
            try:
                with open(t_path, "r", encoding=enc) as f:
                    res["trainers"] = [l.strip().replace(",", "").replace(" ", "").replace("　", "") for l in f if l.strip()]
                break
            except Exception:
                continue

    p_path = get_valid_path(POWER_FILE)
    if p_path:
        try:
            import openpyxl as _openpyxl
            _wb = _openpyxl.load_workbook(p_path, data_only=True)
            _ws = _wb.active
            _headers = [_ws.cell(1, c).value for c in range(1, _ws.max_column + 1)]

            def _col(name, fallback):
                try: return _headers.index(name) + 1
                except ValueError: return fallback

            _venue_col  = 3                       # C列: 競馬場
            _name_col   = _col("騎手名", 2)
            _win_col    = _col("勝率", 9)
            _fuku_col   = _col("連対率", 10)      # J列: 連対率
            _power_col  = _col("騎手パワー", 12)

            for _r in range(2, _ws.max_row + 1):
                _venue = str(_ws.cell(_r, _venue_col).value or "").strip()
                _j_raw = str(_ws.cell(_r, _name_col).value or "").replace(" ", "").replace("　", "").strip()
                if not _venue or not _j_raw:
                    continue

                _j_full = normalize_name(_j_raw, res["jockeys"], priority_set=None)
                if _j_full:
                    res["power_jockeys"].add(_j_full)

                _pw = _ws.cell(_r, _power_col).value
                _wv = _ws.cell(_r, _win_col).value
                _fv = _ws.cell(_r, _fuku_col).value
                val_power = str(int(_pw)) if _pw is not None else "-"
                val_win   = f"{round(float(_wv) * 100, 1)}%" if isinstance(_wv, float) else (str(_wv) if _wv else "-")
                val_fuku  = f"{round(float(_fv) * 100, 1)}%" if isinstance(_fv, float) else (str(_fv) if _fv else "-")

                _key = (_venue, _j_full if _j_full else _j_raw)
                res["power_data"][_key] = {"power": val_power, "win": val_win, "fuku": val_fuku}
        except Exception:
            pass

    return res


# ==================================================
# 5. 南関：馬詳細ページ解析（前走騎手を取る）
# ==================================================
def parse_nankankeiba_detail(html, place_name, resources):
    soup = BeautifulSoup(html, "html.parser")
    data = {"meta": {}, "horses": {}}

    h3 = soup.find("h3", class_="nk23_c-tab1__title")
    data["meta"]["race_name"] = h3.get_text(strip=True) if h3 else ""
    if data["meta"]["race_name"]:
        parts = re.split(r"[ 　]+", data["meta"]["race_name"])
        data["meta"]["grade"] = parts[-1] if len(parts) > 1 else ""

    cond = soup.select_one("a.nk23_c-tab1__subtitle__text.is-blue")
    if cond:
        _cond_text = cond.get_text(strip=True)
        data["meta"]["course"] = f"{place_name} {_cond_text}"
        _dist_m = re.search(r'(\d{1,2},\d{3}|\d{3,4})m', _cond_text)
        data["meta"]["dist"] = _dist_m.group(1).replace(',', '') if _dist_m else ""
    else:
        data["meta"]["course"] = ""
        data["meta"]["dist"] = ""

    # 発走時刻（"20:10発走" 等）。レイアウト変動に強いよう全体テキストから拾う。
    data["meta"]["post_time"] = ""
    try:
        page_text = soup.get_text(" ", strip=True)
        tm = (re.search(r'(\d{1,2}:\d{2})\s*発走', page_text)
              or re.search(r'発走\s*[:：]?\s*(\d{1,2}:\d{2})', page_text))
        if tm:
            data["meta"]["post_time"] = tm.group(1)
    except Exception:
        pass

    shosai_area = soup.select_one("#shosai_aria")
    table = shosai_area.select_one("table.nk23_c-table22__table") if shosai_area else None
    if not table:
        for candidate in soup.select("table.nk23_c-table22__table, table"):
            if candidate.select_one("td.umaban") and candidate.select_one("td[class*='cs-z']"):
                table = candidate
                break
    if not table: return data

    PLACE_MAP = {"船": "船橋", "大": "大井", "川": "川崎", "浦": "浦和", "門": "門別", "盛": "盛岡", "水": "水沢",
                 "笠": "笠松", "名": "名古屋", "園": "園田", "姫": "姫路", "高": "高知", "佐": "佐賀"}
    JRA_PLACES = {"札幌", "函館", "福島", "新潟", "東京", "中山", "中京", "京都", "阪神", "小倉"}
    KNOWN_PLACES = list(PLACE_MAP.values()) + ["JRA"] + list(JRA_PLACES)

    rows = table.select("tbody tr") or table.select("tr")
    for row in rows:
        try:
            # 出走取消・除外馬をスキップ（行クラス・行内テキスト・セルクラスで判定）
            row_classes = " ".join(row.get("class", []))
            if any(cls in row_classes for cls in ["is-disable", "is-cancel", "is-scratched", "is-jogai"]):
                continue

            u_tag = row.select_one("td.umaban") or row.select_one("td.is-col02")
            if not u_tag: continue
            umaban = u_tag.get_text(strip=True)
            if not umaban.isdigit(): continue

            # 馬番セル・馬名セル・騎手セル（前半3セル）で取消/除外を検知
            first_cells = row.select("td")[:5]
            first_cells_text = " ".join(c.get_text(strip=True) for c in first_cells)
            first_cells_classes = " ".join(" ".join(c.get("class", [])) for c in first_cells)
            if any(kw in first_cells_text for kw in ["出走取消", "競走除外", "発走除外", "取消", "除外"]):
                continue
            if "is-disable" in first_cells_classes:
                continue

            h_link = (
                row.select_one("td.is-col03 a.is-link")
                or row.select_one("td.pr-umaName-textRound a.is-link")
                or row.select_one("td.cs-umaR a.is-link")
                or row.select_one("a[href*='/uma_info/']")
            )
            horse_name = h_link.get_text(strip=True) if h_link else "不明"
            # uma_info ID（馬の安定キー。メモを別レースでも引き継ぐために使う）
            uma_id = ""
            if h_link and h_link.get("href"):
                _uid_m = re.search(r"/uma_info/(\d+)", h_link.get("href"))
                if _uid_m:
                    uma_id = _uid_m.group(1)

            jg_td = row.select_one("td.cs-g1")
            j_raw, t_raw = "", ""
            if jg_td:
                links = jg_td.select("a")
                if len(links) >= 1: j_raw = links[0].get_text(strip=True)
                if len(links) >= 2: t_raw = links[1].get_text(strip=True)

            j_full = normalize_name(j_raw, resources["jockeys"], resources["power_jockeys"])
            t_full = normalize_name(t_raw, resources["trainers"], None)

            curr_weight = 0.0
            if jg_td:
                w_span = jg_td.select_one("span.nk23_u-text12")
                if w_span:
                    w_text = w_span.get_text(strip=True)
                    wm = re.search(r'(\d+\.\d+|\d+)', w_text)
                    if wm: curr_weight = float(wm.group(1))

            p_data_curr = lookup_power(place_name, j_full, resources["power_data"])
            curr_power_str = "P:不明"
            if p_data_curr:
                cp = p_data_curr["power"]
                cw = str(p_data_curr["win"]).replace("%", "")
                cf = str(p_data_curr["fuku"]).replace("%", "")
                curr_power_str = f"P:{cp}(勝{cw}%/{cf}%)"

            ai2 = row.select_one("td.cs-ai2 .graph_text_div")
            pair_stats = "-"
            if ai2 and "データ" not in ai2.get_text():
                r_el = ai2.select_one(".is-percent")
                w_el = ai2.select_one(".is-number")
                t_el = ai2.select_one(".is-total")
                if r_el and w_el and t_el:
                    r = r_el.get_text(strip=True)
                    w = w_el.get_text(strip=True)
                    t = t_el.get_text(strip=True)
                    pair_stats = f"勝{r}({w}/{t})"

            history = []
            prev_power_val = None
            prev_jockey_latest = ""

            for i in range(1, 6):
                z = row.select_one(f"td.cs-z{i}")
                if not z: continue
                z_full_text = z.get_text(" ", strip=True)
                if not z_full_text: continue
                
                a_tag = z.select_one("a.is-link.event_stop")
                past_url = ""
                race_name = ""
                if a_tag and a_tag.get("href"):
                    past_url = urljoin("https://www.nankankeiba.com", a_tag.get("href"))
                    race_name = re.sub(r"\s+", " ", a_tag.get_text(" ", strip=True)).strip()
                    if race_name in ("詳細", "結果") or re.fullmatch(r"[\d\.年月日/ -]+", race_name or ""):
                        race_name = ""

                d_txt = ""
                place_short = ""
                source_place = ""
                d_div = z.select_one("p.nk23_u-d-flex")
                if d_div:
                    d_raw = d_div.get_text(" ", strip=True)
                    m_dt = re.search(r"(\d+\.\d+\.\d+)", d_raw)
                    if m_dt: d_txt = m_dt.group(1)
                    rem_text = d_raw.replace(d_txt, "") if d_txt else d_raw

                    for kp in KNOWN_PLACES:
                        if kp in rem_text:
                            source_place = kp
                            place_short = "JRA" if kp in JRA_PLACES else kp
                            break
                    if not place_short:
                        for k, v in PLACE_MAP.items():
                            if k in rem_text:
                                source_place = v
                                place_short = v
                                break

                if not d_txt: d_txt = "不明"
                if not place_short: place_short = "不明"
                if not source_place: source_place = place_short

                # 相対評価では従来どおり中央10場を「JRA」とまとめるが、速度指数では
                # 東京/中山など実競馬場と芝・ダート・障害を分けて基準時計を作る。
                if source_place in JRA_PLACES:
                    if re.search(r'障', z_full_text):
                        surface = "障"
                    elif re.search(r'芝', z_full_text):
                        surface = "芝"
                    elif re.search(r'ダ(?:ート)?', z_full_text):
                        surface = "ダ"
                    else:
                        surface = ""
                else:
                    surface = "ダ"

                dm = re.search(r"(\d{3,4})m?", z_full_text)
                dist = dm.group(1) if dm else ""
                frame_no = _extract_frame_no(z, z_full_text)
                field_size, gate_no = _extract_field_gate(z_full_text)

                rank = ""
                r_tag = z.select_one(".nk23_u-text19")
                if r_tag:
                    rank = r_tag.get_text(strip=True).replace("着", "")
                else:
                    special_tag = z.select_one(".nk23_u-text16")
                    if special_tag:
                        rank = special_tag.get_text(strip=True)

                rank_int = int(rank) if str(rank).isdigit() else None

                pos_p = z.select_one("p.position")
                pas = ""
                if pos_p:
                    pas_spans = [s.get_text(strip=True) for s in pos_p.find_all("span")]
                    pas = "-".join(pas_spans)

                _pre_weight = 0.0
                for _p in z.select("p.nk23_u-text10"):
                    _spans = _p.find_all("span")
                    if len(_spans) >= 2:
                        _wm = re.search(r'(\d+\.\d+)', _spans[1].get_text(strip=True))
                        if _wm:
                            _pre_weight = float(_wm.group(1))
                    break

                p_time = 0.0
                if pas and rank_int is not None:
                    time_m = re.search(r'(\d{1,2}):(\d{2}\.\d)', z_full_text)
                    if time_m:
                        p_time = int(time_m.group(1)) * 60 + float(time_m.group(2))
                    else:
                        for m2 in re.finditer(r'(\d{2,3}\.\d)', z_full_text):
                            candidate = float(m2.group(1))
                            if candidate >= 45.0 and (_pre_weight == 0 or abs(candidate - _pre_weight) > 0.05):
                                p_time = candidate
                                break

                j_prev, pop = "", ""
                p_weight = 0.0
                p_lines = z.select("p.nk23_u-text10")
                for p in p_lines:
                    txt = p.get_text(strip=True)
                    if "人気" in txt:
                        pm = re.search(r"(\d+)人気", txt)
                        if pm: pop = f"{pm.group(1)}人"
                        spans = p.find_all("span")
                        if len(spans) >= 2:
                            j_cand = spans[1].get_text(strip=True)
                            j_prev = re.sub(r"[\d\.]+", "", j_cand)
                            wm_prev = re.search(r'(\d+\.\d+|\d+)', j_cand)
                            p_weight = float(wm_prev.group(1)) if wm_prev else 0.0
                        break

                agari = ""
                ft_elem = z.select_one(".furlongtime")
                if ft_elem: agari = ft_elem.get_text(strip=True) or ""

                j_prev_full = normalize_name(j_prev, resources["jockeys"], resources["power_jockeys"])
                if not j_prev_full and j_prev: j_prev_full = j_prev

                if i == 1:
                    prev_jockey_latest = j_prev_full
                    p_key = (place_short, j_prev_full)
                    p_data_prev = lookup_power(place_short, j_prev_full, resources["power_data"])
                    if p_data_prev: prev_power_val = p_data_prev["power"]

                agari_part = f"({agari})" if agari else ""
                pop_part = f"({pop})" if pop else ""
                if str(rank).isdigit(): rank_part = f"{rank}着"
                elif rank: rank_part = rank
                else: rank_part = "着不明"

                draw_part = ""
                if frame_no:
                    draw_part = f"{frame_no}枠 "
                elif field_size and gate_no:
                    draw_part = f"{field_size}頭{gate_no}番({_gate_group(field_size, gate_no)}) "
                h_str = f"{d_txt} {place_short}{dist} {draw_part}{j_prev_full} {pas}{agari_part}→{rank_part}{pop_part}"
                pop_int = None
                if pop:
                    pm2 = re.search(r'(\d+)', pop)
                    if pm2: pop_int = int(pm2.group(1))

                history.append({
                    "text": h_str,
                    "url": past_url,
                    "race_name": race_name,
                    "place": place_short,
                    "source_place": source_place,
                    "surface": surface,
                    "dist": dist,
                    "jockey": j_prev_full,
                    "pas": pas,
                    "rank": rank_int,
                    "pop": pop_int,
                    "frame": frame_no,
                    "field_size": field_size,
                    "gate_no": gate_no,
                    "weight": p_weight,
                    "time": p_time,
                    "date": d_txt
                })

            if prev_power_val:
                power_line = f"【騎手】{curr_power_str}(前P:{prev_power_val})、 相性:{pair_stats}"
            else:
                power_line = f"【騎手】{curr_power_str}、 相性:{pair_stats}"

            data["horses"][umaban] = {
                "name": horse_name,
                "uma_id": uma_id,
                "jockey": j_full,
                "trainer": t_full,
                "power": curr_power_str,
                "weight": curr_weight,
                "prev_jockey": prev_jockey_latest,
                "compat": pair_stats,
            "hist": history,
                "display_power": power_line,
            }

        except Exception:
            continue

    return data


# ==================================================
# 6. 開催特定など
# ==================================================
def get_nankan_kai_nichi(year, month, day, nk_place_code):
    """
    1. 番組表メニュー(bangumi.do)にアクセス
    2. 指定の日付と場所(nk_place_code)に合致する「番組表」リンクを特定
    3. そのリンク先に遷移して「第◯回」「第◯日」を抽出する
    """
    base_url = "https://www.nankankeiba.com"
    menu_url = f"{base_url}/bangumi_menu/bangumi.do"
    
    sess = get_http_session()
    try:
        # ステップ1: 番組表メニューを取得
        res = sess.get(menu_url, timeout=10)
        res.encoding = "cp932"
        soup = BeautifulSoup(res.text, "html.parser")
        
        # ステップ2: ページ内のリンクから、日付と場所コードが一致するものを探す
        # nk_place_code は '20'(大井), '21'(川崎) など
        target_link = None
        kai_nichi_from_href = None
        target_prefix = f"{year}{str(month).zfill(2)}{str(day).zfill(2)}{str(nk_place_code).zfill(2)}"
        for a in soup.find_all("a", href=re.compile(r"/program/")):
            href = a.get("href")
            # URL内に 日付+場所コード が含まれているかチェック
            # 現行URL例: /program/20260520200303.do
            id_match = re.search(r"/program/(\d{14})\.do", href)
            if id_match and id_match.group(1).startswith(target_prefix):
                program_id = id_match.group(1)
                target_link = urljoin(base_url, href)
                kai_nichi_from_href = (int(program_id[10:12]), int(program_id[12:14]))
                break
        
        if not target_link:
            # リンクが見つからない場合は、予備として推測URLを試す
            target_link = f"{base_url}/program/{target_prefix}.do"

        # ステップ3: 遷移先のページ(出馬表/番組表)を解析
        res_page = sess.get(target_link, timeout=10)
        res_page.encoding = "cp932"
        soup_page = BeautifulSoup(res_page.text, "html.parser")
        
        # ページ全体のテキストとタイトルから正規表現で抽出
        title_text = soup_page.title.get_text(" ", strip=True) if soup_page.title else ""
        full_text = title_text + " " + soup_page.get_text(" ", strip=True)
        
        m_kai = re.search(r"第\s*(\d+)\s*回", full_text)
        m_nichi = re.search(r"第\s*(\d+)\s*日", full_text)
        
        if m_kai and m_nichi:
            return int(m_kai.group(1)), int(m_nichi.group(1))
        if kai_nichi_from_href:
            return kai_nichi_from_href

        return None, None
    except Exception as e:
        print(f"開催情報の取得中にエラー: {e}")
        return None, None

NK_CODE_MAP = {"10": "20", "11": "21", "12": "19", "13": "18"}


def build_nankan_race_id(year, month, day, place_code, r_num, kai=None, nichi=None):
    """南関の16桁レースID（/result/{id}.do や /syousai/{id}.do に使う）を組み立てる。
    kai/nichi 未指定なら get_nankan_kai_nichi でネット取得する（fetch_results の補完用）。"""
    nk_place_code = NK_CODE_MAP.get(str(place_code))
    if not nk_place_code:
        return ""
    if kai is None or nichi is None:
        kai, nichi = get_nankan_kai_nichi(year, month, day, nk_place_code)
    if not kai:
        return ""
    return (
        f"{int(year)}{str(month).zfill(2)}{str(day).zfill(2)}"
        f"{str(nk_place_code).zfill(2)}{int(kai):02d}{int(nichi):02d}{int(r_num):02d}"
    )


def get_kb_url_id(year, month, day, place_code, nichi, race_num):
    return f"{year}{str(month).zfill(2)}{str(place_code).zfill(2)}{str(nichi).zfill(2)}{str(race_num).zfill(2)}{str(month).zfill(2)}{str(day).zfill(2)}"


# ==================================================
# 7. 競馬ブック（談話/調教）
# ==================================================
def _extract_danwa_comment_with_mark(raw_text):
    text = re.sub(r'\s+', ' ', str(raw_text or "").replace("\xa0", " ")).strip()
    if not text:
        return ""
    dash = re.search(r'[―-]+', text)
    if not dash:
        return text
    before = text[:dash.start()]
    after = text[dash.end():].strip()
    mark_m = re.search(r'[◎○◯〇⚪△▲☆★◇◆×✕]', before)
    mark = mark_m.group(0) if mark_m else ""
    if mark and after and not after.startswith(mark):
        return f"{mark} {after}".strip()
    return after or text

def parse_kb_danwa_cyokyo(driver, kb_id):
    d_danwa, d_cyokyo = {}, {}
    try:
        driver.get(f"https://s.keibabook.co.jp/chihou/danwa/1/{kb_id}")
        time.sleep(1)
        
        if "login" in driver.current_url or not driver.find_elements(By.CSS_SELECTOR, "a.logout"):
            if login_keibabook_robust(driver):
                driver.get(f"https://s.keibabook.co.jp/chihou/danwa/1/{kb_id}")
                time.sleep(1)

        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "td.umaban")))
        except Exception:
            pass

        soup = BeautifulSoup(driver.page_source, "html.parser")
        for u_td in soup.select("td.umaban"):
            uma = u_td.get_text(strip=True)
            if not uma.isdigit(): continue
            
            danwa_td = None
            for td in u_td.find_all_next("td"):
                if "umaban" in td.get("class", []): break
                if "danwa" in td.get("class", []):
                    danwa_td = td
                    break
            
            if danwa_td:
                raw_text = danwa_td.get_text(" ", strip=True)
                d_danwa[uma] = _extract_danwa_comment_with_mark(raw_text)

        driver.get(f"https://s.keibabook.co.jp/chihou/cyokyo/1/{kb_id}")
        time.sleep(1)
        
        if "login" in driver.current_url or not driver.find_elements(By.CSS_SELECTOR, "a.logout"):
            if login_keibabook_robust(driver):
                driver.get(f"https://s.keibabook.co.jp/chihou/cyokyo/1/{kb_id}")
                time.sleep(1)

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "td.umaban")))
        except Exception:
            pass

        soup = BeautifulSoup(driver.page_source, "html.parser")
        
        for u_td in soup.select("td.umaban"):
            try:
                uma = u_td.get_text(strip=True)
                if not uma.isdigit(): continue

                base_tr = u_td.find_parent("tr")
                if not base_tr: continue

                tp_td = base_tr.select_one("td.tanpyo")
                tp_txt = tp_td.get_text(strip=True) if tp_td else ""

                content_td = None
                for td in u_td.find_all_next("td"):
                    if "umaban" in td.get("class", []): break
                    if td.has_attr("colspan") and td.select_one("dl.dl-table"):
                        content_td = td
                        break

                if not content_td:
                    d_cyokyo[uma] = f"【短評】{tp_txt}"
                    continue

                cyokyo_lines = []
                dls = content_td.select("dl.dl-table")

                for dl in dls:
                    first_dt = dl.find("dt")
                    first_dt_text = first_dt.get_text(strip=True) if first_dt else ""
                    label = "前走調教" if "(前回)" in first_dt_text else "今走調教"

                    dt_left = dl.select_one("dt.left")
                    info_text = dt_left.get_text(" ", strip=True).replace("\xa0", " ") if dt_left else ""
                    dt_right = dl.select_one("dt.right")
                    cond_text = dt_right.get_text(strip=True).replace("\xa0", " ") if dt_right else ""

                    next_table = None
                    for nxt in dl.find_all_next():
                        if nxt.name == "dl": break
                        if nxt.name == "table" and "cyokyodata" in nxt.get("class", []):
                            next_table = nxt
                            break
                    
                    time_data = ""
                    if next_table:
                        data_rows = []
                        for tr in next_table.select("tr"):
                            cells = [td.get_text(strip=True).replace("\xa0", " ") for td in tr.find_all("td") if td.get_text(strip=True)]
                            if cells: data_rows.append(" ".join(cells))
                        time_data = " / ".join(data_rows)

                    line = f"{label}：{info_text} {cond_text} {time_data}".strip()
                    line = re.sub(r"\s{2,}", " ", line)
                    cyokyo_lines.append(line)

                full_text = f"【短評】{tp_txt}".strip()
                if cyokyo_lines: full_text += "\n" + "\n".join(cyokyo_lines)

                d_cyokyo[uma] = full_text

            except Exception:
                continue

        if not d_cyokyo:
            raw = soup.get_text("\n", strip=True)
            for m in re.finditer(r"(\d+)[^\n]*(?:前走調教|今走調教)：[^\n]+", raw):
                um = m.group(1)
                txt = m.group(0)
                txt = re.sub(r"^\d+\s*", "", txt)
                existing = d_cyokyo.get(um, "")
                d_cyokyo[um] = (existing + "\n" + txt).strip() if existing else txt

    except Exception:
        pass

    return d_danwa, d_cyokyo


# ==================================================
# 8. テキスト正規化・パースユーティリティ
# ==================================================
import unicodedata

def _norm_horse_name(s: str) -> str:
    if s is None: return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = s.strip()
    s = re.sub(r"\s+", "", s.replace("　", " "))
    s = re.sub(r"[（(].*?[）)]", "", s)
    s = re.sub(r"[*_`]", "", s)
    return s


# ==================================================
# 8.5 過去レースからのテン速度取得
# ==================================================
def fetch_past_race_1f_times(url, past_dist=0, driver=None):
    sess = get_http_session()
    try:
        time.sleep(0.5)
        res = sess.get(url, timeout=5)
        res.encoding = "cp932"
        m = re.search(r'([0-9]{1,2}\.[0-9]-[0-9]{1,2}\.[0-9](?:-[0-9]{1,2}\.[0-9])*)', res.text)
        if not m: return None
            
        parts = [float(p) for p in m.group(1).split('-')]
        if not parts: return None
            
        p0 = parts[0]
        
        if past_dist > 0:
            d0 = past_dist % 200
            if d0 == 0: d0 = 200
            
            if d0 < 150 and len(parts) >= 2:
                total_time = p0 + parts[1]
                total_dist = d0 + 200
                return (total_time, total_dist)
            else:
                return (p0, d0)
        else:
            dist_0 = 100 if p0 < 8.0 else 200
            return (p0, dist_0)
    except Exception:
        return None

# ==================================================
# 9. 対戦表（AI評価付き）
# ==================================================
def _matchup_html_has_table(page_html):
    return bool(page_html) and "<table" in page_html and (
        "table08" in page_html or "馬名" in page_html or "対戦" in page_html
    )

def _taisen_cache_path(url):
    """対戦表の生HTMLキャッシュパス。レース終了後にサイトからページが削除されるため、
    取得成功時に保存しておき、過去日の再実行ではキャッシュから復元する。"""
    m_id = re.search(r"/taisen/(\d+)\.do", url or "")
    if not m_id:
        return None
    try:
        os.makedirs("cache", exist_ok=True)
        return os.path.join("cache", f"taisen_{m_id.group(1)}.html")
    except Exception:
        return None


def _load_taisen_cache(cache_path):
    if cache_path and os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cached = f.read()
            if _matchup_html_has_table(cached):
                return cached
        except Exception:
            pass
    return ""


def _save_taisen_cache(cache_path, page_html):
    if cache_path and page_html:
        try:
            with open(cache_path, "w", encoding="utf-8") as f:
                f.write(page_html)
        except Exception:
            pass


def _fetch_matchup_page_html(driver, url):
    last_error = ""
    cache_path = _taisen_cache_path(url)
    page_gone = False

    # 対戦表（/taisen/...do）はJS不要・ログイン不要の静的HTML。
    # requestsで取得するのが最も確実かつ高速なので、まずこちらを試す。
    # （seleniumのヘッドレス描画は待機タイミングで空振りしやすく取得失敗の原因になる）
    for attempt in range(2):
        try:
            sess = get_http_session()
            res = sess.get(url, timeout=15)
            res.encoding = "cp932"
            if res.status_code >= 400:
                last_error = f"HTTP {res.status_code}"
                # 404は分析ページ未生成/期限切れ。リトライしても無駄なので即抜ける
                if res.status_code == 404:
                    page_gone = True
                    break
            elif _matchup_html_has_table(res.text):
                _save_taisen_cache(cache_path, res.text)
                return res.text, ""
            else:
                last_error = "対戦表テーブル未検出(requests)"
        except Exception as e:
            last_error = f"requests {type(e).__name__}: {e}"
        time.sleep(0.5 + attempt * 0.5)

    # 404 = レース終了後にページ削除済み。seleniumで再試行しても無駄なので、
    # 過去取得時のキャッシュがあればそれを使い、無ければ明確な理由を返す。
    if page_gone:
        cached = _load_taisen_cache(cache_path)
        if cached:
            return cached, ""
        return "", "ページ削除済み(HTTP 404: レース終了後は対戦表が公開されず、キャッシュもありません)"

    # フォールバック: selenium（requestsがブロック/失敗した場合の保険）
    for attempt in range(2):
        try:
            driver.get(url)
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: _matchup_html_has_table(d.page_source or "")
                )
            except TimeoutException:
                pass

            page_html = driver.page_source or ""
            if _matchup_html_has_table(page_html):
                _save_taisen_cache(cache_path, page_html)
                return page_html, ""
            last_error = "対戦表テーブル未検出(selenium)"
        except TimeoutException as e:
            last_error = f"TimeoutException: {e}"
            try:
                driver.execute_script("window.stop();")
                page_html = driver.page_source or ""
                if _matchup_html_has_table(page_html):
                    return page_html, ""
            except Exception:
                pass
        except Exception as e:
            last_error = f"{type(e).__name__}: {e}"

        time.sleep(0.6 + attempt * 0.6)

    # 最終フォールバック: 過去取得時の生HTMLキャッシュ
    cached = _load_taisen_cache(cache_path)
    if cached:
        return cached, ""
    return "", last_error

def _fetch_matchup_table_selenium(driver, nankan_id, grades, horses_data=None, current_course="", current_dist=""):
    url = f"https://www.nankankeiba.com/taisen/{nankan_id}.do"
    try:
        page_html, load_error = _fetch_matchup_page_html(driver, url)
        if not page_html:
            return f"(対戦表取得エラー: {load_error or 'ページHTML取得失敗'})"

        soup = BeautifulSoup(page_html, "html.parser")
        
        tbl = soup.find("table", class_=re.compile(r"table08"))
        if not tbl:
            for t in soup.find_all("table"):
                if "馬名" in t.get_text() and t.find("a", href=re.compile(r"(result|race).*\.do")):
                    tbl = t
                    break
        
        if not tbl: return "\n(公式の対戦表データなし)"

        def _cell_span(cell, attr):
            try:
                return max(1, int(cell.get(attr, 1) or 1))
            except (TypeError, ValueError):
                return 1

        grid = {}
        rows = tbl.find_all("tr")
        for r_idx, tr in enumerate(rows):
            c_idx = 0
            for cell in tr.find_all(["td", "th"]):
                while (r_idx, c_idx) in grid:
                    c_idx += 1
                
                rowspan = _cell_span(cell, "rowspan")
                colspan = _cell_span(cell, "colspan")
                
                for r in range(rowspan):
                    for c in range(colspan):
                        grid[(r_idx + r, c_idx + c)] = cell
                
                c_idx += colspan
        
        if not grid: return "\n(公式の対戦表データなし)"
        
        max_col = max([c for (r, c) in grid.keys()])
        max_row = max([r for (r, c) in grid.keys()])

        # 馬名列（uma_infoリンクを含む列）を特定し、レース検出から除外
        horse_info_cols = set()
        for r_idx in range(max_row + 1):
            for c_idx in range(max_col + 1):
                cell = grid.get((r_idx, c_idx))
                if cell and cell.find("a", href=re.compile(r"uma_info")):
                    horse_info_cols.add(c_idx)

        races = []
        for c_idx in range(max_col + 1):
            if c_idx in horse_info_cols:
                continue
            for r_idx in range(min(3, max_row + 1)):
                cell = grid.get((r_idx, c_idx))
                if not cell: continue

                link = cell.find("a", href=re.compile(r"(result|race).*\.do|/\d{10,}"))
                if link:
                    title_elem = cell.find(class_=re.compile(r"detail"))
                    if title_elem:
                        title = title_elem.get_text(" ", strip=True)
                    else:
                        title = cell.get_text(" ", strip=True).replace("競走成績", "").strip()
                        
                    href = link.get("href", "")
                    id_match = re.search(r"(\d{10,})", href)
                    if id_match:
                        full_url = f"https://www.nankankeiba.com/result/{id_match.group(1)}.do"
                    elif href.startswith("/"):
                        full_url = "https://www.nankankeiba.com" + href
                    else:
                        full_url = href

                    if not any(r["url"] == full_url for r in races):
                        races.append({
                            "title": title,
                            "url": full_url,
                            "results": [],
                            "col_index": c_idx
                        })
                    break

        if not races: return "\n(公式の対戦表データなし)"

        current_horse_names = set()
        if horses_data:
            for u, h in horses_data.items():
                cn = _norm_horse_name(h.get("name", ""))
                if cn: current_horse_names.add(cn)

        processed_horses = set()
        safe_grades = {k: _rank_floor_e(v) for k, v in grades.items()} if isinstance(grades, dict) else {}

        for r_idx in range(max_row + 1):
            name = ""
            for c_idx in range(max_col + 1):
                cell = grid.get((r_idx, c_idx))
                if cell:
                    name_link = cell.find("a", href=re.compile(r"uma_info"))
                    if name_link:
                        name = name_link.get_text(strip=True)
                        break
                        
            if not name or name in processed_horses: continue
            processed_horses.add(name)

            name_norm = _norm_horse_name(name)
            grade = safe_grades.get(name_norm, "")
            if not grade and name_norm:
                for k_norm, v in safe_grades.items():
                    if k_norm and (k_norm in name_norm or name_norm in k_norm):
                        grade = v
                        break

            for race_info in races:
                col_idx = race_info["col_index"]
                cell = grid.get((r_idx, col_idx))
                if not cell: continue

                txt = cell.get_text(" ", strip=True)
                if not txt or txt == "-": continue

                rnk = ""
                time_diff = None
                if "｜" in txt or "|" in txt:
                    parts = re.split(r'[｜|]', txt)
                    rnk = parts[0].strip()
                    for p in parts[1:]:
                        tm = re.search(r'[-+]?\d+(?:\.\d+)?', p)
                        if tm:
                            try:
                                time_diff = abs(float(tm.group(0)))
                            except ValueError:
                                time_diff = None
                            break
                else:
                    sp = cell.find("span")
                    if sp: rnk = sp.get_text(strip=True)
                    else: rnk = re.sub(r"\D", "", txt)

                if not rnk and any(x in txt for x in ["除外", "中止", "取消", "除", "中", "取"]):
                    rnk = "除外/中止"

                if rnk:
                    if current_horse_names and name_norm not in current_horse_names:
                        continue
                        
                    race_info["results"].append({
                        "rank": rnk,
                        "name": name,
                        "grade": grade,
                        "time_diff": time_diff,
                        "sort": int(rnk) if str(rnk).isdigit() else 999
                    })

        # 対戦表は常に直近の日付が上に来るよう並べ替える。
        def _race_title_date_key(race):
            t = re.sub(r'[*]', '', race.get("title", ""))
            m = re.search(r'(\d{2,4}[./]\d{1,2}[./]\d{1,2})', t)
            d = parse_date(m.group(1)) if m else datetime.min
            return d if isinstance(d, datetime) else datetime.min
        races.sort(key=_race_title_date_key, reverse=True)

        out = ["\n【対戦表（AI評価付き）】"]
        for r in races:
            if len(r["results"]) < 2: continue
                
            r["results"].sort(key=lambda x: x.get("sort", 999))
            result_times = {}
            result_gates = {}
            if r.get("url"):
                try:
                    result_times, _, _, _, _, result_gates = fetch_race_all_horses(r["url"])
                except Exception:
                    result_times = {}
                    result_gates = {}
            line_parts = []
            for x in r["results"]:
                grade_str = x.get('grade', '')
                g = f"({grade_str})" if grade_str else ""

                curr_umaban = ""
                if horses_data:
                    name_norm = _norm_horse_name(x['name'])
                    for u, h in horses_data.items():
                        if _norm_horse_name(h.get("name", "")) == name_norm:
                            curr_umaban = f"[{u}]"
                            break

                past_gate = result_gates.get(x["name"])
                if past_gate is None:
                    name_norm = _norm_horse_name(x["name"])
                    for gate_name, gate_no in result_gates.items():
                        gate_norm = _norm_horse_name(gate_name)
                        if gate_norm and (gate_norm == name_norm or gate_norm in name_norm or name_norm in gate_norm):
                            past_gate = gate_no
                            break
                past_gate_str = _to_circled(past_gate) if past_gate is not None else ""

                line_parts.append(f"{x['rank']}着 {curr_umaban}{x['name']}{past_gate_str}{g}")

            if line_parts:
                title = r.get("title", "")
                link = f"\nLink: {r['url']}" if r.get("url") else ""
                is_same_both = (current_course and current_dist and current_course in title and current_dist in title)
                is_same_course = (current_course and current_course in title)
                if is_same_both:
                    title_display = f"***{title}***"
                    marker = "❗️"
                elif is_same_course:
                    title_display = f"**{title}**"
                    marker = "◆"
                else:
                    title_display = title
                    marker = "◆"
                out.append(f"{marker} {title_display}\n" + " / ".join(line_parts) + link)

        if len(out) == 1: return "\n(公式の対戦表データなし)"
        return "\n".join(out)
    except Exception as e:
        return f"(対戦表取得エラー: {type(e).__name__}: {e})"


def compute_matchup_points(match_txt, horses_data):
    """対戦表(公式の出走馬同士の先着表)テキストから、現出走馬ごとの先着ポイント(0〜10)を算出する。

    各過去レースで「今走出走馬同士」の先着関係を集計。同場同距離(❗️)の先着を重く(2.0)、
    それ以外(◆)を軽く(1.0)重み付けし、相手1頭につき最良条件の先着を1回だけ計上(同一相手の
    繰り返し先着で過大計上しない)。最後に各馬の合計を 0〜10 にクランプする。降格には使わず加点のみ。
    返り値: {umaban(str): points(int 0..10)}
    """
    out = {str(u): 0 for u in (horses_data or {}).keys()}
    if not match_txt or "着" not in match_txt:
        return out

    current_umaban = set(out.keys())
    # (勝者umaban, 敗者umaban) -> 最良の条件重み
    beats = {}
    lines = match_txt.split("\n")
    i = 0
    n = len(lines)
    while i < n:
        line = lines[i].strip()
        if line.startswith("❗️"):
            weight = 2.0
        elif line.startswith("◆"):
            weight = 1.0
        else:
            i += 1
            continue
        # タイトル行の直後にある結果行(空行・Link行は読み飛ばす)を探す
        results_line = ""
        j = i + 1
        while j < n:
            cand = lines[j].strip()
            if not cand or cand.startswith("Link:"):
                j += 1
                continue
            if cand.startswith("❗️") or cand.startswith("◆"):
                break
            results_line = cand
            break
        if results_line:
            entries = []  # (着順int, umaban_str)
            for part in results_line.split(" / "):
                m_rank = re.match(r'\s*(\d+)着', part)
                m_uma = re.search(r'\[(\d+)\]', part)  # 現出走馬のみ [N] が付く
                if m_rank and m_uma and m_uma.group(1) in current_umaban:
                    entries.append((int(m_rank.group(1)), m_uma.group(1)))
            for ai in range(len(entries)):
                for bi in range(len(entries)):
                    if ai == bi:
                        continue
                    ra, ua = entries[ai]
                    rb, ub = entries[bi]
                    if ra < rb:  # ua が ub に先着
                        key = (ua, ub)
                        if weight > beats.get(key, 0):
                            beats[key] = weight
            i = j
            continue
        i += 1

    totals = {u: 0.0 for u in current_umaban}
    for (winner, _loser), w in beats.items():
        totals[winner] += w
    for u in current_umaban:
        out[u] = int(max(0, min(10, round(totals[u]))))
    return out


def adjust_grades_by_matchups(match_txt, grades, current_course="", current_dist="", internal_verdicts=None):
    if not grades or not match_txt:
        return grades

    rank_val = {"S": 7, "A": 6, "B": 5, "C": 4, "D": 3, "E": 2, "F": 1, "G": 0}
    val_to_rank = {7: "S", 6: "A", 5: "B", 4: "C", 3: "D", 2: "E", 1: "E", 0: "E"}
    new_grades = {k: _rank_floor_e(v) for k, v in grades.items()}

    # 暴走抑制(26/6/10大井1Rで67点の馬がE降格→1着の反省):
    # (1) 内部相対評価(タイム差・外れ値除去・最新同条件優先済み)が「負けていない」と
    #     判定しているペアは、生着順の累積負けカウントでは降格させない。
    # (2) このペアの最新対戦で勝っていれば、古い負けは降格根拠にしない。
    # (3) 本関数による降格は素のランクから最大2段まで(同場同距離の大差負けは
    #     apply_internal_rank_locks の position ロックが無制限に保証する)。
    norm_verdicts = internal_verdicts or {}
    start_vals = {k: rank_val.get(v, 0) for k, v in new_grades.items()}
    MAX_MATCHUP_DEMOTION = 2

    def _internal_blocks_demotion(loser, winner):
        return norm_verdicts.get(loser, {}).get(winner) in ("=", ">")

    def _demotion_floor(loser):
        return max(0, start_vals.get(loser, 0) - MAX_MATCHUP_DEMOTION)

    def _clean_match_name(raw_name):
        name = str(raw_name or "").strip()
        name = re.sub(r'^\[\d{1,2}\]', '', name).strip()
        name = re.sub(r'\([SABCDEFG-]\)\s*$', '', name).strip()
        name = re.sub(r'[\u2460-\u2473]+$', '', name).strip()
        return _norm_horse_name(name)

    def _extract_time_hint(raw_text):
        if "｜" not in str(raw_text) and "|" not in str(raw_text):
            return None
        parts = re.split(r'[｜|]', str(raw_text))
        for p in parts[1:]:
            m = re.search(r'[-+]?\d+(?:\.\d+)?', p)
            if not m:
                continue
            try:
                return abs(float(m.group(0)))
            except ValueError:
                return None
        return None

    def _lookup_norm_value(norm_map, name_norm):
        if not norm_map or not name_norm:
            return None
        if name_norm in norm_map:
            return norm_map[name_norm]
        for k_norm, value in norm_map.items():
            if k_norm and (k_norm == name_norm or k_norm in name_norm or name_norm in k_norm):
                return value
        return None

    def _date_key(dt_text):
        dt_obj = parse_date(dt_text or "")
        return dt_obj.timestamp() if dt_obj != datetime.min else 0

    def _extract_race_place_dist(title_text):
        plain = re.sub(r'[*]', '', str(title_text or ""))
        place = ""
        for kp in ("大井", "川崎", "船橋", "浦和"):
            if kp in plain:
                place = kp
                break
        dist = ""
        dist_matches = re.findall(r'(?<!\d)(\d{3,4})(?!\d)', plain.replace(",", ""))
        if dist_matches:
            known_distances = {
                "大井": {1000, 1200, 1400, 1500, 1600, 1650, 1700, 1800, 2000, 2400, 2600},
                "川崎": {900, 1400, 1500, 1600, 2000, 2100},
                "船橋": {1000, 1200, 1500, 1600, 1700, 1800, 2200, 2400},
                "浦和": {800, 1300, 1400, 1500, 1600, 1900, 2000},
            }
            known = known_distances.get(place, set())
            known_hits = [d for d in dist_matches if int(d) in known]
            current_like = [d for d in dist_matches if 700 <= int(d) <= 2600]
            dist = known_hits[-1] if known_hits else (current_like[-1] if current_like else dist_matches[-1])
        return place, dist

    races = []
    current = None
    for line in str(match_txt).splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith(("◆", "❗")):
            if current:
                races.append(current)
            current = {"title": stripped, "url": "", "results": []}
            continue
        if current is None:
            continue
        if stripped.startswith("Link:"):
            m_url = re.search(r'https?://\S+', stripped)
            if m_url:
                current["url"] = m_url.group(0)
            continue
        if "着" not in stripped:
            continue

        for seg in re.split(r'\s*/\s*', stripped):
            m = re.search(r'(\d+)着\s+(.+)', seg)
            if not m:
                continue
            name = _clean_match_name(m.group(2))
            if name:
                current["results"].append({
                    "rank": int(m.group(1)),
                    "name": name,
                    "time": _extract_time_hint(seg),
                })

    if current:
        races.append(current)

    pair_records = {}
    directed_losses = {}
    same_cond_time_locks = []
    cur_dist_str = str(current_dist or "").replace(",", "")
    cur_dist_int = _dist_to_int(cur_dist_str)
    SAME_COND_LOSER_TIME_CAP = 0.6 if 0 < cur_dist_int <= 1400 else 0.8

    for race in races:
        results = race.get("results", [])
        if len(results) < 2:
            continue
        title = race.get("title", "")
        title_plain = re.sub(r'[*]', '', title)
        race_place, race_dist = _extract_race_place_dist(title_plain)
        if current_course and cur_dist_str and race_dist and not _is_relative_distance_allowed(race_dist, cur_dist_str):
            continue
        is_same_cond = ("***" in title) or ("❗" in title)
        if current_course and cur_dist_str:
            is_same_cond = is_same_cond or (current_course in title_plain and cur_dist_str in title_plain.replace(",", ""))
        is_same_layout = (
            bool(current_course and cur_dist_str and race_place and race_dist)
            and race_place == current_course
            and _is_same_track_layout(current_course, race_dist, cur_dist_str)
        )

        date_match = re.search(r'(\d{2,4}[./]\d{1,2}[./]\d{1,2})', title_plain)
        date_text = date_match.group(1) if date_match else ""

        if race.get("url"):
            try:
                result_times, _, _, _, _, _ = fetch_race_all_horses(race.get("url"))
            except Exception:
                result_times = {}
            time_by_norm = {_norm_horse_name(k): v for k, v in (result_times or {}).items()}
            for item in results:
                if item.get("time") is not None:
                    continue
                item["time"] = _lookup_norm_value(time_by_norm, item.get("name"))

        if is_same_cond:
            timed_results = [x for x in results if x.get("time") is not None]
            for loser_item in timed_results:
                if float(loser_item.get("time") or 0.0) < SAME_COND_LOSER_TIME_CAP:
                    continue
                for winner_item in timed_results:
                    if winner_item["name"] == loser_item["name"]:
                        continue
                    if winner_item["rank"] >= loser_item["rank"]:
                        continue
                    same_cond_time_locks.append({
                        "winner": winner_item["name"],
                        "loser": loser_item["name"],
                        "loser_time": float(loser_item.get("time") or 0.0),
                        "time_diff": round(float(loser_item.get("time") or 0.0) - float(winner_item.get("time") or 0.0), 1),
                        "date": date_text,
                    })

        for i in range(len(results)):
            for j in range(i + 1, len(results)):
                rank_a, name_a = results[i]["rank"], results[i]["name"]
                rank_b, name_b = results[j]["rank"], results[j]["name"]
                if name_a == name_b or rank_a == rank_b:
                    continue
                winner, loser = (name_a, name_b) if rank_a < rank_b else (name_b, name_a)
                rank_diff = abs(rank_a - rank_b)
                time_a, time_b = results[i].get("time"), results[j].get("time")
                time_diff = None
                if time_a is not None and time_b is not None:
                    try:
                        time_diff = abs(float(time_a) - float(time_b))
                    except (TypeError, ValueError):
                        time_diff = None
                pair_key = tuple(sorted((winner, loser)))
                rec = {
                    "winner": winner,
                    "loser": loser,
                    "rank_diff": rank_diff,
                    "time_diff": time_diff,
                    "is_same_cond": is_same_cond,
                    "is_same_layout": is_same_layout,
                    "date": date_text,
                }
                pair_records.setdefault(pair_key, []).append(rec)
                directed_losses.setdefault((loser, winner), []).append(rec)

    # 同条件で一定以上負けている馬は、同レースで先着した今回出走馬より上にしない。
    # 目安は短距離0.6秒以上、中距離以上0.8秒以上。
    # 単発の同条件大敗を「展開や別路線の加点」だけで逆転させないためのハードロック。
    def _has_counter_win(loser, winner, lock_date):
        """同じ相手へ過去または後続で先着があれば、展開依存の可能性としてロックを見送る。"""
        records = pair_records.get(tuple(sorted((loser, winner))), [])
        if not records:
            return False
        lock_key = _date_key(lock_date)
        for rec in records:
            if rec.get("winner") != loser or rec.get("loser") != winner:
                continue
            rec_key = _date_key(rec.get("date", ""))
            if lock_key and rec_key:
                if rec_key != lock_key:
                    return True
            else:
                return True
        return False

    if same_cond_time_locks:
        # 非反復・パス入口値基準: 勝者が本パス内で降格しても敗者を連鎖追随させない
        # (連鎖は点数差で作った評価の階段を E に潰すため)。
        entry_vals = {k: rank_val.get(v, 0) for k, v in new_grades.items()}
        for lock in same_cond_time_locks:
            loser = lock.get("loser")
            winner = lock.get("winner")
            if loser not in new_grades or winner not in new_grades:
                continue
            if _has_counter_win(loser, winner, lock.get("date", "")):
                continue
            target = entry_vals.get(winner, 0)
            if rank_val.get(new_grades[loser], 0) > target:
                new_grades[loser] = val_to_rank.get(target, new_grades[winner])

    # 今回と同競馬場・同コース形態で、同じ相手に「直近2連続」で下に出た馬は相手より上の評価にしない。
    # ※最新の対戦で勝ち返している場合(idx>0の古い連敗)は降格根拠にしない。
    for records in pair_records.values():
        same_layout = [r for r in records if r.get("is_same_layout")]
        same_layout.sort(key=lambda r: _date_key(r["date"]), reverse=True)
        if len(same_layout) >= 2:
            first = same_layout[0]
            second = same_layout[1]
            if first["loser"] == second["loser"] and first["winner"] == second["winner"]:
                loser = first["loser"]
                winner = first["winner"]
                if loser in new_grades and winner in new_grades and not _internal_blocks_demotion(loser, winner):
                    loser_val = rank_val.get(new_grades[loser], 0)
                    winner_val = rank_val.get(new_grades[winner], 0)
                    if loser_val > winner_val:
                        target = max(winner_val, _demotion_floor(loser))
                        if target < loser_val:
                            new_grades[loser] = val_to_rank.get(target, new_grades[winner])

    # 既存の大敗傾向補正も、名前・同条件判定を正規化した結果で継続する。
    def _latest_record_reversed(loser, winner):
        # ペアの最新対戦で loser 側が勝ち返していれば、古い負けは降格根拠にしない。
        records = pair_records.get(tuple(sorted((loser, winner))), [])
        if not records:
            return False
        latest = max(records, key=lambda r: _date_key(r["date"]))
        return latest["winner"] == loser

    for _ in range(2):
        changed = False
        for (loser, winner), losses in directed_losses.items():
            if loser not in new_grades or winner not in new_grades:
                continue
            if _internal_blocks_demotion(loser, winner):
                continue
            if _latest_record_reversed(loser, winner):
                continue

            count_total = len(losses)
            same_cond_losses = [l for l in losses if l["is_same_cond"]]
            cond1 = count_total >= 3 and len(same_cond_losses) >= 1
            cond2 = len(same_cond_losses) >= 2
            cond3 = count_total >= 2 and all(l["rank_diff"] >= 5 for l in losses) and len(same_cond_losses) >= 1
            big_losses = [l for l in losses if l["rank_diff"] >= 3]
            cond4 = len(big_losses) >= 2

            if not (cond1 or cond2 or cond3 or cond4):
                continue

            loser_val = rank_val.get(new_grades[loser], 0)
            winner_val = rank_val.get(new_grades[winner], 0)
            if loser_val > winner_val:
                new_val = max(winner_val, loser_val - 1, _demotion_floor(loser))
                if new_val < loser_val:
                    new_grades[loser] = val_to_rank.get(new_val, "G")
                    changed = True
            elif loser_val == winner_val and (cond1 or cond3 or cond4):
                new_val = max(0, loser_val - 1, _demotion_floor(loser))
                if new_val < loser_val:
                    new_grades[loser] = val_to_rank.get(new_val, "G")
                    changed = True

        if not changed:
            break

    # 引き上げ保証: 今回と同競馬場・同距離で、同じ相手に3ヶ月以内に2回以上先着している馬は、
    # その相手より低い評価にしない（相手の評価まで引き上げる）。
    # 例: 先着を許した馬がC評価なら、2回以上先着している馬はC以上へ調整する。
    today_dt = datetime.now()
    THREE_MONTHS_DAYS = 92

    def _within_3months(date_text):
        d = parse_date(date_text or "")
        if not isinstance(d, datetime) or d == datetime.min:
            return False
        return 0 <= (today_dt - d).days <= THREE_MONTHS_DAYS

    promote_targets = {}  # winner -> 最低保証する rank_val
    for pair_key, records in pair_records.items():
        # 同条件(今回と同場・同距離) かつ 3ヶ月以内 の記録だけを対象にする。
        recent_same = [
            r for r in records
            if r.get("is_same_cond") and _within_3months(r.get("date", ""))
        ]
        if len(recent_same) < 2:
            continue
        # 各方向(どちらが先着したか)ごとに先着回数を数える。
        win_counts = {}
        for r in recent_same:
            win_counts[r["winner"]] = win_counts.get(r["winner"], 0) + 1
        for winner, cnt in win_counts.items():
            if cnt < 2:
                continue
            loser = pair_key[0] if pair_key[1] == winner else pair_key[1]
            # 相手の方が先着回数が多い/同数の矛盾ケースは引き上げない。
            if cnt <= win_counts.get(loser, 0):
                continue
            if winner not in new_grades or loser not in new_grades:
                continue
            winner_val = rank_val.get(new_grades[winner], 0)
            # 相手(先着を許した馬)の評価は「降格前=関数入口」の値を基準にする。
            # 本パス内で相手がこの先着により降格していても、先着馬は本来の相手評価まで引き上げる。
            loser_val = start_vals.get(loser, rank_val.get(new_grades[loser], 0))
            if winner_val < loser_val:
                promote_targets[winner] = max(promote_targets.get(winner, winner_val), loser_val)

    for winner, target_val in promote_targets.items():
        cur_val = rank_val.get(new_grades[winner], 0)
        if target_val > cur_val:
            new_grades[winner] = val_to_rank.get(target_val, new_grades[winner])

    return new_grades

def apply_internal_rank_locks(grades, locks):
    """相対評価モジュールが検出した『今回と同場の明確勝敗』を最終グレードへ反映する
    (JRA app の apply_same_condition_rank_locks 相当)。
    対戦表スクレイプ(adjust_grades_by_matchups)は反復敗戦でしか発火せず、生着順ベースで
    外れ値フィルタを持たないが、こちらは内部の時間差・外れ値除去・逃げ救済を通した判定なので、
    単発の同条件完敗でも『負けた馬が勝った馬を上回らない』を保証できる。

    重要: 降格目標は「ロック適用前のランク(=素点バンド)」を基準に1回だけ計算する。
    勝者自身が別ロックでさらに降格しても、敗者をその連鎖で追随降格させない。
    反復適用は ≤ の連鎖で C/D 帯をまとめて E に潰し、点数差で作った評価の階段を
    破壊するため廃止した(26/6/11 大井1R: 素点50点=バンドA相当の馬が連鎖でE化した反省)。
    """
    if not grades or not locks:
        return grades

    rank_val = {"S": 7, "A": 6, "B": 5, "C": 4, "D": 3, "E": 2, "F": 1, "G": 0}
    val_to_rank = {7: "S", 6: "A", 5: "B", 4: "C", 3: "D", 2: "E", 1: "E", 0: "E"}
    base_vals = {k: rank_val.get(_rank_floor_e(v), 0) for k, v in grades.items()}
    new_vals = dict(base_vals)

    for lock in locks:
        winner = lock.get("winner")
        loser = lock.get("loser")
        if not winner or not loser or winner not in base_vals or loser not in base_vals:
            continue
        w_base = base_vals[winner]
        # clear(明確敗)は勝者と同格まで、strong(完敗 ＜＜)は勝者の1段下まで許容。
        target = max(0, w_base - 1) if lock.get("strength") == "strong" else w_base
        if new_vals[loser] > target:
            new_vals[loser] = target

    return {k: val_to_rank.get(v, "E") for k, v in new_vals.items()}

def adjust_grades_by_jockey_choices(grades, horses_data):
    if not grades or not horses_data: return grades

    rank_val = {"S": 7, "A": 6, "B": 5, "C": 4, "D": 3, "E": 2, "F": 1, "G": 0}
    val_to_rank = {7: "S", 6: "A", 5: "B", 4: "C", 3: "D", 2: "E", 1: "E", 0: "E"}

    def _jockey_p_values(horse):
        p_raw = horse.get("display_power", "")
        m_current = re.search(r'P:(\d+)', p_raw)
        m_prev = re.search(r'前P:(\d+)', p_raw)
        p_current = int(m_current.group(1)) if m_current else 5
        p_prev = int(m_prev.group(1)) if m_prev else p_current
        return p_current, p_prev

    horse_state = {}
    current_by_jockey = {}
    prev_by_jockey = {}

    for umaban, h in horses_data.items():
        name = _norm_horse_name(h.get("name", ""))
        if not name:
            continue

        hist = h.get("hist", [])
        prev_jockey = (h.get("prev_jockey") or "").strip()
        if hist and isinstance(hist[0], dict):
            prev_jockey = prev_jockey or (hist[0].get("jockey") or "").strip()
        current_jockey = (h.get("jockey") or "").strip()
        p_current, p_prev = _jockey_p_values(h)

        horse_state[name] = {
            "horse": h,
            "current_jockey": current_jockey,
            "prev_jockey": prev_jockey,
            "p_current": p_current,
            "p_prev": p_prev,
        }
        if current_jockey:
            current_by_jockey.setdefault(current_jockey, []).append(name)
        if prev_jockey:
            prev_by_jockey.setdefault(prev_jockey, []).append(name)

    new_grades = {k: _rank_floor_e(v) for k, v in grades.items()}

    def _is_jockey_upgrade(name):
        st_info = horse_state.get(name, {})
        return (
            st_info.get("current_jockey")
            and st_info.get("prev_jockey")
            and st_info.get("current_jockey") != st_info.get("prev_jockey")
            and int(st_info.get("p_current", 0) or 0) > int(st_info.get("p_prev", 0) or 0)
        )

    def _cap_unselected_to_chosen(unselected_name, chosen_names, chosen_jockey):
        if unselected_name not in new_grades or _is_jockey_upgrade(unselected_name):
            return

        chosen_vals = [
            rank_val.get(new_grades[ch_name], 0)
            for ch_name in chosen_names
            if ch_name != unselected_name and ch_name in new_grades
        ]
        if not chosen_vals:
            return

        chosen_val = max(chosen_vals)
        if rank_val.get(new_grades[unselected_name], 0) <= chosen_val:
            return

        new_grades[unselected_name] = val_to_rank.get(chosen_val, "E")

        # 騎手Pが上がっていない乗り替わりは、選ばれなかった扱いとして表示側にも残す。
        h = horse_state.get(unselected_name, {}).get("horse", {})
        dp = h.get("display_power", "")
        m_dp = re.search(r'P:(\d+)', dp)
        if m_dp:
            new_p = max(0, int(m_dp.group(1)) - 2)
            dp = dp[:m_dp.start()] + f"P:{new_p}" + dp[m_dp.end():]
            h["display_power"] = dp
        h["misute_jockey"] = chosen_jockey

    # 前走で乗っていた騎手が今回ほかの出走馬を選んだ場合、選ばれなかった馬を選んだ馬より上にしない。
    for name, st_info in horse_state.items():
        prev_jockey = st_info.get("prev_jockey", "")
        current_jockey = st_info.get("current_jockey", "")
        if not prev_jockey or current_jockey == prev_jockey:
            continue
        chosen_names = [n for n in current_by_jockey.get(prev_jockey, []) if n != name]
        _cap_unselected_to_chosen(name, chosen_names, prev_jockey)

    # 同じ騎手が直近で複数の出走馬に乗っていて今回その中から選んだ馬がある場合も同様に扱う。
    for prev_jockey, prev_names in prev_by_jockey.items():
        if len(prev_names) < 2:
            continue
        chosen_names = [n for n in current_by_jockey.get(prev_jockey, []) if n in prev_names]
        if not chosen_names:
            continue
        for name in prev_names:
            if name not in chosen_names:
                _cap_unselected_to_chosen(name, chosen_names, prev_jockey)

    return new_grades


def _circled_num(n):
    try:
        n_int = int(n)
        if 1 <= n_int <= 20: return chr(0x245f + n_int)
    except:
        pass
    return str(n)

def predict_pace_python(horses_data, danwa_data, current_distance_str, horse_patterns=None, chokyo_front_umabans=None):
    predictions = []
    chokyo_front_umabans = {str(u) for u in (chokyo_front_umabans or [])}
    
    dm = re.search(r'(\d{1,3}(?:,\d{3})*|\d+)(?=m)', str(current_distance_str).replace(',', ''))
    curr_dist = int(dm.group(1)) if dm else 1400
    # 今回の頭数(枠の内/外サイド判定に使う。est_pos とテン速度バケツで共用)。
    field_size_today = max([int(u) for u in horses_data.keys() if str(u).isdigit()] or [len(horses_data)])

    # 新馬戦・デビュー戦(全馬に過去走なし)は展開予想を省略する。
    has_any_pas = any(
        isinstance(h, dict) and h.get("pas")
        for d in horses_data.values()
        for h in (d.get("hist") or [])
    )
    if not has_any_pas:
        return "【展開予想】新馬・デビュー戦のため展開予想は省略します。", [], {}

    nankan_places = ["浦和", "船橋", "大井", "川崎"]
    jra_keywords = ["JRA", "東京", "中山", "阪神", "京都", "中京", "小倉", "新潟", "福島", "函館", "札幌"]
    current_place = next((place for place in nankan_places if place in str(current_distance_str)), "")

    for umaban, data in horses_data.items():
        name = data.get("name", "")
        hist = data.get("hist", [])
        danwa = danwa_data.get(str(umaban), danwa_data.get(umaban, ""))
        danwa_mark = _danwa_leading_mark(danwa)
        is_chokyo_front = str(umaban) in chokyo_front_umabans
        
        weighted_positions = []
        ideal_position = None
        has_lead_experience = False
        has_lead_good_record = False
        has_other_good_record = False
        has_lead_top2_record = False
        has_outsider_lead_top3_record = False
        lead_intent_comment = bool(re.search(r'(ハナ|逃げ|先手|先行|前々|前目|前で|積極|自分の形)', danwa))
        lead_reluctant_comment = bool(re.search(r'(ハナ.*こだわらない|控える|溜める|番手.*(いい|競馬)|中団)', danwa))
        # 揉まれたくない/包まれ嫌い → 内枠なら隊列を下げる、ハナ候補から外す。
        dislikes_squeeze = bool(re.search(
            r'(揉ま|包ま|内.{0,3}(苦|弱|ダメ|向かな|不向|嫌)|外.{0,3}(出|回|の方|欲)|砂.{0,3}(被.{0,3}(嫌|苦|ダメ))|もまれ|外々)',
            danwa
        ))
        # 強い逃げ意志 → 隊列を一段押し上げ、ハナ争いを優先。
        strong_lead_intent = bool(re.search(
            r'(逃げ.{0,3}たい|ハナ.{0,3}(切|主張|立|行きた|欲し)|単騎|自分の形|どうしても.{0,3}前)',
            danwa
        ))
        last_dist = None
        is_distance_extension = False
        kickback_intolerant = False  # 砂被り耐性(下のdigitブロックで確定。先に初期化して全馬で参照可に)
        # 過去走の枠サイド別の挙動: 外枠で無理に逃げたか/番手づけしたか、内枠で先手を取れたか、
        # 内枠で被る位置(番手以降)からでも好走できたか(砂被り耐性の実績推定)。
        outer_forced = outer_stalk = inner_forced = inner_sit_ok = inner_sit_bad = inner_sit_placed = 0
        has_makuri = False   # まくり(3-4角で先頭進出)好走歴=中盤でペースを引き上げるタイプ

        for idx, h in enumerate(hist[:5]):
            if not isinstance(h, dict): continue
            pas_str = h.get("pas", "")
            if not pas_str: continue
            pas_parts = pas_str.split("-")
            first_pos_str = pas_parts[0]
            if not first_pos_str.isdigit(): continue
            first_pos = int(first_pos_str)
            rank = h.get("rank")

            try:
                rank_int = int(rank)
            except (TypeError, ValueError):
                rank_int = None
            try:
                pop_int = int(h.get("pop"))
            except (TypeError, ValueError):
                pop_int = None

            if first_pos == 1:
                has_lead_experience = True
                if rank is not None and rank <= 5: has_lead_good_record = True
                if rank_int is not None and rank_int <= 2:
                    has_lead_top2_record = True
                if (
                    rank_int is not None and rank_int <= 3
                    and pop_int is not None and pop_int >= 4
                ):
                    has_outsider_lead_top3_record = True
            elif rank is not None and rank <= 5:
                has_other_good_record = True

            if _style_from_pas(pas_str) == "まくり" and rank is not None and rank <= 4:
                has_makuri = True

            side_h = _draw_side(h.get("field_size"), h.get("gate_no"))
            pop_h = h.get("pop")
            good_h = (rank is not None and rank <= 5) or (
                pop_h is not None and rank is not None and (pop_h - rank) >= 3)
            if side_h == "外":
                if first_pos == 1: outer_forced += 1
                elif first_pos >= 3: outer_stalk += 1
            elif side_h == "内":
                if first_pos == 1:
                    inner_forced += 1
                elif first_pos >= 3:   # 内枠で被る位置から運んだ→砂被り耐性の実績
                    if good_h: inner_sit_ok += 1
                    else: inner_sit_bad += 1
                    # 被りながら複勝圏=明確に「内で溜められる」実績(ボーナス用に厳格判定)
                    if rank is not None and rank <= 3:
                        inner_sit_placed += 1
            
            weight = 3 if idx == 0 else (2 if idx == 1 else 1)
            
            place = h.get("place", "")
            # JRA/重賞からの通過順は格上相手のもの。自己条件に戻ると1〜2列前を取れるので前へ補正。
            jra_shift = -1.5 if _PACE_REFINE else -1
            place_adj = jra_shift if any(kw in place for kw in jra_keywords) else (2 if place and place not in nankan_places else 0)
            
            adjusted_pos = max(1, first_pos + place_adj)
            
            past_dist_str = str(h.get("dist", ""))
            if past_dist_str.isdigit():
                pd = int(past_dist_str)
                if idx == 0:
                    last_dist = pd
                dist_adj = (curr_dist - pd) / 200.0
                if dist_adj > 0:
                    adjusted_pos = max(1, adjusted_pos - min(1.5, dist_adj * 0.35))
                elif dist_adj < 0:
                    adjusted_pos += min(1.0, abs(dist_adj) * 0.25)
            
            for _ in range(weight): weighted_positions.append(adjusted_pos)
            
            pop = h.get("pop")
            if rank is not None:
                is_great_run = False
                if pop is not None and (pop - rank) >= 4: is_great_run = True
                if rank <= 2: is_great_run = True
                if is_great_run and ideal_position is None: ideal_position = adjusted_pos
        
        if weighted_positions:
            avg_pos = sum(weighted_positions) / len(weighted_positions)
            est_pos = (ideal_position * 0.4 + avg_pos * 0.6) if ideal_position is not None and ideal_position < avg_pos else avg_pos
        else:
            est_pos = 6.0

        if last_dist and curr_dist > last_dist:
            is_distance_extension = True
        
        comment_mod = 0.0
        if danwa:
            if re.search(r'(前走|前回).*(逃げ|前).*(苦し|厳し|バテ|甘く)', danwa): comment_mod += 1.0
            if re.search(r'(ハナ.*こだわらない|控える|溜める|番手.(いい|競馬)|中団)', danwa): comment_mod += 2.0
            if re.search(r'(ハナ.(切|行|主張|立)|前.(行け|つけ|行きた))', danwa): comment_mod -= 2.0

        est_pos += comment_mod
        
        curr_w = data.get("weight", 0.0)
        prev_w = hist[0].get("weight", 0.0) if hist and isinstance(hist[0], dict) else 0.0
            
        if curr_w > 0 and prev_w > 0:
            est_pos -= (prev_w - curr_w) * 0.5

        if has_lead_good_record: est_pos -= 1.0
        elif has_lead_experience: est_pos -= 0.5
        if is_distance_extension and (lead_intent_comment or has_lead_experience or est_pos <= 4.0):
            est_pos -= 0.4

        # 強い逃げ意志 → ハナ取りに行くので隊列を更に押し上げ。
        if strong_lead_intent:
            est_pos -= 1.0

        est_pos = max(1.0, est_pos)

        today_side = None
        if str(umaban).isdigit():
            umaban_val = int(umaban)
            today_side = _draw_side(field_size_today, umaban_val)
            waku_mod = max(0, (6 - umaban_val) * 0.15)
            est_pos -= waku_mod
            if umaban_val in [1, 2]: est_pos -= 0.5

            # 砂被り耐性。砂を嫌う馬(談話 or 内で被ると凡走実績)は内枠でも「後方」ではなく
            # 番手外付け(前め・外)に出るので位置は下げない。外枠なら被らず気分よく運べる。
            # 内で被っても溜められる実績馬は内枠を好位確保(溜め)の好材料とみる。
            kickback_intolerant = dislikes_squeeze or (inner_sit_bad >= 2 and inner_sit_ok == 0)
            kickback_tolerant = inner_sit_placed >= 1   # 内で被りつつ複勝圏=溜め可の確実な実績
            if _KICKBACK_LOGIC:
                if kickback_intolerant:
                    if today_side == "外":
                        est_pos -= 0.3                              # 外枠=砂回避で気分よく前へ
                    elif has_lead_experience or strong_lead_intent:
                        est_pos -= 0.2                              # 内でも砂を嫌い前(逃げ/番手外)へ
                    else:
                        est_pos += 0.4                              # 外々で脚を使い小幅後退(旧+1.5は過大)
                elif kickback_tolerant and inner_forced == 0 and today_side in ("内", "中"):
                    est_pos -= 0.2                                  # 内で被っても溜められる→好位を確保(先手型は対象外)
            else:
                # 旧挙動: 内枠ほど一律後退(逆挙動)。
                if dislikes_squeeze:
                    if umaban_val <= 3: est_pos += 1.5
                    elif umaban_val <= 5: est_pos += 0.8
                    else: est_pos += 0.3

            # 短縮/延長×逃げ経験、外枠での過去挙動、内枠先手 を反映。
            if _PACE_DRAW_INTENT:
                is_shorten = bool(last_dist and curr_dist < last_dist)
                # ① 短縮は前に行きにくい(強い逃げ意志が無ければ後退)。延長は既存の-0.4で前進済み。
                if is_shorten and not strong_lead_intent:
                    est_pos += 0.6
                # ② 今回外枠で、過去の外枠走が番手づけ優勢(無理に逃げない)なら先手を主張しないとみて後退。
                #    外からでも先手を主張するタイプ(outer_forced優勢)は据え置き。
                if today_side == "外" and (outer_forced + outer_stalk) >= 2 and outer_stalk >= outer_forced:
                    est_pos += 0.8
                # ③ 今回内枠で、内から先手を取れる馬は先手を確保しやすい(溜めも利く)。
                if today_side == "内" and inner_forced >= 1:
                    est_pos -= 0.5

            est_pos = max(1.0, est_pos)

        # 厩舎談話の先頭が△なら一段控えめ、当日横断の調教3位以内は一段前へ補正する。
        if danwa_mark == "△":
            est_pos += 1.0
        if is_chokyo_front:
            est_pos -= 1.0
        est_pos = max(1.0, est_pos)

        first_corner_profile = _first_corner_draw_profile(
            hist, today_side, current_place=current_place, current_dist=curr_dist
        ) if _PACE_FIRST_CORNER_DRAW else {
            "avg_first_pos": None, "lead_rate": 0.0, "front2_rate": 0.0,
            "side_count": 0, "exact_side_leads": 0, "exact_any_leads": 0,
        }
        
        predictions.append({
            "umaban": umaban,
            "name": name,
            "est_pos": round(est_pos, 2),
            "ideal_pos": ideal_position,
            "hist": hist,
            "has_lead_exp": has_lead_experience,
            "has_lead_good_record": has_lead_good_record,
            "has_other_good_record": has_other_good_record,
            "has_lead_top2_record": has_lead_top2_record,
            "has_outsider_lead_top3_record": has_outsider_lead_top3_record,
            "lead_intent_comment": lead_intent_comment,
            "lead_reluctant_comment": lead_reluctant_comment,
            "wants_lead": ((lead_intent_comment or strong_lead_intent) and not lead_reluctant_comment) or has_lead_experience,
            "is_distance_extension": is_distance_extension,
            "last_dist": last_dist,
            "dislikes_squeeze": dislikes_squeeze,
            "strong_lead_intent": strong_lead_intent,
            "has_makuri": has_makuri,
            "kickback_intolerant": kickback_intolerant,
            "today_side": today_side,
            "first_corner_profile": first_corner_profile,
            "hana_acquisition_score": None,
            "hana_profile_note": "",
            "danwa_mark": danwa_mark,
            "chokyo_front_adjust": is_chokyo_front,
        })

    predictions.sort(key=lambda x: x["est_pos"])
    top_5 = list(predictions[:5])

    # ★/☆（has_lead_exp）が付く馬の「一つ外」も先手候補テン速度を取得する。
    starred_outer_umabans = set()
    for p in predictions:
        if not p.get("has_lead_exp"):
            continue
        try:
            starred_outer_umabans.add(int(p["umaban"]) + 1)
        except (TypeError, ValueError):
            continue
    if starred_outer_umabans:
        top_5_keys = {p["umaban"] for p in top_5}
        for p in predictions:
            try:
                u = int(p["umaban"])
            except (TypeError, ValueError):
                continue
            if u in starred_outer_umabans and p["umaban"] not in top_5_keys:
                top_5.append(p)
                top_5_keys.add(p["umaban"])

    speed_data_list = []
    url_cache = {}
    _DRAW_BUCKET_MIN = 2   # 内/外バケツを採用するのに必要な該当走数(不足時は通算へフォールバック)

    def _base_from(vals):
        # 前目の走りのテン速度群から代表値: 3走以上は上位3の中央値、それ未満は平均。
        if not vals:
            return None
        if len(vals) >= 3:
            return statistics.median(sorted(vals, reverse=True)[:3])
        return sum(vals) / len(vals)

    for p in top_5:
        p["speed_avg"] = None
        speeds = []
        draw_bucket_note = ""
        umaban_int = int(p["umaban"]) if str(p["umaban"]).isdigit() else 8
        is_inner_draw = umaban_int <= 3
        is_outer_draw = umaban_int >= 7
        today_side = _draw_side(field_size_today, umaban_int)  # 今回が内/外
        
        danwa = danwa_data.get(str(p["umaban"]), danwa_data.get(p["umaban"], ""))
        must_lead = bool(re.search(r'(ハナ.*絶対|逃げ.*条件|ハナを切|自分の形|単騎)', danwa)) or p.get("strong_lead_intent", False)
        can_sit_second = bool(re.search(r'(番手.*(いい|競馬|我慢)|控える|溜める|砂を被.*慣れ)', danwa))
        distance_extension = bool(p.get("is_distance_extension"))
        dislikes_squeeze = bool(p.get("dislikes_squeeze"))
        strong_lead_intent = bool(p.get("strong_lead_intent"))
        
        for h in p["hist"][:5]:
            if not isinstance(h, dict) or not h.get("url"): continue
            if h.get("place") not in nankan_places: continue

            pas_str = h.get("pas", "")
            fp_str = pas_str.split('-')[0] if pas_str else ""
            if not fp_str.isdigit():
                continue
            pos = int(fp_str)
            if pos > 5:
                continue

            # この前目の走りを「内枠時/外枠時」のどちらで出したか(今回枠と同基準で分類)。
            past_side = _draw_side(h.get("field_size"), h.get("gate_no"))

            past_url = h["url"]
            past_dist_str = str(h.get("dist", ""))
            pd = int(past_dist_str) if past_dist_str.isdigit() else 0
            
            if past_url in url_cache:
                res_data = url_cache[past_url]
            else:
                res_data = fetch_past_race_1f_times(past_url, past_dist=pd)
                url_cache[past_url] = res_data
                
            if res_data:
                raw_time, dist_0 = res_data
                
                adj_time = raw_time
                adj_time += ((pos - 1) * 0.1) if dist_0 == 200 else (((pos - 1) // 2) * 0.1)
                
                adj_speed = (dist_0 / adj_time) * 3.6
                
                past_place = h.get("place", "")
                
                _STD = {
                    ("浦和",800):48.1, ("浦和",1400):90.7, ("浦和",1500):97.8, ("浦和",2000):134.5,
                    ("船橋",1000):62.5, ("船橋",1200):76.1, ("船橋",1500):98.6,
                    ("船橋",1600):105.1, ("船橋",1800):120.2,
                    ("大井",1000):60.8, ("大井",1200):74.5, ("大井",1400):88.8,
                    ("大井",1600):103.8, ("大井",1800):118.0, ("大井",2000):130.2,
                    ("川崎",900):55.6, ("川崎",1400):92.5, ("川崎",1500):98.3,
                    ("川崎",1600):106.6, ("川崎",2000):136.1,
                }
                _REF_PACE = 88.8 / 7
                
                std_time = _STD.get((past_place, pd))
                if std_time and pd > 0:
                    pace_per_200 = std_time / (pd / 200)
                    adj_speed += (720 / _REF_PACE) - (720 / pace_per_200)
                else:
                    if past_place in ["川崎", "浦和"]: adj_speed += 0.5
                    if pd > 0 and pd <= 1200: adj_speed -= 1.0
                    elif pd >= 1500: adj_speed += 1.0
                
                speeds.append((adj_speed, past_place, past_dist_str, past_side))
                
        if speeds:
            speed_vals = [s[0] for s in speeds]
            pooled_base = _base_from(speed_vals)

            # 内/外バケツ別のテン速度(前目に付けた走りのみ)。今回枠に対応する側を採用。
            matched_vals = [s[0] for s in speeds if today_side is not None and s[3] == today_side]
            draw_bucket_note = ""
            use_empirical_draw = False
            if _TEN_SPEED_DRAW_BUCKET and today_side is not None and len(matched_vals) >= _DRAW_BUCKET_MIN:
                base_speed = _base_from(matched_vals)
                use_empirical_draw = True
                draw_bucket_note = f"今回{today_side}枠:{today_side}枠時テン({len(matched_vals)}走)"
            else:
                base_speed = pooled_base
                if today_side is not None:
                    draw_bucket_note = f"今回{today_side}枠:{today_side}枠実績少→通算"

            max_speed_tuple = max(speeds, key=lambda s: s[0])
            max_speed = max_speed_tuple[0]
            max_speed_info = f"{max_speed_tuple[1]}{max_speed_tuple[2]}"
            final_speed = base_speed

            # 内外の実測バケツを採れた時は、後付けの一律枠補正は重複計上になるので外す。
            # バケツが薄く通算へフォールバックした時だけ、従来の枠ヒューリスティックで補う。
            if not use_empirical_draw:
                if is_inner_draw and p["est_pos"] <= 2.5:
                    final_speed += 1.5 if must_lead else 0.8
                elif is_outer_draw and p["est_pos"] <= 2.5:
                    if must_lead: final_speed += 0.5
                    elif p.get("has_lead_exp"): final_speed += 0.3

                    if can_sit_second and not must_lead:
                        final_speed -= (1.2 if not p.get("has_lead_exp") else 0.6)
            if distance_extension and (must_lead or p.get("wants_lead") or p["est_pos"] <= 3.5):
                final_speed += 0.6

            # ★/☆（逃げ好走★ or 逃げ経験☆）馬はテン速度指数を+2底上げし、先手・前付けを後押しする。
            if p.get("has_lead_good_record") or p.get("has_lead_exp"):
                final_speed += 2.0

            p["speed_avg"] = final_speed
            p["raw_speed_avg"] = base_speed
            p["raw_speed_max"] = max_speed
            p["raw_speed_max_info"] = max_speed_info
            speed_data_list.append({
                "umaban": p["umaban"],
                "name": p["name"],
                "speed": final_speed,
                "raw_speed": base_speed,
                "raw_speed_max": max_speed,
                "raw_speed_max_info": max_speed_info,
                "est_pos": p["est_pos"],
                "has_lead_exp": p.get("has_lead_exp", False),
                "has_lead_good_record": p.get("has_lead_good_record", False),
                "has_other_good_record": p.get("has_other_good_record", False),
                "must_lead": must_lead,
                "can_sit_second": can_sit_second,
                "wants_lead": p.get("wants_lead", False),
                "is_distance_extension": distance_extension,
                "dislikes_squeeze": dislikes_squeeze,
                "strong_lead_intent": strong_lead_intent,
                "hana_tiebreak_bonus": 0.0,
                "pace_note": "",
                "draw_bucket_note": draw_bucket_note
            })
        else:
            speed_data_list.append({
                "umaban": p["umaban"],
                "name": p["name"],
                "speed": -1.0,
                "raw_speed": -1.0,
                "raw_speed_max": -1.0,
                "raw_speed_max_info": "",
                "est_pos": p["est_pos"],
                "has_lead_exp": p.get("has_lead_exp", False),
                "has_lead_good_record": p.get("has_lead_good_record", False),
                "has_other_good_record": p.get("has_other_good_record", False),
                "must_lead": must_lead,
                "can_sit_second": can_sit_second,
                "wants_lead": p.get("wants_lead", False),
                "is_distance_extension": distance_extension,
                "dislikes_squeeze": dislikes_squeeze,
                "strong_lead_intent": strong_lead_intent,
                "hana_tiebreak_bonus": 0.0,
                "pace_note": "",
                "draw_bucket_note": draw_bucket_note
            })
                
    def _get_star_mark(p_dict):
        has_lead = p_dict.get("has_lead_exp", False)
        has_good = p_dict.get("has_lead_good_record", False)
        has_other = p_dict.get("has_other_good_record", False)
    
        if has_good: return "(★)" if has_other else "★"
        if has_lead: return "(☆)" if has_other else "☆"
        return ""

    pred_by_umaban = {str(p["umaban"]): p for p in predictions}
    valid_speed_items = [sd for sd in speed_data_list if sd.get("speed", -1) >= 0]
    if valid_speed_items:
        max_final_speed = max(sd["speed"] for sd in valid_speed_items)
        close_lead_items = [
            sd for sd in valid_speed_items
            if max_final_speed - sd["speed"] <= 2.0
            and (sd.get("must_lead") or sd.get("wants_lead") or sd.get("est_pos", 99) <= 2.5)
            # 揉まれ嫌い × 内枠 はハナ争いを諦める想定で除外
            and not (sd.get("dislikes_squeeze") and str(sd.get("umaban", "")).isdigit() and int(sd["umaban"]) <= 3)
        ]
        if len(close_lead_items) >= 2:
            field_size = max(
                [int(u) for u in horses_data.keys() if str(u).isdigit()] or [len(horses_data)]
            )
            for sd in close_lead_items:
                try:
                    gate_no = int(sd["umaban"])
                except (TypeError, ValueError):
                    gate_no = field_size
                inner_bonus = max(0.0, (field_size + 1 - gate_no) / max(field_size, 1)) * 1.2
                # 揉まれ嫌いは内枠ボーナスを無効化
                if sd.get("dislikes_squeeze"):
                    inner_bonus = 0.0
                extension_bonus = 0.9 if sd.get("is_distance_extension") else 0.0
                must_bonus = 0.4 if sd.get("must_lead") else 0.0
                # 強い逃げ意志は追加加点
                intent_bonus = 0.6 if sd.get("strong_lead_intent") else 0.0
                tiebreak = inner_bonus + extension_bonus + must_bonus + intent_bonus
                notes = []
                if gate_no <= max(3, field_size // 3) and not sd.get("dislikes_squeeze"):
                    notes.append("内枠")
                if sd.get("is_distance_extension"):
                    notes.append("距離延長")
                if sd.get("must_lead"):
                    notes.append("ハナ主張")
                if sd.get("strong_lead_intent"):
                    notes.append("逃げ意志")
                if sd.get("dislikes_squeeze"):
                    notes.append("揉まれ嫌い")
                sd["hana_tiebreak_bonus"] = round(tiebreak, 2)
                if notes:
                    sd["pace_note"] = " / 先手争い:" + "+".join(notes)
                pred = pred_by_umaban.get(str(sd["umaban"]))
                if pred is not None:
                    pred["hana_tiebreak_bonus"] = tiebreak

    # テンの絶対速度だけでなく、今回と同じ内/外側での1コーナー位置を統合。
    # 逃げ未経験馬は、単に内枠だからという理由でハナ候補に格上げしない。
    hana_winner_umaban = None
    if _PACE_FIRST_CORNER_DRAW:
        speed_by_umaban = {str(sd["umaban"]): sd for sd in valid_speed_items}
        hana_candidates = []
        for p in predictions:
            sd = speed_by_umaban.get(str(p["umaban"]))
            if not sd:
                continue
            profile = p.get("first_corner_profile") or {}
            lead_capable = bool(
                p.get("has_lead_exp")
                or p.get("strong_lead_intent")
                or sd.get("must_lead")
            )
            if not lead_capable:
                continue
            if p.get("est_pos", 99) > 3.5 and not profile.get("exact_side_leads"):
                continue
            hana_candidates.append((p, sd, profile))

        max_candidate_speed = max(
            (sd.get("speed", -1) for _, sd, _ in hana_candidates), default=-1
        )
        for p, sd, profile in hana_candidates:
            speed = float(sd.get("speed", 0) or 0)
            avg_first = profile.get("avg_first_pos")
            lead_rate = float(profile.get("lead_rate", 0) or 0)
            front2_rate = float(profile.get("front2_rate", 0) or 0)
            side = p.get("today_side")

            pos_bonus = max(0.0, min(3.0, 4.0 - float(avg_first))) if avg_first is not None else 0.0
            corner_bonus = pos_bonus + lead_rate * 3.0 + front2_rate * 1.5

            # 内は「逃げ経験があり、最速馬から0.6秒以内」の馬だけコースを守る加点。
            # 外は一律加点せず、同場同距離で外から1C先頭を取った実績だけ強く評価。
            tactical_bonus = 0.0
            speed_time = 720.0 / speed if speed > 0 else 99.0
            best_time = 720.0 / max_candidate_speed if max_candidate_speed > 0 else speed_time
            time_gap = speed_time - best_time
            if (
                side == "内"
                and p.get("has_lead_exp")
                and time_gap <= 0.6
                and front2_rate >= 0.35
            ):
                tactical_bonus += 4.0
            exact_side_leads = int(profile.get("exact_side_leads", 0) or 0)
            if exact_side_leads:
                tactical_bonus += 5.0 if side == "外" else 2.0
            elif profile.get("exact_any_leads"):
                tactical_bonus += 0.8

            intent_bonus = 0.0
            if p.get("has_lead_good_record"):
                intent_bonus += 1.5
            elif p.get("has_lead_exp"):
                intent_bonus += 0.8
            if p.get("strong_lead_intent"):
                intent_bonus += 1.5
            if sd.get("must_lead"):
                intent_bonus += 0.8
            if sd.get("can_sit_second") and not sd.get("must_lead"):
                intent_bonus -= 0.8

            hana_score = speed + corner_bonus + tactical_bonus + intent_bonus
            p["hana_acquisition_score"] = round(hana_score, 2)
            sd["hana_acquisition_score"] = round(hana_score, 2)
            side_label = side or "枠不明"
            if avg_first is not None:
                note = (
                    f"1C{side_label}:平均{avg_first:.1f}番手/"
                    f"先頭{lead_rate * 100:.0f}%({int(profile.get('side_count', 0) or 0)}走)"
                )
                if exact_side_leads:
                    note += f"/同場同距離ハナ{exact_side_leads}走"
                p["hana_profile_note"] = note
                sd["hana_profile_note"] = note

        scored_hana_candidates = [
            p for p, _, _ in hana_candidates
            if p.get("hana_acquisition_score") is not None
        ]
        if scored_hana_candidates:
            hana_winner = max(
                scored_hana_candidates,
                key=lambda p: (p["hana_acquisition_score"], -float(p.get("est_pos", 99))),
            )
            hana_winner_umaban = str(hana_winner["umaban"])

    speed_data_list.sort(key=lambda x: x.get("speed", -1), reverse=True)
    
    speeds_log = []
    for i, sd in enumerate(speed_data_list):
        star = _get_star_mark(sd)
        if sd["speed"] >= 0:
            max_s = sd['raw_speed_max']
            max_info = sd.get('raw_speed_max_info', '')
            max_time = 720 / max_s if max_s > 0 else 0
            bucket = sd.get('draw_bucket_note', '')
            bucket_txt = f" [{bucket}]" if bucket else ""
            hana_note = sd.get("hana_profile_note", "")
            hana_score = sd.get("hana_acquisition_score")
            hana_txt = f" / {hana_note} / ハナ指数{hana_score:.1f}" if hana_note and hana_score is not None else ""
            text = f"{_circled_num(sd['umaban'])}{sd['name']}{star}: 最速{max_time:.1f}秒({max_s:.1f}km/h)※{max_info}{bucket_txt}{sd.get('pace_note', '')}{hana_txt}"
        else:
            text = f"{_circled_num(sd['umaban'])}{sd['name']}{star}: テン速度なし(位置)"
        speeds_log.append(text)

    def _pos_band(est_pos):
        if est_pos <= 2.0: return 1
        if est_pos <= 3.5: return 2
        if est_pos <= 6.0: return 3
        return 4

    def final_sort_key(x):
        band = _pos_band(x["est_pos"])
        speed = x.get("speed_avg") or 0
        lead_bonus = 5.0 if x.get("has_lead_good_record") else (2.0 if x.get("has_lead_exp") else 0)
        tie_bonus = x.get("hana_tiebreak_bonus", 0.0) or 0.0
        return (band, -(speed + lead_bonus + tie_bonus), x["est_pos"] - min(0.5, tie_bonus * 0.2))

    predictions.sort(key=final_sort_key)
    if hana_winner_umaban is not None:
        predictions.sort(key=lambda p: 0 if str(p["umaban"]) == hana_winner_umaban else 1)

    escape_count = sum(1 for p in predictions if p["est_pos"] <= 2.0)
    senkou_count = sum(1 for p in predictions if p["est_pos"] <= 3.5)

    if _PACE_REFINE:
        # 「本当に逃げる馬」=見かけ上前(est_pos<=2)かつ逃げ経験あり。逃げ経験の無い馬は
        # 折り合い等のリスクで実際にはハナを取りに行かない(3番手の馬が逃げる事は稀)。
        committed_leaders = sum(1 for p in predictions if p["est_pos"] <= 2.0 and p.get("has_lead_exp"))
        # まくり質=中盤(3-4角)で先頭進出するタイプ。複数いるとペースが上がる(加圧)。
        pace_raisers = sum(1 for p in predictions if p.get("has_makuri"))
        pressure = committed_leaders + 0.5 * min(pace_raisers, 2)

        # 騎手×役割: ペースは番手の馬・騎手が作る。前(est_pos<=3.5)の馬の騎手Pで増減。
        if _PACE_JOCKEY:
            def _jp(p):
                m = re.search(r'P:(\d+)', str(horses_data.get(p["umaban"], {}).get("power", "")))
                return int(m.group(1)) if m else None
            front = [p for p in predictions if p["est_pos"] <= 3.5]
            low_front = sum(1 for p in front if (_jp(p) is not None and _jp(p) <= _PACE_JP_LOW))
            # 上手い騎手の番手(絶対に逃げる馬以外)は流れを壊さない=沈静。
            high_calm = sum(1 for p in front
                            if (_jp(p) is not None and _jp(p) >= _PACE_JP_HIGH)
                            and not (p["est_pos"] <= 2.0 and p.get("has_lead_exp")))
            if low_front >= 2:           # 下手同士が前で競る=煽り
                pressure += 0.5
            pressure -= 0.4 * min(high_calm, 2)
            pressure = max(0.0, pressure)

        # pressure を主軸にS/M/M~H/Hを決定(騎手の沈静/煽りが境界を動かせるように)。
        # 1000m以下は基本的に前が強い。逃げ馬が大量(committed>=4)の超ハイペース時だけHにする。
        # (まくり質や下手騎手の煽りで pressure が上がってもスプリントは前が止まりにくいので除外)
        is_sprint = curr_dist <= 1000
        sprint_super_high = committed_leaders >= 4
        if (committed_leaders >= 2 or pressure >= 2.5) and (not is_sprint or sprint_super_high):
            pace = "ハイペース"
            explanation = "逃げ経験のある先行馬が複数(またはまくり質・下手な先行騎手が加わり)、序盤〜中盤で流れが速くなりやすい。差し馬の台頭に注意。"
        elif committed_leaders == 0 and senkou_count <= 1 and pace_raisers == 0 and pressure < 1.0:
            pace = "スローペース"
            explanation = "明確に逃げ経験のあるハナ主張馬がおらず(上手い騎手の番手で流れも落ち着き)、隊列が落ち着きやすい。前残りに注意。"
        elif pressure >= 1.5 or (committed_leaders >= 1 and (committed_leaders + min(pace_raisers, 2)) >= 2):
            pace = "やや速いペース"
            explanation = "逃げ馬に加え番手・まくり勢が競りかけ、先行争いが予想される。中団からの好位差しが有効。"
        else:
            pace = "ミドルペース"
            explanation = "逃げ経験馬がすんなり隊列を先導し、淀みのない平均的なペースで流れると予想される。"
        # 1000m以下は超ハイペース以外は前残り想定。やや速い(M~H)もミドル(前残り)へ寄せる。
        if is_sprint and not sprint_super_high and pace == "やや速いペース":
            pace = "ミドルペース"
            explanation = "1000m以下で逃げ大量とまではいかず前が止まりにくい。先行有利・前残りに注意。"
    elif escape_count >= 3:
        pace = "ハイペース"
        explanation = "逃げ・先行意欲の高い馬が複数おり、序盤から激しいポジション争いが予想されるため、差し馬の台頭に注意。"
    elif escape_count == 0 and senkou_count <= 1:
        pace = "スローペース"
        explanation = "明確にハナを主張する馬がおらず、押し出されるように隊列が落ち着く可能性が高い。前残りの展開に注意。"
    elif escape_count >= 2:
        pace = "やや速いペース"
        explanation = "逃げたい馬が複数おり、先行争いが予想される。中団からの好位差しが有効になる可能性。"
    else:
        pace = "ミドルペース"
        explanation = "ハナ候補がすんなり隊列を先導し、淀みのない平均的なペースで流れると予想される。"

    leaders = " ".join([f"{_circled_num(h['umaban'])}{h['name']}" for h in predictions[:2]])
    
    def _danwa_for_umaban(umaban):
        return danwa_data.get(str(umaban), danwa_data.get(umaban, ""))

    def _candidate_text(p, with_stable_note=False):
        ideal_mark = _get_star_mark(p)
        text = f"{_circled_num(p['umaban'])}{p['name']}{ideal_mark}"
        extras = []
        if with_stable_note:
            comment = _clean_danwa_text(_danwa_for_umaban(p["umaban"]))
            if comment != "コメントなし":
                extras.append(comment)
        if horse_patterns:
            uid = horses_data.get(str(p["umaban"]), {}).get("uma_id", "")
            ptxt = _format_pattern_short(horse_patterns.get(uid))
            if ptxt:
                extras.append(ptxt)
        if extras:
            text += "\n[[SMALL]]　" + " ／ ".join(extras) + "[[/SMALL]]"
        return text

    pos_groups = {"逃げ": [], "先行": [], "中団": [], "後方": []}
    for p in predictions:
        ep = p["est_pos"]
        if ep <= 1.5: pos_groups["逃げ"].append(_candidate_text(p, with_stable_note=True))
        elif ep <= 3.5: pos_groups["先行"].append(_candidate_text(p, with_stable_note=True))
        elif ep <= 6.0: pos_groups["中団"].append(_candidate_text(p))
        else: pos_groups["後方"].append(_candidate_text(p))
    
    pace_map = {"ハイペース": "H", "スローペース": "S", "やや速いペース": "M~H", "ミドルペース": "M"}
    pace_abbrev = pace_map.get(pace, pace)

    formation_parts = []
    prev_pos = None
    current_group = []
    for p in predictions:
        cn = _circled_num(p['umaban'])
        if prev_pos is not None and (p['est_pos'] - prev_pos) > 1.0:
            formation_parts.append(','.join(current_group))
            current_group = [cn]
        else:
            current_group.append(cn)
        prev_pos = p['est_pos']
    if current_group:
        formation_parts.append(','.join(current_group))
    
    tenkai_text = f"【展開予想】\n"
    tenkai_text += f"◆ペース予想：{pace_abbrev}：{explanation}\n\n"
    tenkai_text += f"◆想定隊列順: \n{' '.join(formation_parts)}\n\n"
    tenkai_text += f"◆ハナ・先行候補：{leaders}\n"

    for label in ["逃げ", "先行", "中団"]:
        if pos_groups[label]:
            if label in ("逃げ", "先行"):
                tenkai_text += f"◆{label}候補：\n" + "\n".join(pos_groups[label]) + "\n"
            else:
                tenkai_text += f"◆{label}候補：{' / '.join(pos_groups[label])}\n"

    if speeds_log:
        tenkai_text += f"◆先手候補テン速度\n"
        tenkai_text += "\n".join(speeds_log) + "\n"
    
    # 展開加点①: S/Mペース(前残り想定)の時だけ、前5頭以内(est_pos上位5頭)に+5。
    pace_bonus_map = {}
    ordered_predictions = sorted(predictions, key=lambda x: x["est_pos"])
    if _PACE_POS_BONUS and pace in ("スローペース", "ミドルペース"):
        for p in ordered_predictions[:5]:
            pace_bonus_map[str(p["umaban"])] = _PACE_POS_BONUS_PTS

    # 展開加点②: 競馬場×距離の逃げ先行勝率に応じ、ペースを問わず想定隊列の前30%に+3〜7。
    front_rate, front_bonus = _course_front_bonus(current_distance_str, curr_dist)
    if front_bonus and ordered_predictions:
        front_count = max(1, int(len(ordered_predictions) * 0.30))
        for p in ordered_predictions[:front_count]:
            uma_key = str(p["umaban"])
            pace_bonus_map[uma_key] = pace_bonus_map.get(uma_key, 0) + front_bonus
        tenkai_text += (
            f"◆競馬場距離加点：想定隊列の前{front_count}頭に+{front_bonus}"
            f"（逃げ先行勝率{front_rate:.1f}%）\n"
        )

    # 展開加点③: 今回ハナと判定した馬に、逃げて2着以内の実績、または
    # 4人気以下で逃げて3着以内の実績があれば、ペースを問わず+2。
    # 逃げ未経験馬や「ハナ候補の2番手」には付けない。
    if hana_winner_umaban is not None:
        hana_pred = pred_by_umaban.get(str(hana_winner_umaban))
        if hana_pred and (
            hana_pred.get("has_lead_top2_record")
            or hana_pred.get("has_outsider_lead_top3_record")
        ):
            uma_key = str(hana_winner_umaban)
            pace_bonus_map[uma_key] = pace_bonus_map.get(uma_key, 0) + 2
            reasons = []
            if hana_pred.get("has_lead_top2_record"):
                reasons.append("逃げ2着以内")
            if hana_pred.get("has_outsider_lead_top3_record"):
                reasons.append("4人気以下で逃げ3着以内")
            tenkai_text += (
                f"◆逃げ実績加点：{_circled_num(uma_key)}{hana_pred.get('name', '')}に+2"
                f"（{'・'.join(reasons)}）\n"
            )

    return tenkai_text, speed_data_list, pace_bonus_map


def build_evaluation_list(grades, horses_data, scored_data=None):
    rank_map = {"S": [], "A": [], "B": [], "C": [], "D": [], "E": [], "無": []}
    
    for u, hd in horses_data.items():
        name_norm = _norm_horse_name(hd["name"])
        grade = grades.get(name_norm, "")
        if not grade:
            for k_norm, v in grades.items():
                if k_norm and (k_norm in name_norm or name_norm in k_norm):
                    grade = v
                    break
        if not grade:
            grade = "無"
        else:
            grade = _rank_floor_e(grade)
        if grade in rank_map: rank_map[grade].append(u)
        else: rank_map[grade] = [u]
            
    eval_text = "【総合評価】"
    parts = []
    for r in ["S", "A", "B", "C", "D", "E", "無"]:
        if rank_map.get(r):
            nums = "".join([f"[{x}]" for x in rank_map[r]])
            if r == "無": parts.append(f"評価なし{nums}")
            else: parts.append(f"{r}{nums}")
    
    eval_text += "  " + "  ".join(parts)

    if scored_data:
        chumoku_items = []
        tataki_items = []
        for u, d in scored_data.items():
            flags = d.get("flags", [])
            if any("絶好調教" in f for f in flags) or any("基準超" in f for f in flags):
                t_info = d.get("chokyo_time_info", "")
                # 基準超(絶好調教)は赤字にしない。赤字はTOP3のみ。実タイムは小さく表示。
                t_str = f"[[SMALL]]({t_info})[[/SMALL]]" if t_info else ""
                chumoku_items.append(f"[{u}]{t_str}")
            if d.get("tataki_second"):
                detail = d.get("tataki_second_detail", "")
                d_str = f"({detail})" if detail else ""
                tataki_items.append(f"[{u}]{d_str}")
        
        # 総合評価直下の🕰️は、基準タイム超と今走調教TOP3だけに絞る。
        chokyo_items = []

        def _item_sort_key(item):
            rk, u_num, _txt = item
            return (rk, u_num)

        for u, d in scored_data.items():
            rk = d.get("chokyo_display_rank") or d.get("chokyo_red_rank")
            if rk:
                try:
                    rk_int = int(rk)
                except (TypeError, ValueError):
                    rk_int = 99
                if rk_int > 3:
                    continue
                total = d.get("chokyo_display_total") or d.get("chokyo_red_total")
                disp = d.get("chokyo_display_disp") or d.get("chokyo_rank_disp") or d.get("chokyo_time_disp") or ""
                rank_note = f"{rk}位/{total}頭" if total else f"{rk}位"
                d_str = f"({disp} {rank_note})" if disp else f"({rank_note})"
                item = f"[{u}][[SMALL]]{d_str}[[/SMALL]]"
                if int(d.get("score_chokyo_global", 0) or 0) > 0:
                    item = f"[[R]]{item}[[/R]]"
                chokyo_items.append((rk_int, int(u) if str(u).isdigit() else 99, item))

        if chumoku_items:
            eval_text += f"\n🕰️： {' '.join(chumoku_items)}"
        if chokyo_items:
            chokyo_items.sort(key=_item_sort_key)
            eval_text += f"\n🕰️： {' '.join(t for _, _, t in chokyo_items)}"
        if tataki_items:
            eval_text += f"\n叩2： {' '.join(tataki_items)}"

        speed_items = []
        for u, d in scored_data.items():
            idx = d.get("common_speed_index")
            rk = d.get("common_speed_rank")
            if idx is None or rk is None or int(rk) > 3:
                continue
            speed_items.append((int(rk), int(u) if str(u).isdigit() else 99, f"[{u}]({float(idx):.1f})"))
        if speed_items:
            speed_items.sort(key=lambda x: (x[0], x[1]))
            eval_text += f"\n⚡速度： {' '.join(x[2] for x in speed_items)}"

    return eval_text


def build_speed_index_table(horses_data, scored_data):
    """指数・同場同距離最高の検証表。総合点には使わない。"""
    rows = []
    for umaban, horse in (horses_data or {}).items():
        sc = (scored_data or {}).get(umaban) or (scored_data or {}).get(str(umaban)) or {}
        idx = sc.get('common_speed_index')
        rank = sc.get('common_speed_rank')
        best = sc.get('common_speed_best_run') or {}
        rows.append({
            'umaban': str(umaban),
            'name': str((horse or {}).get('name') or ''),
            'index': idx,
            'same_max': sc.get('common_speed_same_max'),
            'same_runs': int(sc.get('common_speed_same_runs', 0) or 0),
            'rank': rank,
            'runs': int(sc.get('common_speed_runs', 0) or 0),
            'best': best,
        })
    rows.sort(key=lambda r: (
        r['rank'] is None,
        int(r['rank'] or 999),
        int(r['umaban']) if r['umaban'].isdigit() else 999,
    ))

    body = []
    for row in rows:
        best = row['best']
        if best:
            best_text = (
                f"{best.get('date', '')} {best.get('place', '')}"
                f"{best.get('surface', '')}{best.get('dist', '')}m"
            )
            if best.get('is_transfer'):
                adjust_text = f"{best.get('par_source', '転入元基準')}・日別補正なし"
            else:
                label = best.get('variant_label', '標準')
                adjust_text = f"南関{label}補正" if label != '標準' else "南関標準"
        else:
            best_text = '比較可能な時計なし'
            adjust_text = '-'
        index_text = f"{float(row['index']):.1f}" if row['index'] is not None else '-'
        same_text = (
            f"{float(row['same_max']):.1f} ({row['same_runs']}走)"
            if row['same_max'] is not None else '-'
        )
        if best:
            confidence = best.get('reference_confidence', '')
            samples = int(best.get('reference_samples', 0) or 0)
            adjust_text += f"・信頼{confidence}" if confidence else ''
            adjust_text += f"・{samples}件" if samples else ''
        rank_text = str(row['rank']) if row['rank'] is not None else '-'
        mark = (
            f'<span class="mark-btn" data-uma="{html.escape(row["umaban"])}" '
            f'data-mark="" onclick="cycleMark(this)"></span>'
        )
        body.append(
            '<tr>'
            f'<td>{html.escape(rank_text)}</td>'
            f'<td class="speed-horse">{mark}[{html.escape(row["umaban"])}] {html.escape(row["name"])}</td>'
            f'<td class="speed-value">{html.escape(index_text)}</td>'
            f'<td>{html.escape(same_text)}</td>'
            f'<td>{row["runs"]}</td>'
            f'<td><small>{html.escape(best_text)}</small></td>'
            f'<td><small>{html.escape(adjust_text)}</small></td>'
            '</tr>'
        )
    return (
        '<div class="speed-index-note">指数は平均級70・レコード級100。'
        '同場同距離は取得済み近走内の自己最高です。総合点には加算しません。'
        'JRA・他地方は日別補正なしで、競馬場・芝ダート・距離別基準を使います。</div>'
        '<div class="table-wrap"><table class="speed-index-table">'
        '<thead><tr><th>順位</th><th>馬</th><th>指数</th>'
        '<th>同場同距離</th><th>近走</th><th>最良走</th><th>基準・補正</th></tr></thead>'
        f'<tbody>{"".join(body)}</tbody></table></div>'
    )

def generate_html_output(year, month, day, place_name, r_num, header1, pace_text, eval_list_text, match_txt, ai_out_clean, details_text, index_table_html="(相対評価データなし)", speed_index_html="(指数データなし)"):
    
    def format_rank_text(text):
        t = html.escape(text)
        t = re.sub(r'([SABCDEFG])([①-⑳]*)', lambda m: f'<span class="rank-{m.group(1)}">{m.group(1)}{m.group(2)}</span>', t)
        return t

    def linkify_urls(text):
        return re.sub(
            r'(https?://[^\s<>"]+)',
            r'<a href="\1" target="_blank" rel="noopener" style="color:#2196F3;text-decoration:underline;">\1</a>',
            text
        )

    def inject_marks(text):
        def _repl(m):
            uma_id = ""
            if m.group(1): uma_id = str(ord(m.group(1)) - 0x245f)
            elif m.group(2): uma_id = m.group(2)
            if uma_id:
                return f'<span class="mark-btn" data-uma="{uma_id}" data-mark="" onclick="cycleMark(this)"></span>' + m.group(0)
            return m.group(0)
        return re.sub(r'(?<![\w])([\u2460-\u2473])|\[(\d{1,2})\]', _repl, text)

    def format_detailed_text(text):
        t = html.escape(text)
        # 調教ハイライトのマーカーを先に <strong> へ変換しておく。
        # （あとで実行する SABCDEFG ランク着色が `[[B]]` の B を拾って壊さないため）
        t = t.replace('[[RB]]', '<strong class="chokyo-top3">').replace('[[/RB]]', '</strong>')
        t = t.replace('[[B]]', '<strong class="chokyo-fast">').replace('[[/B]]', '</strong>')
        lines = t.split('\n')
        result = []
        in_chokyo = False
        in_race_content = False

        def close_open_blocks():
            nonlocal in_chokyo, in_race_content
            if in_chokyo:
                result.append('</div>')
                in_chokyo = False
            if in_race_content:
                result.append('</div>')
                in_race_content = False

        def format_jockey_line(line):
            m = re.match(r'^(騎:)(.*?)(　.*)$', line)
            if m:
                prefix, jockey_part, rest = m.groups()
            else:
                m = re.match(r'^(騎:)(.*)$', line)
                if not m:
                    return None
                prefix, jockey_part = m.groups()
                rest = ""

            changed = "←" in jockey_part
            power_up = False
            pm = re.search(r'P:(\d+).*?\(前P:(\d+)\)', line)
            if pm:
                try:
                    power_up = int(pm.group(1)) > int(pm.group(2))
                except ValueError:
                    power_up = False

            classes = ["jockey-name"]
            if changed:
                classes.append("jockey-changed")
            if power_up:
                classes.append("jockey-power-up")
            return f'<span class="jockey-line">{prefix}<span class="{" ".join(classes)}">{jockey_part}</span>{rest}</span>'
        
        for line in lines:
            stripped = line.strip()
            is_horse_header = bool(re.match(r'^[\u2460-\u2473]', stripped)) or bool(re.match(r'^\d{1,2}[\)\uff09]', stripped)) or bool(re.match(r'^\[\d{1,2}\]', stripped))
            
            if is_horse_header:
                close_open_blocks()
                uma_m = re.match(r'^([\u2460-\u2473])', stripped)
                uma_id = str(ord(uma_m.group(1)) - 0x245f) if uma_m else (re.match(r'^\[?(\d{1,2})', stripped).group(1) if re.match(r'^\[?(\d{1,2})', stripped) else '0')
                line_colored = re.sub(r'([SABCDEFG])(?![\w&<>])', lambda m: f'<span class="rank-{m.group(1)}">{m.group(1)}</span>', line)
                mark_btn = f'<span class="mark-btn" data-uma="{uma_id}" data-mark="" onclick="cycleMark(this)"></span>'
                result.append(f'<div class="horse-header">{mark_btn}<span>{line_colored}</span></div>')
            elif '【調教】' in stripped:
                close_open_blocks()
                result.append('<span class="chokyo-label">【調教】</span><div class="chokyo-text">')
                rest = stripped.replace('【調教】', '').strip()
                if rest: result.append(rest)
                in_chokyo = True
            elif stripped.startswith('🐎レース内容'):
                close_open_blocks()
                result.append('<span class="race-content-label">🐎レース内容</span><div class="race-content-text">')
                rest = stripped.replace('🐎レース内容', '').strip()
                if rest: result.append(rest)
                in_race_content = True
            elif stripped.startswith('-' * 10):
                close_open_blocks()
                result.append('<hr>')
            else:
                jockey_line = format_jockey_line(line)
                if jockey_line:
                    result.append(jockey_line)
                else:
                    line_colored = re.sub(r'([SABCDEFG])(?![\w&<>])', lambda m: f'<span class="rank-{m.group(1)}">{m.group(1)}</span>', line)
                    result.append(line_colored)

        close_open_blocks()
        out_html = '\n'.join(result)
        out_html = out_html.replace('[[RB]]', '<strong class="chokyo-top3">').replace('[[/RB]]', '</strong>')
        out_html = out_html.replace('[[B]]', '<strong class="chokyo-fast">').replace('[[/B]]', '</strong>')
        return out_html

    ai_out_clean = _build_kon_chokyo_marks({r_num: ai_out_clean}, num_races_total=1).get(r_num, ai_out_clean)
    details_text = _strip_chokyo_jockey_lines(details_text)

    format_m = linkify_urls(format_rank_text(match_txt.strip()))
    format_m = re.sub(r'\*\*\*(.+?)\*\*\*', r'<b style="color:#e00;">\1</b>', format_m)
    format_m = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', format_m)
    match_txt_html = inject_marks(format_m)
    pace_text_html = _format_small_text_markers(inject_marks(html.escape(pace_text)))
    details_text_html = inject_marks(html.escape(details_text).replace('【近走】', '<hr>【近走】'))

    html_content = f'''<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>{year}/{month}/{day} {place_name}{r_num}R 予想</title>
    <style>
        :root {{
            --primary: #2c3e50; --bg: #f5f7fa; --box-bg: #ffffff; --text: #333333; --border: #e2e8f0;
            --s-color: #d4af37; --a-color: #ff69b4; --b-color: #ff4500; --c-color: #ff8c00; 
            --d-color: #ffd700; --e-color: #32cd32; --f-color: #999999; --g-color: #cccccc;
        }}
        body {{ font-family: "Hiragino Kaku Gothic ProN", "Meiryo", sans-serif; background-color: var(--bg); color: var(--text); margin: 0; padding: 0; font-size: 14px; line-height: 1.5; }}
        .sticky-top {{ position: sticky; top: 0; z-index: 100; box-shadow: 0 2px 4px rgba(0,0,0,0.15); }}
        .header {{ background-color: var(--primary); color: white; padding: 6px 10px; }}
        .header h2 {{ margin: 0; font-size: 1.0em; display: inline; }}
        .header p {{ margin: 0; font-size: 0.78em; opacity: 0.9; display: inline; margin-left: 8px; }}
        
        .rank-S {{ color: var(--s-color); font-weight: bold; font-size: 1.2em; text-shadow: 0 0 1px rgba(0,0,0,0.3); }}
        .rank-A {{ color: var(--a-color); font-weight: bold; font-size: 1.1em; }}
        .rank-B {{ color: var(--b-color); font-weight: bold; font-size: 1.1em; }}
        .rank-C {{ color: var(--c-color); font-weight: bold; font-size: 1.1em; }}
        .rank-D {{ color: var(--d-color); font-weight: bold; font-size: 1.0em; text-shadow: 0 0 1px rgba(0,0,0,0.3); }}
        
        .eval-bar {{ background: #f0f4f8; border-bottom: 1px solid var(--border); padding: 3px 8px; font-size: 0.78em; white-space: nowrap; overflow-x: auto; line-height: 1.4; }}
        .tabs {{ display: flex; background: #fff; border-bottom: 2px solid var(--border); overflow: hidden; }}
        .tab-button {{ flex: 1; background: none; border: none; padding: 7px 2px; font-size: 0.78em; font-weight: bold; color: #666; cursor: pointer; border-bottom: 3px solid transparent; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
        .tab-button.active {{ color: var(--primary); border-bottom: 3px solid var(--primary); }}
        .tab-content {{ display: none; padding: 10px; }}
        .tab-content.active {{ display: block; }}
        .content-box {{ background: var(--box-bg); padding: 12px; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-bottom: 15px; overflow-x: auto; }}
        .speed-index-note {{ margin: 0 0 10px; color: #667085; font-size: 0.82em; line-height: 1.45; }}
        .speed-index-table {{ width: 100%; min-width: 780px; border-collapse: collapse; font-size: 0.9em; }}
        .speed-index-table th, .speed-index-table td {{ padding: 7px 6px; border-bottom: 1px solid var(--border); text-align: center; white-space: nowrap; }}
        .speed-index-table th {{ background: #f5f7fa; color: #667085; }}
        .speed-index-table .speed-horse {{ text-align: left; font-weight: bold; }}
        .speed-index-table .speed-value {{ color: #b42318; font-size: 1.15em; font-weight: bold; font-variant-numeric: tabular-nums; }}
        pre {{ white-space: pre-wrap; font-family: inherit; margin: 0; font-size: 0.95em; }}
        .pace-stable-note {{ display: block; font-size: 0.78em; color: #667085; line-height: 1.35; margin: 0 0 2px 1.6em; }}
        
        .chokyo-label {{ font-size: 0.85em; font-weight: bold; color: #555; }}
        .chokyo-text {{ font-size: 0.8em; color: #666; background: #f8f9fa; padding: 6px; border-radius: 4px; margin: 3px 0 8px; }}
        .chokyo-fast {{ font-weight: bold; color: #111; }}
        .chokyo-top3 {{ font-weight: bold; color: #c0392b; }}
        .chokyo-red {{ font-weight: bold; color: #d0021b; }}
        .chokyo-detail {{ font-size: 0.6em; font-weight: normal; }}
        .race-content-label {{ font-size: 0.85em; font-weight: bold; color: #4d5b2c; }}
        .race-content-text {{ font-size: 0.82em; color: #3f4a2a; background: #fbfcf2; border-left: 3px solid #a3b35c; padding: 6px 8px; border-radius: 4px; margin: 3px 0 8px; }}
        .jockey-line {{ display: block; font-size: 0.82em; color: #555; line-height: 1.35; margin: 1px 0 3px; }}
        .jockey-name {{ font-size: 0.92em; }}
        .jockey-changed {{ font-weight: bold; }}
        .jockey-power-up {{ color: #e74c3c; font-weight: bold; }}
        .horse-header {{ font-size: 1.05em; font-weight: bold; color: #1a1a2e; border-left: 4px solid var(--primary); padding-left: 8px; margin: 15px 0 5px; display: flex; align-items: center; gap: 6px; }}
        .mark-btn {{ display: inline-block; min-width: 20px; height: 20px; line-height: 20px; text-align: center; border: 1.5px solid #ccc; border-radius: 3px; cursor: pointer; font-size: 0.85em; user-select: none; background: #fff; flex-shrink: 0; }}
        .mark-btn:hover {{ background: #f0f0f0; }}
        .mark-btn[data-mark='◎'] {{ color: #e74c3c; border-color: #e74c3c; font-weight: bold; }}
        .mark-btn[data-mark='○'] {{ color: #2196F3; border-color: #2196F3; font-weight: bold; }}
        .mark-btn[data-mark='▲'] {{ color: #ff9800; border-color: #ff9800; font-weight: bold; }}
        .mark-btn[data-mark='△'] {{ color: #4caf50; border-color: #4caf50; }}
        .mark-btn[data-mark='★'] {{ color: #9c27b0; border-color: #9c27b0; }}
        .mark-btn[data-mark='☆'] {{ color: #795548; border-color: #795548; }}
        .mark-btn[data-mark='✓'] {{ color: #00bcd4; border-color: #00bcd4; }}
        .mark-btn[data-mark='消'] {{ color: #999; border-color: #999; text-decoration: line-through; }}

        #mark-menu {{ display: none; position: absolute; background: white; border: 1px solid #ccc; border-radius: 8px; box-shadow: 0 4px 12px rgba(0,0,0,0.15); padding: 8px; z-index: 2000; flex-direction: row; gap: 8px; align-items: center; }}
        .menu-item {{ width: 26px; height: 26px; line-height: 24px; text-align: center; border: 1px solid #eee; border-radius: 4px; cursor: pointer; font-size: 1.0em; display: flex; align-items: center; justify-content: center; }}
        .menu-item:hover {{ background: #f0f0f0; }}
        hr {{ border: 0; border-top: 1px dashed var(--border); margin: 15px 0; }}
        a {{ color: #2196F3; }}
    </style>
</head>
<body>
    <div id="mark-menu">
        <div class="menu-item" onclick="selectMark('◎')" style="color:#e74c3c">◎</div>
        <div class="menu-item" onclick="selectMark('○')" style="color:#2196F3">○</div>
        <div class="menu-item" onclick="selectMark('▲')" style="color:#ff9800">▲</div>
        <div class="menu-item" onclick="selectMark('△')" style="color:#4caf50">△</div>
        <div class="menu-item" onclick="selectMark('☆')" style="color:#795548">☆</div>
        <div class="menu-item" onclick="selectMark('✓')" style="color:#00bcd4">✓</div>
        <div class="menu-item" onclick="selectMark('消')" style="color:#999">消</div>
        <div class="menu-item clear" onclick="selectMark('')" style="color:#333; font-size: 0.7em;">✕</div>
    </div>

    <div class="sticky-top">
        <div class="header"><h2>{year}/{month}/{day} {place_name}{r_num}R</h2><p>{html.escape(header1)}</p></div>
        <div class="eval-bar">{format_rank_text(eval_list_text.replace(chr(10), "　")).replace('[[R]]', '<span class="chokyo-red">').replace('[[/R]]', '</span>').replace('[[SMALL]]', '<small class="chokyo-detail">').replace('[[/SMALL]]', '</small>')}</div>
        <div class="tabs">
            <button class="tab-button active" onclick="openTab(event, 'tab-pace')">展開</button>
            <button class="tab-button" onclick="openTab(event, 'tab-speed-index')">指数</button>
            <button class="tab-button" onclick="openTab(event, 'tab-index')">相対評価</button>
            <button class="tab-button" onclick="openTab(event, 'tab-ai')">出走馬分析</button>
            <button class="tab-button" onclick="openTab(event, 'tab-match')">対戦表</button>
        </div>
    </div>

    <div id="tab-pace" class="tab-content active"><div class="content-box"><pre>{pace_text_html}</pre></div></div>
    <div id="tab-speed-index" class="tab-content"><div class="content-box">{speed_index_html}</div></div>
    <div id="tab-index" class="tab-content"><div class="content-box" style="overflow-x: auto;">{index_table_html}</div></div>
    <div id="tab-ai" class="tab-content"><div class="content-box"><div style="white-space:pre-wrap;font-family:inherit;margin:0;font-size:0.95em;line-height:1.5;">{format_detailed_text(ai_out_clean)}</div></div></div>
    <div id="tab-match" class="tab-content"><div class="content-box"><pre>{match_txt_html}</pre></div></div>

    <script>
        const MARK_PREFIX = '{year}{month}{day}_{place_name}_';
        let currentTargetBtn = null;

        function openTab(evt, tabName) {{
            var i, tabcontent, tablinks;
            tabcontent = document.getElementsByClassName("tab-content");
            for (i = 0; i < tabcontent.length; i++) {{ tabcontent[i].style.display = "none"; tabcontent[i].classList.remove("active"); }}
            tablinks = document.getElementsByClassName("tab-button");
            for (i = 0; i < tablinks.length; i++) {{ tablinks[i].classList.remove("active"); }}
            var targetTab = document.getElementById(tabName);
            if (targetTab) {{ targetTab.style.display = "block"; targetTab.classList.add("active"); }}
            evt.currentTarget.classList.add("active");
            localStorage.setItem('keiba_active_tab_{r_num}', tabName);
        }}
        
        document.addEventListener('DOMContentLoaded', (event) => {{
            const savedTab = localStorage.getItem('keiba_active_tab_{r_num}');
            if (savedTab) {{
                const tabButton = document.querySelector(`button[onclick*="'${{savedTab}}'"]`);
                if (tabButton) tabButton.click();
            }}
            syncAllMarks();
        }});

        function syncAllMarks() {{
            document.querySelectorAll('.mark-btn').forEach(btn => {{
                const uma = btn.dataset.uma;
                const saved = localStorage.getItem(MARK_PREFIX + 'keiba_mark_{r_num}_' + uma);
                btn.dataset.mark = saved || ''; btn.textContent = saved || '';
            }});
        }}

        function openMarkMenu(el, umaId) {{
            currentTargetBtn = el;
            const menu = document.getElementById('mark-menu');
            const rect = el.getBoundingClientRect();
            menu.style.display = 'flex';
            let top = window.scrollY + rect.bottom + 5;
            let left = rect.left;
            if (left + 250 > window.innerWidth) left = window.innerWidth - 260;
            menu.style.top = top + 'px';
            menu.style.left = Math.max(5, left) + 'px';
            if (window.event) window.event.stopPropagation();
        }}

        function selectMark(mark) {{
            if (!currentTargetBtn) return;
            const uma = currentTargetBtn.dataset.uma;
            if (mark) localStorage.setItem(MARK_PREFIX + 'keiba_mark_{r_num}_' + uma, mark);
            else localStorage.removeItem(MARK_PREFIX + 'keiba_mark_{r_num}_' + uma);
            syncAllMarks();
            document.getElementById('mark-menu').style.display = 'none';
        }}

        window.onclick = function(event) {{
            const menu = document.getElementById('mark-menu');
            if (menu.style.display === 'flex') menu.style.display = 'none';
        }};

        function cycleMark(el) {{ openMarkMenu(el, el.dataset.uma); }}
    </script>
</body>
</html>'''
    return html_content


def generate_combined_html(year, month, day, place_name, race_results):
    def _escape(text): return html.escape(str(text)) if text else ""
    def _rank_color(text):
        t = html.escape(str(text))
        return re.sub(r'([SABCDEFG])([①-⑳]*)', lambda m: f'<span class="rank-{m.group(1)}">{m.group(1)}{m.group(2)}</span>', t)
    def _linkify(text):
        return re.sub(r'(https?://[^\s<>"]+)', r'<a href="\1" target="_blank" rel="noopener" style="color:#2196F3;">\1</a>', text)

    def _format_ai(text, race_num='0'):
        t = html.escape(str(text))
        # 調教ハイライトのマーカーを先に <strong> へ変換しておく。
        # （あとで実行する SABCDEFG ランク着色が `[[B]]` の B を拾って壊さないため）
        t = t.replace('[[RB]]', '<strong class="chokyo-top3">').replace('[[/RB]]', '</strong>')
        t = t.replace('[[B]]', '<strong class="chokyo-fast">').replace('[[/B]]', '</strong>')
        lines = t.split('\n')
        result = []
        in_chokyo = False
        in_race_content = False

        def close_open_blocks():
            nonlocal in_chokyo, in_race_content
            if in_chokyo:
                result.append('</div>')
                in_chokyo = False
            if in_race_content:
                result.append('</div>')
                in_race_content = False

        def format_jockey_line(line):
            m = re.match(r'^(騎:)(.*?)(　.*)$', line)
            if m:
                prefix, jockey_part, rest = m.groups()
            else:
                m = re.match(r'^(騎:)(.*)$', line)
                if not m:
                    return None
                prefix, jockey_part = m.groups()
                rest = ""

            changed = "←" in jockey_part
            power_up = False
            pm = re.search(r'P:(\d+).*?\(前P:(\d+)\)', line)
            if pm:
                try:
                    power_up = int(pm.group(1)) > int(pm.group(2))
                except ValueError:
                    power_up = False

            classes = ["jockey-name"]
            if changed:
                classes.append("jockey-changed")
            if power_up:
                classes.append("jockey-power-up")
            return f'<span class="jockey-line">{prefix}<span class="{" ".join(classes)}">{jockey_part}</span>{rest}</span>'

        for line in lines:
            stripped = line.strip()
            is_horse_header = bool(re.match(r'^[\u2460-\u2473]', stripped)) or bool(re.match(r'^\d{1,2}[\)\uff09]', stripped)) or bool(re.match(r'^\[\d{1,2}\]', stripped))
            if is_horse_header:
                close_open_blocks()
                uma_m = re.match(r'^([\u2460-\u2473])', stripped)
                uma_id = str(ord(uma_m.group(1)) - 0x245f) if uma_m else (re.match(r'^\[?(\d{1,2})', stripped).group(1) if re.match(r'^\[?(\d{1,2})', stripped) else '0')
                line_c = re.sub(r'([SABCDEFG])(?![\w&<>])', lambda m: f'<span class="rank-{m.group(1)}">{m.group(1)}</span>', line)
                mark_btn = f'<span class="mark-btn" data-uma="{uma_id}" data-race="{race_num}" data-mark="" onclick="cycleMark(this)"></span>'
                result.append(f'<div class="horse-header">{mark_btn}<span>{line_c}</span></div>')
            elif '【調教】' in stripped:
                close_open_blocks()
                result.append('<span class="chokyo-label">【調教】</span><div class="chokyo-text">')
                rest = stripped.replace('【調教】', '').strip()
                if rest: result.append(rest)
                in_chokyo = True
            elif stripped.startswith('🐎レース内容'):
                close_open_blocks()
                result.append('<span class="race-content-label">🐎レース内容</span><div class="race-content-text">')
                rest = stripped.replace('🐎レース内容', '').strip()
                if rest: result.append(rest)
                in_race_content = True
            elif stripped.startswith('-' * 10):
                close_open_blocks()
                result.append('<hr>')
            else:
                jockey_line = format_jockey_line(line)
                if jockey_line:
                    result.append(jockey_line)
                else:
                    line_c = re.sub(r'([SABCDEFG])(?![\w&<>])', lambda m: f'<span class="rank-{m.group(1)}">{m.group(1)}</span>', line)
                    result.append(line_c)
        close_open_blocks()
        out_html = '\n'.join(result)
        out_html = out_html.replace('[[RB]]', '<strong class="chokyo-top3">').replace('[[/RB]]', '</strong>')
        out_html = out_html.replace('[[B]]', '<strong class="chokyo-fast">').replace('[[/B]]', '</strong>')
        return out_html

    def _inject_marks(text, r_num):
        def _repl(m):
            uma_id = ""
            if m.group(1): uma_id = str(ord(m.group(1)) - 0x245f)
            elif m.group(2): uma_id = m.group(2)
            if uma_id:
                return f'<span class="mark-btn" data-uma="{uma_id}" data-race="{r_num}" data-mark="" onclick="cycleMark(this)"></span>' + m.group(0)
            return m.group(0)
        return re.sub(r'(?<![\w])([\u2460-\u2473])|\[(\d{1,2})\]', _repl, text)

    sorted_races = sorted(race_results.items(), key=lambda x: int(x[0]))

    race_tabs_html = ""
    for i, (r_num, _) in enumerate(sorted_races):
        active = "active" if i == 0 else ""
        race_tabs_html += f'<button class="race-tab {active}" id="btn-race-{r_num}" onclick="openRace(event, \'race-{r_num}\')">{r_num}</button>\n'

    # 各レースのテキストをパース
    parsed_per_race = {}
    for r_num, data in sorted_races:
        text_val = data.get("text", "") if isinstance(data, dict) else str(data)

        markers = [('【展開予想】', 'pace'), ('【総合評価】', 'eval'), ('【評価一覧】', 'eval'), ('【対戦表', 'match'), ('【相対評価】', 'index'), ('【出走馬分析】', 'ai')]
        lines_all = text_val.split('\n')
        header_lines = []
        for ln in lines_all:
            if ln.strip().startswith('【') or ln.strip().startswith('◆'): break
            if ln.strip(): header_lines.append(ln.strip())
        header_part = '\n'.join(header_lines)

        positions = []
        for m_text, m_key in markers:
            idx = text_val.find(m_text)
            if idx >= 0: positions.append((idx, m_key))
        positions.sort()

        pace_part, eval_part, match_part, index_part, ai_part = "", "", "", "", ""
        for pi, (pos, key) in enumerate(positions):
            end_pos = positions[pi + 1][0] if pi + 1 < len(positions) else len(text_val)
            content = text_val[pos:end_pos].strip()
            if key == 'pace': pace_part = content
            elif key == 'eval': eval_part = content
            elif key == 'match': match_part = content
            elif key == 'index': index_part = content.replace('【相対評価】', '').strip()
            elif key == 'ai': ai_part = content.replace('【出走馬分析】', '').strip()
        parsed_per_race[r_num] = {
            "header_part": header_part,
            "pace_part": pace_part,
            "eval_part": eval_part,
            "match_part": match_part,
            "index_part": index_part,
            "ai_part": ai_part,
        }

    # 今走調教 太字/赤太字マーカー付与（全レース横断ランキング）
    num_races_total = len(sorted_races)
    ai_text_map = {r_num: parsed_per_race[r_num]["ai_part"] for r_num in parsed_per_race}
    marked_ai_map = _build_kon_chokyo_marks(ai_text_map, num_races_total=num_races_total)
    for r_num, marked in marked_ai_map.items():
        parsed_per_race[r_num]["ai_part"] = marked

    race_contents_html = ""
    for i, (r_num, data) in enumerate(sorted_races):
        active = "active" if i == 0 else ""
        parts = parsed_per_race[r_num]
        header_part = parts["header_part"]
        pace_part = parts["pace_part"]
        eval_part = parts["eval_part"]
        match_part = parts["match_part"]
        index_part = parts["index_part"]
        ai_part = parts["ai_part"]

        race_contents_html += f'''
        <div id="race-{r_num}" class="race-content {active}" style="display: {'block' if i==0 else 'none'};">
            <div class="race-header-bar">{_escape(header_part)}</div>

            <div class="eval-box" style="padding:6px 12px; border-bottom:1px solid #e2e8f0;">{_inject_marks(_rank_color(eval_part), r_num)}</div>

            <div class="inner-tabs">
                <button class="inner-tab active" onclick="openInnerTab(event, 'race-{r_num}-pace')">展開</button>
                <button class="inner-tab" onclick="openInnerTab(event, 'race-{r_num}-index')">相対評価</button>
                <button class="inner-tab" onclick="openInnerTab(event, 'race-{r_num}-ai')">出走馬分析</button>
                <button class="inner-tab" onclick="openInnerTab(event, 'race-{r_num}-match')">対戦表</button>
            </div>

            <div id="race-{r_num}-pace" class="inner-content active">
                <div class="content-box">
                    <pre>{_format_small_text_markers(_inject_marks(_escape(pace_part), r_num))}</pre>
                </div>
            </div>

            <div id="race-{r_num}-index" class="inner-content">
                <div class="content-box" style="overflow-x: auto;">
                    {index_part}
                </div>
            </div>

            <div id="race-{r_num}-ai" class="inner-content">
                <div class="content-box">
                    <div style="white-space:pre-wrap;font-size:0.95em;line-height:1.5;">{_format_ai(ai_part, r_num)}</div>
                </div>
            </div>

            <div id="race-{r_num}-match" class="inner-content">
                <div class="content-box">
                    <pre>{_inject_marks(_linkify(_rank_color(match_part)), r_num)}</pre>
                </div>
            </div>
        </div>
        '''
    
    combined_html = f'''<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>{year}/{month}/{day} {place_name} 全レース予想</title>
    <style>
        :root {{
            --primary: #2c3e50; --bg: #f5f7fa; --box-bg: #fff; --text: #333; --border: #e2e8f0;
            --s-color: #d4af37; --a-color: #ff69b4; --b-color: #ff4500; --c-color: #ff8c00;
            --d-color: #ffd700; --e-color: #32cd32; --f-color: #999; --g-color: #ccc;
        }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: "Hiragino Kaku Gothic ProN","Meiryo",sans-serif; background: var(--bg); color: var(--text); padding: 0; margin: 0; font-size: 14px; line-height: 1.5; }}
        
        .race-tabs {{ display: flex; flex-wrap: nowrap; gap: 1px; background: #fff; padding: 4px 5px; border-bottom: 2px solid var(--border); position: sticky; top: 0; z-index: 100; box-shadow: 0 2px 4px rgba(0,0,0,0.08); }}
        .race-tab {{ flex: 1 1 0; min-width: 0; padding: 7px 0; border: 1px solid var(--border); border-radius: 4px 4px 0 0; background: #f0f0f0; color: #666; font-weight: bold; font-size: 0.86em; cursor: pointer; border-bottom: none; text-align: center; }}
        .race-tab.active {{ background: var(--primary); color: #fff; border-color: var(--primary); }}
        
        .race-content {{ display: none; }}
        .race-content.active {{ display: block; }}
        .race-header-bar {{ background: #e8ecf1; padding: 10px 12px; font-weight: bold; font-size: 0.95em; border-bottom: 1px solid var(--border); }}
        
        .inner-tabs {{ display: flex; background: #fff; border-bottom: 1px solid var(--border); }}
        .inner-tab {{ flex: 1; padding: 10px 5px; border: none; background: none; font-size: 0.9em; font-weight: bold; color: #888; cursor: pointer; border-bottom: 3px solid transparent; }}
        .inner-tab.active {{ color: var(--primary); border-bottom: 3px solid var(--primary); }}
        
        .inner-content {{ display: none; padding: 10px; }}
        .inner-content.active {{ display: block; }}
        
        .content-box {{ background: #fff; padding: 12px; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); margin-bottom: 10px; overflow-x: auto; }}
        .eval-box {{ font-size: 1.1em; line-height: 1.6; margin-bottom: 10px; word-break: break-all; }}
        pre {{ white-space: pre-wrap; font-family: inherit; margin: 0; font-size: 0.95em; }}
        .pace-stable-note {{ display: block; font-size: 0.78em; color: #667085; line-height: 1.35; margin: 0 0 2px 1.6em; }}
        
        .rank-S {{ color: var(--s-color); font-weight: bold; font-size: 1.2em; text-shadow: 0 0 1px rgba(0,0,0,0.3); }}
        .rank-A {{ color: var(--a-color); font-weight: bold; font-size: 1.1em; }}
        .rank-B {{ color: var(--b-color); font-weight: bold; font-size: 1.1em; }}
        .rank-C {{ color: var(--c-color); font-weight: bold; font-size: 1.1em; }}
        .rank-D {{ color: var(--d-color); font-weight: bold; text-shadow: 0 0 1px rgba(0,0,0,0.3); }}
        
        .horse-header {{ font-size: 1.05em; font-weight: bold; color: #1a1a2e; border-left: 4px solid var(--primary); padding-left: 8px; margin: 15px 0 5px; display: flex; align-items: center; gap: 8px; }}
        .mark-btn {{ display: inline-block; min-width: 20px; height: 20px; line-height: 20px; text-align: center; border: 1.5px solid #ccc; border-radius: 3px; cursor: pointer; background: #fff; font-size: 0.85em; vertical-align: middle; }}
        .mark-btn[data-mark='◎'] {{ color: #e74c3c; border-color: #e74c3c; font-weight: bold; }}
        .mark-btn[data-mark='○'] {{ color: #2196F3; border-color: #2196F3; font-weight: bold; }}
        .mark-btn[data-mark='▲'] {{ color: #ff9800; border-color: #ff9800; font-weight: bold; }}
        .mark-btn[data-mark='△'] {{ color: #4caf50; border-color: #4caf50; }}
        .mark-btn[data-mark='★'] {{ color: #9c27b0; border-color: #9c27b0; }}
        .mark-btn[data-mark='☆'] {{ color: #795548; border-color: #795548; }}
        .mark-btn[data-mark='✓'] {{ color: #00bcd4; border-color: #00bcd4; }}
        .mark-btn[data-mark='消'] {{ color: #999; border-color: #999; text-decoration: line-through; }}

        #mark-menu {{ display: none; position: absolute; background: white; border: 1px solid #ccc; border-radius: 8px; box-shadow: 0 4px 12px rgba(0,0,0,0.15); padding: 8px; z-index: 2000; flex-direction: row; gap: 8px; align-items: center; }}
        .menu-item {{ width: 28px; height: 28px; line-height: 28px; text-align: center; border: 1px solid #eee; border-radius: 4px; cursor: pointer; font-size: 1.1em; }}
        .menu-item:hover {{ background: #f0f0f0; }}
        
        .chokyo-label {{ font-size: 0.85em; font-weight: bold; color: #555; }}
        .chokyo-text {{ font-size: 0.8em; color: #666; background: #f8f9fa; padding: 6px; border-radius: 4px; margin: 3px 0 8px; }}
        .chokyo-fast {{ font-weight: bold; color: #111; }}
        .chokyo-top3 {{ font-weight: bold; color: #c0392b; }}
        .chokyo-red {{ font-weight: bold; color: #d0021b; }}
        .chokyo-detail {{ font-size: 0.6em; font-weight: normal; }}
        .race-content-label {{ font-size: 0.85em; font-weight: bold; color: #4d5b2c; }}
        .race-content-text {{ font-size: 0.82em; color: #3f4a2a; background: #fbfcf2; border-left: 3px solid #a3b35c; padding: 6px 8px; border-radius: 4px; margin: 3px 0 8px; }}
        .jockey-line {{ display: block; font-size: 0.82em; color: #555; line-height: 1.35; margin: 1px 0 3px; }}
        .jockey-name {{ font-size: 0.92em; }}
        .jockey-changed {{ font-weight: bold; }}
        .jockey-power-up {{ color: #e74c3c; font-weight: bold; }}
        hr {{ border: 0; border-top: 1px dashed var(--border); margin: 15px 0; }}
        a {{ color: #2196F3; }}
    </style>
</head>
<body>
    <div id="mark-menu">
        <div class="menu-item" onclick="selectMark('◎')" style="color:#e74c3c">◎</div>
        <div class="menu-item" onclick="selectMark('○')" style="color:#2196F3">○</div>
        <div class="menu-item" onclick="selectMark('▲')" style="color:#ff9800">▲</div>
        <div class="menu-item" onclick="selectMark('△')" style="color:#4caf50">△</div>
        <div class="menu-item" onclick="selectMark('★')" style="color:#9c27b0">★</div>
        <div class="menu-item" onclick="selectMark('☆')" style="color:#795548">☆</div>
        <div class="menu-item" onclick="selectMark('✓')" style="color:#00bcd4">✓</div>
        <div class="menu-item" onclick="selectMark('消')" style="color:#999">消</div>
        <div class="menu-item" onclick="selectMark('')" style="color:#333;font-size:0.8em">--</div>
    </div>

    <div class="race-tabs">
        {race_tabs_html}
    </div>
    
    {race_contents_html}
    
    <script>
        const MARK_PREFIX = '{year}{month}{day}_{place_name}_';
        let currentTargetBtn = null;

        function openRace(evt, raceId) {{
            document.querySelectorAll('.race-content').forEach(el => {{ el.classList.remove('active'); el.style.display = 'none'; }});
            document.querySelectorAll('.race-tab').forEach(el => el.classList.remove('active'));
            var target = document.getElementById(raceId);
            if (target) {{ target.style.display = 'block'; target.classList.add('active'); }}
            evt.currentTarget.classList.add('active');
            localStorage.setItem('keiba_combined_race', raceId);
        }}
        
        function openInnerTab(evt, tabId) {{
            var parent = evt.currentTarget.closest('.race-content');
            parent.querySelectorAll('.inner-content').forEach(el => {{ el.classList.remove('active'); el.style.display = 'none'; }});
            parent.querySelectorAll('.inner-tab').forEach(el => el.classList.remove('active'));
            var target = document.getElementById(tabId);
            if (target) {{ target.style.display = 'block'; target.classList.add('active'); }}
            evt.currentTarget.classList.add('active');
            localStorage.setItem('keiba_inner_' + parent.id, tabId);
        }}
        
        function syncAllMarks() {{
            document.querySelectorAll('.mark-btn').forEach(btn => {{
                const r = btn.dataset.race; const u = btn.dataset.uma;
                const s = localStorage.getItem(MARK_PREFIX + 'keiba_mark_' + r + '_' + u);
                btn.dataset.mark = s || ''; btn.textContent = s || '';
            }});
        }}

        function cycleMark(el) {{
            currentTargetBtn = el;
            const menu = document.getElementById('mark-menu');
            const rect = el.getBoundingClientRect();
            menu.style.display = 'flex';
            menu.style.top = (window.scrollY + rect.bottom + 5) + 'px';
            menu.style.left = Math.min(window.innerWidth - 300, rect.left) + 'px';
            event.stopPropagation();
        }}

        function selectMark(mark) {{
            if (!currentTargetBtn) return;
            const r = currentTargetBtn.dataset.race; const u = currentTargetBtn.dataset.uma;
            if (mark) localStorage.setItem(MARK_PREFIX + 'keiba_mark_' + r + '_' + u, mark);
            else localStorage.removeItem(MARK_PREFIX + 'keiba_mark_' + r + '_' + u);
            syncAllMarks();
            document.getElementById('mark-menu').style.display = 'none';
        }}

        window.onclick = function() {{ document.getElementById('mark-menu').style.display = 'none'; }};

        document.addEventListener('DOMContentLoaded', () => {{
            const savedRace = localStorage.getItem('keiba_combined_race');
            if (savedRace && document.getElementById(savedRace)) {{
                const btn = document.querySelector(`.race-tab[onclick*="'${{savedRace}}'"]`);
                if (btn) btn.click();
            }} else {{
                const first = document.querySelector('.race-tab');
                if (first) first.click();
            }}
            
            document.querySelectorAll('.race-content').forEach(rc => {{
                const savedInner = localStorage.getItem('keiba_inner_' + rc.id);
                if (savedInner) {{
                    const btn = rc.querySelector(`.inner-tab[onclick*="'${{savedInner}}'"]`);
                    if (btn) btn.click();
                }}
            }});
            
            syncAllMarks();
        }});
    </script>
</body>
</html>'''
    return combined_html

# ==================================================
# 9. 指数テーブル生成
# ==================================================
def generate_index_table(horses_data, pace_speed_list):
    if not pace_speed_list: return ""

    max_comp_si = max([si for _, si, _ in pace_speed_list if si is not None], default=0)

    table_lines = ["\n【指数テーブル】"]
    table_lines.append("| 馬番 | 馬名 | 総合指数 | 展開 |")
    table_lines.append("|---|---|---|---|")

    horse_indices = {str(h_num): {"name": h_data["name"], "speed_index": None, "pace_type": None}
                     for h_num, h_data in horses_data.items()}

    for uma_num, speed_index, pace_type in pace_speed_list:
        if str(uma_num) in horse_indices:
            horse_indices[str(uma_num)]["speed_index"] = speed_index
            horse_indices[str(uma_num)]["pace_type"] = pace_type

    for uma_num_str in sorted(horse_indices.keys(), key=int):
        data = horse_indices[uma_num_str]
        name = data["name"]
        speed_index = data["speed_index"]
        pace_type = data["pace_type"]

        si_display = f"{speed_index:.1f}" if speed_index is not None else "-"
        pace_display = pace_type if pace_type else "-"

        if speed_index is not None and max_comp_si > 0 and speed_index >= max_comp_si * 0.8:
            si_display = f"**{si_display}**"
        
        table_lines.append(f"| {uma_num_str} | {name} | {si_display} | {pace_display} |")

    return "\n".join(table_lines)


# ==================================================
# 10. ジェネレータ（全レース処理）
# ==================================================
def _cleanup_old_cache(days=7):
    """予想生成時に、days日より前のローカルcacheファイルを削除する。
    データはSupabaseに保存済みで、ローカルcacheは短期用途(1時間の再生成スキップ／同期ステージ)
    に留めるため。ファイル名中の8桁日付(YYYYMMDD)で判定し、cutoff日以前(<=)を削除。
    日付が読めないファイルは残す。例: 7/8生成時は cutoff=7/1 → 7/1以前を削除。"""
    import glob as _glob
    import datetime as _dt
    try:
        import store as _st
        cache_dir = _st.CACHE_DIR
    except Exception:
        cache_dir = "cache"
    if not os.path.isdir(cache_dir):
        return 0
    jst_today = (_dt.datetime.utcnow() + _dt.timedelta(hours=9)).date()
    cutoff = (jst_today - _dt.timedelta(days=days)).strftime("%Y%m%d")
    removed = 0
    for path in _glob.glob(os.path.join(cache_dir, "*")):
        m = re.search(r'(20\d{6})', os.path.basename(path))
        if not m or m.group(1) > cutoff:
            continue
        try:
            os.remove(path)
            removed += 1
        except Exception:
            pass
    return removed


def run_races_iter(year, month, day, place_code, target_races, mode="dify", manual_kai_nichi=None, bypass_cache=False, **kwargs):
    resources = load_resources()

    # 生成のたびに7日より前の古いローカルcacheを掃除（Supabaseにあるのでローカルは短期のみ）。
    try:
        _removed = _cleanup_old_cache(7)
        if _removed:
            yield {"type": "status", "data": f"🧹 7日より前の古いキャッシュ {_removed} 件を削除"}
    except Exception:
        pass

    kb_input_map = {"10": "大井", "11": "川崎", "12": "船橋", "13": "浦和"}
    nk_code_map = {"10": "20", "11": "21", "12": "19", "13": "18"}

    place_name = kb_input_map.get(place_code, "地方")
    nk_place_code = nk_code_map.get(place_code)

    driver = get_driver()

    try:
        kai, nichi = None, None
        if manual_kai_nichi and manual_kai_nichi.get("kai") and manual_kai_nichi.get("nichi"):
            # 文字列が渡されても確実に整数化する
            kai = int(manual_kai_nichi["kai"])
            nichi = int(manual_kai_nichi["nichi"])
            yield {"type": "status", "data": f"📅 手動指定を使用: {place_name} 第{kai}回 {nichi}日目"}
        else:
            yield {"type": "status", "data": f"📅 開催特定中 ({place_name})..."}
            kai, nichi = get_nankan_kai_nichi(year, month, day, nk_place_code)
            
        if not kai:
            yield {"type": "error", "data": "開催特定失敗（左のメニューから手動で第何回・何日目か入力してください）"}
            return

        yield {"type": "status", "data": f"✅ {place_name} 第{kai}回 {nichi}日目"}

        yield {"type": "status", "data": "🔑 競馬ブック ログイン中..."}
        login_keibabook_robust(driver)

        nankan_base = "https://www.nankankeiba.com"
        nankan_race_prefix = (
            f"{year}{str(month).zfill(2)}{str(day).zfill(2)}"
            f"{str(nk_place_code).zfill(2)}{int(kai):02d}{int(nichi):02d}"
        )
        prog_url = f"{nankan_base}/program/{nankan_race_prefix}.do"
        driver.get(prog_url)
        soup = BeautifulSoup(driver.page_source, "html.parser")

        r_nums = []
        race_detail_urls = {}
        for a in soup.find_all("a", href=True):
            href = a["href"]
            id_match = re.search(r"/syousai/(\d{16})\.do", href)
            if id_match and id_match.group(1).startswith(nankan_race_prefix):
                race_id = id_match.group(1)
                race_num = int(race_id[14:16])
                r_nums.append(race_num)
                race_detail_urls.setdefault(race_num, urljoin(nankan_base, href))
        r_nums = sorted(set(r_nums)) or list(range(1, 13))

        os.makedirs("cache", exist_ok=True)
        day_chokyo = []          # 全体TOP3・レース内TOP3の今走調教ランキング素材
        day_chokyo_display = []  # 表示用: 今走/前走調教の同日順位素材
        day_chokyo_pending = {}  # r_num -> 再採点・再描画用の素材一式
        for r_num in r_nums:
            if target_races and r_num not in target_races:
                continue

            cache_file = f"cache/{year}{month}{day}_{place_code}_{r_num}_{mode}.json"
            dify_cache_file = f"cache/{year}{month}{day}_{place_code}_{r_num}_dify.json"
            if mode == "relative":
                cache_file = dify_cache_file
            if not bypass_cache and os.path.exists(cache_file):
                if time.time() - os.path.getmtime(cache_file) < 3600:
                    try:
                        with open(cache_file, "r", encoding="utf-8") as f:
                            cached_data = json.load(f)
                        
                        data_text = cached_data.get("data_text", "")
                        data_html = cached_data.get("data_html", "")
                        cache_needs_refresh = (
                            mode in ("dify", "normal", "relative", "pace")
                            and (
                                "jockey-line" not in data_html
                                or "race-content-text" not in data_html
                                or "pace-stable-note" not in data_html
                            )
                        )
                        cache_needs_refresh = cache_needs_refresh or (
                            mode in ("dify", "normal", "relative", "pace")
                            and cached_data.get("score_policy") != SCORE_POLICY_VERSION
                        )
                        if any(x in data_text for x in ["（回答生成エラー）", "⚠️", "パースエラー", "Error"]):
                            pass
                        elif cache_needs_refresh:
                            pass
                        else:
                            yield {"type": "status", "data": f"⚡ {r_num}R キャッシュから高速読み込み中... (前回実行から1時間以内)"}
                            
                            yield {
                                "type": "result",
                                "race_num": r_num,
                                "data_text": data_text,
                                "data_html": data_html
                            }
                            time.sleep(0.5)
                            continue
                    except Exception:
                        pass

            yield {"type": "status", "data": f"🏇 {r_num}R データ解析中..."}

            try:
                nk_id = f"{year}{str(month).zfill(2)}{str(day).zfill(2)}{str(nk_place_code).zfill(2)}{str(kai).zfill(2)}{str(nichi).zfill(2)}{str(r_num).zfill(2)}"
                kb_id = get_kb_url_id(year, month, day, place_code, nichi, r_num)

                danwa, cyokyo = parse_kb_danwa_cyokyo(driver, kb_id)

                detail_url = race_detail_urls.get(r_num) or f"{nankan_base}/syousai/{nk_id}.do"
                detail_url_candidates = []
                for candidate in (
                    detail_url,
                    f"{nankan_base}/syousai/{nk_id}.do?type=uma_shosai",
                    f"{nankan_base}/uma_shosai/{nk_id}.do",
                ):
                    if candidate not in detail_url_candidates:
                        detail_url_candidates.append(candidate)

                sess = get_http_session()
                nk_data = {"meta": {}, "horses": {}}
                last_detail_error = ""
                for candidate_url in detail_url_candidates:
                    try:
                        detail_res = sess.get(candidate_url, timeout=15)
                        detail_res.encoding = "cp932"
                        if detail_res.status_code >= 400:
                            last_detail_error = f"HTTP {detail_res.status_code}: {candidate_url}"
                            continue
                        nk_data = parse_nankankeiba_detail(detail_res.text, place_name, resources)
                        if nk_data["horses"]:
                            detail_url = candidate_url
                            break
                    except Exception as e:
                        last_detail_error = f"{candidate_url}: {e}"

                if not nk_data["horses"]:
                    try:
                        driver.get(detail_url)
                        time.sleep(2)
                        for attempt in range(3):
                            try:
                                driver.execute_script("if(typeof changeShosai === 'function'){ changeShosai('s1'); }")
                                time.sleep(1.5)
                                break
                            except Exception:
                                time.sleep(1)
                        nk_data = parse_nankankeiba_detail(driver.page_source, place_name, resources)
                    except Exception:
                        pass

                if not nk_data["horses"]:
                    suffix = f": {last_detail_error}" if last_detail_error else ""
                    yield {"type": "error", "data": f"{r_num}R データなし (HTML解析失敗{suffix})"}
                    continue

                header = (
                    f"レース名:{r_num}R {nk_data['meta'].get('race_name','')} "
                    f"格:{nk_data['meta'].get('grade','')} "
                    f"コース:{nk_data['meta'].get('course','')}"
                )

                horse_texts = []
                for u in sorted(nk_data["horses"].keys(), key=int):
                    h = nk_data["horses"][u]
                    power_line = h.get("display_power", f"【騎手】{h.get('power','P:不明')}、 相性:{h.get('compat','-')}")

                    curr_j = (h.get("jockey") or "").strip()
                    prev_j = (h.get("prev_jockey") or "").strip()

                    if prev_j:
                        if curr_j == prev_j:
                            jockey_disp = f"{curr_j}(同)"
                        else:
                            jockey_disp = f"{curr_j}←{prev_j}"
                    else:
                        jockey_disp = curr_j

                    block = [
                        f"[{u}]{h['name']} 騎:{jockey_disp} 師:{h.get('trainer','')}",
                        f"話:{danwa.get(u,'なし')}",
                        f"調:{cyokyo.get(u,'データなし')}",
                        power_line,
                        "【近走】",
                    ]
                    for hs in h.get("hist", []):
                        hs_text = hs.get("text", "") if isinstance(hs, dict) else hs
                        block.append(hs_text)

                    horse_texts.append("\n".join(block))

                full_prompt = header + "\n\n" + "\n\n".join(horse_texts)

                # 出走各馬の好走パターン(ユーザーメモ)を取得（採点±5と展開の小表示に使用）
                horse_patterns = _fetch_horse_patterns(nk_data["horses"])

                if mode == "raw":
                    yield {"type": "status", "data": f"🔍 {r_num}R 対戦データを取得中..."}
                    header1 = f"{nk_data['meta'].get('race_name','')}  {nk_data['meta'].get('course','')}  {nk_data['meta'].get('grade','')}"
                    course_match = re.search(r'(浦和|川崎|大井|船橋)', header1)
                    current_course = course_match.group(1) if course_match else place_name
                    _meta_dist = nk_data['meta'].get('dist', '')
                    if _meta_dist and str(_meta_dist).isdigit():
                        current_dist = _meta_dist
                    else:
                        _dist_m = re.search(r'(\d{1,2},\d{3}|\d{3,4})m', nk_data['meta'].get('course', '') + ' ' + header1)
                        current_dist = _dist_m.group(1).replace(',', '') if _dist_m else "1400"
                    match_txt = _fetch_matchup_table_selenium(driver, nk_id, grades={}, horses_data=nk_data["horses"], current_course=current_course, current_dist=current_dist)
                    
                    pace_text, pace_speed_list, pace_bonus_map = predict_pace_python(nk_data["horses"], danwa, nk_data['meta'].get('course',''), horse_patterns=horse_patterns)
                    details_text = _strip_chokyo_jockey_lines("\n\n".join(horse_texts))

                    final_text = (
                        f"📅 {year}/{month}/{day} {place_name}{r_num}R\n\n"
                        f"{header1}\n\n"
                        f"{pace_text}\n\n"
                        f"{match_txt}\n\n"
                        f"{details_text}"
                    )
                    final_html = generate_html_output(year, month, day, place_name, r_num, header1, pace_text, "【総合評価】  (AI未実行)", match_txt, "(AI未実行)", details_text, "(相対評価未実行)")

                    try:
                        with open(cache_file, "w", encoding="utf-8") as f:
                            json.dump({
                                "data_text": final_text,
                                "data_html": final_html,
                                "score_policy": SCORE_POLICY_VERSION
                            }, f, ensure_ascii=False)
                    except Exception:
                        pass

                    yield {"type": "result", "race_num": r_num, "data_text": final_text, "data_html": final_html}
                    time.sleep(1)
                    continue

                header1 = f"{nk_data['meta'].get('race_name','')}  {nk_data['meta'].get('course','')}  {nk_data['meta'].get('grade','')}"
                pace_text, pace_speed_list, pace_bonus_map = predict_pace_python(nk_data["horses"], danwa, nk_data['meta'].get('course',''), horse_patterns=horse_patterns)
                details_text = _strip_chokyo_jockey_lines("\n\n".join(horse_texts))

                if mode == "pace":
                    yield {"type": "status", "data": f"⏱️ {r_num}R 展開予想を生成中..."}
                    final_text = (
                        f"📅 {year}/{month}/{day} {place_name}{r_num}R\n\n"
                        f"{header1}\n\n"
                        f"{pace_text}\n\n"
                        f"{details_text}"
                    )
                    final_html = generate_html_output(
                        year, month, day, place_name, r_num, header1, pace_text,
                        "【総合評価】 (展開のみモードのため省略)",
                        "【対戦表】 (展開のみモードのため省略)",
                        "(展開のみモードのため省略)",
                        details_text,
                        "(相対評価未実行)"
                    )

                    try:
                        with open(cache_file, "w", encoding="utf-8") as f:
                            json.dump({
                                "data_text": final_text,
                                "data_html": final_html,
                                "score_policy": SCORE_POLICY_VERSION
                            }, f, ensure_ascii=False)
                    except Exception:
                        pass

                    yield {"type": "result", "race_num": r_num, "data_text": final_text, "data_html": final_html}
                    time.sleep(1)
                    continue

                course_match = re.search(r'(浦和|川崎|大井|船橋)', header1)
                current_course = course_match.group(1) if course_match else place_name
                
                _meta_dist = nk_data['meta'].get('dist', '')
                if _meta_dist and str(_meta_dist).isdigit():
                    current_dist = _meta_dist
                else:
                    _dist_m = re.search(r'(\d{1,2},\d{3}|\d{3,4})m', nk_data['meta'].get('course', '') + ' ' + header1)
                    current_dist = _dist_m.group(1).replace(',', '') if _dist_m else "1400"
                
                is_young = bool(re.search(r'(2歳|3歳)', header1))
                has_any_history = any(
                    any(isinstance(hs, dict) and hs.get("rank") is not None for hs in h.get("hist", []))
                    for h in nk_data["horses"].values()
                )
                is_newcomer = ("新馬" in header1) or (not has_any_history)
                
                m_class = re.search(r'([A-C][1-3]|A[12]|[23]歳)', header1)
                race_class_str = m_class.group(1).replace('Ａ','A').replace('Ｂ','B').replace('Ｃ','C') if m_class else "C2"

                scored_data, index_table_html, rank_locks, internal_verdicts = calculate_horse_scores(
                    nk_data["horses"], cyokyo, current_course, current_dist, is_young, race_class_str, r_num, pace_speed_list, is_newcomer=is_newcomer
                )

                # 展開加点(⑦): S/Mの前5頭+5と、競馬場×距離別の前30%+3〜7を合計へ加算。
                for _u, _b in (pace_bonus_map or {}).items():
                    if str(_u) in scored_data:
                        scored_data[str(_u)]["score_pace"] = _b
                    elif _u in scored_data:
                        scored_data[_u]["score_pace"] = _b

                # 対戦表(公式の出走馬同士の先着表)を取得し、先着状況で⑥(0〜10)を加点してから合計・バンド分け。
                match_txt = _fetch_matchup_table_selenium(driver, nk_id, {}, horses_data=nk_data["horses"], current_course=current_course, current_dist=current_dist)
                matchup_points = compute_matchup_points(match_txt, nk_data["horses"])
                for u_key, pts in matchup_points.items():
                    if u_key in scored_data:
                        scored_data[u_key]["score_6"] = pts
                    elif u_key.isdigit() and int(u_key) in scored_data:
                        scored_data[int(u_key)]["score_6"] = pts

                # 好走パターン(ユーザーメモ)：今回の想定位置・枠に一致する◯=+5/✕=-5を合計へ反映。
                apply_pattern_bonus(scored_data, nk_data["horses"], pace_speed_list, horse_patterns)

                yield {"type": "status", "data": f"📊 {r_num}R 採点中..."}
                ai_out_clean, grades = format_deterministic_evaluation(nk_data["horses"], cyokyo, scored_data, danwa)

                # 順序: 素点バンド → 内部ロック(同場の確定勝敗) → 騎手選択 → 点数下限 → S/A保証。
                # 対戦表は⑥加点で取り込んだので降格パス(adjust_grades_by_matchups)は廃止。
                grades = apply_internal_rank_locks(grades, rank_locks)
                grades = adjust_grades_by_jockey_choices(grades, nk_data["horses"])
                grades = apply_score_floor_to_grades(grades, nk_data["horses"], scored_data)
                grades = ensure_top_grade_presence(grades)
                grades = apply_display_grade_overrides(grades, nk_data["horses"], scored_data)

                ai_out_clean = _apply_final_grades_to_ai_detail(ai_out_clean, grades)
                ai_out_clean = ensure_race_content_in_ai_detail(ai_out_clean, nk_data["horses"])

                match_txt = _fetch_matchup_table_selenium(driver, nk_id, grades, horses_data=nk_data["horses"], current_course=current_course, current_dist=current_dist)
                eval_list_text = build_evaluation_list(grades, nk_data["horses"], scored_data)

                # 当日各コース横断の調教上位(赤字＋加点)の集計用に、今走調教の順位キーを控える。
                for _u, _sd in scored_data.items():
                    entries = _sd.get("chokyo_rank_entries") or []
                    if entries:
                        for _e in entries:
                            _item = {
                                "r_num": int(r_num), "umaban": str(_u),
                                "key": (_e.get("date"), _e.get("course_key"), _e.get("metric")),
                                "time": _e.get("time"),
                                "adj_time": _e.get("adj_time", _e.get("time")),
                                "disp": _e.get("disp"),
                                "line": _e.get("line"),
                                "metric": _e.get("metric"),
                                "label": _e.get("label", "今走調教"),
                            }
                            day_chokyo.append(_item)
                            day_chokyo_display.append(dict(_item))
                    elif _sd.get("chokyo_rank_key") is not None:
                        _item = {
                            "r_num": int(r_num), "umaban": str(_u),
                            "key": _sd.get("chokyo_rank_key"),
                            "time": _sd.get("chokyo_rank_time"),
                            "disp": _sd.get("chokyo_rank_disp"),
                            "label": "今走調教",
                        }
                        day_chokyo.append(_item)
                        day_chokyo_display.append(dict(_item))
                    for _e in (_sd.get("zen_chokyo_rank_entries") or []):
                        day_chokyo_display.append({
                            "r_num": int(r_num), "umaban": str(_u),
                            "key": (_e.get("date"), _e.get("course_key"), _e.get("metric")),
                            "time": _e.get("time"),
                            "adj_time": _e.get("adj_time", _e.get("time")),
                            "disp": _e.get("disp"),
                            "line": _e.get("line"),
                            "metric": _e.get("metric"),
                            "label": _e.get("label", "前走調教"),
                        })
                # 確定後にループ後段で再採点・再描画するための素材を控える（赤字加点の先読み用）。
                day_chokyo_pending[int(r_num)] = {
                    "race_key": f"{year}{month}{day}_{place_code}_{r_num}",
                    "cache_file": cache_file, "nk_id": nk_id,
                    "year": year, "month": month, "day": day,
                    "place_name": place_name, "place_code": place_code, "r_num": r_num,
                    "header1": header1, "pace_text": pace_text, "details_text": details_text,
                    "course_text": nk_data['meta'].get('course', ''), "horse_patterns": horse_patterns,
                    "index_table_html": index_table_html, "match_txt": match_txt,
                    "horses": nk_data["horses"], "cyokyo": cyokyo, "danwa": danwa,
                    "scored_data": scored_data, "rank_locks": rank_locks,
                    "current_course": current_course, "current_dist": current_dist,
                    "post_time": nk_data['meta'].get('post_time', ''),
                }

                final_text = (
                    f"📅 {year}/{month}/{day} {place_name}{r_num}R\n\n"
                    f"{header1}\n\n"
                    f"{pace_text}\n\n"
                    f"{eval_list_text}\n\n"
                    f"{match_txt}\n\n"
                    f"【相対評価】\n{index_table_html}\n\n"
                    f"【出走馬分析】\n{ai_out_clean}\n\n"
                    f"{details_text}"
                )

                final_html = generate_html_output(
                    year, month, day, place_name, r_num, header1, pace_text, eval_list_text,
                    match_txt, ai_out_clean, details_text, index_table_html,
                    build_speed_index_table(nk_data["horses"], scored_data),
                )

                # 馬名→uma_id（メモを別レースに引き継ぐためのキー）
                uma_ids_map = {
                    h.get("name"): h.get("uma_id", "")
                    for h in nk_data["horses"].values() if h.get("name")
                }

                try:
                    if not any(x in ai_out_clean for x in ["（回答生成エラー）", "⚠️", "パースエラー", "Error"]):
                        with open(cache_file, "w", encoding="utf-8") as f:
                            json.dump({
                                "data_text": final_text,
                                "data_html": final_html,
                                "ai_out_clean": ai_out_clean,
                                "grades": grades,
                                "eval_list_text": eval_list_text,
                                "score_policy": SCORE_POLICY_VERSION,
                                "race_id": nk_id,
                                "course": current_course,
                                "dist": str(current_dist),
                                "post_time": nk_data['meta'].get('post_time', ''),
                                "uma_ids": uma_ids_map,
                            }, f, ensure_ascii=False)
                        # DB台帳へ登録（結果取得・アーカイブ・メモ連携用）。失敗しても生成は止めない。
                        try:
                            import store as _db
                            _db.register_race(
                                race_key=f"{year}{month}{day}_{place_code}_{r_num}",
                                date=f"{year}{month}{day}", place_code=str(place_code),
                                place_name=place_name, race_num=int(r_num),
                                race_id=nk_id, course=current_course, dist=str(current_dist),
                                race_name=header1.strip(), grades=grades,
                                eval_list_text=eval_list_text, uma_ids=uma_ids_map,
                                post_time=nk_data['meta'].get('post_time', ''),
                                data_html=final_html, data_text=final_text,
                            )
                            # 調教データを馬(uma_id)単位で蓄積（メモと同じく永続）。失敗しても生成は止めない。
                            try:
                                _db.upsert_training(_build_training_rows(
                                    nk_data["horses"], cyokyo,
                                    f"{year}{month}{day}_{place_code}_{r_num}"))
                            except Exception as _et:
                                print(f"  [DB] upsert_training skipped: {_et}")
                        except Exception as _e:
                            print(f"  [DB] register_race skipped: {_e}")
                except Exception:
                    pass

                yield {"type": "result", "race_num": r_num, "data_text": final_text, "data_html": final_html}

            except Exception as e:
                yield {"type": "error", "data": f"{r_num}R Error: {e}"}

        # === 全体TOP3＋レース内TOP3を確定 → 該当レースを再採点・再描画して上書き ===
        # 失敗しても try で握りつぶし、ループ内で出した暫定結果(現行どおり)をそのまま残す。
        try:
            red_ranks = _compute_day_chokyo_red(day_chokyo)
            race_ranks = _compute_race_chokyo_bonus(day_chokyo)
            display_ranks = _compute_day_chokyo_ranks(day_chokyo_display)
            affected = sorted(
                {rn for (rn, _ub) in red_ranks}
                | {rn for (rn, _ub) in race_ranks}
                | {rn for (rn, _ub, _label, _line) in display_ranks}
            )
            if affected:
                import store as _db
                applied_races = 0
                # 1レースの失敗で他レースの反映が巻き添えにならないよう、各レースを独立処理する。
                for rn in affected:
                    try:
                        p = day_chokyo_pending.get(rn)
                        if not p:
                            continue
                        red_here = {ub: info for (r2, ub), info in red_ranks.items() if r2 == rn}
                        race_here = {ub: info for (r2, ub), info in race_ranks.items() if r2 == rn}
                        rank_here = _rank_infos_for_race(display_ranks, rn)
                        merged_rank_here = _merge_chokyo_rank_infos(red_here, rank_here)
                        for sd in p["scored_data"].values():
                            sd["score_chokyo_global"] = 0
                            sd["score_chokyo_race"] = 0
                            sd["score_chokyo_top"] = 0
                        for ub, infos in merged_rank_here.items():
                            sd = p["scored_data"].get(ub) or p["scored_data"].get(str(ub))
                            if sd is not None:
                                for info in infos:
                                    rk = info.get("rank")
                                    total = info.get("total")
                                    label = info.get("label") or "今走調教"
                                    if not rk:
                                        continue
                                    if label == "前走調教":
                                        prev = sd.get("zen_chokyo_display_rank")
                                        if not prev or int(rk) < int(prev):
                                            sd["zen_chokyo_display_rank"] = rk
                                            sd["zen_chokyo_display_total"] = total
                                            if info.get("disp"):
                                                sd["zen_chokyo_display_disp"] = info.get("disp")
                                        continue

                                    prev = sd.get("chokyo_display_rank")
                                    if not prev or int(rk) < int(prev):
                                        sd["chokyo_display_rank"] = rk
                                        sd["chokyo_display_total"] = total
                                        if info.get("disp"):
                                            sd["chokyo_display_disp"] = info.get("disp")
                                    pts = int(info.get("points", 0) or 0)
                                    if pts > 0:
                                        sd["score_chokyo_global"] = pts
                                        sd["chokyo_red_rank"] = rk
                                        sd["chokyo_red_total"] = total
                                        if info.get("disp"):
                                            sd["chokyo_rank_disp"] = info.get("disp")
                        for ub, info in race_here.items():
                            sd = p["scored_data"].get(ub) or p["scored_data"].get(str(ub))
                            if sd is None:
                                continue
                            sd["score_chokyo_race"] = int(info.get("points", 0) or 0)
                            sd["chokyo_race_rank"] = info.get("rank")
                            sd["chokyo_race_total"] = info.get("total")
                            sd["chokyo_race_disp"] = info.get("disp")
                        for sd in p["scored_data"].values():
                            global_pts = int(sd.get("score_chokyo_global", 0) or 0)
                            race_pts = int(sd.get("score_chokyo_race", 0) or 0)
                            sd["score_chokyo_top"] = _combine_chokyo_rank_points(global_pts, race_pts)
                        rb = _rebuild_race_with_chokyo(p, red_here, rank_here)
                        uma_ids_map = {h.get("name"): h.get("uma_id", "") for h in p["horses"].values() if h.get("name")}
                        try:
                            with open(p["cache_file"], "w", encoding="utf-8") as f:
                                json.dump({
                                    "data_text": rb["text"], "data_html": rb["html"],
                                    "ai_out_clean": rb["ai"], "grades": rb["grades"],
                                    "eval_list_text": rb["eval"], "score_policy": SCORE_POLICY_VERSION,
                                    "race_id": p["nk_id"], "course": p["current_course"],
                                    "dist": str(p["current_dist"]), "post_time": p["post_time"],
                                    "uma_ids": uma_ids_map,
                                }, f, ensure_ascii=False)
                        except Exception:
                            pass
                        try:
                            _db.register_race(
                                race_key=p["race_key"], date=f"{p['year']}{p['month']}{p['day']}",
                                place_code=str(p["place_code"]), place_name=p["place_name"], race_num=int(p["r_num"]),
                                race_id=p["nk_id"], course=p["current_course"], dist=str(p["current_dist"]),
                                race_name=p["header1"].strip(), grades=rb["grades"],
                                eval_list_text=rb["eval"], uma_ids=uma_ids_map, post_time=p["post_time"],
                                data_html=rb["html"], data_text=rb["text"],
                            )
                        except Exception as _e2:
                            print(f"  [調教上位] DB上書きスキップ {rn}R: {_e2}")
                        applied_races += 1
                        yield {"type": "result", "race_num": p["r_num"], "data_text": rb["text"], "data_html": rb["html"]}
                    except Exception as _ern:
                        import traceback
                        print(f"  [調教上位] {rn}R 再描画スキップ: {_ern}")
                        traceback.print_exc()
                if applied_races:
                    yield {"type": "status", "data": f"🏇 全体・レース内の調教順位を{applied_races}レースに反映しました"}
        except Exception as _e:
            print(f"  [調教上位] 反映スキップ: {_e}")

    except Exception as e:
        yield {"type": "error", "data": f"Fatal: {e}"}
    finally:
        try:
            driver.quit()
        except Exception:
            pass
